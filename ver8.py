from typing import List
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell import cell
from openpyxl.cell.cell import ERROR_CODES
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, alignment
from openpyxl.styles.fills import fills
from openpyxl.worksheet.dimensions import SheetDimension
from openpyxl.utils import get_column_letter
from tkinter import * 
from tkinter import ttk
from ttkthemes import ThemedTk
import time
import os.path
from os import error, path
from PIL import Image
from tkinter import messagebox
import traceback
import sys
import datetime
# import tab6
# from tab6 import addFullQual
# from gui import *
import gui



ft1 = Font(name='Arial',bold=True, size=14)
align = Alignment(horizontal='center')
thick_border = Border(left=Side(style='thick'), 
                     right=Side(style='thick'), 
                     top=Side(style='thick'), 
                     bottom=Side(style='thick'))

thick_border_blue = Border(left=Side(style='thick',color='0066CC'), 
                     right=Side(style='thick',color='0066CC')   
                  ) 

thick_border_blue_topBottom = Border(top=Side(style='thick',color='0066CC'), 
                     bottom=Side(style='thick',color='0066CC')   
                  )                    

thin_border_all = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

thin_border_all_grey = Border(left=Side(style='thin',color="DDDDDD"), 
                     right=Side(style='thin',color="DDDDDD"), 
                     top=Side(style='thin',color="DDDDDD"), 
                     bottom=Side(style='thin',color="DDDDDD"))

thin_border_sides = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style=None), 
                     bottom=Side(style=None))


thin_border_sides_Bottom = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style=None), 
                     bottom=Side(style='thin'))


#GLOBAL ERROR VARIABLES
fileerrorLabel = None
qualerrorLabel = None
blockerrorLabel = None
starterrorLabel = None
enderrorLabel = None
monthCheck = False

#GLOBAL MIRS
#QUAL1
global qual1mirs
qual1block1mirs=["3 MIRS / 1:6 / 3 HRS","0 MIRS","1 MIR / 1:12 / 3 HRS","0 MIRS"]
global qual2mirs
qual1block2mirs=["0 MIRS","0 MIRS","0 MIRS","1 MIR / 1:12 / 7 HRS","1 MIR / 1:12 / 4 HRS","1 MIR / 1:12 / 3 HRS","1 MIR / 1:12 / 2 HRS","1 MIR / 1:12 / 3.5 HRS","0 MIRS","1 MIR / 1:12 / 1.5HRS","1 MIR / 1:12 / 7.5HRS","1 MIR / 1:12 / 7.5HRS","1 MIR / 1:12 / 7.5HRS","0 MIRS","1 MIR / 1:12 / 7.5HRS","1 MIR / 1:12 / 7.5HRS","0 MIRS","0 MIRS"]
global qual1block3mirs
qual1block3mirs = [" 2 MIR / 1:8 / 4 HRS"," 2 MIR / 1:8 / 4.5 HRS"," 2 MIR / 1:8 / 5.5 HRS"," 2 MIR / 1:8 / 2.5 HRS"," 2 MIR / 1:8 / 4.5 HRS","0 MIRS"]
global qual1block4mirs
qual1block4mirs = ['2 MIRS / 1:8 / 6 HRS','2 MIRS/ 1:8 / 7 HRS','2 MIRS / 1:8 / 6 HRS','2 MIRS / 1:8 / 6.25 HRS','2 MIRS / 1:8 / 2 HRS','0 MIRS','2 MIRS / 1:8 / 5.5 HRS']
global qual1block5mirs
qual1block5mirs = ['0 MIRS','2 MIRS/ 1:8 / 3 HRS','2 MIRS/ 1:8 / 5.5 HRS','2 MIRS/ 1:8 / 5 HRS','2 MIRS/ 1:8 / 3.25 HRS','0 MIRS']
global qual1block6mirs
qual1block6mirs = ['1 MIRS / 1:12 / 3.5 HRS','1 MIRS / 1:12 / 5.75 HRS','1 MIRS / 1:12 / 4 HRS','1 MIRS / 1:12 / 6 HRS','1 MIRS / 1:12 / 5.25 HRS','1 MIRS / 1:12 / 4 HRS','1 MIRS / 1:12 / 6 HRS']
#QUAL2
global qual2block1mirs
qual2block1mirs =['1 MIR / 1:12 / 4 HRS','1 MIR / 1:12 / 5 HRS','3 MIRS / 1:6 / 8 HRS','3 MIRS / 1:6 / 8 HRS','3 MIRS / 1:6 / 8 HRS','3 MIRS / 1:6 / 8 HRS','3 MIRS / 1:6 / 8 HRS','2 MIRS / 1:8 / 8 HRS']
global qual2block2mirs
qual2block2mirs =['1 MIR / 1:12 / 3 HRS','1 MIR / 1:12 / 1 HR','3 MIRS / 1:6 / 3 HRS','3 MIRS / 1:6 / 8 HRS','3 MIRS / 1:6 / 8 HRS','3 MIR / 1:6 / 3 HRS','1 MIR / 1:12 / 2 HRS','2 MIRS / 1:8 / 6.5 HRS']
#QUAL3
global qual3block1mirs
qual3block1mirs = ['0 MIRS','1 MIRS / 1:12 / 4 HRS','1 MIRS / 1:12 / 8 HRS','1 MIRS / 1:12 / 3 HRS','1 MIRS / 1:12 / 5 HRS','2 MIRS / 1:8 / 8 HRS','2 MIRS / 1:8 / 8 HRS','1 MIRS / 1:12 / 8 HRS','1 MIRS / 1:12 / 8 HRS','1 MIRS / 1:12 / 6 HRS','1 MIRS / 1:12 / 6 HRS','1 MIRS / 1:12 / 1 HRS','0 MIRS']
global qual3block2mirs
qual3block2mirs =['1 MIRS / 1:12 / 3 HRS','1 MIRS / 1:12 / 8 HRS','1 MIRS / 1:12 / 8 HRS','1 MIRS / 1:12 / 8 HRS','1 MIRS / 1:12 / 8 HRS']
global qual3block3mirs
qual3block3mirs = ['1 MIRS / 1:12 / 1 HRS','1 MIRS / 1:12 / 8 HRS','1 MIRS / 1:12 / 5 HRS','1 MIRS / 1:12 / 8 HRS','1 MIRS / 1:12 / 4 HRS','2 MIRS / 1:8 / 6 HRS','2 MIRS / 1:8 / 8 HRS','2 MIRS / 1:8 / 4 HRS','0 MIRS']



#============================================================================================================================#
#============================================================================================================================#
#============================================================================================================================#
    
    
def Workbook_ReadIn():
    global workbook_Title
    workbook_Title = gui.tab3e2.get()
    print(workbook_Title)
    global workbook
    if path.exists(str(gui.tab3e2.get())+".xlsx") == True: 
        workbook = load_workbook(filename=workbook_Title+".xlsx")
        # workbook.add_named_style(date_style)
    else:
        fileerrorLabel = Label(gui.tab3,text="File Does Not Exist",font="Helvetica 10 bold", fg="red")
        fileerrorLabel.grid(row=11,column=0)
        return
    # print(path.exists(str(gui.e1.get())+".xlsx"))
    global sheet
    sheet = workbook.active
    print(workbook)
    global sheetIndex
    #Check Months
    startMonth = startdatecellref.value
    startMonth = startMonth[0:2]
    print(startMonth)
    endMonth = enddatecellref.value
    endMonth = endMonth[0:2]
    print(endMonth)
    global endDateReadIn
    endDateReadIn = enddatecellref.value
    endDateReadIn = endDateReadIn[3:5]
    print("FIRST END DATE READ IN: "+endDateReadIn)
    global monthCheck

    # month = str(gui.e4.get())
    # print("Month: "+str(month[0:2]))
    print("Start Month: "+str(startMonth))
    if str(startMonth[0:2]) =="01":
        if'JAN' in workbook.sheetnames:
            sheet = workbook["JAN"]
            sheetIndex = 1
        elif 'Jan' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("JAN")
            sheet = ws1
            sheetIndex = 1

    if str(startMonth[0:2]) =="02":
        if'FEB' in workbook.sheetnames:
            sheet = workbook["FEB"]
            sheetIndex =2
        elif 'Feb' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("FEB")
            sheet = ws1
            sheetIndex = 2
    if str(startMonth[0:2]) =="03":
        if'MAR' in workbook.sheetnames:
            sheet = workbook["MAR"]
            sheetIndex = 3
        elif 'MAR' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("MAR")
            sheet = ws1
            sheetIndex = 3
    if str(startMonth[0:2]) =="04":
        if'APR' in workbook.sheetnames:
            sheet = workbook["APR"]
            sheetIndex = 4
        elif 'APR' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("APR")
            sheet = ws1
            sheetIndex = 4
    if str(startMonth[0:2]) =="05":
        if'MAY' in workbook.sheetnames:
            sheet = workbook["MAY"]
            sheetIndex = 5
        elif 'MAY' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("MAY")
            sheet = ws1
            sheetIndex = 5

    if str(startMonth[0:2]) =="06":
        if'JUNE' in workbook.sheetnames:
            sheet = workbook["JUNE"]
            sheetIndex = 6
        elif 'JUNE' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("JUNE")
            sheet = ws1
            sheetIndex = 6

    if str(startMonth[0:2]) =="07":
        if'JULY' in workbook.sheetnames:
            sheet = workbook["JULY"]
            sheetIndex = 7
        elif 'JULY' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("JULY")
            sheet = ws1
            sheetIndex = 7

    if str(startMonth[0:2]) =="08":
        if'AUG' in workbook.sheetnames:
            sheet = workbook["AUG"]
            sheetIndex = 8
        elif 'AUG' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("AUG")
            sheet = ws1
            sheetIndex = 8

    if str(startMonth[0:2]) =="09":
        if'SEPT' in workbook.sheetnames:
            sheet = workbook["SEPT"]
            sheetIndex = 9
        elif 'SEPT' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("SEPT")
            sheet = ws1
            sheetIndex = 9
    if str(startMonth[0:2]) =="10":
        if'OCT' in workbook.sheetnames:
            sheet = workbook["OCT"]
            sheetIndex = 10
        elif 'OCT' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("OCT")
            sheet = ws1
            sheetIndex = 10
    if str(startMonth[0:2]) =="11":
        if'NOV' in workbook.sheetnames:
            sheet = workbook["NOV"]
            sheetIndex = 11
        elif 'NOV' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("NOV")
            sheet = ws1
            sheetIndex = 11

    if str(startMonth[0:1]) =="12":
        if'DEC' in workbook.sheetnames:
            sheet = workbook["DEC"]
            sheetIndex = 12
        elif 'DEC' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("DEC")
            sheet = ws1
            sheetIndex = 12

    if(startMonth == endMonth):
        None
        tab3add_qual()
    elif(startMonth != endMonth):
        endDateReadIn = "30"
        monthCheck =True
        tab3add_qual() 


  



def readFromExcel():
    try:
        tab3Label3 = Label(gui.gui.tab3frame3,text = "FINISHED",fg="grey26",bg="grey26",font="Helvetica 10 bold").grid(row=4,column=0)
        # global gui.my_progress
        gui.gui.my_progress.grid(row=0,column=0,pady=(50,10))

        #GETTING READ FROM FILE
        global readExcelFile
        readExcelFile = gui.tab3e1.get()
        print(readExcelFile)
        if path.exists(str(gui.tab3e1.get())+".xlsx") == True: 
            readFromBook = load_workbook(filename =readExcelFile+ '.xlsx')
        else:
            fileerrorLabel = Label(gui.tab3,text="File Does Not Exist",font="Helvetica 10 bold", fg="red")
            fileerrorLabel.grid(row=5,column=0)
            return

        
        sheet = readFromBook.active
        print(sheet.title)
        startingRow = 2
        startingCol = 1
        global qualcellref
        global blockcellref
        global startdatecellref
        global enddatecellref
        global classcellref
        global instructorcellref
        qualcellref = sheet.cell(row=startingRow,column=startingCol)
        print(qualcellref.value)
        blockcellref = sheet.cell(row=startingRow,column=startingCol+1)
        print(blockcellref.value)
        startdatecellref = sheet.cell(row=startingRow,column=startingCol+2)
        print(startdatecellref.value)
        enddatecellref = sheet.cell(row=startingRow,column=startingCol+3)
        classcellref = sheet.cell(row=startingRow,column=startingCol+4)
        instructorcellref = sheet.cell(row=startingRow,column=startingCol+5)
        
        print(qualcellref.value)
        skittle = 1
        while qualcellref.value:
            
            gui.gui.my_progress['value']+=30
                
            gui.tab3.update_idletasks() 
            print("QUALCELLREF VALUE: " +str(qualcellref.value))
            Workbook_ReadIn()
            print(skittle)
            startingRow+=1
            qualcellref = sheet.cell(row=startingRow,column=startingCol)
            print(qualcellref.value)
            blockcellref = sheet.cell(row=startingRow,column=startingCol+1)
            print(blockcellref.value)
            startdatecellref = sheet.cell(row=startingRow,column=startingCol+2)
            print(startdatecellref.value)
            enddatecellref = sheet.cell(row=startingRow,column=startingCol+3)
            classcellref = sheet.cell(row=startingRow,column=startingCol+4)
            instructorcellref = sheet.cell(row=startingRow,column=startingCol+5)
            skittle+=1
        tab3Label3 = Label(gui.gui.tab3frame3,text = "FINISHED",fg="white",bg="grey26",font="Helvetica 10 bold").grid(row=4,column=0)
        
        gui.gui.my_progress.stop()
    except():
        print(traceback.format_exc())
        messagebox.showwarning(title="Error Occured", message="something went wrong in READ FROM EXCEL. Check your entries and try again")
        if path.exists("errors.txt") == TRUE:
            ct = datetime.datetime.now() 
            with open("errors.txt", "a") as file:
                file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))
        else:
            ct = datetime.datetime.now() 
            with open("errors.txt", "x") as file:
                file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))
 
        




def existing_WORKBOOK():
    try:
        finishedLabel = Label(gui.tab1frame3,text="FINISHED",font="Helvetica 10 bold", fg="grey26",bg="grey26").grid(row=4,column=0) 
        # global gui.my_progress2
        gui.my_progress2.grid(row=0,column=0)
        global workbook_Title
        workbook_Title = gui.e1.get()
        global workbook
        if path.exists(str(gui.e1.get())+".xlsx") == True: 
            workbook = load_workbook(filename=workbook_Title+".xlsx")
            # workbook.add_named_style(date_style)
        else:
            fileerrorLabel = Label(gui.tab1frame3,text="File Does Not Exist",font="Helvetica 10 bold", fg="red")
            fileerrorLabel.grid(row=2,column=0)
            return
        # print(path.exists(str(gui.e1.get())+".xlsx"))
        global sheet
        sheet = workbook.active
        print(workbook)
        global sheetIndex
        #Check Months
        global startMonth
        startMonth = gui.e4.get()
        startMonth = startMonth[0:2]
        endMonth = gui.e5.get()
        endMonth = endMonth[0:2]
        global endDate
        endDate = gui.e5.get()
        endDate = endDate[3:5]
        global monthCheck

        # month = str(gui.e4.get())
        # print("Month: "+str(month[0:2]))
        print("Start Month: "+str(startMonth))
        gui.tab1.update_idletasks() 
        if str(startMonth[0:2]) =="01":
            if'JAN' in workbook.sheetnames:
                sheet = workbook["JAN"]
                sheetIndex = 1
            elif 'Jan' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("JAN")
                sheet = ws1
                sheetIndex = 1

        if str(startMonth[0:2]) =="02":
            if'FEB' in workbook.sheetnames:
                sheet = workbook["FEB"]
                sheetIndex =2
            elif 'Feb' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("FEB")
                sheet = ws1
                sheetIndex = 2
        if str(startMonth[0:2]) =="03":
            if'MAR' in workbook.sheetnames:
                sheet = workbook["MAR"]
                sheetIndex = 3
            elif 'MAR' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("MAR")
                sheet = ws1
                sheetIndex = 3
        if str(startMonth[0:2]) =="04":
            if'APR' in workbook.sheetnames:
                sheet = workbook["APR"]
                sheetIndex = 4
            elif 'APR' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("APR")
                sheet = ws1
                sheetIndex = 4
        if str(startMonth[0:2]) =="05":
            if'MAY' in workbook.sheetnames:
                sheet = workbook["MAY"]
                sheetIndex = 5
            elif 'MAY' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("MAY")
                sheet = ws1
                sheetIndex = 5

        if str(startMonth[0:2]) =="06":
            if'JUNE' in workbook.sheetnames:
                sheet = workbook["JUNE"]
                sheetIndex = 6
            elif 'JUNE' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("JUNE")
                sheet = ws1
                sheetIndex = 6

        if str(startMonth[0:2]) =="07":
            if'JULY' in workbook.sheetnames:
                sheet = workbook["JULY"]
                sheetIndex = 7
            elif 'JULY' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("JULY")
                sheet = ws1
                sheetIndex = 7

        if str(startMonth[0:2]) =="08":
            if'AUG' in workbook.sheetnames:
                sheet = workbook["AUG"]
                sheetIndex = 8
            elif 'AUG' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("AUG")
                sheet = ws1
                sheetIndex = 8

        if str(startMonth[0:2]) =="09":
            if'SEPT' in workbook.sheetnames:
                sheet = workbook["SEPT"]
                sheetIndex = 9
            elif 'SEPT' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("SEPT")
                sheet = ws1
                sheetIndex = 9
        if str(startMonth[0:2]) =="10":
            if'OCT' in workbook.sheetnames:
                sheet = workbook["OCT"]
                sheetIndex = 10
            elif 'OCT' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("OCT")
                sheet = ws1
                sheetIndex = 10
        if str(startMonth[0:2]) =="11":
            if'NOV' in workbook.sheetnames:
                sheet = workbook["NOV"]
                sheetIndex = 11
            elif 'NOV' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("NOV")
                sheet = ws1
                sheetIndex = 11

        if str(startMonth[0:1]) =="12":
            if'DEC' in workbook.sheetnames:
                sheet = workbook["DEC"]
                sheetIndex = 12
            elif 'DEC' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("DEC")
                sheet = ws1
                sheetIndex = 12

        if(startMonth == endMonth):
            None
            add_qual() 
        elif(startMonth != endMonth):
            endDate = 30
            monthCheck =True
            add_qual()
        finishedLabel = Label(gui.tab1frame3,text="FINISHED",font="Helvetica 10 bold", fg="white",bg="grey26").grid(row=4,column=0) 
        gui.my_progress.stop() 
    except():
        print(traceback.format_exc())
        messagebox.showwarning(title="Error Occured", message="something went wrong in EXISTING BOOK. Check your entries and try again")
        if path.exists("errors.txt") == TRUE:
            ct = datetime.datetime.now() 
            with open("errors.txt", "a") as file:
                file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))
        else:
            ct = datetime.datetime.now() 
            with open("errors.txt", "x") as file:
                file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))



def new_WORKBOOK():
    try:
        global workbook_Title
        workbook_Title = gui.newe1.get()
        if path.exists(str(gui.newe1.get())+".xlsx") == TRUE:
            fileerrorLabel = Label(gui.tab2mainframe,text="File Already Exists",font="Helvetica 10 bold", fg="red")
            fileerrorLabel.grid(row=5,column=0)
            return
        else:
            global workbook 
            workbook = Workbook()
            # sheet = workbook.active
            ws1 = workbook.create_sheet("JAN")
            ws2 = workbook.create_sheet("FEB")
            ws3 = workbook.create_sheet("MAR")
            ws4 = workbook.create_sheet("APR")
            ws5 = workbook.create_sheet("MAY")
            ws6 = workbook.create_sheet("JUNE")
            ws7 = workbook.create_sheet("JULY")
            ws8 = workbook.create_sheet("AUG")
            ws9 = workbook.create_sheet("SEPT")
            ws10 = workbook.create_sheet("OCT")
            ws11 = workbook.create_sheet("NOV")
            ws12 = workbook.create_sheet("DEC")

            #ADD DAYS TO THE SHEET
        
        
        def sheetDates():
            
            startmonth=""
            dayList = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
            global day_index
            day_index=4
            def addDaysAndDates():
                for col in range(1,32):
                    column_letter = get_column_letter(col)
                    # print(column_letter)
                    sheet.column_dimensions[column_letter].width = 42
                print(sheet.title)
                global day_index
                col_index=1
                print(endDay)
                print("Day INdex: "+str(day_index))
                while col_index<endDay:
                    if day_index ==6 and col_index<endDay:
                        daycellref=sheet.cell(row=2, column=col_index) 
                        daycellref.value = (dayList[day_index])
                        daycellref.border = thick_border_blue_topBottom
                        day_index = 0
                        col_index+=1
                        # workbook.save(filename=workbook_Title+".xlsx")
                    else:
                        daycellref=sheet.cell(row=2, column=col_index) 
                        daycellref.value = (dayList[day_index])
                        day_index+=1
                        col_index+=1
                        # workbook.save(filename=workbook_Title+".xlsx")
                    
                # workbook.save(filename=workbook_Title+".xlsx")
                
                for i in range(1,endDay):
                    print("Enday: "+str(endDay))
                    daycellref=sheet.cell(row=2, column=i)
                    daycellref.fill = PatternFill("solid", fgColor="DDDDDD")
                    if daycellref.value == "Saturday":
                        column_letter = get_column_letter(i)
                        sheet.column_dimensions[column_letter].width = 0.1
                        for a in range(5,100):
                            daycellref2 = sheet.cell(row=a,column=i)
                            daycellref2.font = Font(size=1)
                        # workbook.save(filename=workbook_Title+".xlsx")
                    elif daycellref.value == "Sunday":
                        column_letter = get_column_letter(i)
                        sheet.column_dimensions[column_letter].width = 0.1
                        for a in range(5,100):
                            daycellref2 = sheet.cell(row=a,column=i)
                            daycellref2.font = Font(size=1)
                        # workbook.save(filename=workbook_Title+".xlsx")

            for sheet in workbook.worksheets:
                endDay = 32
                print(sheet.title)
                if str(sheet.title)=="JAN" or str(sheet.title)=="MAR"or str(sheet.title)=="MAY"or str(sheet.title)=="JULY"or str(sheet.title)=="AUG"or str(sheet.title)=="OCT"or str(sheet.title)=="DEC":
                    endDay = 32
                    addDaysAndDates()
                elif str(sheet.title) == "APR"or str(sheet.title)=="JUNE"or str(sheet.title)=="SEPT"or str(sheet.title)=="NOV":
                    endDay = 31
                    addDaysAndDates()
                elif str(sheet.title) == "FEB":
                    endDay = 29
                    addDaysAndDates()
                else:
                    continue

                for i in range(1,endDay):
                    if str(sheet.title) == "JAN":
                        startmonth = "01"
                    elif str(sheet.title) == "FEB":
                        startmonth = "02"
                    elif str(sheet.title) == "MAR":
                        startmonth = "03"
                    elif str(sheet.title) == "APR":
                        startmonth = "04"
                    elif str(sheet.title) == "MAY":
                        startmonth = "05"
                    elif str(sheet.title) == "JUNE":
                        startmonth = "06"
                    elif str(sheet.title) == "JULY":
                        startmonth = "07"
                    elif str(sheet.title) == "AUG":
                        startmonth = "08"
                    elif str(sheet.title) == "SEPT":
                        startmonth = "09"
                    elif str(sheet.title) == "OCT":
                        startmonth = "10"
                    elif str(sheet.title) == "NOV":
                        startmonth = "11"
                    elif str(sheet.title) == "DEC":
                        startmonth = "12"
                    datecellref=sheet.cell(row=1, column=i)
                    datecellref.fill = PatternFill("solid", fgColor="DDDDDD")
                    if(i<10):
                        datecellref.value=str(startmonth)+"/0"+str(i)+"/21"
                    else:
                        datecellref.value=str(startmonth)+"/"+str(i)+"/21"

                #SETTING TITLE
                sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=32)
                sheet.row_dimensions[3].height = 60
                title_Cell = sheet['A3']
                title_Cell.border = thick_border
                monthTitle = str(sheet.title)
                # print(monthTitle)
                title_Cell.value = monthTitle+" ELECTRICAL SYSTEMS INSTRUCTOR SCHEDULE"
                title_Cell.font = Font(size=60,name="Times New Roman",bold=True)

                #SETTING DIVIDER
                for i in range(1,32):
                    datecellref2=sheet.cell(row=4, column=i)
                    datecellref2.fill = PatternFill("solid", fgColor="000000")
                    datecellref2.border = thick_border_blue_topBottom
                    datecellref2.value="blank"  
                    gui.tab1.update_idletasks() 


    #CEAT AND 7LVL
                navyUniqueRow = 20
                for i in range(1,32):
                                ceatRef = sheet.cell(row=navyUniqueRow,column=i) 
                                ceatRef.fill = PatternFill("solid", fgColor="B8CCE4")
                                ceatRef.value = "Navy Unique"

                ceatRow = navyUniqueRow+5
                for i in range(1,32):
                    ceatRef = sheet.cell(row=ceatRow,column=i) 
                    ceatRef.fill = PatternFill("solid", fgColor="B8CCE4")
                    ceatRef.value = "CEAT"
                ceatRow2 = ceatRow+5
                for i in range(1,32):
                    ceatRef2 = sheet.cell(row=ceatRow2,column=i) 
                    ceatRef2.fill = PatternFill("solid", fgColor="B8CCE4")
                    ceatRef2.value = "CEAT"
                distMxRow = ceatRow2+5
                for i in range(1,32):
                    ceatRef2 = sheet.cell(row=distMxRow,column=i) 
                    ceatRef2.fill = PatternFill("solid", fgColor="B8CCE4")
                    ceatRef2.value = "dISTR MX"
                sevenLRow = distMxRow+5
                for i in range(1,32):
                    ceatRef2 = sheet.cell(row=sevenLRow,column=i) 
                    ceatRef2.fill = PatternFill("solid", fgColor="B8CCE4")
                    ceatRef2.value = "7 LEVEL BLK 1/10"
                airfieldRow = sevenLRow+5
                for i in range(1,32):
                    ceatRef2 = sheet.cell(row=airfieldRow,column=i) 
                    ceatRef2.fill = PatternFill("solid", fgColor="B8CCE4")
                    ceatRef2.value = "AIRFIELD"
                #LEAVE
                tab5leaverow = airfieldRow+5
                for i in range(1,32):
                    leaveref = sheet.cell(row=tab5leaverow,column=i) 
                    leaveref.fill = PatternFill("solid", fgColor="E6B8B7")
                    leaveref.value = "LEAVE"
                #ADDITIONAL DUTIES
                additionaldutiesRow = tab5leaverow+7
                for i in range(1,32):
                    additionaldutiesref = sheet.cell(row=additionaldutiesRow,column=i) 
                    additionaldutiesref.fill = PatternFill("solid", fgColor="BFBFBF")
                    additionaldutiesref.value = "ADDITIONAL DUTIES"
                endRow = additionaldutiesRow+7
                for i in range(1,32):
                    additionaldutiesref = sheet.cell(row=endRow,column=i) 
                    additionaldutiesref.fill = PatternFill("solid", fgColor="000000")
                    additionaldutiesref.value = "END"


        sheetDates()
        myLabel0 = Label(gui.tab2mainframe,text = "Created",font='Helvetica 16 bold',bg="grey26").grid(row=5,column=0,ipadx=20)
        workbook.save(filename=workbook_Title+".xlsx")
    except:
            print(traceback.format_exc())
            messagebox.showwarning(title="Error Occured", message="something went wrong in NEW BOOK. Check your entries and try again")
            if path.exists("errors.txt") == TRUE:
                ct = datetime.datetime.now() 
                with open("errors.txt", "a") as file:
                    file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))
            else:
                ct = datetime.datetime.now() 
                with open("errors.txt", "x") as file:
                    file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))     



def tab3add_qual():
        try:

        
    #NONE DECLARED ERROR LABELS USED
            global fileerrorLabel
            global qualerrorLabel
            global blockerrorLabel
            global starterrorLabel
            global enderrorLabel
            global qual1block1mirs
            global qual1block2mirs      
            sheet
        #SET COLUMN WIDTH
            # for col in range(1,32):
            #     column_letter = get_column_letter(col)
            #     # print(column_letter)
            #     sheet.column_dimensions[column_letter].width = 42

        #SET TITLE
            sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=32)
            sheet.row_dimensions[3].height = 60
            title_Cell = sheet['A3']
            title_Cell.border = thick_border
            monthTitle = str(sheet.title)
            # print(monthTitle)
            title_Cell.value = monthTitle+" ELECTRICAL SYSTEMS INSTRUCTOR SCHEDULE"
            title_Cell.font = Font(size=60,name="Times New Roman",bold=True)

            
            # global startMonth
            #SET DATES
            # for i in range(1,32):
            #     datecellref=sheet.cell(row=1, column=i)
            #     datecellref.fill = PatternFill("solid", fgColor="DDDDDD")
                
            #     if(i<10):
            #         datecellref.value=str(startMonth)+"/0"+str(i)+"/2021"
            #     else:
            #         datecellref.value=str(startMonth)+"/"+str(i)+"/2021"


            #SET DIVIDER      

            for i in range(1,32):
                datecellref2=sheet.cell(row=4, column=i)
                datecellref2.fill = PatternFill("solid", fgColor="000000")
                datecellref2.border = thick_border_blue_topBottom
                datecellref2.value="blank"  
                gui.tab1.update_idletasks() 


            #GETTING INPUT VALUES FROM USER
            qualName = str(qualcellref.value)
            instructor = str(instructorcellref.value)
            if str(qualName) == "FAMILY DAY":
                None
            else:
                classNum = int(classcellref.value)
                if int(qualName) > 3 and str(qualName) !="FAMILY DAY":
                    qualerrorLabel = Label(gui.tab1frame2,text = "Qual Num Doesn't Exist",font="Helvetica 10 bold", fg="red")
                    qualerrorLabel.grid(row=11,column=0)
                    gui.e2.delete(0, END)
                    return 
                blockName = str(blockcellref.value)
                if int(blockName) > 6 and str(qualName) !="FAMILY DAY":
                    blockerrorLabel = Label(gui.tab1frame2,text = "Block Number Doesn't Exist",font="Helvetica 10 bold", fg="red")
                    blockerrorLabel.grid(row=11,column=0)
                    gui.e3.delete(0, END)
                    return
            startDate = str(startdatecellref.value)
            startDate = startDate[3:5]
            print(str(startDate))
            if int(startDate) > 32:
                starterrorLabel = Label(gui.tab1frame2,text = "Start Date Number too big",font="Helvetica 10 bold", fg="red")
                starterrorLabel.grid(row=11,column=0)
                gui.e4.delete(0, END)
                return
            if int(endDateReadIn) > 32:
                enderrorLabel = Label(gui.tab1frame2,text = "End Date Number too big",font="Helvetica 10 bold", fg="red")
                enderrorLabel.grid(row=11,column=0)
                gui.e5.delete(0, END)
                return


            

            #Start Date
            col_index = int(startDate)
            row_index=5
            # print("First row index: "+str(row_index))

            #THE DASHED CELLS ITERATORS
            # blank_col_index = col_index
            # blank_row_index=row_index+1
            
            #TITLE CELL
            cellref = sheet.cell(row=row_index,column=col_index)
            cellref.border = thick_border
            blockcheckcell = sheet.cell(row=row_index, column = 1)
        

            #DASHED CELLS
            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)


            #Setting Styles
            # cellref.font=ft1
            # cellref.alignment=align
            # blank_cellref.font=ft1
            # blank_cellref.alignment=align
            # blank_cellref2.font=ft1
            # blank_cellref2.alignment=align
            # blank_cellref3.font=ft1
            # blank_cellref3.alignment=align
            # blank_cellref4.font=ft1
            # blank_cellref4.alignment=align

            #End Date
            #check if months match


            #SETTING ACTIVE MIRS
            activeMirs=[]
            
            if qualName =="1" and blockName == "1":
                for items in qual1block1mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="2":
                for items in qual1block2mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="3":
                for items in qual1block3mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="4":
                for items in qual1block4mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="5":
                for items in qual1block5mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="6":
                for items in qual1block6mirs:
                    activeMirs.append(items)
            elif qualName =="2" and blockName =="1":
                for items in qual2block1mirs:
                    activeMirs.append(items)
            elif qualName =="2" and blockName =="2":
                for items in qual2block2mirs:
                    activeMirs.append(items)
            elif qualName =="3" and blockName =="1":
                for items in qual3block1mirs:
                    activeMirs.append(items)
            elif qualName =="3" and blockName =="2":
                for items in qual3block2mirs:
                    activeMirs.append(items)
            elif qualName =="3" and blockName =="3":
                for items in qual3block3mirs:
                    activeMirs.append(items)
        
                    



            des_col=int(endDateReadIn)+1

            
            
            total_index = 0

            mirsIndex = 0

            

            # print(tab5onleavelist[0])
            # print(tab5onleavelist[1])
            # print(tab5onleavelist[2])
            # for person in tab5onleavelist:
            #     print(person)
           
            
            
            
            
        
        
            q1check = False
            global colCheck
            colCheck = False


            #COLUMN ITERATOR LOOP
            while col_index < des_col:
                colCheck = False
                tab5leavecheck = False
                tab5onleavelist = []
                tab5leaverow = 5
                leavecellref = sheet.cell(row=tab5leaverow,column=col_index)
                while leavecellref.value !="LEAVE":
                    tab5leaverow+=1
                    leavecellref = sheet.cell(row=tab5leaverow,column=col_index)
                if leavecellref.value =="LEAVE":
                    tab5leaverow+=1
                    leavecellref = sheet.cell(row=tab5leaverow,column=col_index)
                while leavecellref.value:
                    person = str(leavecellref.value)
                    print(person)
                    tab5onleavelist.append(person)
                    tab5leaverow+=1
                    leavecellref = sheet.cell(row=tab5leaverow,column=col_index)

                for person in tab5onleavelist:
                    if instructor == person:
                        tab5leavecheck = True 
                
                #IF THERES A VALUE IN THE CELL
                while cellref.value:
                    colCheck = False
                    print(cellref.value)
                    if cellref.value == "Navy Unique":
                        sheet.insert_rows(row_index,5)
                        if col_index==1:
                            colCheck = True
                        else:
                            col_index-=1
                        # row_index-=5
                        cellref = sheet.cell(row=row_index,column=col_index)
                        # cellref.value="NEW"
                        # print(cellref.value)
                        workbook.save(filename=workbook_Title+".xlsx")
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        break
                    # print(str(cellref.value))
                    if str(cellref.value) == "FAMILY DAY":
                        col_index+=1
                        cellref = sheet.cell(row=row_index,column=col_index)
                        None
                    
                    elif str(cellref.value)!= "FAMILY DAY" and str(qualName)=="FAMILY DAY":
                        while row_index <=100:
                            cellref.value = "FAMILY DAY"
                            cellref.font = Font(color="FF0000")
                            cellref.fill = PatternFill("solid","FFFFFF")
                            cellref.border = thin_border_all_grey
                            row_index+=1
                            cellref = sheet.cell(row=row_index,column=col_index)

                    #First: IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK IS LESS THAN OR EQUAL AND THE CLASS NUM ENTERED IS GREATER
                    elif(int(qualName) <= int(cellref.value[1]) and int(blockName) <= int(cellref.value[10]) and int(qualName) == 1 and int(blockName)==1):
                        print("Start Date: "+str(startDate))
                        print("Start Date Type: "+str(type(startDate)))
                        if startDate!="01" and (int(classNum) >= int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999"):
                            sheet.insert_rows(5,5)
                            q1check = True
                            for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(startDate)-1):
                                for cell in rows:
                                    cell.value = "Q1  Block:1 Class:99999"
                                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid",fgColor="white")
                                    cell.border = thin_border_all_grey
                                    cell.font = Font(color="FFFFFF")
                            # row_index=4
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                        elif startDate!="01" and (int(classNum) < int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999"):
                            sheet.insert_rows(5,5)
                            q1check = True
                            for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(startDate)-1):
                                for cell in rows:
                                    cell.value = "Q1  Block:1 Class:99999"
                                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid",fgColor="white")
                                    cell.border = thin_border_all_grey
                                    cell.font = Font(color="FFFFFF")
                                    
                            row_index=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                        elif startDate=="1" and int(classNum) > int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999" :
                            sheet.insert_rows(5,5)
                            row_index=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        elif startDate=="1" and int(classNum) < int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999" :
                            row_index+=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                        #Second:IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK NUM IS GREATER OR EQUAL 
                    elif(int(qualName) <= int(cellref.value[1]) and int(blockName) > int(cellref.value[10])):
                        row_index+=5
                        print("Row index: " + str(row_index))
                        print(cellref.value)
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        

                        #Third:IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK IS LESS THAN OR EQUAL
                    elif(int(qualName) <= int(cellref.value[1]) and int(blockName) <= int(cellref.value[10])):
                        if int(classNum) > int(cellref.value[18:23]):
                            print("Cell Coordinate: "+str(cellref))
                            print("Last Blank Cell Coordinate: "+str(blank_cellref4))
                            sheet.insert_rows(row_index,5)
                            workbook.save(filename=workbook_Title+".xlsx")
                            newrowref = cellref.row
                            print("Newrowref: "+str(newrowref))
                            row_index=newrowref-5
                            print("row index: "+str(row_index))
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        else:

                            row_index+=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)


                        #FOURTH: IF THE ENTERED QUAL IS GREATER THAN THE CELL VALUE
                    elif(int(qualName) > int(cellref.value[1])):
                        row_index+=5
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)


                    elif(int(qualName)<=int(cellref.value[1])):
                        sheet.move_range(""+str(cellref.coordinate)+":"+str(blank_cellref4.coordinate)+"", rows=5)
                        newrowref = cellref.row
                        row_index=newrowref-5
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        
                else:
                    if str(qualName)=="FAMILY DAY":
                        while row_index <=100:
                            cellref.value = "FAMILY DAY"
                            cellref.font = Font(color="FF0000")
                            cellref.fill = PatternFill("solid","FFFFFF")
                            cellref.border = thin_border_all_grey
                            row_index+=1
                            cellref = sheet.cell(row=row_index,column=col_index)
                        
                    elif(int(qualName)==1 and int(blockName)==1 and startDate!="01" and q1check ==False):
                        sheet.insert_rows(5,5)
                        print("Row Index: "+str(row_index))
                        q1check =True
                        for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(startDate)-1):
                                for cell in rows:
                                    cell.value = "Q1  Block:1 Class:99999"
                                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid",fgColor="white")
                                    cell.border = thin_border_all_grey
                                    cell.font = Font(color="FFFFFF")
                                
                                    print("col index: "+str(col_index))
                    # row_index=4
                    print("Row Index: "+str(row_index))
                    cellref = sheet.cell(row=row_index,column=col_index)
                    blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                    blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                    blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                    blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                    if str(qualName)=="FAMILY DAY":
                        None    
                    else:
                        cellref.value="Q"+str(qualName)+" " +" Block:"+str(blockName)+" Class:"+str(classNum)
                        cellref.font = ft1
                        cellref.alignment=align
                    # print("col index: "+str(col_index))
                        #TITLE COLOR CODING
                        #QUAL 1
                        cellref.border = thick_border
                        if (qualName == "1" and blockName == "1"):
                            cellref.fill = PatternFill("solid", fgColor="00CCFF")
                        elif (qualName == "1"and blockName == "2"):
                            cellref.fill = PatternFill("solid", fgColor="33CCCC")
                        elif (qualName == "1" and blockName == "3"):
                            cellref.fill = PatternFill("solid", fgColor="92D050")
                        elif (qualName == "1"and blockName == "4"):
                            cellref.fill = PatternFill("solid", fgColor="FFFF00")
                        elif (qualName == "1" and blockName == "5"):
                            cellref.fill = PatternFill("solid", fgColor="FFC000")
                        elif (qualName == "1"and blockName == "6"):
                            cellref.fill = PatternFill("solid", fgColor="FF0000")

                        #QUAL2
                        elif (qualName == "2" and blockName == "1"):
                            cellref.fill = PatternFill("solid", fgColor="00B0F0")
                        elif (qualName == "2"and blockName == "2"):
                            cellref.fill = PatternFill("solid", fgColor="92D050")

                        #QUAL3
                        elif (qualName == "3" and blockName == "1"):
                            cellref.fill = PatternFill("solid", fgColor="FFFF00")
                        elif (qualName == "3"and blockName == "2"):
                            cellref.fill = PatternFill("solid", fgColor="9BBB59")
                        elif (qualName == "3"and blockName == "3"):
                            cellref.fill = PatternFill("solid", fgColor="FFC000")
                        


                        #SETTING BLANK VALUES BORDERS
                        blank_cellref.border = thin_border_sides
                        blank_cellref2.border = thin_border_sides
                        blank_cellref3.border = thin_border_sides
                        blank_cellref4.border = thin_border_sides_Bottom

                        #SETTING BLANK VALUES
                        if tab5leavecheck == True:
                            messagebox.showwarning(title="Instructor on Leave!", message="The entered instructor is on leave for one of the days")
                            blank_cellref.value="Instructor: Instructor is on leave"
                            tab5leavecheck ==False
                        else:
                            blank_cellref.value="Instructor: "+instructor
                        blank_cellref.font = ft1
                        blank_cellref.alignment=align
                        blank_cellref2.value="------"
                        blank_cellref2.font = ft1
                        blank_cellref2.alignment=align
                        blank_cellref3.value="------"
                        blank_cellref3.font = ft1
                        blank_cellref3.alignment=align
                        if mirsIndex > len(activeMirs)-1:
                            blank_cellref4.value = "NonPop"
                        else:
                            blank_cellref4.value=str(activeMirs[mirsIndex])
                        blank_cellref4.font = ft1
                        blank_cellref4.alignment=align
                    total_index+=1
                   
                    # print("Title = " + cellref.coordinate)
                    # print("blank = " + blank_cellref.coordinate)
                
                if colCheck ==True:
                    None
                else:
                    col_index+=1
                mirsIndex+=1
                # print("column index: "+str(col_index))   


                #This line controls if they shoot up the rows or nor
                # row_index=4
                # print("row index"+str(row_index))   
                cellref = sheet.cell(row=row_index,column=col_index)
                blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                print("Row Index: "+str(row_index))
                print("cellref: "+str(cellref))
                workbook.save(filename=workbook_Title+".xlsx")


                
                
                global submitTotal
                submittedLabel = Label(gui.tab3frame3,text = "Submitted ",font="Helvetica 10 bold", fg="black",bg="grey26")
                datesLabel = Label(gui.tab3frame3,text = "Total Number of Dates Affected: "+str(submitTotal),font="Helvetica 10 bold", fg="black",bg="grey26")
                submittedLabel.grid(row=1,column=0)
                datesLabel.grid(row=2,column=0)
                submitTotal+=1
                gui.e2.delete(0, END)
                gui.e3.delete(0, END)
                gui.e4.delete(0, END)
                gui.e5.delete(0, END)

            second_Month_Check_ReadIn() 
        except:
            print(traceback.format_exc()) 
            messagebox.showwarning(title="Error Occured", message="something went wrong in Add Qual ReadIN. Check your entries and try again")
            if path.exists("errors.txt") == TRUE:
                ct = datetime.datetime.now() 
                with open("errors.txt", "a") as file:
                    file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))
            else:
                ct = datetime.datetime.now() 
                with open("errors.txt", "x") as file:
                    file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))
            





def add_qual():
        try:

        
    #NONE DECLARED ERROR LABELS USED
            global fileerrorLabel
            global qualerrorLabel
            global blockerrorLabel
            global starterrorLabel
            global enderrorLabel
            global qual1block1mirs
            global qual1block2mirs      
            sheet
        #SET COLUMN WIDTH
            # for col in range(1,32):
            #     column_letter = get_column_letter(col)
            #     # print(column_letter)
            #     sheet.column_dimensions[column_letter].width = 42

        #SET TITLE
            sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=32)
            sheet.row_dimensions[3].height = 60
            title_Cell = sheet['A3']
            title_Cell.border = thick_border
            monthTitle = str(sheet.title)
            # print(monthTitle)
            title_Cell.value = monthTitle+" ELECTRICAL SYSTEMS INSTRUCTOR SCHEDULE"
            title_Cell.font = Font(size=60,name="Times New Roman",bold=True)

            
            # global startMonth
            #SET DATES
            # for i in range(1,32):
            #     datecellref=sheet.cell(row=1, column=i)
            #     datecellref.fill = PatternFill("solid", fgColor="DDDDDD")
                
            #     if(i<10):
            #         datecellref.value=str(startMonth)+"/0"+str(i)+"/2021"
            #     else:
            #         datecellref.value=str(startMonth)+"/"+str(i)+"/2021"


            #SET DIVIDER      

            for i in range(1,32):
                datecellref2=sheet.cell(row=4, column=i)
                datecellref2.fill = PatternFill("solid", fgColor="000000")
                datecellref2.border = thick_border_blue_topBottom
                datecellref2.value="blank"  
                gui.tab1.update_idletasks() 


            #GETTING INPUT VALUES FROM USER
            instructor = str(gui.e7.get())
            qualName = gui.e2.get()
            if str(qualName) == "FAMILY DAY":
                None
            else:
                qualName = gui.e2.get()
                if int(qualName) > 3 and str(qualName) !="FAMILY DAY":
                    qualerrorLabel = Label(gui.tab1frame2,text = "Qual Num Doesn't Exist",font="Helvetica 10 bold", fg="red")
                    qualerrorLabel.grid(row=11,column=0)
                    gui.e2.delete(0, END)
                    return 
                blockName = gui.e3.get()
                if int(blockName) > 6 and str(qualName) !="FAMILY DAY":
                    blockerrorLabel = Label(gui.tab1frame2,text = "Block Number Doesn't Exist",font="Helvetica 10 bold", fg="red")
                    blockerrorLabel.grid(row=11,column=0)
                    gui.e3.delete(0, END)
                    return
            startDate = gui.e4.get()
            startDate = startDate[3:5]
            print(str(startDate))
            if int(startDate) > 32:
                starterrorLabel = Label(gui.tab1frame2,text = "Start Date Number too big",font="Helvetica 10 bold", fg="red")
                starterrorLabel.grid(row=11,column=0)
                gui.e4.delete(0, END)
                return
            if int(endDate) > 32:
                enderrorLabel = Label(gui.tab1frame2,text = "End Date Number too big",font="Helvetica 10 bold", fg="red")
                enderrorLabel.grid(row=11,column=0)
                gui.e5.delete(0, END)
                return

            classNum = gui.e6.get()

            

            #Start Date
            col_index = int(startDate)
            row_index=5
            # print("First row index: "+str(row_index))

            #THE DASHED CELLS ITERATORS
            # blank_col_index = col_index
            # blank_row_index=row_index+1
            
            #TITLE CELL
            cellref = sheet.cell(row=row_index,column=col_index)
            cellref.border = thick_border
            blockcheckcell = sheet.cell(row=row_index, column = 1)
        

            #DASHED CELLS
            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)


            #Setting Styles
            # cellref.font=ft1
            # cellref.alignment=align
            # blank_cellref.font=ft1
            # blank_cellref.alignment=align
            # blank_cellref2.font=ft1
            # blank_cellref2.alignment=align
            # blank_cellref3.font=ft1
            # blank_cellref3.alignment=align
            # blank_cellref4.font=ft1
            # blank_cellref4.alignment=align

            #End Date
            #check if months match


            #SETTING ACTIVE MIRS
            activeMirs=[]
            
            if qualName =="1" and blockName == "1":
                for items in qual1block1mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="2":
                for items in qual1block2mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="3":
                for items in qual1block3mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="4":
                for items in qual1block4mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="5":
                for items in qual1block5mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="6":
                for items in qual1block6mirs:
                    activeMirs.append(items)
            elif qualName =="2" and blockName =="1":
                for items in qual2block1mirs:
                    activeMirs.append(items)
            elif qualName =="2" and blockName =="2":
                for items in qual2block2mirs:
                    activeMirs.append(items)
            elif qualName =="3" and blockName =="1":
                for items in qual3block1mirs:
                    activeMirs.append(items)
            elif qualName =="3" and blockName =="2":
                for items in qual3block2mirs:
                    activeMirs.append(items)
            elif qualName =="3" and blockName =="3":
                for items in qual3block3mirs:
                    activeMirs.append(items)
        
                    



            des_col=int(endDate)+1

            
            
            total_index = 0

            mirsIndex = 0

            

            # print(tab5onleavelist[0])
            # print(tab5onleavelist[1])
            # print(tab5onleavelist[2])
            # for person in tab5onleavelist:
            #     print(person)
           
            
            
            
            
        
        
            q1check = False
            global colCheck
            colCheck = False


            #COLUMN ITERATOR LOOP
            while col_index < des_col:
                # print(col_index)
                # column_letter = get_column_letter(col_index)
                # print(column_letter)
                # if sheet.column_dimensions[column_letter].width < 1:
                #     col_index+=1
                #     continue
                colCheck = False
                tab5leavecheck = False
                tab5onleavelist = []
                tab5leaverow = 5
                leavecellref = sheet.cell(row=tab5leaverow,column=col_index)
                while leavecellref.value !="LEAVE":
                    tab5leaverow+=1
                    leavecellref = sheet.cell(row=tab5leaverow,column=col_index)
                if leavecellref.value =="LEAVE":
                    tab5leaverow+=1
                    leavecellref = sheet.cell(row=tab5leaverow,column=col_index)
                while leavecellref.value:
                    person = str(leavecellref.value)
                    print(person)
                    tab5onleavelist.append(person)
                    tab5leaverow+=1
                    leavecellref = sheet.cell(row=tab5leaverow,column=col_index)

                for person in tab5onleavelist:
                    if instructor == person:
                        tab5leavecheck = True 
                
                #IF THERES A VALUE IN THE CELL
                while cellref.value:
                    colCheck = False
                    print(cellref.value)
                    if cellref.value == "Navy Unique":
                        sheet.insert_rows(row_index,5)
                        if col_index==1:
                            colCheck = True
                        else:
                            col_index-=1
                        # row_index-=5
                        cellref = sheet.cell(row=row_index,column=col_index)
                        # cellref.value="NEW"
                        # print(cellref.value)
                        workbook.save(filename=workbook_Title+".xlsx")
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        break
                    # print(str(cellref.value))
                    if str(cellref.value) == "FAMILY DAY":
                        col_index+=1
                        cellref = sheet.cell(row=row_index,column=col_index)
                        None
                    
                    elif str(cellref.value)!= "FAMILY DAY" and str(qualName)=="FAMILY DAY":
                        while row_index <=100:
                            cellref.value = "FAMILY DAY"
                            cellref.font = Font(color="FF0000")
                            cellref.fill = PatternFill("solid","FFFFFF")
                            cellref.border = thin_border_all_grey
                            row_index+=1
                            cellref = sheet.cell(row=row_index,column=col_index)

                    #First: IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK IS LESS THAN OR EQUAL AND THE CLASS NUM ENTERED IS GREATER
                    elif(int(qualName) <= int(cellref.value[1]) and int(blockName) <= int(cellref.value[10]) and int(qualName) == 1 and int(blockName)==1):
                        print("Start Date: "+str(startDate))
                        print("Start Date Type: "+str(type(startDate)))
                        if startDate!="01" and (int(classNum) >= int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999"):
                            sheet.insert_rows(5,5)
                            q1check = True
                            for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(startDate)-1):
                                for cell in rows:
                                    cell.value = "Q1  Block:1 Class:99999"
                                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid",fgColor="white")
                                    cell.border = thin_border_all_grey
                                    cell.font = Font(color="FFFFFF")
                            # row_index=4
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                        elif startDate!="01" and (int(classNum) < int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999"):
                            sheet.insert_rows(5,5)
                            q1check = True
                            for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(startDate)-1):
                                for cell in rows:
                                    cell.value = "Q1  Block:1 Class:99999"
                                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid",fgColor="white")
                                    cell.border = thin_border_all_grey
                                    cell.font = Font(color="FFFFFF")
                                    
                            row_index=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                        elif startDate=="1" and int(classNum) > int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999" :
                            sheet.insert_rows(5,5)
                            row_index=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        elif startDate=="1" and int(classNum) < int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999" :
                            row_index+=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                        #Second:IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK NUM IS GREATER OR EQUAL 
                    elif(int(qualName) <= int(cellref.value[1]) and int(blockName) > int(cellref.value[10])):
                        row_index+=5
                        print("Row index: " + str(row_index))
                        print(cellref.value)
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        

                        #Third:IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK IS LESS THAN OR EQUAL
                    elif(int(qualName) <= int(cellref.value[1]) and int(blockName) <= int(cellref.value[10])):
                        if int(classNum) > int(cellref.value[18:23]):
                            print("Cell Coordinate: "+str(cellref))
                            print("Last Blank Cell Coordinate: "+str(blank_cellref4))
                            sheet.insert_rows(row_index,5)
                            workbook.save(filename=workbook_Title+".xlsx")
                            newrowref = cellref.row
                            print("Newrowref: "+str(newrowref))
                            row_index=newrowref-5
                            print("row index: "+str(row_index))
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        else:

                            row_index+=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)


                        #FOURTH: IF THE ENTERED QUAL IS GREATER THAN THE CELL VALUE
                    elif(int(qualName) > int(cellref.value[1])):
                        row_index+=5
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)


                    elif(int(qualName)<=int(cellref.value[1])):
                        sheet.move_range(""+str(cellref.coordinate)+":"+str(blank_cellref4.coordinate)+"", rows=5)
                        newrowref = cellref.row
                        row_index=newrowref-5
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        
                else:
                    if str(qualName)=="FAMILY DAY":
                        while row_index <=100:
                            cellref.value = "FAMILY DAY"
                            cellref.font = Font(color="FF0000")
                            cellref.fill = PatternFill("solid","FFFFFF")
                            cellref.border = thin_border_all_grey
                            row_index+=1
                            cellref = sheet.cell(row=row_index,column=col_index)
                        
                    elif(int(qualName)==1 and int(blockName)==1 and startDate!="01" and q1check ==False):
                        sheet.insert_rows(5,5)
                        print("Row Index: "+str(row_index))
                        q1check =True
                        for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(startDate)-1):
                                for cell in rows:
                                    cell.value = "Q1  Block:1 Class:99999"
                                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid",fgColor="white")
                                    cell.border = thin_border_all_grey
                                    cell.font = Font(color="FFFFFF")
                                
                                    print("col index: "+str(col_index))
                    # row_index=4
                    print("Row Index: "+str(row_index))
                    cellref = sheet.cell(row=row_index,column=col_index)
                    blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                    blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                    blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                    blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                    if str(qualName)=="FAMILY DAY":
                        None    
                    else:
                        cellref.value="Q"+str(qualName)+" " +" Block:"+str(blockName)+" Class:"+str(classNum)
                        cellref.font = ft1
                        cellref.alignment=align
                    # print("col index: "+str(col_index))
                        #TITLE COLOR CODING
                        #QUAL 1
                        cellref.border = thick_border
                        if (qualName == "1" and blockName == "1"):
                            cellref.fill = PatternFill("solid", fgColor="00CCFF")
                        elif (qualName == "1"and blockName == "2"):
                            cellref.fill = PatternFill("solid", fgColor="33CCCC")
                        elif (qualName == "1" and blockName == "3"):
                            cellref.fill = PatternFill("solid", fgColor="92D050")
                        elif (qualName == "1"and blockName == "4"):
                            cellref.fill = PatternFill("solid", fgColor="FFFF00")
                        elif (qualName == "1" and blockName == "5"):
                            cellref.fill = PatternFill("solid", fgColor="FFC000")
                        elif (qualName == "1"and blockName == "6"):
                            cellref.fill = PatternFill("solid", fgColor="FF0000")

                        #QUAL2
                        elif (qualName == "2" and blockName == "1"):
                            cellref.fill = PatternFill("solid", fgColor="00B0F0")
                        elif (qualName == "2"and blockName == "2"):
                            cellref.fill = PatternFill("solid", fgColor="92D050")

                        #QUAL3
                        elif (qualName == "3" and blockName == "1"):
                            cellref.fill = PatternFill("solid", fgColor="FFFF00")
                        elif (qualName == "3"and blockName == "2"):
                            cellref.fill = PatternFill("solid", fgColor="9BBB59")
                        elif (qualName == "3"and blockName == "3"):
                            cellref.fill = PatternFill("solid", fgColor="FFC000")
                        


                        #SETTING BLANK VALUES BORDERS
                        blank_cellref.border = thin_border_sides
                        blank_cellref2.border = thin_border_sides
                        blank_cellref3.border = thin_border_sides
                        blank_cellref4.border = thin_border_sides_Bottom

                        #SETTING BLANK VALUES
                        if tab5leavecheck == True:
                            messagebox.showwarning(title="Instructor on Leave!", message="The entered instructor is on leave for one of the days")
                            blank_cellref.value="Instructor: Instructor is on leave"
                            tab5leavecheck ==False
                        else:
                            blank_cellref.value="Instructor: "+instructor
                        blank_cellref.font = ft1
                        blank_cellref.alignment=align
                        blank_cellref2.value="------"
                        blank_cellref2.font = ft1
                        blank_cellref2.alignment=align
                        blank_cellref3.value="------"
                        blank_cellref3.font = ft1
                        blank_cellref3.alignment=align
                        if mirsIndex > len(activeMirs)-1:
                            blank_cellref4.value = "NonPop"
                        else:
                            blank_cellref4.value=str(activeMirs[mirsIndex])
                        blank_cellref4.font = ft1
                        blank_cellref4.alignment=align
                    total_index+=1
                   
                    # print("Title = " + cellref.coordinate)
                    # print("blank = " + blank_cellref.coordinate)
                
                if colCheck ==True:
                    None
                else:
                    col_index+=1
                mirsIndex+=1
                # print("column index: "+str(col_index))   


                #This line controls if they shoot up the rows or nor
                # row_index=4
                # print("row index"+str(row_index))   
                cellref = sheet.cell(row=row_index,column=col_index)
                blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                print("Row Index: "+str(row_index))
                print("cellref: "+str(cellref))
                workbook.save(filename=workbook_Title+".xlsx")


                
                
                global submitTotal
                submittedLabel = Label(gui.tab1frame3,text = "Submitted ",font="Helvetica 10 bold", fg="black",bg="grey26")
                datesLabel = Label(gui.tab1frame3,text = "Total Number of Dates Affected: "+str(submitTotal),font="Helvetica 8 bold", fg="black",bg="grey26")
                submittedLabel.grid(row=1,column=0,ipadx=50)
                datesLabel.grid(row = 2, column=0)
                submitTotal+=1
                # gui.e2.delete(0, END)
                # gui.e3.delete(0, END)
                # gui.e4.delete(0, END)
                # gui.e5.delete(0, END)
                gui.my_progress2['value']+=30
                
                gui.tab1.update_idletasks()

            second_Month_Check()
            gui.my_progress2['value']=0
            
        except:
            print(traceback.format_exc())
            messagebox.showwarning(title="Error Occured", message="something went wrong in ADD QUAL.\nCheck your entries and try again.\nFor a more detailed explanation check errors txt file")
            if path.exists("errors.txt") == TRUE:
                ct = datetime.datetime.now() 
                with open("errors.txt", "a") as file:
                    file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))
            else:
                ct = datetime.datetime.now() 
                with open("errors.txt", "x") as file:
                    file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))




def second_Month_Check():
    if monthCheck == True:
        sheets = workbook.sheetnames
        # for i in sheets:
        #     print(i)
        global sheet
        print(str(sheet))
        sheet = workbook[sheets[sheetIndex+1]]
        print(str(sheet.title))
        # sheet = workbook.active
        global newstartDate
        newstartDate = "1"
        global newendDate
        newendDate = endDate = gui.e5.get()
        newendDate = endDate[3:5]
        second_add_qual()
    else:
        None 

              

def second_Month_Check_ReadIn():
    if monthCheck == True:
        sheets = workbook.sheetnames
        # for i in sheets:
        #     print(i)
        global sheet
        # print(str(sheet))
        sheet = workbook[sheets[sheetIndex+1]]
        # print(str(sheet.title))
        # sheet = workbook.active
        global newstartDateReadIn
        newstartDateReadIn = "1"
        global newendDateReadIn
        newendDateReadIn = endDateReadIn=enddatecellref.value
        print("NEW END DATE Read In: "+str(endDateReadIn))
        newendDateReadIn = endDateReadIn[3:5]
        print(newendDateReadIn)
        second_add_qualReadIn()
    else:
        None                 
        

def second_add_qualReadIn():
        try:

        
    #NONE DECLARED ERROR LABELS USED
            global fileerrorLabel
            global qualerrorLabel
            global blockerrorLabel
            global starterrorLabel
            global enderrorLabel
            global qual1block1mirs
            global qual1block2mirs      
            sheet
        #SET COLUMN WIDTH
            # for col in range(1,32):
            #     column_letter = get_column_letter(col)
            #     # print(column_letter)
            #     sheet.column_dimensions[column_letter].width = 42

        #SET TITLE
            sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=32)
            sheet.row_dimensions[3].height = 60
            title_Cell = sheet['A3']
            title_Cell.border = thick_border
            monthTitle = str(sheet.title)
            # print(monthTitle)
            title_Cell.value = monthTitle+" ELECTRICAL SYSTEMS INSTRUCTOR SCHEDULE"
            title_Cell.font = Font(size=60,name="Times New Roman",bold=True)

            
            # global startMonth
            #SET DATES
            # for i in range(1,32):
            #     datecellref=sheet.cell(row=1, column=i)
            #     datecellref.fill = PatternFill("solid", fgColor="DDDDDD")
                
            #     if(i<10):
            #         datecellref.value=str(startMonth)+"/0"+str(i)+"/2021"
            #     else:
            #         datecellref.value=str(startMonth)+"/"+str(i)+"/2021"


            #SET DIVIDER      

            for i in range(1,32):
                datecellref2=sheet.cell(row=4, column=i)
                datecellref2.fill = PatternFill("solid", fgColor="000000")
                datecellref2.border = thick_border_blue_topBottom
                datecellref2.value="blank"  
                gui.tab1.update_idletasks() 


            #GETTING INPUT VALUES FROM USER
            qualName = str(qualcellref.value)
            instructor = str(instructorcellref.value)
            if str(qualName) == "FAMILY DAY":
                None
            else:
                classNum = int(classcellref.value)
                if int(qualName) > 3 and str(qualName) !="FAMILY DAY":
                    qualerrorLabel = Label(gui.tab1frame2,text = "Qual Num Doesn't Exist",font="Helvetica 10 bold", fg="red")
                    qualerrorLabel.grid(row=11,column=0)
                    gui.e2.delete(0, END)
                    return 
                blockName = str(blockcellref.value)
                if int(blockName) > 6 and str(qualName) !="FAMILY DAY":
                    blockerrorLabel = Label(gui.tab1frame2,text = "Block Number Doesn't Exist",font="Helvetica 10 bold", fg="red")
                    blockerrorLabel.grid(row=11,column=0)
                    gui.e3.delete(0, END)
                    return
            # startDate = gui.e4.get()
            # startDate = startDate[3:5]
            # print(str(startDate))
            if int(newstartDateReadIn) > 32:
                starterrorLabel = Label(gui.tab1frame2,text = "Start Date Number too big",font="Helvetica 10 bold", fg="red")
                starterrorLabel.grid(row=11,column=0)
                gui.e4.delete(0, END)
                return
            if int(newendDateReadIn) > 32:
                enderrorLabel = Label(gui.tab1frame2,text = "End Date Number too big",font="Helvetica 10 bold", fg="red")
                enderrorLabel.grid(row=11,column=0)
                gui.e5.delete(0, END)
                return

            classNum = gui.e6.get()

            

            #Start Date
            col_index = int(newstartDateReadIn)
            row_index=5
            # print("First row index: "+str(row_index))

            #THE DASHED CELLS ITERATORS
            # blank_col_index = col_index
            # blank_row_index=row_index+1
            
            #TITLE CELL
            cellref = sheet.cell(row=row_index,column=col_index)
            cellref.border = thick_border
            blockcheckcell = sheet.cell(row=row_index, column = 1)
        

            #DASHED CELLS
            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)


            #Setting Styles
            # cellref.font=ft1
            # cellref.alignment=align
            # blank_cellref.font=ft1
            # blank_cellref.alignment=align
            # blank_cellref2.font=ft1
            # blank_cellref2.alignment=align
            # blank_cellref3.font=ft1
            # blank_cellref3.alignment=align
            # blank_cellref4.font=ft1
            # blank_cellref4.alignment=align

            #End Date
            #check if months match


            #SETTING ACTIVE MIRS
            activeMirs=[]
            
            if qualName =="1" and blockName == "1":
                for items in qual1block1mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="2":
                for items in qual1block2mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="3":
                for items in qual1block3mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="4":
                for items in qual1block4mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="5":
                for items in qual1block5mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="6":
                for items in qual1block6mirs:
                    activeMirs.append(items)
            elif qualName =="2" and blockName =="1":
                for items in qual2block1mirs:
                    activeMirs.append(items)
            elif qualName =="2" and blockName =="2":
                for items in qual2block2mirs:
                    activeMirs.append(items)
            elif qualName =="3" and blockName =="1":
                for items in qual3block1mirs:
                    activeMirs.append(items)
            elif qualName =="3" and blockName =="2":
                for items in qual3block2mirs:
                    activeMirs.append(items)
            elif qualName =="3" and blockName =="3":
                for items in qual3block3mirs:
                    activeMirs.append(items)
        
                    



            des_col=int(newendDateReadIn)+1

            
            
            total_index = 0

            mirsIndex = 0

            

            # print(tab5onleavelist[0])
            # print(tab5onleavelist[1])
            # print(tab5onleavelist[2])
            # for person in tab5onleavelist:
            #     print(person)
           
            
            
            
            
        
        
            q1check = False
            global colCheck
            colCheck = False


            #COLUMN ITERATOR LOOP
            while col_index < des_col:
                colCheck = False
                tab5leavecheck = False
                tab5onleavelist = []
                tab5leaverow = 5
                leavecellref = sheet.cell(row=tab5leaverow,column=col_index)
                while leavecellref.value !="LEAVE":
                    tab5leaverow+=1
                    leavecellref = sheet.cell(row=tab5leaverow,column=col_index)
                if leavecellref.value =="LEAVE":
                    tab5leaverow+=1
                    leavecellref = sheet.cell(row=tab5leaverow,column=col_index)
                while leavecellref.value:
                    person = str(leavecellref.value)
                    print(person)
                    tab5onleavelist.append(person)
                    tab5leaverow+=1
                    leavecellref = sheet.cell(row=tab5leaverow,column=col_index)

                for person in tab5onleavelist:
                    if instructor == person:
                        tab5leavecheck = True 
                
                #IF THERES A VALUE IN THE CELL
                while cellref.value:
                    colCheck = False
                    print(cellref.value)
                    if cellref.value == "Navy Unique":
                        sheet.insert_rows(row_index,5)
                        if col_index==1:
                            colCheck = True
                        else:
                            col_index-=1
                        # row_index-=5
                        cellref = sheet.cell(row=row_index,column=col_index)
                        # cellref.value="NEW"
                        # print(cellref.value)
                        workbook.save(filename=workbook_Title+".xlsx")
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        break
                    # print(str(cellref.value))
                    if str(cellref.value) == "FAMILY DAY":
                        col_index+=1
                        cellref = sheet.cell(row=row_index,column=col_index)
                        None
                    
                    elif str(cellref.value)!= "FAMILY DAY" and str(qualName)=="FAMILY DAY":
                        while row_index <=100:
                            cellref.value = "FAMILY DAY"
                            cellref.font = Font(color="FF0000")
                            cellref.fill = PatternFill("solid","FFFFFF")
                            cellref.border = thin_border_all_grey
                            row_index+=1
                            cellref = sheet.cell(row=row_index,column=col_index)

                    #First: IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK IS LESS THAN OR EQUAL AND THE CLASS NUM ENTERED IS GREATER
                    elif(int(qualName) <= int(cellref.value[1]) and int(blockName) <= int(cellref.value[10]) and int(qualName) == 1 and int(blockName)==1):
                        print("Start Date: "+str(newstartDate))
                        print("Start Date Type: "+str(type(newstartDate)))
                        if newstartDate!="01" and (int(classNum) >= int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999"):
                            sheet.insert_rows(5,5)
                            q1check = True
                            for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(newstartDate)-1):
                                for cell in rows:
                                    cell.value = "Q1  Block:1 Class:99999"
                                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid",fgColor="white")
                                    cell.border = thin_border_all_grey
                                    cell.font = Font(color="FFFFFF")
                            # row_index=4
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                        elif newstartDate!="01" and (int(classNum) < int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999"):
                            sheet.insert_rows(5,5)
                            q1check = True
                            for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(newstartDate)-1):
                                for cell in rows:
                                    cell.value = "Q1  Block:1 Class:99999"
                                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid",fgColor="white")
                                    cell.border = thin_border_all_grey
                                    cell.font = Font(color="FFFFFF")
                                    
                            row_index=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                        elif newstartDate=="1" and int(classNum) > int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999" :
                            sheet.insert_rows(5,5)
                            row_index=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        elif newstartDate=="1" and int(classNum) < int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999" :
                            row_index+=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                        #Second:IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK NUM IS GREATER OR EQUAL 
                    elif(int(qualName) <= int(cellref.value[1]) and int(blockName) > int(cellref.value[10])):
                        row_index+=5
                        print("Row index: " + str(row_index))
                        print(cellref.value)
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        

                        #Third:IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK IS LESS THAN OR EQUAL
                    elif(int(qualName) <= int(cellref.value[1]) and int(blockName) <= int(cellref.value[10])):
                        if int(classNum) > int(cellref.value[18:23]):
                            print("Cell Coordinate: "+str(cellref))
                            print("Last Blank Cell Coordinate: "+str(blank_cellref4))
                            sheet.insert_rows(row_index,5)
                            workbook.save(filename=workbook_Title+".xlsx")
                            newrowref = cellref.row
                            print("Newrowref: "+str(newrowref))
                            row_index=newrowref-5
                            print("row index: "+str(row_index))
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        else:

                            row_index+=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)


                        #FOURTH: IF THE ENTERED QUAL IS GREATER THAN THE CELL VALUE
                    elif(int(qualName) > int(cellref.value[1])):
                        row_index+=5
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)


                    elif(int(qualName)<=int(cellref.value[1])):
                        sheet.move_range(""+str(cellref.coordinate)+":"+str(blank_cellref4.coordinate)+"", rows=5)
                        newrowref = cellref.row
                        row_index=newrowref-5
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        
                else:
                    if str(qualName)=="FAMILY DAY":
                        while row_index <=100:
                            cellref.value = "FAMILY DAY"
                            cellref.font = Font(color="FF0000")
                            cellref.fill = PatternFill("solid","FFFFFF")
                            cellref.border = thin_border_all_grey
                            row_index+=1
                            cellref = sheet.cell(row=row_index,column=col_index)
                        
                    elif(int(qualName)==1 and int(blockName)==1 and newstartDate!="01" and q1check ==False):
                        sheet.insert_rows(5,5)
                        print("Row Index: "+str(row_index))
                        q1check =True
                        for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(newstartDate)-1):
                                for cell in rows:
                                    cell.value = "Q1  Block:1 Class:99999"
                                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid",fgColor="white")
                                    cell.border = thin_border_all_grey
                                    cell.font = Font(color="FFFFFF")
                                
                                    print("col index: "+str(col_index))
                    # row_index=4
                    print("Row Index: "+str(row_index))
                    cellref = sheet.cell(row=row_index,column=col_index)
                    blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                    blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                    blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                    blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                    if str(qualName)=="FAMILY DAY":
                        None    
                    else:
                        cellref.value="Q"+str(qualName)+" " +" Block:"+str(blockName)+" Class:"+str(classNum)
                        cellref.font = ft1
                        cellref.alignment=align
                    # print("col index: "+str(col_index))
                        #TITLE COLOR CODING
                        #QUAL 1
                        cellref.border = thick_border
                        if (qualName == "1" and blockName == "1"):
                            cellref.fill = PatternFill("solid", fgColor="00CCFF")
                        elif (qualName == "1"and blockName == "2"):
                            cellref.fill = PatternFill("solid", fgColor="33CCCC")
                        elif (qualName == "1" and blockName == "3"):
                            cellref.fill = PatternFill("solid", fgColor="92D050")
                        elif (qualName == "1"and blockName == "4"):
                            cellref.fill = PatternFill("solid", fgColor="FFFF00")
                        elif (qualName == "1" and blockName == "5"):
                            cellref.fill = PatternFill("solid", fgColor="FFC000")
                        elif (qualName == "1"and blockName == "6"):
                            cellref.fill = PatternFill("solid", fgColor="FF0000")

                        #QUAL2
                        elif (qualName == "2" and blockName == "1"):
                            cellref.fill = PatternFill("solid", fgColor="00B0F0")
                        elif (qualName == "2"and blockName == "2"):
                            cellref.fill = PatternFill("solid", fgColor="92D050")

                        #QUAL3
                        elif (qualName == "3" and blockName == "1"):
                            cellref.fill = PatternFill("solid", fgColor="FFFF00")
                        elif (qualName == "3"and blockName == "2"):
                            cellref.fill = PatternFill("solid", fgColor="9BBB59")
                        elif (qualName == "3"and blockName == "3"):
                            cellref.fill = PatternFill("solid", fgColor="FFC000")
                        


                        #SETTING BLANK VALUES BORDERS
                        blank_cellref.border = thin_border_sides
                        blank_cellref2.border = thin_border_sides
                        blank_cellref3.border = thin_border_sides
                        blank_cellref4.border = thin_border_sides_Bottom

                        #SETTING BLANK VALUES
                        if tab5leavecheck == True:
                            messagebox.showwarning(title="Instructor on Leave!", message="The entered instructor is on leave for one of the days")
                            blank_cellref.value="Instructor: Instructor is on leave"
                            tab5leavecheck ==False
                        else:
                            blank_cellref.value="Instructor: "+instructor
                        blank_cellref.font = ft1
                        blank_cellref.alignment=align
                        blank_cellref2.value="------"
                        blank_cellref2.font = ft1
                        blank_cellref2.alignment=align
                        blank_cellref3.value="------"
                        blank_cellref3.font = ft1
                        blank_cellref3.alignment=align
                        if mirsIndex > len(activeMirs)-1:
                            blank_cellref4.value = "NonPop"
                        else:
                            blank_cellref4.value=str(activeMirs[mirsIndex])
                        blank_cellref4.font = ft1
                        blank_cellref4.alignment=align
                    total_index+=1
                   
                    # print("Title = " + cellref.coordinate)
                    # print("blank = " + blank_cellref.coordinate)
                
                if colCheck ==True:
                    None
                else:
                    col_index+=1
                mirsIndex+=1
                # print("column index: "+str(col_index))   


                #This line controls if they shoot up the rows or nor
                # row_index=4
                # print("row index"+str(row_index))   
                cellref = sheet.cell(row=row_index,column=col_index)
                blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                print("Row Index: "+str(row_index))
                print("cellref: "+str(cellref))
                workbook.save(filename=workbook_Title+".xlsx")


                
                
                global submitTotal
                submittedLabel = Label(gui.tab3frame3,text = "Submitted ",font="Helvetica 10 bold", fg="black",bg="grey26")
                datesLabel = Label(gui.tab3frame3,text = "Total Number of Dates Affected: "+str(submitTotal),font="Helvetica 10 bold", fg="black",bg="grey26")
                submittedLabel.grid(row=1,column=0,ipadx=50)
                datesLabel.grid(row = 2, column=0)
                submitTotal+=1
                gui.e2.delete(0, END)
                gui.e3.delete(0, END)
                gui.e4.delete(0, END)
                gui.e5.delete(0, END)

            # if monthCheck == True:
            #     global sheet
            #     sheet = workbook.active[2]
            #     startDate = 1
            #     endDate = endDate = gui.e5.get()
            #     endDate = endDate[3:5]
            #     second_add_qual()
            # else:
            #     None
        except:
            print(traceback.format_exc())
            messagebox.showwarning(title="Error Occured", message="something went wrong in Second Qual Add ReadIN try again")
            if path.exists("errors.txt") == TRUE:
                ct = datetime.datetime.now() 
                with open("errors.txt", "a") as file:
                    file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))
            else:
                ct = datetime.datetime.now() 
                with open("errors.txt", "x") as file:
                    file.write("\n"+str(ct)+"\n"+str(traceback.format_exc())) 
            



def second_add_qual():
        try:

        
    #NONE DECLARED ERROR LABELS USED
            global fileerrorLabel
            global qualerrorLabel
            global blockerrorLabel
            global starterrorLabel
            global enderrorLabel
            global qual1block1mirs
            global qual1block2mirs      
            sheet
        #SET COLUMN WIDTH
            # for col in range(1,32):
            #     column_letter = get_column_letter(col)
            #     # print(column_letter)
            #     sheet.column_dimensions[column_letter].width = 42

        #SET TITLE
            sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=32)
            sheet.row_dimensions[3].height = 60
            title_Cell = sheet['A3']
            title_Cell.border = thick_border
            monthTitle = str(sheet.title)
            # print(monthTitle)
            title_Cell.value = monthTitle+" ELECTRICAL SYSTEMS INSTRUCTOR SCHEDULE"
            title_Cell.font = Font(size=60,name="Times New Roman",bold=True)

            
            # global startMonth
            #SET DATES
            # for i in range(1,32):
            #     datecellref=sheet.cell(row=1, column=i)
            #     datecellref.fill = PatternFill("solid", fgColor="DDDDDD")
                
            #     if(i<10):
            #         datecellref.value=str(startMonth)+"/0"+str(i)+"/2021"
            #     else:
            #         datecellref.value=str(startMonth)+"/"+str(i)+"/2021"


            #SET DIVIDER      

            for i in range(1,32):
                datecellref2=sheet.cell(row=4, column=i)
                datecellref2.fill = PatternFill("solid", fgColor="000000")
                datecellref2.border = thick_border_blue_topBottom
                datecellref2.value="blank"  
                gui.tab1.update_idletasks() 


            #GETTING INPUT VALUES FROM USER
            instructor = str(gui.e7.get())
            qualName = gui.e2.get()
            if str(qualName) == "FAMILY DAY":
                None
            else:
                qualName = gui.e2.get()
                if int(qualName) > 3 and str(qualName) !="FAMILY DAY":
                    qualerrorLabel = Label(gui.tab1frame2,text = "Qual Num Doesn't Exist",font="Helvetica 10 bold", fg="red")
                    qualerrorLabel.grid(row=11,column=0)
                    gui.e2.delete(0, END)
                    return 
                blockName = gui.e3.get()
                if int(blockName) > 6 and str(qualName) !="FAMILY DAY":
                    blockerrorLabel = Label(gui.tab1frame2,text = "Block Number Doesn't Exist",font="Helvetica 10 bold", fg="red")
                    blockerrorLabel.grid(row=11,column=0)
                    gui.e3.delete(0, END)
                    return
            # startDate = gui.e4.get()
            # startDate = startDate[3:5]
            # print(str(startDate))
            if int(newstartDate) > 32:
                starterrorLabel = Label(gui.tab1frame2,text = "Start Date Number too big",font="Helvetica 10 bold", fg="red")
                starterrorLabel.grid(row=11,column=0)
                gui.e4.delete(0, END)
                return
            if int(newendDate) > 32:
                enderrorLabel = Label(gui.tab1frame2,text = "End Date Number too big",font="Helvetica 10 bold", fg="red")
                enderrorLabel.grid(row=11,column=0)
                gui.e5.delete(0, END)
                return

            classNum = gui.e6.get()

            

            #Start Date
            col_index = int(newstartDate)
            row_index=5
            # print("First row index: "+str(row_index))

            #THE DASHED CELLS ITERATORS
            # blank_col_index = col_index
            # blank_row_index=row_index+1
            
            #TITLE CELL
            cellref = sheet.cell(row=row_index,column=col_index)
            cellref.border = thick_border
            blockcheckcell = sheet.cell(row=row_index, column = 1)
        

            #DASHED CELLS
            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)


            #Setting Styles
            # cellref.font=ft1
            # cellref.alignment=align
            # blank_cellref.font=ft1
            # blank_cellref.alignment=align
            # blank_cellref2.font=ft1
            # blank_cellref2.alignment=align
            # blank_cellref3.font=ft1
            # blank_cellref3.alignment=align
            # blank_cellref4.font=ft1
            # blank_cellref4.alignment=align

            #End Date
            #check if months match


            #SETTING ACTIVE MIRS
            activeMirs=[]
            
            if qualName =="1" and blockName == "1":
                for items in qual1block1mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="2":
                for items in qual1block2mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="3":
                for items in qual1block3mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="4":
                for items in qual1block4mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="5":
                for items in qual1block5mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="6":
                for items in qual1block6mirs:
                    activeMirs.append(items)
            elif qualName =="2" and blockName =="1":
                for items in qual2block1mirs:
                    activeMirs.append(items)
            elif qualName =="2" and blockName =="2":
                for items in qual2block2mirs:
                    activeMirs.append(items)
            elif qualName =="3" and blockName =="1":
                for items in qual3block1mirs:
                    activeMirs.append(items)
            elif qualName =="3" and blockName =="2":
                for items in qual3block2mirs:
                    activeMirs.append(items)
            elif qualName =="3" and blockName =="3":
                for items in qual3block3mirs:
                    activeMirs.append(items)
        
                    



            des_col=int(newendDate)+1

            
            
            total_index = 0

            mirsIndex = 0

            

            # print(tab5onleavelist[0])
            # print(tab5onleavelist[1])
            # print(tab5onleavelist[2])
            # for person in tab5onleavelist:
            #     print(person)
           
            
            
            
            
        
        
            q1check = False
            global colCheck
            colCheck = False


            #COLUMN ITERATOR LOOP
            while col_index < des_col:
                colCheck = False
                tab5leavecheck = False
                tab5onleavelist = []
                tab5leaverow = 5
                leavecellref = sheet.cell(row=tab5leaverow,column=col_index)
                while leavecellref.value !="LEAVE":
                    tab5leaverow+=1
                    leavecellref = sheet.cell(row=tab5leaverow,column=col_index)
                if leavecellref.value =="LEAVE":
                    tab5leaverow+=1
                    leavecellref = sheet.cell(row=tab5leaverow,column=col_index)
                while leavecellref.value:
                    person = str(leavecellref.value)
                    print(person)
                    tab5onleavelist.append(person)
                    tab5leaverow+=1
                    leavecellref = sheet.cell(row=tab5leaverow,column=col_index)

                for person in tab5onleavelist:
                    if instructor == person:
                        tab5leavecheck = True 
                
                #IF THERES A VALUE IN THE CELL
                while cellref.value:
                    colCheck = False
                    print(cellref.value)
                    if cellref.value == "Navy Unique":
                        sheet.insert_rows(row_index,5)
                        if col_index==1:
                            colCheck = True
                        else:
                            col_index-=1
                        # row_index-=5
                        cellref = sheet.cell(row=row_index,column=col_index)
                        # cellref.value="NEW"
                        # print(cellref.value)
                        workbook.save(filename=workbook_Title+".xlsx")
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        break
                    # print(str(cellref.value))
                    if str(cellref.value) == "FAMILY DAY":
                        col_index+=1
                        cellref = sheet.cell(row=row_index,column=col_index)
                        None
                    
                    elif str(cellref.value)!= "FAMILY DAY" and str(qualName)=="FAMILY DAY":
                        while row_index <=100:
                            cellref.value = "FAMILY DAY"
                            cellref.font = Font(color="FF0000")
                            cellref.fill = PatternFill("solid","FFFFFF")
                            cellref.border = thin_border_all_grey
                            row_index+=1
                            cellref = sheet.cell(row=row_index,column=col_index)

                    #First: IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK IS LESS THAN OR EQUAL AND THE CLASS NUM ENTERED IS GREATER
                    elif(int(qualName) <= int(cellref.value[1]) and int(blockName) <= int(cellref.value[10]) and int(qualName) == 1 and int(blockName)==1):
                        print("Start Date: "+str(newstartDate))
                        print("Start Date Type: "+str(type(newstartDate)))
                        if newstartDate!="01" and (int(classNum) >= int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999"):
                            sheet.insert_rows(5,5)
                            q1check = True
                            for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(newstartDate)-1):
                                for cell in rows:
                                    cell.value = "Q1  Block:1 Class:99999"
                                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid",fgColor="white")
                                    cell.border = thin_border_all_grey
                                    cell.font = Font(color="FFFFFF")
                            # row_index=4
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                        elif newstartDate!="01" and (int(classNum) < int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999"):
                            sheet.insert_rows(5,5)
                            q1check = True
                            for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(newstartDate)-1):
                                for cell in rows:
                                    cell.value = "Q1  Block:1 Class:99999"
                                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid",fgColor="white")
                                    cell.border = thin_border_all_grey
                                    cell.font = Font(color="FFFFFF")
                                    
                            row_index=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                        elif newstartDate=="1" and int(classNum) > int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999" :
                            sheet.insert_rows(5,5)
                            row_index=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        elif newstartDate=="1" and int(classNum) < int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999" :
                            row_index+=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                        #Second:IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK NUM IS GREATER OR EQUAL 
                    elif(int(qualName) <= int(cellref.value[1]) and int(blockName) > int(cellref.value[10])):
                        row_index+=5
                        print("Row index: " + str(row_index))
                        print(cellref.value)
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        

                        #Third:IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK IS LESS THAN OR EQUAL
                    elif(int(qualName) <= int(cellref.value[1]) and int(blockName) <= int(cellref.value[10])):
                        if int(classNum) > int(cellref.value[18:23]):
                            print("Cell Coordinate: "+str(cellref))
                            print("Last Blank Cell Coordinate: "+str(blank_cellref4))
                            sheet.insert_rows(row_index,5)
                            workbook.save(filename=workbook_Title+".xlsx")
                            newrowref = cellref.row
                            print("Newrowref: "+str(newrowref))
                            row_index=newrowref-5
                            print("row index: "+str(row_index))
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        else:

                            row_index+=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)


                        #FOURTH: IF THE ENTERED QUAL IS GREATER THAN THE CELL VALUE
                    elif(int(qualName) > int(cellref.value[1])):
                        row_index+=5
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)


                    elif(int(qualName)<=int(cellref.value[1])):
                        sheet.move_range(""+str(cellref.coordinate)+":"+str(blank_cellref4.coordinate)+"", rows=5)
                        newrowref = cellref.row
                        row_index=newrowref-5
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        
                else:
                    if str(qualName)=="FAMILY DAY":
                        while row_index <=100:
                            cellref.value = "FAMILY DAY"
                            cellref.font = Font(color="FF0000")
                            cellref.fill = PatternFill("solid","FFFFFF")
                            cellref.border = thin_border_all_grey
                            row_index+=1
                            cellref = sheet.cell(row=row_index,column=col_index)
                        
                    elif(int(qualName)==1 and int(blockName)==1 and newstartDate!="01" and q1check ==False):
                        sheet.insert_rows(5,5)
                        print("Row Index: "+str(row_index))
                        q1check =True
                        for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(newstartDate)-1):
                                for cell in rows:
                                    cell.value = "Q1  Block:1 Class:99999"
                                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid",fgColor="white")
                                    cell.border = thin_border_all_grey
                                    cell.font = Font(color="FFFFFF")
                                
                                    print("col index: "+str(col_index))
                    # row_index=4
                    print("Row Index: "+str(row_index))
                    cellref = sheet.cell(row=row_index,column=col_index)
                    blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                    blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                    blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                    blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                    if str(qualName)=="FAMILY DAY":
                        None    
                    else:
                        cellref.value="Q"+str(qualName)+" " +" Block:"+str(blockName)+" Class:"+str(classNum)
                        cellref.font = ft1
                        cellref.alignment=align
                    # print("col index: "+str(col_index))
                        #TITLE COLOR CODING
                        #QUAL 1
                        cellref.border = thick_border
                        if (qualName == "1" and blockName == "1"):
                            cellref.fill = PatternFill("solid", fgColor="00CCFF")
                        elif (qualName == "1"and blockName == "2"):
                            cellref.fill = PatternFill("solid", fgColor="33CCCC")
                        elif (qualName == "1" and blockName == "3"):
                            cellref.fill = PatternFill("solid", fgColor="92D050")
                        elif (qualName == "1"and blockName == "4"):
                            cellref.fill = PatternFill("solid", fgColor="FFFF00")
                        elif (qualName == "1" and blockName == "5"):
                            cellref.fill = PatternFill("solid", fgColor="FFC000")
                        elif (qualName == "1"and blockName == "6"):
                            cellref.fill = PatternFill("solid", fgColor="FF0000")

                        #QUAL2
                        elif (qualName == "2" and blockName == "1"):
                            cellref.fill = PatternFill("solid", fgColor="00B0F0")
                        elif (qualName == "2"and blockName == "2"):
                            cellref.fill = PatternFill("solid", fgColor="92D050")

                        #QUAL3
                        elif (qualName == "3" and blockName == "1"):
                            cellref.fill = PatternFill("solid", fgColor="FFFF00")
                        elif (qualName == "3"and blockName == "2"):
                            cellref.fill = PatternFill("solid", fgColor="9BBB59")
                        elif (qualName == "3"and blockName == "3"):
                            cellref.fill = PatternFill("solid", fgColor="FFC000")
                        


                        #SETTING BLANK VALUES BORDERS
                        blank_cellref.border = thin_border_sides
                        blank_cellref2.border = thin_border_sides
                        blank_cellref3.border = thin_border_sides
                        blank_cellref4.border = thin_border_sides_Bottom

                        #SETTING BLANK VALUES
                        if tab5leavecheck == True:
                            messagebox.showwarning(title="Instructor on Leave!", message="The entered instructor is on leave for one of the days")
                            blank_cellref.value="Instructor: Instructor is on leave"
                            tab5leavecheck ==False
                        else:
                            blank_cellref.value="Instructor: "+instructor
                        blank_cellref.font = ft1
                        blank_cellref.alignment=align
                        blank_cellref2.value="------"
                        blank_cellref2.font = ft1
                        blank_cellref2.alignment=align
                        blank_cellref3.value="------"
                        blank_cellref3.font = ft1
                        blank_cellref3.alignment=align
                        if mirsIndex > len(activeMirs)-1:
                            blank_cellref4.value = "NonPop"
                        else:
                            blank_cellref4.value=str(activeMirs[mirsIndex])
                        blank_cellref4.font = ft1
                        blank_cellref4.alignment=align
                    total_index+=1
                   
                    # print("Title = " + cellref.coordinate)
                    # print("blank = " + blank_cellref.coordinate)
                
                if colCheck ==True:
                    None
                else:
                    col_index+=1
                mirsIndex+=1
                # print("column index: "+str(col_index))   


                #This line controls if they shoot up the rows or nor
                # row_index=4
                # print("row index"+str(row_index))   
                cellref = sheet.cell(row=row_index,column=col_index)
                blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                print("Row Index: "+str(row_index))
                print("cellref: "+str(cellref))
                workbook.save(filename=workbook_Title+".xlsx")


                
                
                global submitTotal
                submittedLabel = Label(gui.tab1frame3,text = "Submitted ",font="Helvetica 10 bold", fg="black",bg="grey26")
                datesLabel = Label(gui.tab1frame3,text = "Total Number of Dates Affected: "+str(submitTotal),font="Helvetica 8 bold", fg="black",bg="grey26")
                submittedLabel.grid(row=1,column=0,ipadx=50)
                datesLabel.grid(row = 2, column=0)
                submitTotal+=1
                gui.e2.delete(0, END)
                gui.e3.delete(0, END)
                gui.e4.delete(0, END)
                gui.e5.delete(0, END)

            # if monthCheck == True:
            #     global sheet
            #     sheet = workbook.active[2]
            #     startDate = 1
            #     endDate = endDate = gui.e5.get()
            #     endDate = endDate[3:5]
            #     second_add_qual()
            # else:
            #     None
        except():
            print(traceback.format_exc())
            messagebox.showwarning(title="Error Occured", message="something went wrong in Second Add Month. Check your entries and try again")
            if path.exists("errors.txt") == TRUE:
                ct = datetime.datetime.now() 
                with open("errors.txt", "a") as file:
                    file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))
            else:
                ct = datetime.datetime.now() 
                with open("errors.txt", "x") as file:
                    file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))
            
             
            


#======================================================================================================#
#======================================================================================================#
#======================================================================================================#
#======================================================================================================#
#======================================================================================================#
#======================================================================================================#
#======================================================================================================#       







