from tab6 import Worker3
from newBook import Worker
from singleEntry import Worker2
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell import cell
from openpyxl.cell.cell import ERROR_CODES
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, alignment
from openpyxl.styles.fills import fills
from openpyxl.worksheet.dimensions import SheetDimension
from openpyxl.utils import get_column_letter
import sys
import os
from PyQt5 import QtCore, QtGui,QtWidgets,uic
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtCore import QObject, QThread, pyqtSignal
from PyQt5.QtCore import QPropertyAnimation, QPoint
from os import path
from PyQt5.QtGui import QMovie
import logging
import threading
import time
import globalVars
import pyautogui
import subprocess
from random import randint
import datetime
import traceback





class AppDemo(QMainWindow):
    def __init__(self):
        super().__init__()
        uic.loadUi('curtis_program - Testing Seperation/QTDESIGNER/ver1.ui',self)
        # global finishedLabel
        # global failedLabel
        global statusWidget
        statusWidget = self.stackedWidget_2
        global movie
        self.movie = QMovie("curtis_program - Testing Seperation/QTDESIGNER/images/spinning.gif")
        self.label_10.setMovie(self.movie)
        self.movie.start()
        global movie2
        self.movie2 = QMovie("curtis_program - Testing Seperation/QTDESIGNER/images/loading4.gif")
        self.label_6.setMovie(self.movie2)
        self.movie2.start()
        self.stackedWidget_2.setCurrentWidget(self.page_8)
        self.singleEntryButton.clicked.connect(self.showPage1)
        self.newWorkbookButton.clicked.connect(self.showPage2)
        # self.dateTimeEdit.setDateTime(QtCore.QDateTime.currentDateTime())
        self.readFromFileButton.clicked.connect(self.showPage3)
        self.addFullQualButton.clicked.connect(self.showPage4)
        self.addMirsButton.clicked.connect(self.daKinter)

        
        self.scrollArea.setWidgetResizable(False)
        self.scrollArea_2.widgetResizable()
        self.scrollArea_2.setWidgetResizable(True)
        
        # datetime = QDateTime.currentDateTime()
        # # datetime.setDisplayFormat("dd/MM/yy hh:mm")
        # text = datetime.toString()
        # self.timeLabel.setText(text)
        self.toolButton.clicked.connect(self.launchDialog)
        self.toolButton_2.clicked.connect(self.launchDialog2)
        # statusWidget.setHidden(True)
        self.page2Submit.clicked.connect(self.passNewWorkbook)
        self.page1Submit.clicked.connect(self.singleEntry)
        self.stackedWidget_2.setCurrentWidget(self.page_9)
        
        self.thumbListWidget.setSortingEnabled(True)
        # self.populateButton.clicked.connect(self.daKinter)
        # self.moveButton.clicked.connect(self.select)
        self.toggleFrame.setHidden(True)
        self.hideToggle.clicked.connect(self.showToggle)
        self.toggleButton.clicked.connect(self.showToggle)
        self.progressBar.setValue(globalVars.progressbarValue)
        self.progressBar.setHidden(True)
        self.filterWidget.setHidden(True)
        self.tabWidget.setHidden(True)
        self.settingsButton.clicked.connect(self.showSettings)
        # self.testButton.clicked.connect(self.newReportProgress)

        self.opacity_effect = QGraphicsOpacityEffect()
        self.opacity_effect.setOpacity(0.7)
        self.filterWidget.setGraphicsEffect(self.opacity_effect)

        self.opacity_effect2 = QGraphicsOpacityEffect()
        self.opacity_effect2.setOpacity(1)
        self.tabWidget.setGraphicsEffect(self.opacity_effect2)
        self.toolButton_3.clicked.connect(self.launchDialogPage4)
        self.toolButton_4.clicked.connect(self.launchDialogPage4_2)
        self.page4Submit.clicked.connect(self.FullQual)
        # self.clearButton.clicked.connect(self.clearProgressFrame)
        self.radioButton.toggled.connect(lambda:self.btnstate(self.radioButton))
        self.radioButton_2.toggled.connect(lambda:self.btnstate(self.radioButton_2))
        self.radioButton_3.toggled.connect(lambda:self.btnstate(self.radioButton_3))
        self.timer = QTimer()


        global mainemployeeslist
        with open('curtis_program - Testing Seperation/QTDESIGNER/textFiles/employees.txt') as f:
            mainemployeeslist = [line.rstrip() for line in f]
        for item in mainemployeeslist:
            self.thumbListWidget.insertItem(0,item)

        #DATE AND TIME STUFF
        timer = QTimer(self)
        timer.timeout.connect(self.showtime)
        timer.start()




    def clearFrame(self):
        self.progressFrame = self.layout
        while self.layout.count():
            child = self.layout.takeAt(0)
            if child.widget() is not None:
                child.widget().deleteLater()
            elif child.layout() is not None:
                self.clearLayout(child.layout())

    def btnstate(self,b):
        if self.radioButton.isChecked() == True:
            self.sideBar.setStyleSheet('background-color: rgb(0, 150, 225);')
            self.topbar.setStyleSheet('background-color: rgb(117, 117, 117);')
            self.bottombar.setStyleSheet('background-color: rgb(117, 117, 117);')
            self.page.setStyleSheet('background-color: rgb(232, 232, 232);')
            self.singleEntryButton.setStyleSheet('background-color: rgb(0, 150, 225);')
            self.singleEntryButton.setStyleSheet("QPushButton::hover""{""background-color :rgb(255, 62, 65);""}")
            self.newWorkbookButton.setStyleSheet('background-color: rgb(0, 150, 225);')
            self.newWorkbookButton.setStyleSheet("QPushButton::hover""{""background-color :rgb(255, 62, 65);""}")
            self.readFromFileButton.setStyleSheet('background-color: rgb(0, 150, 225);')
            self.readFromFileButton.setStyleSheet("QPushButton::hover""{""background-color :rgb(255, 62, 65);""}")
            self.addFullQualButton.setStyleSheet('background-color: rgb(0, 150, 225);')
            self.addFullQualButton.setStyleSheet("QPushButton::hover""{""background-color :rgb(255, 62, 65);""}")
            self.addMirsButton.setStyleSheet('background-color: rgb(0, 150, 225);')
            self.addMirsButton.setStyleSheet("QPushButton::hover""{""background-color :rgb(255, 62, 65);""}")
            self.page_2.setStyleSheet('background-color: rgb(232, 232, 232);')
            self.page_3.setStyleSheet('background-color: rgb(232, 232, 232);')
            self.page_4.setStyleSheet('background-color: rgb(232, 232, 232);')
        elif self.radioButton_2.isChecked() == True:
            self.sideBar.setStyleSheet('background-color: rgb(83, 0, 89);')
            self.topbar.setStyleSheet('background-color: rgb(34, 34, 34);')
            self.bottombar.setStyleSheet('background-color: rgb(34, 34, 34);')
            self.page.setStyleSheet('background-color: black;')
            self.singleEntryButton.setStyleSheet('background-color: rgb(83, 0, 89);')
            self.singleEntryButton.setStyleSheet("QPushButton::hover""{""background-color: rgb(184, 81, 174);""}")
            self.newWorkbookButton.setStyleSheet('background-color: rgb(83, 0, 89);')
            self.newWorkbookButton.setStyleSheet("QPushButton::hover""{""background-color: rgb(184, 81, 174);""}")
            self.readFromFileButton.setStyleSheet('background-color: rgb(83, 0, 89);')
            self.readFromFileButton.setStyleSheet("QPushButton::hover""{""background-color: rgb(184, 81, 174);""}")
            self.addFullQualButton.setStyleSheet('background-color: rgb(83, 0, 89);')
            self.addFullQualButton.setStyleSheet("QPushButton::hover""{""background-color: rgb(184, 81, 174);""}")
            self.addMirsButton.setStyleSheet('background-color: rgb(83, 0, 89);')
            self.addMirsButton.setStyleSheet("QPushButton::hover""{""background-color: rgb(184, 81, 174);""}")
            self.page_2.setStyleSheet('background-color: black;')
            self.page_3.setStyleSheet('background-color: black;')
            self.page_4.setStyleSheet('background-color: black;')
        elif self.radioButton_3.isChecked() == True:
            self.sideBar.setStyleSheet('background-color: rgb(111, 0, 0);')
            self.topbar.setStyleSheet('background-color: rgb(34, 34, 34);')
            self.bottombar.setStyleSheet('background-color: rgb(34, 34, 34);')
            self.page.setStyleSheet('background-color: rgb(0, 79, 91);')
            self.singleEntryButton.setStyleSheet('background-color: rgb(111, 0, 0);')
            self.singleEntryButton.setStyleSheet("QPushButton::hover""{""background-color: rgb(0, 255, 247);""}")
            self.newWorkbookButton.setStyleSheet('background-color: rgb(111, 0, 0);')
            self.newWorkbookButton.setStyleSheet("QPushButton::hover""{""background-color: rgb(0, 255, 247);""}")
            self.readFromFileButton.setStyleSheet('background-color: rgb(111, 0, 0);')
            self.readFromFileButton.setStyleSheet("QPushButton::hover""{""background-color: rgb(0, 255, 247);""}")
            self.addFullQualButton.setStyleSheet('background-color: rgb(111, 0, 0);')
            self.addFullQualButton.setStyleSheet("QPushButton::hover""{""background-color: rgb(0, 255, 247);""}")
            self.addMirsButton.setStyleSheet('background-color: rgb(111, 0, 0);')
            self.addMirsButton.setStyleSheet("QPushButton::hover""{""background-color: rgb(0, 255, 247);""}")
            self.page_2.setStyleSheet('background-color: rgb(0, 79, 91);')
            self.page_3.setStyleSheet('background-color: rgb(0, 79, 91);')
            self.page_4.setStyleSheet('background-color: rgb(0, 79, 91);')
            
        
        # self.sideBar.setStyleSheet('background-color: black;')
        




    def showtime(self):
        datetime = QDateTime.currentDateTime()
        text = datetime.toString()
        self.timeLabel.setText(text)

        
    def showToggle(self, checked):
        if self.toggleFrame.isHidden():
            self.toggleFrame.setHidden(False)
            self.filterWidget.setHidden(False)
            
        else:
            self.toggleFrame.setHidden(True)
            self.filterWidget.setHidden(True)
            self.tabWidget.setHidden(True)
            

    def showSettings(self):
        #skittle
        if self.tabWidget.isHidden():
            self.tabWidget.setHidden(False)
            self.anim = QPropertyAnimation(self.tabWidget, b"pos")
            effect = QGraphicsOpacityEffect(self.tabWidget)
            self.tabWidget.setGraphicsEffect(effect)
            self.anim.setStartValue(QPoint(240, 35))
            self.anim.setEndValue(QPoint(240,45 ))
            self.anim.setDuration(250)
            self.anim_2 = QPropertyAnimation(effect, b"opacity")
            self.anim_2.setStartValue(0)
            self.anim_2.setEndValue(1)
            self.anim_2.setDuration(500)
            self.anim_group = QParallelAnimationGroup()
            self.anim_group.addAnimation(self.anim)
            self.anim_group.addAnimation(self.anim_2)
            self.anim_group.start()
            
        else:
            self.tabWidget.setHidden(True)
            # self.filterWidget.setHidden(True)

      
        


  

       



#PAGE2====================================NEW BOOK======================================================####
    def passNewWorkbook(self):
        globalVars.progressbarValue =0
        self.progressBar.setValue(globalVars.progressbarValue)
        # statusWidget.setVisible(True)
        # self.label_6.setVisible(True)
        print(globalVars.Page2fileName)
        globalVars.Page2fileName = self.fileNamePage2.text()
        globalVars.Page2fileName = str(globalVars.Page2fileName)
        print(globalVars.Page2fileName)
        if globalVars.Page2fileName == '':
            print("your entry is blank")
            self.stackedWidget_2.setCurrentWidget(self.page_12)
            return
        if path.exists(str(globalVars.Page2fileName)+".xlsx") == True:
            self.stackedWidget_2.setCurrentWidget(self.page_7)
            # statusWidget.setHidden(False)
            return
        else:
            # self.label_6.setHidden(False)
            print("FIRST CHECK: ")
            print(statusWidget.isVisible())
            print("FIRST CHECK: ")
            print(self.label_6.isVisible())
            self.stackedWidget_2.setCurrentWidget(self.page_8)
            statusWidget.setHidden(False)
            # self.label_6.setVisible(True)
            global movie2
            self.movie2.start()
            self.thread = QThread()
            # Step 3: Create a worker object
            self.worker = Worker()
            # Step 4: Move worker to the thread
            self.worker.moveToThread(self.thread)
            # Step 5: Connect signals and slots
            self.thread.started.connect(self.worker.new_WORKBOOK)
            self.worker.finished.connect(self.thread.quit)
            self.worker.finished.connect(self.worker.deleteLater)
            self.thread.finished.connect(self.thread.deleteLater)
            self.worker.progress.connect(self.reportProgress)
            # Step 6: Start the thread
            self.thread.start()
            self.thread.finished.connect(self.runFinished)



#PAGE1==========================================SINGLE ENTRY=============================================
    global initY
    initY = 1
    def newProgressBar(self):
        globalVars.progressbarValue = 0
        global initY
        self.progress = QProgressBar(self.scrollAreaWidgetContents_2)
        self.progress.setGeometry(0, initY, 380,35 )
        # self.progress.setValue(globalVars.progressbarValue)
        self.progress.show()
        initY+=35.5

    
    # def clearProgressFrame(self):
    #         self.progress.deleteLater()
    #         self.progress.deleteLater()

    def page4ReportProgress(self):
        # self.progress.setHidden(False)
        self.progress.setValue(globalVars.progressbarValue)
        # animation = QPropertyAnimation(self.progressBar, "value")
        # animation.setDuration(2000)
        # animation.setStartValue(0)
        # animation.setEndValue(100)
        # animation.start()


    def reportProgress(self):
        self.progressBar.setHidden(False)
        self.progressBar.setValue(globalVars.progressbarValue)
        # animation = QPropertyAnimation(self.progressBar, "value")
        # animation.setDuration(2000)
        # animation.setStartValue(0)
        # animation.setEndValue(100)
        # animation.start()





#==============================================SINGLE ENTRY=====================================================#
    def singleEntry(self):
        # statusWidget.setVisible(True)
        # self.label_6.setVisible(True)

        #E1
        # print(globalVars.page1e1)
        globalVars.page1e1 = self.fileNameLine.text()
        globalVars.page1e1 = str(globalVars.page1e1)
        print(globalVars.page1e1)

        #e2
        # print(globalVars.page1e2)
        globalVars.page1e2 = self.instructorLine.text()
        globalVars.page1e2 = str(globalVars.page1e2)
        print(globalVars.page1e2)

        #e3
        # print(globalVars.page1e3)
        globalVars.page1e3 = self.classLine.text()
        globalVars.page1e3 = str(globalVars.page1e3)
        print(globalVars.page1e3)

        #qualNum
        # print(globalVars.page1Qual)
        if self.isHoliday.isChecked()==True:
            globalVars.page1Qual = "FAMILY DAY"
            
        else:
            globalVars.page1Qual = self.qualBox.text()
            globalVars.page1Qual = str(globalVars.page1Qual)
            print(globalVars.page1Qual)

        #blockNum
        # print(globalVars.page1Block)
        globalVars.page1Block = self.blockBox.text()
        globalVars.page1Block = str(globalVars.page1Block)
        print(globalVars.page1Block)

        #start Date
        # print(globalVars.page1StartDate)
        globalVars.page1StartDate = self.page1StartDate.text()
        globalVars.page1StartDate = str(globalVars.page1StartDate)
        print(globalVars.page1StartDate)

        #End Date
        # print(globalVars.page1EndDate)
        globalVars.page1EndDate = self.page1EndDate.text()
        globalVars.page1EndDate = str(globalVars.page1EndDate)
        print(globalVars.page1EndDate)

        if globalVars.page1e1 == '':
            self.stackedWidget_2.setCurrentWidget(self.page_12)
            return

        if globalVars.page1Qual == '0' or globalVars.blockName == '0' and self.isHoliday.isChecked() == False:
            self.stackedWidget_2.setCurrentWidget(self.page_12)
            return
        # if globalVars.blockName == 0 and self.isHoliday.isChecked() ==False:
        #     self.stackedWidget_2.setCurrentWidget(self.page_12)
        #     return 

        if path.exists(str(globalVars.page1e1)) != True:
            self.stackedWidget_2.setCurrentWidget(self.page_7)
            # statusWidget.setHidden(False)
            return
        else:
            self.isHoliday.setChecked(False)
            # self.label_6.setHidden(False)
            print("FIRST CHECK: ")
            print(statusWidget.isVisible())
            print("FIRST CHECK: ")
            print(self.label_6.isVisible())
            self.stackedWidget_2.setCurrentWidget(self.page_8)
            statusWidget.setHidden(False)
            # self.label_6.setVisible(True)
            global movie2
            self.movie2.start()
            self.thread = QThread()
            # Step 3: Create a worker object
            self.worker = Worker2()
            # Step 4: Move worker to the thread
            self.worker.moveToThread(self.thread)
            # Step 5: Connect signals and slots
            self.thread.started.connect(self.worker.existing_WORKBOOK)
            self.worker.finished.connect(self.thread.quit)
            self.worker.finished.connect(self.worker.deleteLater)
            self.thread.finished.connect(self.thread.deleteLater)
            # self.worker.progress.connect(self.reportProgress)
            # Step 6: Start the thread
            self.thread.start()
            self.thread.finished.connect(self.runFinished)

#PAGE4=================================================================================================


    def page4Error(self,error):
        print("error is running")
        self.msg = QMessageBox()
        self.msg.setIcon(QMessageBox.Critical)
        self.msg.setText("Error When Adding Qual")
        if error == 'special':
            self.msg.setInformativeText("It seems one of your entries has special characters in it. Please check your entries.")
        elif error == 'nostartdate':
            self.msg.setInformativeText("It seems one of the Start Date cells in your read From file is empty. Please check your entries.")
        elif error == 'general':
            self.msg.setInformativeText("An unknown general error occured when adding full Qual. Please check your entries.")
        elif error == 'readFrom':
            self.msg.setInformativeText("An unknown general error occured when reading your read From file. Please check your entries.")
        elif error == 'instructorOnLeave':
            self.msg.setIcon(QMessageBox.Information)
            self.msg.setInformativeText("And instructor is on leave for one of the dates you scheduled them for.")
            self.msg.setWindowTitle("INSTRUCTOR ON LEAVE")
        self.msg.setWindowTitle("FULL ADD QUAL ERROR")
        # self.msg.setDetailedText("The details are as follows:")
        self.msg.show()
        

        
        



    def isActive(self):
        self.thread.quit
        self.page4Error()

    def FullQual(self):
        print("=============================================START=======================================================")
        # checkBox = self.checkBox
        if self.checkBox.isChecked() ==True:
            globalVars.isChecked = True
            print("Checked is True")
            print(globalVars.isChecked)
        
        # statusWidget.setVisible(True)
        # self.label_6.setVisible(True)
        print(globalVars.page4ReadFromFile)
        globalVars.page4ReadFromFile = ("curtis_program - Testing Seperation/QTDESIGNER/inputExcel/"+self.page4ReadFrom.text())
        globalVars.page4ReadFromFile = str(globalVars.page4ReadFromFile)
        print(globalVars.page4ReadFromFile)

        print(globalVars.page4DesFile)
        globalVars.page4DesFile = ("curtis_program - Testing Seperation/QTDESIGNER/outputExcel/"+self.page4DesFile.text())
        globalVars.page4DesFile = str(globalVars.page4DesFile)
        print(globalVars.page4DesFile)


        
        
        if path.exists(str(globalVars.page4ReadFromFile)) != True:
            self.stackedWidget_2.setCurrentWidget(self.page_10)
            # statusWidget.setHidden(False)
            return
        if path.exists(str(globalVars.page4DesFile)) != True:
            self.stackedWidget_2.setCurrentWidget(self.page_11)
            # statusWidget.setHidden(False)
            return
        else:
            self.stackedWidget_2.setCurrentWidget(self.page_8)
            # Worker3.addFullQual(self)
            statusWidget.setHidden(False)
            # self.label_6.setVisible(True)
            global movie2
            self.movie2.start()
            self.thread = QThread()
            # Step 3: Create a worker object
            self.worker = Worker3()
            # Step 4: Move worker to the thread
            self.worker.moveToThread(self.thread)
            # self.timer2 = QTimer()
            # self.timer2.start(100)
            # self.timer2.timeout.connect(self.isActive)
            # self.timer2.moveToThread(self.thread)

            # self.timer2.moveToThread(self.thread)
            
            # Step 5: Connect signals and slots
            self.thread.started.connect(self.worker.addFullQual)
            self.worker.finished.connect(self.thread.quit)
            self.worker.finished.connect(self.worker.deleteLater)
            self.thread.finished.connect(self.thread.deleteLater)
            self.worker.newBar.connect(self.newProgressBar)
            self.worker.progress.connect(self.page4ReportProgress)
            self.worker.error.connect(self.page4Error)
            # Step 6: Start the thread
            self.thread.start()
            self.thread.finished.connect(self.runFinished)





#PAGE5======================================ADD MIRS======================================================
    def daKinter(self):
        subprocess.Popen(['python' , 'ver8.py'])

#EXTRAS===================================================================================================


            
    def runFinished(self):
        globalVars.progressbarValue=0
        self.progressBar.setHidden(True)
        print("Second CHECK: ")
        print(statusWidget.isVisible())
        print("Second CHECK: ")
        print(self.label_6.isVisible())
        self.stackedWidget_2.setCurrentWidget(self.page_6)
        # statusWidget.setHidden(False)
        # effect = QGraphicsOpacityEffect(self.label_19)
        # self.label_19.setGraphicsEffect(effect)
        # self.anim_2 = QPropertyAnimation(effect, b"opacity")
        # self.anim_2.setStartValue(1)
        # self.anim_2.setEndValue(0)
        # self.anim_2.setDuration(5000)
        # self.anim_2.start()
        self.fileNamePage2.setText("")
        self.movie2.stop()
        
        

    def launchDialog(self):
        os.chdir(r'curtis_program - Testing Seperation/QTDESIGNER/outputExcel')
        file_Filter = 'Data File (*.xlsx *.csv *.dat);; Excel File (*.xlsx *.xls)'
        filePath,_ = QFileDialog.getOpenFileName(
            parent=self,
            caption="Select File",
            directory=os.getcwd(),
            filter=file_Filter,
            initialFilter='Excel File (*.xlsx *.xls)'
        )
        # fileName = QFileInfo(filePath).fileName()
        url = QUrl.fromLocalFile(filePath)
        global fileName
        fileName = str(url.fileName())
        print(url.fileName())
        self.fileNameLine.setText(fileName)

    def launchDialog2(self):
        file_Filter = 'Data File (*.xlsx *.csv *.dat);; Excel File (*.xlsx *.xls)'
        filePath,_ = QFileDialog.getOpenFileName(
            parent=self,
            caption="Select File",
            directory=os.getcwd(),
            filter=file_Filter,
            initialFilter='Excel File (*.xlsx *.xls)'
        )
        # fileName = QFileInfo(filePath).fileName()
        url = QUrl.fromLocalFile(filePath)
        global fileName
        fileName = str(url.fileName())
        print(url.fileName())
        self.fileNameLine_2.setText(fileName)
        

    def launchDialogPage4(self):
        os.chdir(r'curtis_program - Testing Seperation/QTDESIGNER/inputExcel')
        file_Filter = 'Data File (*.xlsx *.csv *.dat);; Excel File (*.xlsx *.xls)'
        filePath,_ = QFileDialog.getOpenFileName(
            parent=self,
            caption="Select File",
            directory=os.getcwd(),
            filter=file_Filter,
            initialFilter='Excel File (*.xlsx *.xls)'
        )
        # fileName = QFileInfo(filePath).fileName()
        url = QUrl.fromLocalFile(filePath)
        global ReadFrom
        ReadFrom = str(url.fileName())
        print(url.fileName())
        self.page4ReadFrom.setText(ReadFrom)

    def launchDialogPage4_2(self):
        os.chdir(r'curtis_program - Testing Seperation/QTDESIGNER/outputExcel')
        file_Filter = 'Data File (*.xlsx *.csv *.dat);; Excel File (*.xlsx *.xls)'
        filePath,_ = QFileDialog.getOpenFileName(
            parent=self,
            caption="Select File",
            directory=os.getcwd(),
            filter=file_Filter,
            initialFilter='Excel File (*.xlsx *.xls)'
        )
        # fileName = QFileInfo(filePath).fileName()
        url = QUrl.fromLocalFile(filePath)
        global page4DesFile
        DesFile = str(url.fileName())
        print(url.fileName())
        self.page4DesFile.setText(DesFile)

        
       

    def showPage1(self):
        print("Page1")
        self.anim = QPropertyAnimation(self.stackedWidget, b"pos")
        effect = QGraphicsOpacityEffect(self.stackedWidget)
        self.stackedWidget.setGraphicsEffect(effect)
        self.anim.setStartValue(QPoint(200, 180))
        self.anim.setEndValue(QPoint(200, 150))
        self.anim.setDuration(500)
        self.anim_2 = QPropertyAnimation(effect, b"opacity")
        self.anim_2.setStartValue(0)
        self.anim_2.setEndValue(1)
        self.anim_2.setDuration(500)
        self.anim_group = QParallelAnimationGroup()
        self.anim_group.addAnimation(self.anim)
        self.anim_group.addAnimation(self.anim_2)
        self.anim_group.start()
        self.stackedWidget.setCurrentWidget(self.page)


    def showPage2(self):
        print("Page2")
        self.anim = QPropertyAnimation(self.stackedWidget, b"pos")
        effect = QGraphicsOpacityEffect(self.stackedWidget)
        self.stackedWidget.setGraphicsEffect(effect)
        self.anim.setStartValue(QPoint(200, 100))
        self.anim.setEndValue(QPoint(200, 150))
        self.anim.setEasingCurve(QEasingCurve.OutBounce)
        self.anim.setDuration(500)
        self.anim_2 = QPropertyAnimation(effect, b"opacity")
        self.anim_2.setStartValue(0)
        self.anim_2.setEndValue(1)
        self.anim_2.setDuration(500)
        self.anim_group = QParallelAnimationGroup()
        self.anim_group.addAnimation(self.anim)
        self.anim_group.addAnimation(self.anim_2)
        self.anim_group.start()
        self.stackedWidget.setCurrentWidget(self.page_2)

    def showPage3(self):
        print("Page3")
        self.anim = QPropertyAnimation(self.stackedWidget, b"pos")
        effect = QGraphicsOpacityEffect(self.stackedWidget)
        self.stackedWidget.setGraphicsEffect(effect)
        self.anim.setStartValue(QPoint(200, 100))
        self.anim.setEndValue(QPoint(200, 150))
        self.anim.setEasingCurve(QEasingCurve.OutBounce)
        self.anim.setDuration(500)
        self.anim_2 = QPropertyAnimation(effect, b"opacity")
        self.anim_2.setStartValue(0)
        self.anim_2.setEndValue(1)
        self.anim_2.setDuration(500)
        self.anim_group = QParallelAnimationGroup()
        self.anim_group.addAnimation(self.anim)
        self.anim_group.addAnimation(self.anim_2)
        self.anim_group.start()
        self.stackedWidget.setCurrentWidget(self.page_3)

    def showPage4(self):
        print("Page4")
        self.anim = QPropertyAnimation(self.stackedWidget, b"pos")
        effect = QGraphicsOpacityEffect(self.stackedWidget)
        self.stackedWidget.setGraphicsEffect(effect)
        self.anim.setStartValue(QPoint(200, 100))
        self.anim.setEndValue(QPoint(200, 150))
        self.anim.setEasingCurve(QEasingCurve.OutBounce)
        self.anim.setDuration(500)
        self.anim_2 = QPropertyAnimation(effect, b"opacity")
        self.anim_2.setStartValue(0)
        self.anim_2.setEndValue(1)
        self.anim_2.setDuration(500)
        self.anim_group = QParallelAnimationGroup()
        self.anim_group.addAnimation(self.anim)
        self.anim_group.addAnimation(self.anim_2)
        self.anim_group.start()
        self.stackedWidget.setCurrentWidget(self.page_4)

    def showPage5(self):
        print("Page5")
        self.anim = QPropertyAnimation(self.stackedWidget, b"pos")
        effect = QGraphicsOpacityEffect(self.stackedWidget)
        self.stackedWidget.setGraphicsEffect(effect)
        self.anim.setStartValue(QPoint(200, 100))
        self.anim.setEndValue(QPoint(200, 150))
        self.anim.setEasingCurve(QEasingCurve.OutBounce)
        self.anim.setDuration(500)
        self.anim_2 = QPropertyAnimation(effect, b"opacity")
        self.anim_2.setStartValue(0)
        self.anim_2.setEndValue(1)
        self.anim_2.setDuration(500)
        self.anim_group = QParallelAnimationGroup()
        self.anim_group.addAnimation(self.anim)
        self.anim_group.addAnimation(self.anim_2)
        self.anim_group.start()
        self.stackedWidget.setCurrentWidget(self.page_5)
        
        

       
        


if __name__ == "__main__":
    app = QApplication(sys.argv)
    demo = AppDemo()
    demo.show()
    

    try:
        sys.exit(app.exec_())
    except SystemExit:
        print("closing window...")




