from typing import List
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell import cell
from openpyxl.cell.cell import ERROR_CODES
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, alignment
from openpyxl.styles.fills import fills
from openpyxl.worksheet.dimensions import SheetDimension
from openpyxl.utils import get_column_letter
import time
import os.path
from os import error, path
# from tkinter import messagebox
import traceback
import sys
import datetime
from PyQt5 import QtCore, QtGui,QtWidgets,uic
from PyQt5.QtWidgets import QMainWindow,QApplication, QWidget
from PyQt5.QtCore import QObject, QThread, pyqtSignal,pyqtSlot
import logging
import threading
import time
import globalVars


thick_border_blue = Border(left=Side(style='thick',color='0066CC'), 
                    right=Side(style='thick',color='0066CC')   
                ) 

thick_border_blue_topBottom = Border(top=Side(style='thick',color='0066CC'), 
                    bottom=Side(style='thick',color='0066CC')   
                )                    

thin_border_all = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))

thin_border_all_grey = Border(left=Side(style='thin',color="DDDDDD"), 
                    right=Side(style='thin',color="DDDDDD"), 
                    top=Side(style='thin',color="DDDDDD"), 
                    bottom=Side(style='thin',color="DDDDDD"))

thin_border_sides = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style=None), 
                    bottom=Side(style=None))


thin_border_sides_Bottom = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style=None), 
                    bottom=Side(style='thin'))
ft1 = Font(name='Arial',bold=True, size=14)

align = Alignment(horizontal='center')
thick_border = Border(left=Side(style='thick'), 
                    right=Side(style='thick'), 
                    top=Side(style='thick'), 
                    bottom=Side(style='thick'))
# import main


    #GLOBAL ERROR VARIABLES
fileerrorLabel = None
qualerrorLabel = None
blockerrorLabel = None
starterrorLabel = None
enderrorLabel = None
monthCheck = False
class Worker(QObject):
    
    finished = pyqtSignal()
    progress = pyqtSignal(int)
        
    def new_WORKBOOK(self):
        try:
            Page2fileName = globalVars.Page2fileName
            print(type(Page2fileName))
            print("PAGE2FILE NAME: ",Page2fileName)
            saved_args = locals()
            print("Second SAVED ARGS:",saved_args)
            global workbook_Title
            workbook_Title = Page2fileName
            # print(fileName)

            global workbook 
            workbook = Workbook()
            # sheet = workbook.active
            ws1 = workbook.create_sheet("JAN")
            ws2 = workbook.create_sheet("FEB")
            ws3 = workbook.create_sheet("MAR")
            ws4 = workbook.create_sheet("APR")
            ws5 = workbook.create_sheet("MAY")
            ws6 = workbook.create_sheet("JUNE")
            ws7 = workbook.create_sheet("JULY")
            ws8 = workbook.create_sheet("AUG")
            ws9 = workbook.create_sheet("SEPT")
            ws10 = workbook.create_sheet("OCT")
            ws11 = workbook.create_sheet("NOV")
            ws12 = workbook.create_sheet("DEC")

            #ADD DAYS TO THE SHEET


            def sheetDates():
                # time.sleep(3)
                startmonth=""
                dayList = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
                global day_index
                day_index=4
                def addDaysAndDates():
                    for col in range(1,32):
                        column_letter = get_column_letter(col)
                        # print(column_letter)
                        sheet.column_dimensions[column_letter].width = 42
                    # print(sheet.title)
                    global day_index
                    col_index=1
                    # print(endDay)
                    # print("Day INdex: "+str(day_index))
                    while col_index<endDay:
                        if day_index ==6 and col_index<endDay:
                            daycellref=sheet.cell(row=2, column=col_index) 
                            daycellref.value = (dayList[day_index])
                            daycellref.border = thick_border_blue_topBottom
                            day_index = 0
                            col_index+=1
                            # workbook.save(filename=workbook_Title+".xlsx")
                        else:
                            daycellref=sheet.cell(row=2, column=col_index) 
                            daycellref.value = (dayList[day_index])
                            day_index+=1
                            col_index+=1
                            # workbook.save(filename=workbook_Title+".xlsx")
                        
                    # workbook.save(filename=workbook_Title+".xlsx")
                    
                    for i in range(1,endDay):
                        # print("Enday: "+str(endDay))
                        daycellref=sheet.cell(row=2, column=i)
                        daycellref.fill = PatternFill("solid", fgColor="DDDDDD")
                        if daycellref.value == "Saturday":
                            column_letter = get_column_letter(i)
                            sheet.column_dimensions[column_letter].width = 0.1
                            for a in range(5,100):
                                daycellref2 = sheet.cell(row=a,column=i)
                                daycellref2.font = Font(size=1)
                            # workbook.save(filename=workbook_Title+".xlsx")
                        elif daycellref.value == "Sunday":
                            column_letter = get_column_letter(i)
                            sheet.column_dimensions[column_letter].width = 0.1
                            for a in range(5,100):
                                daycellref2 = sheet.cell(row=a,column=i)
                                daycellref2.font = Font(size=1)
                            # workbook.save(filename=workbook_Title+".xlsx")

                for sheet in workbook.worksheets:
                    i=0
                    self.progress.emit(i + 12)
                    globalVars.progressbarValue+=12
                    time.sleep(0.1)
                    
                    sheetIndex = 0
                    endDay = 32
                    # print(sheet.title)
                    if str(sheet.title)=="JAN" or str(sheet.title)=="MAR"or str(sheet.title)=="MAY"or str(sheet.title)=="JULY"or str(sheet.title)=="AUG"or str(sheet.title)=="OCT"or str(sheet.title)=="DEC":
                        endDay = 32
                        addDaysAndDates()
                    elif str(sheet.title) == "APR"or str(sheet.title)=="JUNE"or str(sheet.title)=="SEPT"or str(sheet.title)=="NOV":
                        endDay = 31
                        addDaysAndDates()
                    elif str(sheet.title) == "FEB":
                        endDay = 29
                        addDaysAndDates()
                    else:
                        continue

                    for i in range(1,endDay):
                        if str(sheet.title) == "JAN":
                            startmonth = "01"
                        elif str(sheet.title) == "FEB":
                            startmonth = "02"
                        elif str(sheet.title) == "MAR":
                            startmonth = "03"
                        elif str(sheet.title) == "APR":
                            startmonth = "04"
                        elif str(sheet.title) == "MAY":
                            startmonth = "05"
                        elif str(sheet.title) == "JUNE":
                            startmonth = "06"
                        elif str(sheet.title) == "JULY":
                            startmonth = "07"
                        elif str(sheet.title) == "AUG":
                            startmonth = "08"
                        elif str(sheet.title) == "SEPT":
                            startmonth = "09"
                        elif str(sheet.title) == "OCT":
                            startmonth = "10"
                        elif str(sheet.title) == "NOV":
                            startmonth = "11"
                        elif str(sheet.title) == "DEC":
                            startmonth = "12"
                        datecellref=sheet.cell(row=1, column=i)
                        datecellref.fill = PatternFill("solid", fgColor="DDDDDD")
                        if(i<10):
                            datecellref.value=str(startmonth)+"/0"+str(i)+"/21"
                        else:
                            datecellref.value=str(startmonth)+"/"+str(i)+"/21"

                    #SETTING TITLE
                    sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=32)
                    sheet.row_dimensions[3].height = 60
                    title_Cell = sheet['A3']
                    title_Cell.border = thick_border
                    monthTitle = str(sheet.title)
                    # print(monthTitle)
                    title_Cell.value = monthTitle+" ELECTRICAL SYSTEMS INSTRUCTOR SCHEDULE"
                    title_Cell.font = Font(size=60,name="Times New Roman",bold=True)

                    #SETTING DIVIDER
                    for i in range(1,32):
                        datecellref2=sheet.cell(row=4, column=i)
                        datecellref2.fill = PatternFill("solid", fgColor="000000")
                        datecellref2.border = thick_border_blue_topBottom
                        datecellref2.value="blank"  
                        # gui.tab1.update_idletasks() 
                        # print("SHEET INDEX: ",sheetIndex)
                        sheetIndex+=1



        #CEAT AND 7LVL
                    navyUniqueRow = 20
                    for i in range(1,32):
                                    ceatRef = sheet.cell(row=navyUniqueRow,column=i) 
                                    ceatRef.fill = PatternFill("solid", fgColor="B8CCE4")
                                    ceatRef.value = "Navy Unique"

                    ceatRow = navyUniqueRow+5
                    for i in range(1,32):
                        ceatRef = sheet.cell(row=ceatRow,column=i) 
                        ceatRef.fill = PatternFill("solid", fgColor="B8CCE4")
                        ceatRef.value = "CEAT"
                    ceatRow2 = ceatRow+5
                    for i in range(1,32):
                        ceatRef2 = sheet.cell(row=ceatRow2,column=i) 
                        ceatRef2.fill = PatternFill("solid", fgColor="B8CCE4")
                        ceatRef2.value = "CEAT"
                    distMxRow = ceatRow2+5
                    for i in range(1,32):
                        ceatRef2 = sheet.cell(row=distMxRow,column=i) 
                        ceatRef2.fill = PatternFill("solid", fgColor="B8CCE4")
                        ceatRef2.value = "dISTR MX"
                    sevenLRow = distMxRow+5
                    for i in range(1,32):
                        ceatRef2 = sheet.cell(row=sevenLRow,column=i) 
                        ceatRef2.fill = PatternFill("solid", fgColor="B8CCE4")
                        ceatRef2.value = "7 LEVEL BLK 1/10"
                    airfieldRow = sevenLRow+5
                    for i in range(1,32):
                        ceatRef2 = sheet.cell(row=airfieldRow,column=i) 
                        ceatRef2.fill = PatternFill("solid", fgColor="B8CCE4")
                        ceatRef2.value = "AIRFIELD"
                    #LEAVE
                    tab5leaverow = airfieldRow+5
                    for i in range(1,32):
                        leaveref = sheet.cell(row=tab5leaverow,column=i) 
                        leaveref.fill = PatternFill("solid", fgColor="E6B8B7")
                        leaveref.value = "LEAVE"
                    #ADDITIONAL DUTIES
                    additionaldutiesRow = tab5leaverow+7
                    for i in range(1,32):
                        additionaldutiesref = sheet.cell(row=additionaldutiesRow,column=i) 
                        additionaldutiesref.fill = PatternFill("solid", fgColor="BFBFBF")
                        additionaldutiesref.value = "ADDITIONAL DUTIES"
                    endRow = additionaldutiesRow+7
                    for i in range(1,32):
                        additionaldutiesref = sheet.cell(row=endRow,column=i) 
                        additionaldutiesref.fill = PatternFill("solid", fgColor="000000")
                        additionaldutiesref.value = "END"
            sheetDates()
            workbook.save("curtis_program - Testing Seperation/QTDESIGNER/outputExcel/"+workbook_Title+".xlsx")
            self.finished.emit()

        # label_19.setHidden(False)
        # myLabel0 = Label(gui.tab2mainframe,text = "Created",font='Helvetica 16 bold',bg="grey26").grid(row=5,column=0,ipadx=20)
        # workbook.save(filename=workbook_Title)
        except:
                print(traceback.format_exc())
            # messagebox.showwarning(title="Error Occured", message="something went wrong in NEW BOOK. Check your entries and try again")
            # if path.exists("curtis_program - Testing Seperation/QTDESIGNER/textFiles/errors.txt") == True:
            #     ct = datetime.datetime.now() 
            #     with open("curtis_program - Testing Seperation/QTDESIGNER/textFiles/errors.txt", "a") as file:
            #         file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))
            # else:
            #     ct = datetime.datetime.now() 
            #     with open("curtis_program - Testing Seperation/QTDESIGNER/textFiles/errors.txt", "x") as file:
            #         file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))     