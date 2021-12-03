from typing import List, get_origin
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell import cell
from openpyxl.cell.cell import ERROR_CODES
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, alignment
from openpyxl.styles.fills import fills
from openpyxl.worksheet.dimensions import SheetDimension
from openpyxl.utils import get_column_letter
import time
import os.path
from os import error, path
# from tkinter import messagebox
import traceback
import sys
import datetime
from PyQt5 import QtCore, QtGui,QtWidgets,uic
from PyQt5.QtWidgets import QMainWindow,QApplication, QMessageBox, QWidget
from PyQt5.QtCore import QObject, QThread, pyqtSignal,pyqtSlot,QTimer
import logging
import threading
import time
import globalVars



ft1 = Font(name='Arial',bold=True, size=14)
align = Alignment(horizontal='center')
thick_border = Border(left=Side(style='thick'), 
                     right=Side(style='thick'), 
                     top=Side(style='thick'), 
                     bottom=Side(style='thick'))

thick_border_blue = Border(left=Side(style='thick',color='0066CC'), 
                     right=Side(style='thick',color='0066CC')   
                  ) 

thick_border_blue_topBottom = Border(top=Side(style='thick',color='0066CC'), 
                     bottom=Side(style='thick',color='0066CC')   
                  )                    

thin_border_all = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

thin_border_all_grey = Border(left=Side(style='thin',color="DDDDDD"), 
                     right=Side(style='thin',color="DDDDDD"), 
                     top=Side(style='thin',color="DDDDDD"), 
                     bottom=Side(style='thin',color="DDDDDD"))

thin_border_sides = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style=None), 
                     bottom=Side(style=None))


thin_border_sides_Bottom = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style=None), 
                     bottom=Side(style='thin'))


#GLOBAL ERROR VARIABLES
fileerrorLabel = None
qualerrorLabel = None
blockerrorLabel = None
starterrorLabel = None
enderrorLabel = None
monthCheck = False


class Worker3(QObject):

    
    # Worker3.addFullQual(self)
    error = pyqtSignal(str)
    finished = pyqtSignal()
    progress = pyqtSignal(int)
    newBar = pyqtSignal()
    
    
    def secondMonthCheck(self):
        print("NUMBER OF HOLIDAYS: "+str(globalVars.familyDays))
        endDate = globalVars.endDate
        print("END DATE IN MONTH CHECK:"+str(endDate))
        if endDate >= globalVars.monthEndDays[globalVars.monthEndDaysIndex]:
            newEndDate = globalVars.monthEndDays[globalVars.monthEndDaysIndex]
            newStartDate = 1
            print("New End Date: "+str(newEndDate))
            # endDate = endDate-(globalVars.monthEndDays[globalVars.monthEndDaysIndex])+globalVars.familyDays
            print("END DATE IN MONTH CHECK:"+str(endDate))
            print("New Current Block Total Day Count: "+str(endDate))
            Worker3.tab7add_qual(self,globalVars.qualName,globalVars.currentBlock,globalVars.startDate,newEndDate)
            globalVars.sheetIndex+=1
            global sheet
            sheets = workbook.sheetnames
            print(str(sheet))
            sheet = workbook[sheets[globalVars.sheetIndex]]
            endDate = endDate-(globalVars.monthEndDays[globalVars.monthEndDaysIndex])+globalVars.familyDays
            Worker3.tab7add_qual(self,globalVars.qualName,globalVars.currentBlock,newStartDate,endDate)
            globalVars.endDate = newEndDate
            globalVars.monthEndDaysIndex+=1
            globalVars.holidayCheck=False
            globalVars.familyDays = 0
        else:
            Worker3.tab7add_qual(self,globalVars.qualName,globalVars.currentBlock,globalVars.startDate,endDate)
        if globalVars.holidayCheck ==True:
            print("AMOUNT OF FAMILY DAYS: "+str(globalVars.familyDays))
            globalVars.startDate = endDate+1+globalVars.familyDays
        else:
            globalVars.startDate = endDate+1
        print("START DATE: "+str(globalVars.startDate))
        print("END DATE: "+str(globalVars.endDate))
        globalVars.blockTotalDayCountIndex+=1
        globalVars.qual1BlockIndex+=1
        globalVars.weekendCount=0
        globalVars.holidayCheck=False
        globalVars.familyDays = 0

        print("Next Start Date: "+str(globalVars.startDate))




    def checkHolidays(self):
        holidayColIndex = 1
        holidayRowIndex=49
        holidayCellRef = sheet.cell(row=holidayRowIndex,column=holidayColIndex)
        while holidayColIndex <= 29:
            if holidayCellRef.value == "FAMILY DAY":
                while holidayCellRef.value !='blank':
                    print(holidayCellRef.coordinate)
                    print("VALUE: "+str(holidayCellRef.value))
                    holidayCellRef.value = "FAMILY DAY"
                    holidayCellRef.font = Font(color="FF0000")
                    holidayCellRef.fill = PatternFill("solid","FFFFFF")
                    holidayCellRef.border = thin_border_all_grey
                    holidayRowIndex-=1
                    holidayCellRef = sheet.cell(row=holidayRowIndex,column=holidayColIndex)
                holidayRowIndex=49
            else:
                holidayColIndex+=1
                holidayCellRef = sheet.cell(row=holidayRowIndex,column=holidayColIndex)
        # workbook.save(workbook_Title)
        



        


    




    def readFullQualIn(self):
        print("readFullQual is First")
        global lastDay
        global lastDate
        global workbook_Title
        workbook_Title = globalVars.page4DesFile
        print(workbook_Title)
        global workbook
        if path.exists(str(globalVars.page4DesFile)) == True: 
            workbook = load_workbook(filename=workbook_Title)
            print("Second File Found")
            # workbook.add_named_style(date_style)
        else:
            print("Second File Not Found")
            # fileerrorLabel = Label(gui.tab7,text="File Does Not Exist",font="Helvetica 10 bold", fg="red")
            # fileerrorLabel.grid(row=11,column=0)
            return
        # print(path.exists(str(e1.get())+".xlsx"))
        global sheet
        sheet = workbook.active
        print(workbook)
        # globalVars.sheetIndex
        #Check Months
        startMonth = startdatecellref.value
        if startMonth is None:
            print("Start month is none")
        print(startMonth)
        startMonth = startMonth[0:2]
        print(startMonth)
        global monthCheck
        for ws in workbook.worksheets:
            ws.sheet_view.zoomScale = 60
        workbook.save(workbook_Title)
        # month = str(e4.get())
        # print("Month: "+str(month[0:2]))
        print("Start Month: "+str(startMonth))
        if str(startMonth[0:2]) =="01":
            if'JAN' in workbook.sheetnames:
                sheet = workbook["JAN"]
                globalVars.sheetIndex = 1
                lastDay =globalVars.monthEndDays[0]
                globalVars.monthEndDaysIndex = 0
            elif 'Jan' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("JAN")
                sheet = ws1
                globalVars.sheetIndex = 1
                lastDay =31
                globalVars.monthEndDaysIndex = 0

        if str(startMonth[0:2]) =="02":
            if'FEB' in workbook.sheetnames:
                sheet = workbook["FEB"]
                globalVars.sheetIndex =2
                lastDay =globalVars.monthEndDays[1]
                globalVars.monthEndDaysIndex = 1
            elif 'Feb' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("FEB")
                sheet = ws1
                globalVars.sheetIndex = 2
                lastDay =26
                globalVars.monthEndDaysIndex = 1
        if str(startMonth[0:2]) =="03":
            if'MAR' in workbook.sheetnames:
                sheet = workbook["MAR"]
                globalVars.sheetIndex = 3
                globalVars.monthEndDaysIndex = 2

                lastDay =globalVars.monthEndDays[2]
            elif 'MAR' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("MAR")
                sheet = ws1
                globalVars.sheetIndex = 3
                globalVars.monthEndDaysIndex = 2

                lastDay=31
        if str(startMonth[0:2]) =="04":
            if'APR' in workbook.sheetnames:
                sheet = workbook["APR"]
                globalVars.sheetIndex = 4
                globalVars.monthEndDaysIndex = 3

                lastDay =globalVars.monthEndDays[3]
            elif 'APR' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("APR")
                sheet = ws1
                globalVars.sheetIndex = 4
                globalVars.monthEndDaysIndex = 3

                lastDay=30
        if str(startMonth[0:2]) =="05":
            if'MAY' in workbook.sheetnames:
                sheet = workbook["MAY"]
                globalVars.sheetIndex = 5
                globalVars.monthEndDaysIndex = 4

                lastDay =globalVars.monthEndDays[4]
            elif 'MAY' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("MAY")
                sheet = ws1
                globalVars.monthEndDaysIndex = 4

                globalVars.sheetIndex = 5
                lastDay=31

        if str(startMonth[0:2]) =="06":
            if'JUNE' in workbook.sheetnames:
                sheet = workbook["JUNE"]
                globalVars.sheetIndex = 6
                lastDay =globalVars.monthEndDays[5]
                globalVars.monthEndDaysIndex = 5
            elif 'JUNE' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("JUNE")
                sheet = ws1
                globalVars.sheetIndex = 6
                lastDay=30
                globalVars.monthEndDaysIndex = 5

        if str(startMonth[0:2]) =="07":
            if'JULY' in workbook.sheetnames:
                sheet = workbook["JULY"]
                globalVars.sheetIndex = 7
                lastDay =globalVars.monthEndDays[6]
                globalVars.monthEndDaysIndex = 6
            elif 'JULY' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("JULY")
                sheet = ws1
                globalVars.sheetIndex = 7
                lastDay=31
                globalVars.monthEndDaysIndex = 6

        if str(startMonth[0:2]) =="08":
            if'AUG' in workbook.sheetnames:
                sheet = workbook["AUG"]
                globalVars.sheetIndex = 8
                lastDay =globalVars.monthEndDays[7]
                globalVars.monthEndDaysIndex = 7
            elif 'AUG' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("AUG")
                sheet = ws1
                globalVars.sheetIndex = 8
                lastDay=31
                globalVars.monthEndDaysIndex = 7

        if str(startMonth[0:2]) =="09":
            if'SEPT' in workbook.sheetnames:
                sheet = workbook["SEPT"]
                globalVars.sheetIndex = 9
                lastDay =globalVars.monthEndDays[8]
                globalVars.monthEndDaysIndex = 8
            elif 'SEPT' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("SEPT")
                sheet = ws1
                globalVars.sheetIndex = 9
                lastDay=30
                globalVars.monthEndDaysIndex = 8
        if str(startMonth[0:2]) =="10":
            if'OCT' in workbook.sheetnames:
                sheet = workbook["OCT"]
                globalVars.sheetIndex = 10
                lastDay =globalVars.monthEndDays[9]
                globalVars.monthEndDaysIndex = 9
            elif 'OCT' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("OCT")
                sheet = ws1
                globalVars.sheetIndex = 10
                lastDay=31
                globalVars.monthEndDaysIndex = 9
        if str(startMonth[0:2]) =="11":
            if'NOV' in workbook.sheetnames:
                sheet = workbook["NOV"]
                globalVars.sheetIndex = 11
                lastDay =globalVars.monthEndDays[10]
                globalVars.monthEndDaysIndex = 10
            elif 'NOV' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("NOV")
                sheet = ws1
                globalVars.sheetIndex = 11
                lastDay=30
                globalVars.monthEndDaysIndex = 10

        if str(startMonth[0:1]) =="12":
            if'DEC' in workbook.sheetnames:
                sheet = workbook["DEC"]
                globalVars.sheetIndex = 12
                lastDay =globalVars.monthEndDays[11]
                globalVars.monthEndDaysIndex = 11
            elif 'DEC' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("DEC")
                sheet = ws1
                globalVars.sheetIndex = 12
                lastDay=31
                globalVars.monthEndDaysIndex = 11

        # if(startMonth == endMonth):
        #     None
        #     tab7add_qual()
        # elif(startMonth != endMonth):
        #     endDate = "30"
        #     monthCheck =True
        # print(qualcellref.value)
        # print(type(qualcellref.value))
        qualValue = str(qualcellref.value)
        globalVars.startDate = int(startdatecellref.value[3:5])
        print("Start Date first: "+str(globalVars.startDate))
        columnIndex = globalVars.startDate


        # global secondMonthCheck
        # secondMonthCheck = False

#========================================QUAL1================================================================
#========================================######==============================================================
        if qualValue == "1":
            globalVars.qualName = 1
            
#=======================================BLOCK1================================================================
            monthEndDate = globalVars.monthEndDays[globalVars.monthEndDaysIndex]
            currentBlock = globalVars.qual1Block[globalVars.qual1BlockIndex]
            globalVars.currentBlock = currentBlock
            currentBlockTotalDayCount = globalVars.blockTotalDayCount[globalVars.blockTotalDayCountIndex]
            globalVars.currentblockTotalDayCount = currentBlockTotalDayCount
            endDate = globalVars.startDate+currentBlockTotalDayCount+(globalVars.weekendCount)
            globalVars.endDate = endDate
            # print("Start Date: "+str(globalVars.startDate))
            # print("End Date before Weekends: "+str(endDate))
            # print("Current Block: "+str(currentBlock))
            # print("Current Sheet: "+sheet.title)
            Worker3.weekendCheck(self)
            Worker3.checkHolidays(self)
            print("Weekend count: "+str(globalVars.weekendCount))

            print("Current Block: "+str(currentBlock))
            print("Current Sheet: "+sheet.title)
            print("Start Date: "+str(globalVars.startDate))
            endDate = globalVars.startDate+currentBlockTotalDayCount+(globalVars.weekendCount)+globalVars.familyDays
            globalVars.endDate = endDate
            print("After Weekends End Day: "+str(endDate))
            Worker3.secondMonthCheck(self)
            i=1
            globalVars.progressbarValue+=16.67
            self.progress.emit(i)
            
            print("Progress Value: "+str(globalVars.progressbarValue))
            i+=1
                

            #====================================BLOCK2-========================================================
            monthEndDate = globalVars.monthEndDays[globalVars.monthEndDaysIndex]
            currentBlock = globalVars.qual1Block[globalVars.qual1BlockIndex]
            globalVars.currentBlock = currentBlock
            currentBlockTotalDayCount = globalVars.blockTotalDayCount[globalVars.blockTotalDayCountIndex]
            globalVars.currentblockTotalDayCount = currentBlockTotalDayCount
            endDate = globalVars.startDate+currentBlockTotalDayCount+(globalVars.weekendCount)
            globalVars.endDate = endDate
            print("Start Date BLOCK 2 START: "+str(globalVars.startDate))
            print("End Date before Weekends: "+str(endDate))
            print("Current Block: "+str(currentBlock))
            print("Current Sheet: "+sheet.title)
            Worker3.weekendCheck(self)
            print("Weekend count: "+str(globalVars.weekendCount))

            print("Current Block: "+str(currentBlock))
            print("Current Sheet: "+sheet.title)
            print("Start Date BLOCK2: "+str(globalVars.startDate))
            endDate = globalVars.startDate+currentBlockTotalDayCount+(globalVars.weekendCount)
            globalVars.endDate = endDate
            print("After Weekends End Day: "+str(endDate))
            Worker3.secondMonthCheck(self)
            i=0
            globalVars.progressbarValue+=16.67
            print("Progress Value: "+str(globalVars.progressbarValue))
            self.progress.emit(i + 16.67)
            
            #====================================BLOCK3-========================================================
            monthEndDate = globalVars.monthEndDays[globalVars.monthEndDaysIndex]
            currentBlock = globalVars.qual1Block[globalVars.qual1BlockIndex]
            globalVars.currentBlock = currentBlock
            currentBlockTotalDayCount = globalVars.blockTotalDayCount[globalVars.blockTotalDayCountIndex]
            globalVars.currentblockTotalDayCount = currentBlockTotalDayCount
            endDate = globalVars.startDate+currentBlockTotalDayCount+(globalVars.weekendCount)
            globalVars.endDate = endDate
            print("Start Date: "+str(globalVars.startDate))
            print("End Date before Weekends: "+str(endDate))
            print("Current Block: "+str(currentBlock))
            print("Current Sheet: "+sheet.title)
            Worker3.weekendCheck(self)
            print("Weekend count: "+str(globalVars.weekendCount))

            print("Current Block: "+str(currentBlock))
            print("Current Sheet: "+sheet.title)
            print("Start Date: "+str(globalVars.startDate))
            endDate = globalVars.startDate+currentBlockTotalDayCount+(globalVars.weekendCount)
            globalVars.endDate = endDate
            print("After Weekends End Day: "+str(endDate))
            Worker3.secondMonthCheck(self)
            i=0
            globalVars.progressbarValue+=16.67
            print("Progress Value: "+str(globalVars.progressbarValue))
            self.progress.emit(i + 16.67)
            
            #====================================BLOCK4-========================================================
            monthEndDate = globalVars.monthEndDays[globalVars.monthEndDaysIndex]
            currentBlock = globalVars.qual1Block[globalVars.qual1BlockIndex]
            globalVars.currentBlock = currentBlock
            currentBlockTotalDayCount = globalVars.blockTotalDayCount[globalVars.blockTotalDayCountIndex]
            globalVars.currentblockTotalDayCount = currentBlockTotalDayCount
            endDate = globalVars.startDate+currentBlockTotalDayCount+(globalVars.weekendCount)
            globalVars.endDate = endDate
            print("Start Date: "+str(globalVars.startDate))
            print("End Date before Weekends: "+str(endDate))
            print("Current Block: "+str(currentBlock))
            print("Current Sheet: "+sheet.title)
            Worker3.weekendCheck(self)
            print("Weekend count: "+str(globalVars.weekendCount))

            print("Current Block: "+str(currentBlock))
            print("Current Sheet: "+sheet.title)
            print("Start Date: "+str(globalVars.startDate))
            endDate = globalVars.startDate+currentBlockTotalDayCount+(globalVars.weekendCount)
            globalVars.endDate = endDate
            print("After Weekends End Day: "+str(endDate))
            Worker3.secondMonthCheck(self)
            i=0
            globalVars.progressbarValue+=16.67
            print("Progress Value: "+str(globalVars.progressbarValue))
            self.progress.emit(i + 16.67)
            
            #====================================BLOCK5-========================================================
            monthEndDate = globalVars.monthEndDays[globalVars.monthEndDaysIndex]
            currentBlock = globalVars.qual1Block[globalVars.qual1BlockIndex]
            globalVars.currentBlock = currentBlock
            currentBlockTotalDayCount = globalVars.blockTotalDayCount[globalVars.blockTotalDayCountIndex]
            globalVars.currentblockTotalDayCount = currentBlockTotalDayCount
            endDate = globalVars.startDate+currentBlockTotalDayCount+(globalVars.weekendCount)
            globalVars.endDate = endDate
            print("Start Date: "+str(globalVars.startDate))
            print("End Date before Weekends: "+str(endDate))
            print("Current Block: "+str(currentBlock))
            print("Current Sheet: "+sheet.title)
            Worker3.weekendCheck(self)
            print("Weekend count: "+str(globalVars.weekendCount))

            print("Current Block: "+str(currentBlock))
            print("Current Sheet: "+sheet.title)
            print("Start Date: "+str(globalVars.startDate))
            endDate = globalVars.startDate+currentBlockTotalDayCount+(globalVars.weekendCount)
            globalVars.endDate = endDate
            print("After Weekends End Day: "+str(endDate))
            Worker3.secondMonthCheck(self)
            i=0
            globalVars.progressbarValue+=16.67
            print("Progress Value: "+str(globalVars.progressbarValue))
            self.progress.emit(i + 16.67)
            
            #====================================BLOCK6-========================================================
            monthEndDate = globalVars.monthEndDays[globalVars.monthEndDaysIndex]
            currentBlock = globalVars.qual1Block[globalVars.qual1BlockIndex]
            globalVars.currentBlock = currentBlock
            currentBlockTotalDayCount = globalVars.blockTotalDayCount[globalVars.blockTotalDayCountIndex]
            globalVars.currentblockTotalDayCount = currentBlockTotalDayCount
            endDate = globalVars.startDate+currentBlockTotalDayCount+(globalVars.weekendCount)
            globalVars.endDate = endDate
            print("Start Date: "+str(globalVars.startDate))
            print("End Date before Weekends: "+str(endDate))
            print("Current Block: "+str(currentBlock))
            print("Current Sheet: "+sheet.title)
            Worker3.weekendCheck(self)
            print("Weekend count: "+str(globalVars.weekendCount))

            print("Current Block: "+str(currentBlock))
            print("Current Sheet: "+sheet.title)
            print("Start Date: "+str(globalVars.startDate))
            endDate = globalVars.startDate+currentBlockTotalDayCount+(globalVars.weekendCount)
            globalVars.endDate = endDate
            print("After Weekends End Day: "+str(endDate))
            Worker3.secondMonthCheck(self)
            i=0
            globalVars.progressbarValue+=16.67
            print("Progress Value: "+str(globalVars.progressbarValue))
            self.progress.emit(i)


#========================================QUAL2================================================================
#========================================######==============================================================

        elif qualValue =="2":
            globalVars.qualName = 2
            
#=======================================BLOCK1================================================================
            monthEndDate = globalVars.monthEndDays[globalVars.monthEndDaysIndex]
            currentBlock = globalVars.qual1Block[globalVars.qual1BlockIndex]
            globalVars.currentBlock = currentBlock
            currentBlockTotalDayCount = globalVars.qual2BlockTotalDayCount[globalVars.blockTotalDayCountIndex]
            globalVars.currentblockTotalDayCount = currentBlockTotalDayCount
            endDate = globalVars.startDate+currentBlockTotalDayCount+(globalVars.weekendCount)
            globalVars.endDate = endDate
            print("Start Date: "+str(globalVars.startDate))
            print("End Date before Weekends: "+str(endDate))
            print("Current Block: "+str(currentBlock))
            print("Current Sheet: "+sheet.title)
            Worker3.weekendCheck(self)
            print("Weekend count: "+str(globalVars.weekendCount))

            print("Current Block: "+str(currentBlock))
            print("Current Sheet: "+sheet.title)
            print("Start Date: "+str(globalVars.startDate))
            endDate = globalVars.startDate+currentBlockTotalDayCount+(globalVars.weekendCount)
            globalVars.endDate = endDate
            print("After Weekends End Day: "+str(endDate))
            Worker3.secondMonthCheck(self)
            i=1
            globalVars.progressbarValue+=50
            self.progress.emit(i)
            
            # print("Progress Value: "+str(globalVars.progressbarValue))
            i+=1
                

            #====================================BLOCK2-========================================================
            monthEndDate = globalVars.monthEndDays[globalVars.monthEndDaysIndex]
            currentBlock = globalVars.qual1Block[globalVars.qual1BlockIndex]
            globalVars.currentBlock = currentBlock
            currentBlockTotalDayCount = globalVars.qual2BlockTotalDayCount[globalVars.blockTotalDayCountIndex]
            globalVars.currentblockTotalDayCount = currentBlockTotalDayCount
            endDate = globalVars.startDate+currentBlockTotalDayCount+(globalVars.weekendCount)
            globalVars.endDate = endDate
            print("Start Date: "+str(globalVars.startDate))
            print("End Date before Weekends: "+str(endDate))
            print("Current Block: "+str(currentBlock))
            print("Current Sheet: "+sheet.title)
            Worker3.weekendCheck(self)
            print("Weekend count: "+str(globalVars.weekendCount))

            print("Current Block: "+str(currentBlock))
            print("Current Sheet: "+sheet.title)
            print("Start Date: "+str(globalVars.startDate))
            endDate = globalVars.startDate+currentBlockTotalDayCount+(globalVars.weekendCount)
            globalVars.endDate = endDate
            print("After Weekends End Day: "+str(endDate))
            Worker3.secondMonthCheck(self)
            i=0
            globalVars.progressbarValue+=50
            print("Progress Value: "+str(globalVars.progressbarValue))
            self.progress.emit(i + 16.67)

#========================================QUAL3================================================================
#========================================######==============================================================

        elif qualValue =="3":
            globalVars.qualName = 3
            
#=======================================BLOCK1================================================================
            monthEndDate = globalVars.monthEndDays[globalVars.monthEndDaysIndex]
            currentBlock = globalVars.qual1Block[globalVars.qual1BlockIndex]
            globalVars.currentBlock = currentBlock
            currentBlockTotalDayCount = globalVars.qual3BlockTotalDayCount[globalVars.blockTotalDayCountIndex]
            globalVars.currentblockTotalDayCount = currentBlockTotalDayCount
            endDate = globalVars.startDate+currentBlockTotalDayCount+(globalVars.weekendCount)
            globalVars.endDate = endDate
            print("Start Date: "+str(globalVars.startDate))
            print("End Date before Weekends: "+str(endDate))
            print("Current Block: "+str(currentBlock))
            print("Current Sheet: "+sheet.title)
            Worker3.weekendCheck(self)
            print("Weekend count: "+str(globalVars.weekendCount))

            print("Current Block: "+str(currentBlock))
            print("Current Sheet: "+sheet.title)
            print("Start Date: "+str(globalVars.startDate))
            endDate = globalVars.startDate+currentBlockTotalDayCount+(globalVars.weekendCount)
            globalVars.endDate = endDate
            print("After Weekends End Day: "+str(endDate))
            Worker3.secondMonthCheck(self)
            i=1
            globalVars.progressbarValue+=33.34
            self.progress.emit(i)
            
            print("Progress Value: "+str(globalVars.progressbarValue))
            i+=1
                

            #====================================BLOCK2-========================================================
            monthEndDate = globalVars.monthEndDays[globalVars.monthEndDaysIndex]
            currentBlock = globalVars.qual1Block[globalVars.qual1BlockIndex]
            globalVars.currentBlock = currentBlock
            currentBlockTotalDayCount = globalVars.qual3BlockTotalDayCount[globalVars.blockTotalDayCountIndex]
            globalVars.currentblockTotalDayCount = currentBlockTotalDayCount
            endDate = globalVars.startDate+currentBlockTotalDayCount+(globalVars.weekendCount)
            globalVars.endDate = endDate
            print("Start Date: "+str(globalVars.startDate))
            print("End Date before Weekends: "+str(endDate))
            print("Current Block: "+str(currentBlock))
            print("Current Sheet: "+sheet.title)
            Worker3.weekendCheck(self)
            print("Weekend count: "+str(globalVars.weekendCount))

            print("Current Block: "+str(currentBlock))
            print("Current Sheet: "+sheet.title)
            print("Start Date: "+str(globalVars.startDate))
            endDate = globalVars.startDate+currentBlockTotalDayCount+(globalVars.weekendCount)
            globalVars.endDate = endDate
            print("After Weekends End Day: "+str(endDate))
            Worker3.secondMonthCheck(self)
            i=0
            globalVars.progressbarValue+=33.34
            print("Progress Value: "+str(globalVars.progressbarValue))
            self.progress.emit(i + 16.67)
        #====================================BLOCK3-========================================================
            monthEndDate = globalVars.monthEndDays[globalVars.monthEndDaysIndex]
            currentBlock = globalVars.qual1Block[globalVars.qual1BlockIndex]
            globalVars.currentBlock = currentBlock
            currentBlockTotalDayCount = globalVars.qual3BlockTotalDayCount[globalVars.blockTotalDayCountIndex]
            globalVars.currentblockTotalDayCount = currentBlockTotalDayCount
            endDate = globalVars.startDate+currentBlockTotalDayCount+(globalVars.weekendCount)
            globalVars.endDate = endDate
            print("Start Date: "+str(globalVars.startDate))
            print("End Date before Weekends: "+str(endDate))
            print("Current Block: "+str(currentBlock))
            print("Current Sheet: "+sheet.title)
            Worker3.weekendCheck(self)
            print("Weekend count: "+str(globalVars.weekendCount))

            print("Current Block: "+str(currentBlock))
            print("Current Sheet: "+sheet.title)
            print("Start Date: "+str(globalVars.startDate))
            endDate = globalVars.startDate+currentBlockTotalDayCount+(globalVars.weekendCount)
            globalVars.endDate = endDate
            print("After Weekends End Day: "+str(endDate))
            Worker3.secondMonthCheck(self)
            i=0
            globalVars.progressbarValue+=33.34
            print("Progress Value: "+str(globalVars.progressbarValue))
            self.progress.emit(i + 16.67)


           


            #RESET
        

        globalVars.qual1BlockIndex =1

        
        globalVars.blockTotalDayCountIndex=1

        globalVars.startDate = 0
        globalVars.endDate = 0
        globalVars.currentblockTotalDayCount = 0
        globalVars.weekendCount=0

        globalVars.qualName = 0
        globalVars.blockName = 0
        globalVars.currentBlock = 0
        globalVars.sheetIndex = 0


            

    def weekendCheck(self):
        lastDayIsWeekend=False
        
        startDate = globalVars.startDate
        endDate = globalVars.endDate
        print("In Weekend Check Start Date: "+str(startDate))
        print("In Weekend Check Start Date: "+str(endDate))
        # globalVars.endDate = globalVars.startDate+blockDayTotal
        while startDate <= endDate:
            # print("start date: "+str(startDate))
            # print("End Date: "+str(endDate))
            # print("Starting Weekend Count: "+str(globalVars.weekendCount))

            weekendCell = sheet.cell(row=2,column=startDate)
            if (weekendCell.value == "Saturday" or weekendCell.value == "Sunday"):
                if((weekendCell.value=="Saturday"or weekendCell.value =="Sunday") and startDate>globalVars.monthEndDays[globalVars.monthEndDaysIndex]):
                    lastDayIsWeekend = True
                    globalVars.weekendCount = globalVars.weekendCount-1
                globalVars.weekendCount+=1
                endDate+=1
                startDate+=1
                weekendCell = sheet.cell(row=2,column=startDate)
            else:
                startDate+=1
                weekendCell = sheet.cell(row=2,column=startDate)
        print("WEEKEND COUNT: "+str(globalVars.weekendCount))

        if endDate > globalVars.monthEndDays[globalVars.monthEndDaysIndex]:
            newSheetIndex = globalVars.sheetIndex
            newSheetIndex+=1
            sheets = workbook.sheetnames
            newSheet = workbook[sheets[newSheetIndex]]
            print ("New Sheet: "+str(newSheet.title))
            newStartDate = 1
            if lastDayIsWeekend == True:
                newEndDate = (endDate -globalVars.monthEndDays[globalVars.monthEndDaysIndex])-2#+globalVars.weekendCount
            else:
                newEndDate = (endDate -globalVars.monthEndDays[globalVars.monthEndDaysIndex])
            print("New Start Date: "+str(newStartDate))
            print("New End Date: "+str(newEndDate))
            while newStartDate <= newEndDate:
                print("Weekend Count: "+str(globalVars.weekendCount))
                print("New Start Date: "+str(newStartDate))
                print("New End Date: "+str(newEndDate))

                weekendCell = newSheet.cell(row=2,column=newStartDate)
                

                print("Weekend Value: "+str(weekendCell.value))
                if weekendCell.value == "Saturday" or weekendCell.value == "Sunday":
                    globalVars.weekendCount+=1
                    print("WEEKEND COUNT: "+str(globalVars.weekendCount))

                    newEndDate+=1
                    newStartDate+=1
                    weekendCell = sheet.cell(row=2,column=newStartDate)
                else:
                    newStartDate+=1
                    weekendCell = sheet.cell(row=2,column=newStartDate)
             

    
        
        # msgBox = QMessageBox()
        # msgBox.setIcon(QMessageBox.Critical)
        # msgBox.setText("Error while trying to read from Read In File! Check your read from Entries")
        # msgBox.setWindowTitle("ERROR IN ADD FULL QUAL")
        # msgBox.show()

    def addFullQual(self):
        try:
            
            
            # tab7Label3 = Label(gui.tab7frame3,text = "FINISHED",fg="grey26",bg="grey26",font="Helvetica 10 bold").grid(row=4,column=0)
            # global tab7my_progress
            # gui.tab7my_progress.grid(row=0,column=0,pady=(50,10))

            #GETTING READ FROM FILE
            global tab7readExcelFile
            tab7readExcelFile = globalVars.page4ReadFromFile
            # print(tab7readExcelFile)
            if path.exists(str(globalVars.page4ReadFromFile)) == True: 
                readFromBook = load_workbook(filename =tab7readExcelFile)
                print("First file found")
            else:
                print("First file Not found")
                # fileerrorLabel = Label(gui.tab7,text="File Does Not Exist",font="Helvetica 10 bold", fg="red")
                # fileerrorLabel.grid(row=5,column=0)

            
            sheet = readFromBook.active
            # print(sheet.title)
            startingRow = 2
            startingCol = 1
            global qualcellref
            # global blockcellref
            global startdatecellref
            # global enddatecellref
            global classcellref
            global instructorcellref
            qualcellref = sheet.cell(row=startingRow,column=startingCol)
            print("QUAL VALUE: "+str(qualcellref.value))
            
            startdatecellref = sheet.cell(row=startingRow,column=startingCol+1)
            classcellref = sheet.cell(row=startingRow,column=startingCol+2)
            instructorcellref = sheet.cell(row=startingRow,column=startingCol+3)

            
            
            # print(qualcellref.value)
            while qualcellref.value:
                self.newBar.emit()
                qualstrValue = str(qualcellref.value)
                if  qualstrValue.isalnum()==False:
                    print("SpecialChracters Error")
                    self.finished.emit()
                    self.error.emit('special')
                    return
                     
                if qualcellref.value and startdatecellref.value is None:
                    print("Timer Error")
                    self.finished.emit()
                    self.error.emit('nostartdate')
                    return
                print("Start Date: "+str(startdatecellref.value))
                i=0
                print(str(qualcellref.value))
                print(qualcellref.coordinate)
                print("CLASS: "+str(classcellref.value))
                print(str(classcellref.coordinate))
                # gui.tab7my_progress['value']+=30
                    
                # gui.tab7.update_idletasks() 
                # print("QUALCELLREF VALUE: " +str(qualcellref.value))
                Worker3.readFullQualIn(self)
                startingRow+=1
                qualcellref = sheet.cell(row=startingRow,column=startingCol)
                # print(qualcellref.value)
                # blockcellref = sheet.cell(row=startingRow,column=startingCol+1)
                # print(blockcellref.value)
                startdatecellref = sheet.cell(row=startingRow,column=startingCol+1)
                print(startdatecellref.value)
                # enddatecellref = sheet.cell(row=startingRow,column=startingCol+3)
                classcellref = sheet.cell(row=startingRow,column=startingCol+2)
                print("CLASS: "+str(classcellref.value))
                print(str(classcellref.coordinate))
                
                instructorcellref = sheet.cell(row=startingRow,column=startingCol+3)
                globalVars.instructor = str(instructorcellref.value)
            self.finished.emit()
            print("Finished")
            # tab7Label3 = Label(gui.tab7frame3,text = "FINISHED",fg="white",bg="grey26",font="Helvetica 10 bold").grid(row=4,column=0)
            
            # gui.tab7my_progress.stop()
        except():
            # print(traceback.format_exc())
            self.finished.emit()
            
            self.error.emit('readFrom')
            print("Error emitted")
            if path.exists("../textFiles/errors.txt") == True:
                ct = datetime.datetime.now() 
                with open("curtis_program - Testing Seperation/QTDESIGNER/textFiles/errors.txt", "a") as file:
                    file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))
            else:
                ct = datetime.datetime.now() 
                with open("curtis_program - Testing Seperation/QTDESIGNER/textFiles/errors.txt", "x") as file:
                    file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))  


    def tab7add_qual(self,qualName,blockName,startDate,endDate):
            try:
                print("Is Checked: "+str(globalVars.isChecked))
                print("Starting to add qual")
                print("QUAL: "+str(qualName))
                print("Block: "+str(blockName))
                print("Starting Date: "+str(startDate))
                print("End Date: "+str(endDate))
            
        #NONE DECLARED ERROR LABELS USED
                global fileerrorLabel
                global qualerrorLabel
                global blockerrorLabel
                global starterrorLabel
                global enderrorLabel
                global qual1block1mirs
                global qual1block2mirs      
            #SET COLUMN WIDTH
                # for col in range(1,32):
                #     column_letter = get_column_letter(col)
                #     # print(column_letter)
                #     sheet.column_dimensions[column_letter].width = 42

            #SET TITLE
                sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=32)
                sheet.row_dimensions[3].height = 60
                title_Cell = sheet['A3']
                title_Cell.border = thick_border
                monthTitle = str(sheet.title)
                # print(monthTitle)
                title_Cell.value = monthTitle+" ELECTRICAL SYSTEMS INSTRUCTOR SCHEDULE"
                title_Cell.font = Font(size=60,name="Times New Roman",bold=True)

                
                # global startMonth
                #SET DATES
                # for i in range(1,32):
                #     datecellref=sheet.cell(row=1, column=i)
                #     datecellref.fill = PatternFill("solid", fgColor="DDDDDD")
                    
                #     if(i<10):
                #         datecellref.value=str(startMonth)+"/0"+str(i)+"/2021"
                #     else:
                #         datecellref.value=str(startMonth)+"/"+str(i)+"/2021"


                #SET DIVIDER      

                for i in range(1,32):
                    datecellref2=sheet.cell(row=4, column=i)
                    datecellref2.fill = PatternFill("solid", fgColor="000000")
                    datecellref2.border = thick_border_blue_topBottom
                    datecellref2.value="blank"  
                    # gui.tab7.update_idletasks() 


                #GETTING INPUT VALUES FROM USER
                # qualName = str(qualcellref.value)
                instructor = str(instructorcellref.value)
                if str(qualName) == "FAMILY DAY":
                    None
                else:
                    classNum = int(classcellref.value)
                    if int(qualName) > 3 and str(qualName) !="FAMILY DAY":
                        # qualerrorLabel = Label(gui.tab7frame2,text = "Qual Num Doesn't Exist",font="Helvetica 10 bold", fg="red")
                        # qualerrorLabel.grid(row=11,column=0)
                        # e2.delete(0, END)
                        return 
                    # blockName = str(blockcellref.value)
                    if int(blockName) > 6 and str(qualName) !="FAMILY DAY":
                        # blockerrorLabel = Label(gui.tab7frame2,text = "Block Number Doesn't Exist",font="Helvetica 10 bold", fg="red")
                        # blockerrorLabel.grid(row=11,column=0)
                        # e3.delete(0, END)
                        return
                # startDate = str(startdatecellref.value)
                # startDate = startDate[3:5]
                # print(str(startDate))
                if int(startDate) > 32:
                    # starterrorLabel = Label(gui.tab7frame2,text = "Start Date Number too big",font="Helvetica 10 bold", fg="red")
                    # starterrorLabel.grid(row=11,column=0)
                    # e4.delete(0, END)
                    return
                if int(endDate) > 32:
                    # enderrorLabel = Label(gui.tab7frame2,text = "End Date Number too big",font="Helvetica 10 bold", fg="red")
                    # enderrorLabel.grid(row=11,column=0)
                    # e5.delete(0, END)
                    return


                

                #Start Date
                col_index = int(startDate)
                row_index=5
                # print("First row index: "+str(row_index))

                #THE DASHED CELLS ITERATORS
                # blank_col_index = col_index
                # blank_row_index=row_index+1
                
                #TITLE CELL
                cellref = sheet.cell(row=row_index,column=col_index)
                cellref.border = thick_border
                blockcheckcell = sheet.cell(row=row_index, column = 1)
            

                #DASHED CELLS
                blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)


                #Setting Styles
                # cellref.font=ft1
                # cellref.alignment=align
                # blank_cellref.font=ft1
                # blank_cellref.alignment=align
                # blank_cellref2.font=ft1
                # blank_cellref2.alignment=align
                # blank_cellref3.font=ft1
                # blank_cellref3.alignment=align
                # blank_cellref4.font=ft1
                # blank_cellref4.alignment=align

                #End Date
                #check if months match


                #SETTING ACTIVE MIRS
                activeMirs=[]
                
                if qualName ==1 and blockName == 1:
                    for items in globalVars.qual1block1mirs:
                        activeMirs.append(items)
                elif qualName ==1 and blockName ==2:
                    for items in globalVars.qual1block2mirs:
                        activeMirs.append(items)
                elif qualName ==1 and blockName ==3:
                    for items in globalVars.qual1block3mirs:
                        activeMirs.append(items)
                elif qualName ==1 and blockName ==4:
                    for items in globalVars.qual1block4mirs:
                        activeMirs.append(items)
                elif qualName ==1 and blockName ==5:
                    for items in globalVars.qual1block5mirs:
                        activeMirs.append(items)
                elif qualName ==1 and blockName ==6:
                    for items in globalVars.qual1block6mirs:
                        activeMirs.append(items)
                elif qualName ==2 and blockName ==1:
                    for items in globalVars.qual2block1mirs:
                        activeMirs.append(items)
                elif qualName ==2 and blockName ==2:
                    for items in globalVars.qual2block2mirs:
                        activeMirs.append(items)
                elif qualName ==3 and blockName ==1:
                    for items in globalVars.qual3block1mirs:
                        activeMirs.append(items)
                elif qualName ==3 and blockName ==2:
                    for items in globalVars.qual3block2mirs:
                        activeMirs.append(items)
                elif qualName ==3 and blockName ==3:
                    for items in globalVars.qual3block3mirs:
                        activeMirs.append(items)
            
                        



                des_col=int(endDate)+1

                
                
                total_index = 0

                mirsIndex = 0

                

                # print(tab5onleavelist[0])
                # print(tab5onleavelist[1])
                # print(tab5onleavelist[2])
                # for person in tab5onleavelist:
                #     print(person)
            
                
                
                
                
            
            
                q1check = False
                global colCheck
                colCheck = False


                #COLUMN ITERATOR LOOP
                # print(type(des_col))
                # print(des_col)
                while col_index < des_col:
                    colCheck = False
                    tab5leavecheck = False
                    tab5onleavelist = []
                    tab5leaverow = 5
                    leavecellref = sheet.cell(row=tab5leaverow,column=col_index)
                    while leavecellref.value !="LEAVE":
                        tab5leaverow+=1
                        leavecellref = sheet.cell(row=tab5leaverow,column=col_index)
                    if leavecellref.value =="LEAVE":
                        tab5leaverow+=1
                        leavecellref = sheet.cell(row=tab5leaverow,column=col_index)
                    while leavecellref.value:
                        person = str(leavecellref.value)
                        # print(person)
                        tab5onleavelist.append(person)
                        tab5leaverow+=1
                        leavecellref = sheet.cell(row=tab5leaverow,column=col_index)

                    for person in tab5onleavelist:
                        if instructor == person:
                            tab5leavecheck = True 
                    
                    #IF THERES A VALUE IN THE CELL
                    while cellref.value:
                        colCheck = False
                        # print(cellref.value)
                        if cellref.value == "Navy Unique":
                            sheet.insert_rows(row_index,5)
                            if col_index==1:
                                colCheck = True
                            else:
                                col_index-=1
                            # row_index-=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            # cellref.value="NEW"
                            # print(cellref.value)
                            # workbook.save(filename=workbook_Title)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                            break
                        # print(str(cellref.value))
                        if str(cellref.value) == "FAMILY DAY":
                            col_index+=1
                            des_col+=1
                            globalVars.holidayCheck = True
                            globalVars.endDate+=1
                            globalVars.familyDays +=1
                            # globalVars.startDate = globalVars.endDate+1
                            print("START DATE IN QUAL: "+str(globalVars.startDate))
                            print("END DATE IN QUAL: "+str(globalVars.endDate))
                            cellref = sheet.cell(row=row_index,column=col_index)
                            None
                        
                        elif str(cellref.value)!= "FAMILY DAY" and str(qualName)=="FAMILY DAY":
                            while row_index <=100:
                                cellref.value = "FAMILY DAY"
                                cellref.font = Font(color="FF0000")
                                cellref.fill = PatternFill("solid","FFFFFF")
                                cellref.border = thin_border_all_grey
                                row_index+=1
                                cellref = sheet.cell(row=row_index,column=col_index)

                        #First: IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK IS LESS THAN OR EQUAL AND THE CLASS NUM ENTERED IS GREATER
                        elif(int(qualName) <= int(cellref.value[1]) and int(blockName) <= int(cellref.value[10]) and int(qualName) == 1 and int(blockName)==1):
                            # print("Start Date: "+str(startDate))
                            # print("Start Date Type: "+str(type(startDate)))
                            if startDate!=1 and (int(classNum) >= int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999"):
                                if globalVars.isChecked == True:
                                    for ws in workbook.worksheets:
                                        ws.insert_rows(5,5)
                                        Worker3.checkHolidays(self)
                                else:
                                    sheet.insert_rows(5,5)
                                    Worker3.checkHolidays(self)
                                q1check = True
                                for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(startDate)-1):
                                    for cell in rows:
                                        cell.value = "Q1  Block:1 Class:99999"
                                        cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid",fgColor="white")
                                        cell.border = thin_border_all_grey
                                        cell.font = Font(color="FFFFFF")
                                # row_index=4
                                cellref = sheet.cell(row=row_index,column=col_index)
                                blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                                blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                                blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                                blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                            elif startDate!=1 and (int(classNum) < int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999"):
                                if globalVars.isChecked == True:
                                    for ws in workbook.worksheets:
                                        ws.insert_rows(5,5)
                                        Worker3.checkHolidays(self)
                                else:
                                    sheet.insert_rows(5,5)
                                    Worker3.checkHolidays(self)
                                
                                q1check = True
                                for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(startDate)-1):
                                    for cell in rows:
                                        cell.value = "Q1  Block:1 Class:99999"
                                        cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid",fgColor="white")
                                        cell.border = thin_border_all_grey
                                        cell.font = Font(color="FFFFFF")
                                        
                                row_index=5
                                cellref = sheet.cell(row=row_index,column=col_index)
                                blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                                blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                                blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                                blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                            elif startDate==1 and int(classNum) > int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999" :
                                if globalVars.isChecked == True:
                                    for ws in workbook.worksheets:
                                        ws.insert_rows(5,5)
                                        Worker3.checkHolidays(self)
                                else:
                                    sheet.insert_rows(5,5)
                                    Worker3.checkHolidays(self)
                                row_index=5
                                cellref = sheet.cell(row=row_index,column=col_index)
                                blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                                blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                                blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                                blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                            elif startDate==1 and (int(classNum) < int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999") :
                                row_index+=5
                                cellref = sheet.cell(row=row_index,column=col_index)
                                blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                                blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                                blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                                blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                            #Second:IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK NUM IS GREATER OR EQUAL 
                        elif(int(qualName) <= int(cellref.value[1]) and int(blockName) > int(cellref.value[10])):
                            row_index+=5
                            # print("Row index: " + str(row_index))
                            # print(cellref.value)
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                            

                            #Third:IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK IS LESS THAN OR EQUAL
                        elif(int(qualName) <= int(cellref.value[1]) and int(blockName) <= int(cellref.value[10])):
                            if int(classNum) > int(cellref.value[18:23]):
                                # print("Cell Coordinate: "+str(cellref))
                                # print("Last Blank Cell Coordinate: "+str(blank_cellref4))
                                sheet.insert_rows(row_index,5)
                                Worker3.checkHolidays(self)
                                newrowref = cellref.row
                                # print("Newrowref: "+str(newrowref))
                                row_index=newrowref-5
                                # print("row index: "+str(row_index))
                                cellref = sheet.cell(row=row_index,column=col_index)
                                blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                                blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                                blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                                blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                            else:

                                row_index+=5
                                cellref = sheet.cell(row=row_index,column=col_index)
                                blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                                blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                                blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                                blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)


                            #FOURTH: IF THE ENTERED QUAL IS GREATER THAN THE CELL VALUE
                        elif(int(qualName) > int(cellref.value[1])):
                            row_index+=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)


                        elif(int(qualName)<=int(cellref.value[1])):
                            sheet.move_range(""+str(cellref.coordinate)+":"+str(blank_cellref4.coordinate)+"", rows=5)
                            newrowref = cellref.row
                            row_index=newrowref-5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                            
                    else:
                        if str(qualName)=="FAMILY DAY":
                            while row_index <=100:
                                cellref.value = "FAMILY DAY"
                                cellref.font = Font(color="FF0000")
                                cellref.fill = PatternFill("solid","FFFFFF")
                                cellref.border = thin_border_all_grey
                                row_index+=1
                                cellref = sheet.cell(row=row_index,column=col_index)
                            
                        elif(int(qualName)==1 and int(blockName)==1 and int(startDate!=1) and q1check ==False):
                            sheet.insert_rows(5,5)
                            Worker3.checkHolidays(self)
                            # print("Row Index: "+str(row_index))
                            q1check =True
                            for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(startDate)-1):
                                    for cell in rows:
                                        cell.value = "Q1  Block:1 Class:99999"
                                        cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid",fgColor="white")
                                        cell.border = thin_border_all_grey
                                        cell.font = Font(color="FFFFFF")
                                    
                                        # print("col index: "+str(col_index))
                        # row_index=4
                        # print("Row Index: "+str(row_index))
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        if str(qualName)=="FAMILY DAY":
                            None    
                        else:
                            cellref.value="Q"+str(qualName)+" " +" Block:"+str(blockName)+" Class:"+str(classNum)
                            cellref.font = ft1
                            cellref.alignment=align
                        # print("col index: "+str(col_index))
                            #TITLE COLOR CODING
                            #QUAL 1
                            cellref.border = thick_border
                            if (qualName == 1 and blockName == 1):
                                cellref.fill = PatternFill("solid", fgColor="00CCFF")
                            elif (qualName == 1 and blockName == 2):
                                cellref.fill = PatternFill("solid", fgColor="33CCCC")
                            elif (qualName == 1 and blockName == 3):
                                cellref.fill = PatternFill("solid", fgColor="92D050")
                            elif (qualName == 1 and blockName == 4):
                                cellref.fill = PatternFill("solid", fgColor="FFFF00")
                            elif (qualName == 1 and blockName == 5):
                                cellref.fill = PatternFill("solid", fgColor="FFC000")
                            elif (qualName == 1 and blockName == 6):
                                cellref.fill = PatternFill("solid", fgColor="FF0000")

                            #QUAL2
                            elif (qualName == 2 and blockName == 1):
                                cellref.fill = PatternFill("solid", fgColor="00B0F0")
                            elif (qualName == 2 and blockName == 2):
                                cellref.fill = PatternFill("solid", fgColor="92D050")

                            #QUAL3
                            elif (qualName == 3 and blockName == 1):
                                cellref.fill = PatternFill("solid", fgColor="FFFF00")
                            elif (qualName == 3 and blockName == 2):
                                cellref.fill = PatternFill("solid", fgColor="9BBB59")
                            elif (qualName == 3 and blockName == 3):
                                cellref.fill = PatternFill("solid", fgColor="FFC000")
                            


                            #SETTING BLANK VALUES BORDERS
                            blank_cellref.border = thin_border_sides
                            blank_cellref2.border = thin_border_sides
                            blank_cellref3.border = thin_border_sides
                            blank_cellref4.border = thin_border_sides_Bottom

                            # SETTING BLANK VALUES
                            if tab5leavecheck == True:
                                # messagebox.showwarning(title="Instructor on Leave!", message="The entered instructor is on leave for one of the days")
                                # blank_cellref.value="Instructor: Instructor is on leave"
                                self.error.emit('instructorOnLeave')
                                tab5leavecheck ==False
                            else:
                                blank_cellref.value="Instructor: "+globalVars.instructor
                            blank_cellref.font = ft1
                            blank_cellref.alignment=align
                            blank_cellref2.value="------"
                            blank_cellref2.font = ft1
                            blank_cellref2.alignment=align
                            blank_cellref3.value="------"
                            blank_cellref3.font = ft1
                            blank_cellref3.alignment=align
                            if mirsIndex > len(activeMirs)-1:
                                blank_cellref4.value = "NonPop"
                            else:
                                blank_cellref4.value=str(activeMirs[mirsIndex])
                            blank_cellref4.font = ft1
                            blank_cellref4.alignment=align
                        total_index+=1
                    
                        # print("Title = " + cellref.coordinate)
                        # print("blank = " + blank_cellref.coordinate)
                    
                    if colCheck ==True:
                        None
                    else:
                        col_index+=1
                    mirsIndex+=1
                    # print("column index: "+str(col_index))   


                    #This line controls if they shoot up the rows or nor
                    # row_index=4
                    # print("row index"+str(row_index))   
                    cellref = sheet.cell(row=row_index,column=col_index)
                    blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                    blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                    blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                    blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                    # print("Row Index: "+str(row_index))
                    # print("cellref: "+str(cellref))
                Worker3.checkHolidays(self)
                workbook.save(filename=workbook_Title)
                


                    
                    
                  

                # second_Month_Check_ReadIn() 
            except:
                self.error.emit('general')
                print(traceback.format_exc()) 