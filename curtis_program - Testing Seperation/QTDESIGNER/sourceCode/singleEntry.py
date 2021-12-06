from typing import List
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell import cell
from openpyxl.cell.cell import ERROR_CODES
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, alignment
from openpyxl.styles.fills import fills
from openpyxl.worksheet.dimensions import SheetDimension
from openpyxl.utils import get_column_letter
import time
import os.path
from os import error, path
# from tkinter import messagebox
import traceback
import sys
import datetime
from PyQt5 import QtCore, QtGui,QtWidgets,uic
from PyQt5.QtWidgets import QMainWindow,QApplication, QWidget
from PyQt5.QtCore import QObject, QThread, pyqtSignal,pyqtSlot
import logging
import threading
import time
import globalVars


thick_border_blue = Border(left=Side(style='thick',color='0066CC'), 
                    right=Side(style='thick',color='0066CC')   
                ) 

thick_border_blue_topBottom = Border(top=Side(style='thick',color='0066CC'), 
                    bottom=Side(style='thick',color='0066CC')   
                )                    

thin_border_all = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))

thin_border_all_grey = Border(left=Side(style='thin',color="DDDDDD"), 
                    right=Side(style='thin',color="DDDDDD"), 
                    top=Side(style='thin',color="DDDDDD"), 
                    bottom=Side(style='thin',color="DDDDDD"))

thin_border_sides = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style=None), 
                    bottom=Side(style=None))


thin_border_sides_Bottom = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style=None), 
                    bottom=Side(style='thin'))
ft1 = Font(name='Arial',bold=True, size=14)

align = Alignment(horizontal='center')
thick_border = Border(left=Side(style='thick'), 
                    right=Side(style='thick'), 
                    top=Side(style='thick'), 
                    bottom=Side(style='thick'))
# import main


    #GLOBAL ERROR VARIABLES
fileerrorLabel = None
qualerrorLabel = None
blockerrorLabel = None
starterrorLabel = None
enderrorLabel = None
monthCheck = False


class Worker2(QObject):
    finished = pyqtSignal()
    progress = pyqtSignal(int)

    def second_add_qual(self):
        try:

        
    #NONE DECLARED ERROR LABELS USED
            global fileerrorLabel
            global qualerrorLabel
            global blockerrorLabel
            global starterrorLabel
            global enderrorLabel
            global qual1block1mirs
            global qual1block2mirs      
            sheet
        #SET COLUMN WIDTH
            # for col in range(1,32):
            #     column_letter = get_column_letter(col)
            #     # print(column_letter)
            #     sheet.column_dimensions[column_letter].width = 42

        #SET TITLE
            sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=32)
            sheet.row_dimensions[3].height = 60
            title_Cell = sheet['A3']
            title_Cell.border = thick_border
            monthTitle = str(sheet.title)
            # print(monthTitle)
            title_Cell.value = monthTitle+" ELECTRICAL SYSTEMS INSTRUCTOR SCHEDULE"
            title_Cell.font = Font(size=60,name="Times New Roman",bold=True)

            
            # global startMonth
            #SET DATES
            # for i in range(1,32):
            #     datecellref=sheet.cell(row=1, column=i)
            #     datecellref.fill = PatternFill("solid", fgColor="DDDDDD")
                
            #     if(i<10):
            #         datecellref.value=str(startMonth)+"/0"+str(i)+"/2021"
            #     else:
            #         datecellref.value=str(startMonth)+"/"+str(i)+"/2021"


            #SET DIVIDER      

            for i in range(1,32):
                datecellref2=sheet.cell(row=4, column=i)
                datecellref2.fill = PatternFill("solid", fgColor="000000")
                datecellref2.border = thick_border_blue_topBottom
                datecellref2.value="blank"  
                # gui.tab1.update_idletasks() 


            #GETTING INPUT VALUES FROM USER
            instructor = globalVars.page1e2
            qualName = globalVars.page1Qual
            if str(qualName) == "FAMILY DAY":
                None
            else:
                qualName = globalVars.page1Qual
                if int(qualName) > 3 and str(qualName) !="FAMILY DAY":
                    # qualerrorLabel = Label(gui.tab1frame2,text = "Qual Num Doesn't Exist",font="Helvetica 10 bold", fg="red")
                    # qualerrorLabel.grid(row=11,column=0)
                    # gui.e2.delete(0, END)
                    return 
                blockName = globalVars.page1Block
                if int(blockName) > 6 and str(qualName) !="FAMILY DAY":
                    # blockerrorLabel = Label(gui.tab1frame2,text = "Block Number Doesn't Exist",font="Helvetica 10 bold", fg="red")
                    # blockerrorLabel.grid(row=11,column=0)
                    # gui.e3.delete(0, END)
                    return
            # startDate = gui.e4.get()
            # startDate = startDate[3:5]
            # print(str(startDate))
            if int(newstartDate) > 32:
                # starterrorLabel = Label(gui.tab1frame2,text = "Start Date Number too big",font="Helvetica 10 bold", fg="red")
                # starterrorLabel.grid(row=11,column=0)
                # gui.e4.delete(0, END)
                return
            if int(newendDate) > 32:
                # enderrorLabel = Label(gui.tab1frame2,text = "End Date Number too big",font="Helvetica 10 bold", fg="red")
                # enderrorLabel.grid(row=11,column=0)
                # gui.e5.delete(0, END)
                return

            classNum = globalVars.page1e3

            

            #Start Date
            col_index = int(newstartDate)
            row_index=5
            # print("First row index: "+str(row_index))

            #THE DASHED CELLS ITERATORS
            # blank_col_index = col_index
            # blank_row_index=row_index+1
            
            #TITLE CELL
            cellref = sheet.cell(row=row_index,column=col_index)
            cellref.border = thick_border
            blockcheckcell = sheet.cell(row=row_index, column = 1)
        

            #DASHED CELLS
            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)


            #Setting Styles
            # cellref.font=ft1
            # cellref.alignment=align
            # blank_cellref.font=ft1
            # blank_cellref.alignment=align
            # blank_cellref2.font=ft1
            # blank_cellref2.alignment=align
            # blank_cellref3.font=ft1
            # blank_cellref3.alignment=align
            # blank_cellref4.font=ft1
            # blank_cellref4.alignment=align

            #End Date
            #check if months match


            #SETTING ACTIVE MIRS
            activeMirs=[]
            
            if qualName =="1" and blockName == "1":
                for items in globalVars.qual1block1mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="2":
                for items in globalVars.qual1block2mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="3":
                for items in globalVars.qual1block3mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="4":
                for items in globalVars.qual1block4mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="5":
                for items in globalVars.qual1block5mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="6":
                for items in globalVars.qual1block6mirs:
                    activeMirs.append(items)
            elif qualName =="2" and blockName =="1":
                for items in globalVars.qual2block1mirs:
                    activeMirs.append(items)
            elif qualName =="2" and blockName =="2":
                for items in globalVars.qual2block2mirs:
                    activeMirs.append(items)
            elif qualName =="3" and blockName =="1":
                for items in globalVars.qual3block1mirs:
                    activeMirs.append(items)
            elif qualName =="3" and blockName =="2":
                for items in globalVars.qual3block2mirs:
                    activeMirs.append(items)
            elif qualName =="3" and blockName =="3":
                for items in globalVars.qual3block3mirs:
                    activeMirs.append(items)
        
                    



            des_col=int(newendDate)+1

            
            
            total_index = 0

            mirsIndex = 0

            

            # print(tab5onleavelist[0])
            # print(tab5onleavelist[1])
            # print(tab5onleavelist[2])
            # for person in tab5onleavelist:
            #     print(person)
           
            
            
            
            
        
        
            q1check = False
            global colCheck
            colCheck = False


            #COLUMN ITERATOR LOOP
            while col_index < des_col:
                colCheck = False
                tab5leavecheck = False
                tab5onleavelist = []
                tab5leaverow = 5
                leavecellref = sheet.cell(row=tab5leaverow,column=col_index)
                while leavecellref.value !="LEAVE":
                    tab5leaverow+=1
                    leavecellref = sheet.cell(row=tab5leaverow,column=col_index)
                if leavecellref.value =="LEAVE":
                    tab5leaverow+=1
                    leavecellref = sheet.cell(row=tab5leaverow,column=col_index)
                while leavecellref.value:
                    person = str(leavecellref.value)
                    print(person)
                    tab5onleavelist.append(person)
                    tab5leaverow+=1
                    leavecellref = sheet.cell(row=tab5leaverow,column=col_index)

                for person in tab5onleavelist:
                    if instructor == person:
                        tab5leavecheck = True 
                
                #IF THERES A VALUE IN THE CELL
                while cellref.value:
                    colCheck = False
                    print(cellref.value)
                    if cellref.value == "Navy Unique":
                        sheet.insert_rows(row_index,5)
                        if col_index==1:
                            colCheck = True
                        else:
                            col_index-=1
                        # row_index-=5
                        cellref = sheet.cell(row=row_index,column=col_index)
                        # cellref.value="NEW"
                        # print(cellref.value)
                        workbook.save(filename=workbook_Title)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        break
                    # print(str(cellref.value))
                    if str(cellref.value) == "FAMILY DAY":
                        col_index+=1
                        des_col+=1
                        cellref = sheet.cell(row=row_index,column=col_index)
                        None
                    
                    elif str(cellref.value)!= "FAMILY DAY" and str(qualName)=="FAMILY DAY":
                        while cellref.value !='LEAVE':
                            cellref.value = "FAMILY DAY"
                            cellref.font = Font(color="FF0000")
                            cellref.fill = PatternFill("solid","FFFFFF")
                            cellref.border = thin_border_all_grey
                            row_index+=1
                            cellref = sheet.cell(row=row_index,column=col_index)

                    #First: IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK IS LESS THAN OR EQUAL AND THE CLASS NUM ENTERED IS GREATER
                    elif(int(qualName) <= int(cellref.value[1]) and int(blockName) <= int(cellref.value[10]) and int(qualName) == 1 and int(blockName)==1):
                        print("Start Date: "+str(newstartDate))
                        print("Start Date Type: "+str(type(newstartDate)))
                        if newstartDate!="01" and (int(classNum) >= int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999"):
                            sheet.insert_rows(5,5)
                            q1check = True
                            for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(newstartDate)-1):
                                for cell in rows:
                                    cell.value = "Q1  Block:1 Class:99999"
                                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid",fgColor="white")
                                    cell.border = thin_border_all_grey
                                    cell.font = Font(color="FFFFFF")
                            # row_index=4
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                        elif newstartDate!="01" and (int(classNum) < int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999"):
                            sheet.insert_rows(5,5)
                            q1check = True
                            for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(newstartDate)-1):
                                for cell in rows:
                                    cell.value = "Q1  Block:1 Class:99999"
                                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid",fgColor="white")
                                    cell.border = thin_border_all_grey
                                    cell.font = Font(color="FFFFFF")
                                    
                            row_index=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                        elif newstartDate=="1" and int(classNum) > int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999" :
                            sheet.insert_rows(5,5)
                            row_index=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        elif newstartDate=="1" and int(classNum) < int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999" :
                            row_index+=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                        #Second:IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK NUM IS GREATER OR EQUAL 
                    elif(int(qualName) <= int(cellref.value[1]) and int(blockName) > int(cellref.value[10])):
                        row_index+=5
                        print("Row index: " + str(row_index))
                        print(cellref.value)
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        

                        #Third:IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK IS LESS THAN OR EQUAL
                    elif(int(qualName) <= int(cellref.value[1]) and int(blockName) <= int(cellref.value[10])):
                        if int(classNum) > int(cellref.value[18:23]):
                            print("Cell Coordinate: "+str(cellref))
                            print("Last Blank Cell Coordinate: "+str(blank_cellref4))
                            sheet.insert_rows(row_index,5)
                            workbook.save(filename=workbook_Title)
                            newrowref = cellref.row
                            print("Newrowref: "+str(newrowref))
                            row_index=newrowref-5
                            print("row index: "+str(row_index))
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        else:

                            row_index+=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)


                        #FOURTH: IF THE ENTERED QUAL IS GREATER THAN THE CELL VALUE
                    elif(int(qualName) > int(cellref.value[1])):
                        row_index+=5
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)


                    elif(int(qualName)<=int(cellref.value[1])):
                        sheet.move_range(""+str(cellref.coordinate)+":"+str(blank_cellref4.coordinate)+"", rows=5)
                        newrowref = cellref.row
                        row_index=newrowref-5
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        
                else:
                    if str(qualName)=="FAMILY DAY":
                        while cellref.value != 'LEAVE':
                            cellref.value = "FAMILY DAY"
                            cellref.font = Font(color="FF0000")
                            cellref.fill = PatternFill("solid","FFFFFF")
                            cellref.border = thin_border_all_grey
                            row_index+=1
                            cellref = sheet.cell(row=row_index,column=col_index)
                        
                    elif(int(qualName)==1 and int(blockName)==1 and newstartDate!="01" and q1check ==False):
                        sheet.insert_rows(5,5)
                        print("Row Index: "+str(row_index))
                        q1check =True
                        for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(newstartDate)-1):
                                for cell in rows:
                                    cell.value = "Q1  Block:1 Class:99999"
                                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid",fgColor="white")
                                    cell.border = thin_border_all_grey
                                    cell.font = Font(color="FFFFFF")
                                
                                    print("col index: "+str(col_index))
                    # row_index=4
                    print("Row Index: "+str(row_index))
                    cellref = sheet.cell(row=row_index,column=col_index)
                    blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                    blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                    blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                    blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                    if str(qualName)=="FAMILY DAY":
                        None    
                    else:
                        cellref.value="Q"+str(qualName)+" " +" Block:"+str(blockName)+" Class:"+str(classNum)
                        cellref.font = ft1
                        cellref.alignment=align
                    # print("col index: "+str(col_index))
                        #TITLE COLOR CODING
                        #QUAL 1
                        cellref.border = thick_border
                        if (qualName == "1" and blockName == "1"):
                            cellref.fill = PatternFill("solid", fgColor="00CCFF")
                        elif (qualName == "1"and blockName == "2"):
                            cellref.fill = PatternFill("solid", fgColor="33CCCC")
                        elif (qualName == "1" and blockName == "3"):
                            cellref.fill = PatternFill("solid", fgColor="92D050")
                        elif (qualName == "1"and blockName == "4"):
                            cellref.fill = PatternFill("solid", fgColor="FFFF00")
                        elif (qualName == "1" and blockName == "5"):
                            cellref.fill = PatternFill("solid", fgColor="FFC000")
                        elif (qualName == "1"and blockName == "6"):
                            cellref.fill = PatternFill("solid", fgColor="FF0000")

                        #QUAL2
                        elif (qualName == "2" and blockName == "1"):
                            cellref.fill = PatternFill("solid", fgColor="00B0F0")
                        elif (qualName == "2"and blockName == "2"):
                            cellref.fill = PatternFill("solid", fgColor="92D050")

                        #QUAL3
                        elif (qualName == "3" and blockName == "1"):
                            cellref.fill = PatternFill("solid", fgColor="FFFF00")
                        elif (qualName == "3"and blockName == "2"):
                            cellref.fill = PatternFill("solid", fgColor="9BBB59")
                        elif (qualName == "3"and blockName == "3"):
                            cellref.fill = PatternFill("solid", fgColor="FFC000")
                        


                        #SETTING BLANK VALUES BORDERS
                        blank_cellref.border = thin_border_sides
                        blank_cellref2.border = thin_border_sides
                        blank_cellref3.border = thin_border_sides
                        blank_cellref4.border = thin_border_sides_Bottom

                        #SETTING BLANK VALUES
                        if tab5leavecheck == True:
                            # messagebox.showwarning(title="Instructor on Leave!", message="The entered instructor is on leave for one of the days")
                            blank_cellref.value="Instructor: Instructor is on leave"
                            tab5leavecheck ==False
                        else:
                            blank_cellref.value="Instructor: "+instructor
                        blank_cellref.font = ft1
                        blank_cellref.alignment=align
                        blank_cellref2.value="------"
                        blank_cellref2.font = ft1
                        blank_cellref2.alignment=align
                        blank_cellref3.value="------"
                        blank_cellref3.font = ft1
                        blank_cellref3.alignment=align
                        if mirsIndex > len(activeMirs)-1:
                            blank_cellref4.value = "NonPop"
                        else:
                            blank_cellref4.value=str(activeMirs[mirsIndex])
                        blank_cellref4.font = ft1
                        blank_cellref4.alignment=align
                    total_index+=1
                   
                    # print("Title = " + cellref.coordinate)
                    # print("blank = " + blank_cellref.coordinate)
                
                if colCheck ==True:
                    None
                else:
                    col_index+=1
                mirsIndex+=1
                # print("column index: "+str(col_index))   


                #This line controls if they shoot up the rows or nor
                # row_index=4
                # print("row index"+str(row_index))   
                cellref = sheet.cell(row=row_index,column=col_index)
                blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                print("Row Index: "+str(row_index))
                print("cellref: "+str(cellref))
                workbook.save(filename=workbook_Title)


                
                
                global submitTotal
                # submittedLabel = Label(gui.tab1frame3,text = "Submitted ",font="Helvetica 10 bold", fg="black",bg="grey26")
                # datesLabel = Label(gui.tab1frame3,text = "Total Number of Dates Affected: "+str(submitTotal),font="Helvetica 8 bold", fg="black",bg="grey26")
                # submittedLabel.grid(row=1,column=0,ipadx=50)
                # datesLabel.grid(row = 2, column=0)
                submitTotal+=1

            # if monthCheck == True:
            #     global sheet
            #     sheet = workbook.active[2]
            #     startDate = 1
            #     endDate = endDate = gui.e5.get()
            #     endDate = endDate[3:5]
            #     second_add_qual()
            # else:
            #     None
        except():
            print(traceback.format_exc())
            # messagebox.showwarning(title="Error Occured", message="something went wrong in Second Add Month. Check your entries and try again")
            # if path.exists("curtis_program - Testing Seperation/QTDESIGNER/textFiles/errors.txt") == TRUE:
            #     ct = datetime.datetime.now() 
            #     with open("curtis_program - Testing Seperation/QTDESIGNER/textFiles/errors.txt", "a") as file:
            #         file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))
            # else:
            #     ct = datetime.datetime.now() 
            #     with open("curtis_program - Testing Seperation/QTDESIGNER/textFiles/errors.txt", "x") as file:
            #         file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))




    def second_Month_Check(self):
        if monthCheck == True:
            sheets = workbook.sheetnames
            # for i in sheets:
            #     print(i)
            global sheet
            print(str(sheet))
            sheet = workbook[sheets[sheetIndex+1]]
            print(str(sheet.title))
            # sheet = workbook.active
            global newstartDate
            newstartDate = "1"
            global newendDate
            newendDate = endDate = globalVars.page1EndDate
            newendDate = endDate[3:5]
            Worker2.second_add_qual()
        else:
            None
            self.finished.emit() 


    def add_qual(self):
        try:

        
    #NONE DECLARED ERROR LABELS USED
            global fileerrorLabel
            global qualerrorLabel
            global blockerrorLabel
            global starterrorLabel
            global enderrorLabel
            global qual1block1mirs
            global qual1block2mirs      
            sheet
        #SET COLUMN WIDTH
            # for col in range(1,32):
            #     column_letter = get_column_letter(col)
            #     # print(column_letter)
            #     sheet.column_dimensions[column_letter].width = 42

        #SET TITLE
            sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=32)
            sheet.row_dimensions[3].height = 60
            title_Cell = sheet['A3']
            title_Cell.border = thick_border
            monthTitle = str(sheet.title)
            # print(monthTitle)
            title_Cell.value = monthTitle+" ELECTRICAL SYSTEMS INSTRUCTOR SCHEDULE"
            title_Cell.font = Font(size=60,name="Times New Roman",bold=True)

            
            # global startMonth
            #SET DATES
            # for i in range(1,32):
            #     datecellref=sheet.cell(row=1, column=i)
            #     datecellref.fill = PatternFill("solid", fgColor="DDDDDD")
                
            #     if(i<10):
            #         datecellref.value=str(startMonth)+"/0"+str(i)+"/2021"
            #     else:
            #         datecellref.value=str(startMonth)+"/"+str(i)+"/2021"


            #SET DIVIDER      

            for i in range(1,32):
                datecellref2=sheet.cell(row=4, column=i)
                datecellref2.fill = PatternFill("solid", fgColor="000000")
                datecellref2.border = thick_border_blue_topBottom
                datecellref2.value="blank"  
                # gui.tab1.update_idletasks() 


            #GETTING INPUT VALUES FROM USER
            instructor = globalVars.page1e2
            qualName = globalVars.page1Qual
            if str(qualName) == "FAMILY DAY":
                None
            else:
                qualName = globalVars.page1Qual
                if int(qualName) > 3 and str(qualName) !="FAMILY DAY":
                    # qualerrorLabel = Label(gui.tab1frame2,text = "Qual Num Doesn't Exist",font="Helvetica 10 bold", fg="red")
                    # qualerrorLabel.grid(row=11,column=0)
                    # gui.e2.delete(0, END)
                    return 
                blockName = globalVars.page1Block
                if int(blockName) > 6 and str(qualName) !="FAMILY DAY":
                    # blockerrorLabel = Label(gui.tab1frame2,text = "Block Number Doesn't Exist",font="Helvetica 10 bold", fg="red")
                    # blockerrorLabel.grid(row=11,column=0)
                    # gui.e3.delete(0, END)
                    return
            startDate = globalVars.page1StartDate
            startDate = startDate[3:5]
            print(str(startDate))
            if int(startDate) > 32:
                # starterrorLabel = Label(gui.tab1frame2,text = "Start Date Number too big",font="Helvetica 10 bold", fg="red")
                # starterrorLabel.grid(row=11,column=0)
                # gui.e4.delete(0, END)
                return
            if int(endDate) > 32:
                # enderrorLabel = Label(gui.tab1frame2,text = "End Date Number too big",font="Helvetica 10 bold", fg="red")
                # enderrorLabel.grid(row=11,column=0)
                # gui.e5.delete(0, END)
                return

            classNum = globalVars.page1e3

            

            #Start Date
            col_index = int(startDate)
            row_index=5
            # print("First row index: "+str(row_index))

            #THE DASHED CELLS ITERATORS
            # blank_col_index = col_index
            # blank_row_index=row_index+1
            
            #TITLE CELL
            cellref = sheet.cell(row=row_index,column=col_index)
            cellref.border = thick_border
            blockcheckcell = sheet.cell(row=row_index, column = 1)
        

            #DASHED CELLS
            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)


            #Setting Styles
            # cellref.font=ft1
            # cellref.alignment=align
            # blank_cellref.font=ft1
            # blank_cellref.alignment=align
            # blank_cellref2.font=ft1
            # blank_cellref2.alignment=align
            # blank_cellref3.font=ft1
            # blank_cellref3.alignment=align
            # blank_cellref4.font=ft1
            # blank_cellref4.alignment=align

            #End Date
            #check if months match


            #SETTING ACTIVE MIRS
            activeMirs=[]
            
            if qualName =="1" and blockName == "1":
                for items in globalVars.qual1block1mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="2":
                for items in globalVars.qual1block2mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="3":
                for items in globalVars.qual1block3mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="4":
                for items in globalVars.qual1block4mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="5":
                for items in globalVars.qual1block5mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="6":
                for items in globalVars.qual1block6mirs:
                    activeMirs.append(items)
            elif qualName =="2" and blockName =="1":
                for items in globalVars.qual2block1mirs:
                    activeMirs.append(items)
            elif qualName =="2" and blockName =="2":
                for items in globalVars.qual2block2mirs:
                    activeMirs.append(items)
            elif qualName =="3" and blockName =="1":
                for items in globalVars.qual3block1mirs:
                    activeMirs.append(items)
            elif qualName =="3" and blockName =="2":
                for items in globalVars.qual3block2mirs:
                    activeMirs.append(items)
            elif qualName =="3" and blockName =="3":
                for items in globalVars.qual3block3mirs:
                    activeMirs.append(items)
        
                    



            des_col=int(endDate)+1

            
            
            total_index = 0

            mirsIndex = 0

            

            # print(tab5onleavelist[0])
            # print(tab5onleavelist[1])
            # print(tab5onleavelist[2])
            # for person in tab5onleavelist:
            #     print(person)
           
            
            
            
            
        
        
            q1check = False
            global colCheck
            colCheck = False


            #COLUMN ITERATOR LOOP
            while col_index < des_col:
                # print(col_index)
                # column_letter = get_column_letter(col_index)
                # print(column_letter)
                # if sheet.column_dimensions[column_letter].width < 1:
                #     col_index+=1
                #     continue
                colCheck = False
                tab5leavecheck = False
                tab5onleavelist = []
                tab5leaverow = 5
                leavecellref = sheet.cell(row=tab5leaverow,column=col_index)
                while leavecellref.value !="LEAVE":
                    tab5leaverow+=1
                    leavecellref = sheet.cell(row=tab5leaverow,column=col_index)
                if leavecellref.value =="LEAVE":
                    tab5leaverow+=1
                    leavecellref = sheet.cell(row=tab5leaverow,column=col_index)
                while leavecellref.value:
                    person = str(leavecellref.value)
                    print(person)
                    tab5onleavelist.append(person)
                    tab5leaverow+=1
                    leavecellref = sheet.cell(row=tab5leaverow,column=col_index)

                for person in tab5onleavelist:
                    if instructor == person:
                        tab5leavecheck = True 
                
                #IF THERES A VALUE IN THE CELL
                while cellref.value:
                    colCheck = False
                    print(cellref.value)
                    if cellref.value == "Navy Unique":
                        sheet.insert_rows(row_index,5)
                        if col_index==1:
                            colCheck = True
                        else:
                            col_index-=1
                        # row_index-=5
                        cellref = sheet.cell(row=row_index,column=col_index)
                        # cellref.value="NEW"
                        # print(cellref.value)
                        workbook.save(filename=workbook_Title)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        break
                    # print(str(cellref.value))
                    if str(cellref.value) == "FAMILY DAY":
                        col_index+=1
                        des_col+=1
                        cellref = sheet.cell(row=row_index,column=col_index)
                        None
                    
                    elif str(cellref.value)!= "FAMILY DAY" and str(qualName)=="FAMILY DAY":
                        while cellref.value !="LEAVE":
                            cellref.value = "FAMILY DAY"
                            cellref.font = Font(color="FF0000")
                            cellref.fill = PatternFill("solid","FFFFFF")
                            cellref.border = thin_border_all_grey
                            row_index+=1
                            cellref = sheet.cell(row=row_index,column=col_index)

                    #First: IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK IS LESS THAN OR EQUAL AND THE CLASS NUM ENTERED IS GREATER
                    elif(int(qualName) <= int(cellref.value[1]) and int(blockName) <= int(cellref.value[10]) and int(qualName) == 1 and int(blockName)==1):
                        print("Start Date: "+str(startDate))
                        print("Start Date Type: "+str(type(startDate)))
                        if startDate!="01" and (int(classNum) >= int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999"):
                            sheet.insert_rows(5,5)
                            q1check = True
                            for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(startDate)-1):
                                for cell in rows:
                                    cell.value = "Q1  Block:1 Class:99999"
                                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid",fgColor="white")
                                    cell.border = thin_border_all_grey
                                    cell.font = Font(color="FFFFFF")
                            # row_index=4
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                        elif startDate!="01" and (int(classNum) < int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999"):
                            sheet.insert_rows(5,5)
                            q1check = True
                            for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(startDate)-1):
                                for cell in rows:
                                    cell.value = "Q1  Block:1 Class:99999"
                                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid",fgColor="white")
                                    cell.border = thin_border_all_grey
                                    cell.font = Font(color="FFFFFF")
                                    
                            row_index=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                        elif startDate=="1" and int(classNum) > int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999" :
                            sheet.insert_rows(5,5)
                            row_index=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        elif startDate=="1" and int(classNum) < int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999" :
                            row_index+=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                        #Second:IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK NUM IS GREATER OR EQUAL 
                    elif(int(qualName) <= int(cellref.value[1]) and int(blockName) > int(cellref.value[10])):
                        row_index+=5
                        print("Row index: " + str(row_index))
                        print(cellref.value)
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        

                        #Third:IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK IS LESS THAN OR EQUAL
                    elif(int(qualName) <= int(cellref.value[1]) and int(blockName) <= int(cellref.value[10])):
                        if int(classNum) > int(cellref.value[18:23]):
                            print("Cell Coordinate: "+str(cellref))
                            print("Last Blank Cell Coordinate: "+str(blank_cellref4))
                            sheet.insert_rows(row_index,5)
                            workbook.save(filename=workbook_Title)
                            newrowref = cellref.row
                            print("Newrowref: "+str(newrowref))
                            row_index=newrowref-5
                            print("row index: "+str(row_index))
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        else:

                            row_index+=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)


                        #FOURTH: IF THE ENTERED QUAL IS GREATER THAN THE CELL VALUE
                    elif(int(qualName) > int(cellref.value[1])):
                        row_index+=5
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)


                    elif(int(qualName)<=int(cellref.value[1])):
                        sheet.move_range(""+str(cellref.coordinate)+":"+str(blank_cellref4.coordinate)+"", rows=5)
                        newrowref = cellref.row
                        row_index=newrowref-5
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        
                else:
                    if str(qualName)=="FAMILY DAY":
                        while cellref.value !='LEAVE':
                            cellref.value = "FAMILY DAY"
                            cellref.font = Font(color="FF0000")
                            cellref.fill = PatternFill("solid","FFFFFF")
                            cellref.border = thin_border_all_grey
                            row_index+=1
                            cellref = sheet.cell(row=row_index,column=col_index)
                        
                    elif(int(qualName)==1 and int(blockName)==1 and startDate!="01" and q1check ==False):
                        sheet.insert_rows(5,5)
                        print("Row Index: "+str(row_index))
                        q1check =True
                        for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(startDate)-1):
                                for cell in rows:
                                    cell.value = "Q1  Block:1 Class:99999"
                                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid",fgColor="white")
                                    cell.border = thin_border_all_grey
                                    cell.font = Font(color="FFFFFF")
                                
                                    print("col index: "+str(col_index))
                    # row_index=4
                    print("Row Index: "+str(row_index))
                    cellref = sheet.cell(row=row_index,column=col_index)
                    blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                    blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                    blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                    blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                    if str(qualName)=="FAMILY DAY":
                        None    
                    else:
                        cellref.value="Q"+str(qualName)+" " +" Block:"+str(blockName)+" Class:"+str(classNum)
                        cellref.font = ft1
                        cellref.alignment=align
                    # print("col index: "+str(col_index))
                        #TITLE COLOR CODING
                        #QUAL 1
                        cellref.border = thick_border
                        if (qualName == "1" and blockName == "1"):
                            cellref.fill = PatternFill("solid", fgColor="00CCFF")
                        elif (qualName == "1"and blockName == "2"):
                            cellref.fill = PatternFill("solid", fgColor="33CCCC")
                        elif (qualName == "1" and blockName == "3"):
                            cellref.fill = PatternFill("solid", fgColor="92D050")
                        elif (qualName == "1"and blockName == "4"):
                            cellref.fill = PatternFill("solid", fgColor="FFFF00")
                        elif (qualName == "1" and blockName == "5"):
                            cellref.fill = PatternFill("solid", fgColor="FFC000")
                        elif (qualName == "1"and blockName == "6"):
                            cellref.fill = PatternFill("solid", fgColor="FF0000")

                        #QUAL2
                        elif (qualName == "2" and blockName == "1"):
                            cellref.fill = PatternFill("solid", fgColor="00B0F0")
                        elif (qualName == "2"and blockName == "2"):
                            cellref.fill = PatternFill("solid", fgColor="92D050")

                        #QUAL3
                        elif (qualName == "3" and blockName == "1"):
                            cellref.fill = PatternFill("solid", fgColor="FFFF00")
                        elif (qualName == "3"and blockName == "2"):
                            cellref.fill = PatternFill("solid", fgColor="9BBB59")
                        elif (qualName == "3"and blockName == "3"):
                            cellref.fill = PatternFill("solid", fgColor="FFC000")
                        


                        #SETTING BLANK VALUES BORDERS
                        blank_cellref.border = thin_border_sides
                        blank_cellref2.border = thin_border_sides
                        blank_cellref3.border = thin_border_sides
                        blank_cellref4.border = thin_border_sides_Bottom

                        #SETTING BLANK VALUES
                        if tab5leavecheck == True:
                            # messagebox.showwarning(title="Instructor on Leave!", message="The entered instructor is on leave for one of the days")
                            blank_cellref.value="Instructor: Instructor is on leave"
                            tab5leavecheck ==False
                        else:
                            blank_cellref.value="Instructor: "+instructor
                        blank_cellref.font = ft1
                        blank_cellref.alignment=align
                        blank_cellref2.value="------"
                        blank_cellref2.font = ft1
                        blank_cellref2.alignment=align
                        blank_cellref3.value="------"
                        blank_cellref3.font = ft1
                        blank_cellref3.alignment=align
                        if mirsIndex > len(activeMirs)-1:
                            blank_cellref4.value = "NonPop"
                        else:
                            blank_cellref4.value=str(activeMirs[mirsIndex])
                        blank_cellref4.font = ft1
                        blank_cellref4.alignment=align
                    total_index+=1
                   
                    # print("Title = " + cellref.coordinate)
                    # print("blank = " + blank_cellref.coordinate)
                
                if colCheck ==True:
                    None
                else:
                    col_index+=1
                mirsIndex+=1
                # print("column index: "+str(col_index))   


                #This line controls if they shoot up the rows or nor
                # row_index=4
                # print("row index"+str(row_index))   
                cellref = sheet.cell(row=row_index,column=col_index)
                blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                print("Row Index: "+str(row_index))
                print("cellref: "+str(cellref))
                workbook.save(filename=workbook_Title)


                
                
                global submitTotal
                # submittedLabel = Label(gui.tab1frame3,text = "Submitted ",font="Helvetica 10 bold", fg="black",bg="grey26")
                # datesLabel = Label(gui.tab1frame3,text = "Total Number of Dates Affected: "+str(submitTotal),font="Helvetica 8 bold", fg="black",bg="grey26")
                # submittedLabel.grid(row=1,column=0,ipadx=50)
                # datesLabel.grid(row = 2, column=0)
                # submitTotal+=1
                # gui.e2.delete(0, END)
                # gui.e3.delete(0, END)
                # gui.e4.delete(0, END)
                # gui.e5.delete(0, END)
                # gui.my_progress2['value']+=30
                
                # gui.tab1.update_idletasks()

            Worker2.second_Month_Check(self)
            # gui.my_progress2['value']=0
            
        except:
            print(traceback.format_exc())
            # messagebox.showwarning(title="Error Occured", message="something went wrong in ADD QUAL.\nCheck your entries and try again.\nFor a more detailed explanation check errors txt file")
            # if path.exists("curtis_program - Testing Seperation/QTDESIGNER/textFiles/errors.txt") == TRUE:
            #     ct = datetime.datetime.now() 
            #     with open("curtis_program - Testing Seperation/QTDESIGNER/textFiles/errors.txt", "a") as file:
            #         file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))
            # else:
            #     ct = datetime.datetime.now() 
            #     with open("curtis_program - Testing Seperation/QTDESIGNER/textFiles/errors.txt", "x") as file:
            #         file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))



    def existing_WORKBOOK(self):
        try:
            # finishedLabel = Label(gui.tab1frame3,text="FINISHED",font="Helvetica 10 bold", fg="grey26",bg="grey26").grid(row=4,column=0) 
            # global gui.my_progress2
            # gui.my_progress2.grid(row=0,column=0)
            global workbook_Title
            workbook_Title = globalVars.page1e1
            print(workbook_Title)
            global workbook
            workbook = load_workbook(filename=workbook_Title)
           
                
                # workbook.add_named_style(date_style)
            # print(path.exists(str(gui.e1.get())+".xlsx"))
            global sheet
            sheet = workbook.active
            print(workbook)
            global sheetIndex
            #Check Months
            global startMonth
            startMonth = globalVars.page1StartDate
            startMonth = startMonth[0:2]
            endMonth = globalVars.page1EndDate
            endMonth = endMonth[0:2]
            global endDate
            endDate = globalVars.page1EndDate
            endDate = endDate[3:5]
            global monthCheck

            # month = str(gui.e4.get())
            # print("Month: "+str(month[0:2]))
            print("Start Month: "+str(startMonth))
            if str(startMonth[0:2]) =="01":
                if'JAN' in workbook.sheetnames:
                    sheet = workbook["JAN"]
                    sheetIndex = 1
                elif 'Jan' not in workbook.sheetnames:
                    ws1 = workbook.create_sheet("JAN")
                    sheet = ws1
                    sheetIndex = 1

            if str(startMonth[0:2]) =="02":
                if'FEB' in workbook.sheetnames:
                    sheet = workbook["FEB"]
                    sheetIndex =2
                elif 'Feb' not in workbook.sheetnames:
                    ws1 = workbook.create_sheet("FEB")
                    sheet = ws1
                    sheetIndex = 2
            if str(startMonth[0:2]) =="03":
                if'MAR' in workbook.sheetnames:
                    sheet = workbook["MAR"]
                    sheetIndex = 3
                elif 'MAR' not in workbook.sheetnames:
                    ws1 = workbook.create_sheet("MAR")
                    sheet = ws1
                    sheetIndex = 3
            if str(startMonth[0:2]) =="04":
                if'APR' in workbook.sheetnames:
                    sheet = workbook["APR"]
                    sheetIndex = 4
                elif 'APR' not in workbook.sheetnames:
                    ws1 = workbook.create_sheet("APR")
                    sheet = ws1
                    sheetIndex = 4
            if str(startMonth[0:2]) =="05":
                if'MAY' in workbook.sheetnames:
                    sheet = workbook["MAY"]
                    sheetIndex = 5
                elif 'MAY' not in workbook.sheetnames:
                    ws1 = workbook.create_sheet("MAY")
                    sheet = ws1
                    sheetIndex = 5

            if str(startMonth[0:2]) =="06":
                if'JUNE' in workbook.sheetnames:
                    sheet = workbook["JUNE"]
                    sheetIndex = 6
                elif 'JUNE' not in workbook.sheetnames:
                    ws1 = workbook.create_sheet("JUNE")
                    sheet = ws1
                    sheetIndex = 6

            if str(startMonth[0:2]) =="07":
                if'JULY' in workbook.sheetnames:
                    sheet = workbook["JULY"]
                    sheetIndex = 7
                elif 'JULY' not in workbook.sheetnames:
                    ws1 = workbook.create_sheet("JULY")
                    sheet = ws1
                    sheetIndex = 7

            if str(startMonth[0:2]) =="08":
                if'AUG' in workbook.sheetnames:
                    sheet = workbook["AUG"]
                    sheetIndex = 8
                elif 'AUG' not in workbook.sheetnames:
                    ws1 = workbook.create_sheet("AUG")
                    sheet = ws1
                    sheetIndex = 8

            if str(startMonth[0:2]) =="09":
                if'SEPT' in workbook.sheetnames:
                    sheet = workbook["SEPT"]
                    sheetIndex = 9
                elif 'SEPT' not in workbook.sheetnames:
                    ws1 = workbook.create_sheet("SEPT")
                    sheet = ws1
                    sheetIndex = 9
            if str(startMonth[0:2]) =="10":
                if'OCT' in workbook.sheetnames:
                    sheet = workbook["OCT"]
                    sheetIndex = 10
                elif 'OCT' not in workbook.sheetnames:
                    ws1 = workbook.create_sheet("OCT")
                    sheet = ws1
                    sheetIndex = 10
            if str(startMonth[0:2]) =="11":
                if'NOV' in workbook.sheetnames:
                    sheet = workbook["NOV"]
                    sheetIndex = 11
                elif 'NOV' not in workbook.sheetnames:
                    ws1 = workbook.create_sheet("NOV")
                    sheet = ws1
                    sheetIndex = 11

            if str(startMonth[0:1]) =="12":
                if'DEC' in workbook.sheetnames:
                    sheet = workbook["DEC"]
                    sheetIndex = 12
                elif 'DEC' not in workbook.sheetnames:
                    ws1 = workbook.create_sheet("DEC")
                    sheet = ws1
                    sheetIndex = 12

            if(startMonth == endMonth):
                None
                Worker2.add_qual(self)
                # self.finished.emit() 
                
            elif(startMonth != endMonth):
                endDate = 30
                monthCheck =True
                Worker2.add_qual(self)
                # self.finished.emit()
            # finishedLabel = Label(gui.tab1frame3,text="FINISHED",font="Helvetica 10 bold", fg="white",bg="grey26").grid(row=4,column=0) 
            # gui.my_progress.stop()
      
        except():
            print(traceback.format_exc())
            # messagebox.showwarning(title="Error Occured", message="something went wrong in EXISTING BOOK. Check your entries and try again")
            # if path.exists("curtis_program - Testing Seperation/QTDESIGNER/textFiles/errors.txt") == TRUE:
            #     ct = datetime.datetime.now() 
            #     with open("curtis_program - Testing Seperation/QTDESIGNER/textFiles/errors.txt", "a") as file:
            #         file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))
            # else:
            #     ct = datetime.datetime.now() 
            #     with open("curtis_program - Testing Seperation/QTDESIGNER/textFiles/errors.txt", "x") as file:
            #         file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))
   

    