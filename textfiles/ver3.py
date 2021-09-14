from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell import cell
from openpyxl.cell.cell import ERROR_CODES
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.dimensions import SheetDimension
from openpyxl.utils import get_column_letter
from tkinter import *
from tkinter import ttk
from ttkthemes import ThemedTk
import time
import os.path
from os import path


thick_border = Border(left=Side(style='thick'), 
                     right=Side(style='thick'), 
                     top=Side(style='thick'), 
                     bottom=Side(style='thick'))

thick_border_blue = Border(left=Side(style='thick',color='0066CC'), 
                     right=Side(style='thick',color='0066CC')   
                  ) 

thick_border_blue_topBottom = Border(top=Side(style='thick',color='0066CC'), 
                     bottom=Side(style='thick',color='0066CC')   
                  )                    

thin_border_all = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

thin_border_sides = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style=None), 
                     bottom=Side(style=None))


thin_border_sides_Bottom = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style=None), 
                     bottom=Side(style='thin'))


#GLOBAL ERROR VARIABLES
fileerrorLabel = None
qualerrorLabel = None
blockerrorLabel = None
starterrorLabel = None
enderrorLabel = None
monthCheck = False

def changeOnHover(button, colorOnHover, colorOnLeave):
  
    # adjusting backgroung of the widget
    # background on entering widget
    button.bind("<Enter>", func=lambda e: button.config(
        background=colorOnHover))
  
    # background color on leving widget
    button.bind("<Leave>", func=lambda e: button.config(
        background=colorOnLeave))

#============================================================================================================================#
#============================================================================================================================#
#============================================================================================================================#



def existing_WORKBOOK():
    global workbook_Title
    workbook_Title = e1.get()
    global workbook
    if path.exists(str(e1.get())+".xlsx") == True: 
        workbook = load_workbook(filename=workbook_Title+".xlsx")
        # workbook.add_named_style(date_style)
    else:
        fileerrorLabel = Label(tab1,text="File Does Not Exist",font="Helvetica 10 bold", fg="red")
        fileerrorLabel.grid(row=11,column=0)
        return
    # print(path.exists(str(e1.get())+".xlsx"))
    global sheet
    sheet = workbook.active
    print(workbook)
    global sheetIndex
    #Check Months
    startMonth = e4.get()
    startMonth = startMonth[0:2]
    endMonth = e5.get()
    endMonth = endMonth[0:2]
    global endDate
    endDate = e5.get()
    endDate = endDate[3:5]
    global monthCheck

    # month = str(e4.get())
    # print("Month: "+str(month[0:2]))
    print("Start Month: "+str(startMonth))
    if str(startMonth[0:2]) =="01":
        if'JAN' in workbook.sheetnames:
            sheet = workbook["JAN"]
            sheetIndex = 1
        elif 'Jan' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("JAN")
            sheet = ws1
            sheetIndex = 1

    if str(startMonth[0:2]) =="02":
        if'FEB' in workbook.sheetnames:
            sheet = workbook["FEB"]
            sheetIndex =2
        elif 'Feb' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("FEB")
            sheet = ws1
            sheetIndex = 2
    if str(startMonth[0:2]) =="03":
        if'MAR' in workbook.sheetnames:
            sheet = workbook["MAR"]
            sheetIndex = 3
        elif 'MAR' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("MAR")
            sheet = ws1
            sheetIndex = 3
    if str(startMonth[0:2]) =="04":
        if'APR' in workbook.sheetnames:
            sheet = workbook["APR"]
            sheetIndex = 4
        elif 'APR' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("APR")
            sheet = ws1
            sheetIndex = 4
    if str(startMonth[0:2]) =="05":
        if'MAY' in workbook.sheetnames:
            sheet = workbook["MAY"]
            sheetIndex = 5
        elif 'MAY' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("MAY")
            sheet = ws1
            sheetIndex = 5

    if str(startMonth[0:2]) =="06":
        if'JUNE' in workbook.sheetnames:
            sheet = workbook["JUNE"]
            sheetIndex = 6
        elif 'JUNE' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("JUNE")
            sheet = ws1
            sheetIndex = 6

    if str(startMonth[0:2]) =="07":
        if'JULY' in workbook.sheetnames:
            sheet = workbook["JULY"]
            sheetIndex = 7
        elif 'JULY' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("JULY")
            sheet = ws1
            sheetIndex = 7

    if str(startMonth[0:2]) =="08":
        if'AUG' in workbook.sheetnames:
            sheet = workbook["AUG"]
            sheetIndex = 8
        elif 'AUG' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("AUG")
            sheet = ws1
            sheetIndex = 8

    if str(startMonth[0:2]) =="09":
        if'SEPT' in workbook.sheetnames:
            sheet = workbook["SEPT"]
            sheetIndex = 9
        elif 'SEPT' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("SEPT")
            sheet = ws1
            sheetIndex = 9
    if str(startMonth[0:2]) =="10":
        if'OCT' in workbook.sheetnames:
            sheet = workbook["OCT"]
            sheetIndex = 10
        elif 'OCT' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("OCT")
            sheet = ws1
            sheetIndex = 10
    if str(startMonth[0:2]) =="11":
        if'NOV' in workbook.sheetnames:
            sheet = workbook["NOV"]
            sheetIndex = 11
        elif 'NOV' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("NOV")
            sheet = ws1
            sheetIndex = 11

    if str(startMonth[0:1]) =="12":
        if'DEC' in workbook.sheetnames:
            sheet = workbook["DEC"]
            sheetIndex = 12
        elif 'DEC' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("DEC")
            sheet = ws1
            sheetIndex = 12

    if(startMonth == endMonth):
        None
        add_qual()
    elif(startMonth != endMonth):
        endDate = 30
        monthCheck =True
        add_qual()  




def new_WORKBOOK():
    global workbook_Title
    workbook_Title = newe1.get()
    if path.exists(str(newe1.get())+".xlsx") == TRUE:
        fileerrorLabel = Label(tab2,text="File Already Exists",font="Helvetica 10 bold", fg="red")
        fileerrorLabel.grid(row=5,column=0)
        return
    else:
        global workbook 
        workbook = Workbook()
        # sheet = workbook.active
        ws1 = workbook.create_sheet("JAN")
        ws2 = workbook.create_sheet("FEB")
        ws3 = workbook.create_sheet("MAR")
        ws4 = workbook.create_sheet("APR")
        ws5 = workbook.create_sheet("MAY")
        ws6 = workbook.create_sheet("JUNE")
        ws7 = workbook.create_sheet("JULY")
        ws8 = workbook.create_sheet("AUG")
        ws9 = workbook.create_sheet("SEPT")
        ws10 = workbook.create_sheet("OCT")
        ws11 = workbook.create_sheet("NOV")
        ws12 = workbook.create_sheet("DEC")

        #ADD DAYS TO THE SHEET
    # dayList = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
  
    # for sheet in workbook.worksheets:
    #     col_index=1
    #     day_index=int(newe2.get())
    #     while col_index<31:
    #         daycellref=sheet.cell(row=1, column=col_index) 
    #         daycellref.value = (dayList[day_index])
    #         day_index+=1
    #         col_index+=1
    #         if day_index ==6 and col_index<31:
    #             daycellref.value = (dayList[day_index])
    #             day_index = 0

    #     # workbook.save(filename=workbook_Title+".xlsx")
    #     for i in range(1,31):
    #         daycellref=sheet.cell(row=1, column=i)
    #         if daycellref.value == "Saturday":
    #             sheet.delete_cols(i)
    #             # workbook.save(filename=workbook_Title+".xlsx")
    #         elif daycellref.value == "Sunday":
    #             sheet.delete_cols(i)
    #             # workbook.save(filename=workbook_Title+".xlsx")


    # col_index = 1
    # day_index = 0
    # while col_index<31:
    #     for i in dayList:
    #         daycellref=sheet.cell(row=1, column=col_index) 
    #         daycellref.value = i
    #         print(i)
    #         print(str(daycellref.value))
    #         col_index+=1
    #         workbook.save(filename=workbook_Title+".xlsx")

    # def add_days():
    #     for i in dayList:
    #         daycellref=sheet.cell(row=1, column=col_index) 
    #         daycellref.value = i
    #         print(i)
    #         print(str(daycellref.value))
    #         workbook.save(filename=workbook_Title+".xlsx")
    #         day_index+=1
    #         col_index+=1
    #         if day_index == 6 and col_index<31:
    #             daycellref.value = i
    #             i=0
    #             day_index == 0
                
           
             
    # for col in range(1,31):
    #         column_letter = get_column_letter(col)
    #         # print(column_letter)
    #         sheet.column_dimensions[column_letter].width = 42

    #SET TITLE
    # sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=31)
    # sheet.row_dimensions[2].height = 60
    # title_Cell = sheet['A2']
    # title_Cell.border = thick_border
    # title_Cell.value = "SEPTEMBER ELECTRICAL SYSTEMS INSTRUCTOR SCHEDULE"
    # title_Cell.font = Font(size=60,name="Times New Roman",bold=True)

        

        #SET DATES
    # for i in range(1,31):
    #     datecellref=sheet.cell(row=1, column=i)
    #     datecellref.fill = PatternFill("solid", fgColor="DDDDDD")
        
    #     if(i<10):
    #         datecellref.value="01/0"+str(i)+"/2021"
    #     else:
    #         datecellref.value="01/"+str(i)+"/2021"

    #     #SET DIVIDER      

    # for i in range(1,31):
    #     datecellref2=sheet.cell(row=3, column=i)
    #     datecellref2.fill = PatternFill("solid", fgColor="000000")
    #     datecellref2.value="blank" 
    myLabel0 = Label(tab2,text = "Created",font='Helvetica 16 bold').grid(row=5,column=0)
    workbook.save(filename=workbook_Title+".xlsx")
    








def add_qual():
        
    #NONE DECLARED ERROR LABELS USED
        global fileerrorLabel
        global qualerrorLabel
        global blockerrorLabel
        global starterrorLabel
        global enderrorLabel      
        sheet
    #SET COLUMN WIDTH
        for col in range(1,31):
            column_letter = get_column_letter(col)
            # print(column_letter)
            sheet.column_dimensions[column_letter].width = 42

    #SET TITLE
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=31)
        sheet.row_dimensions[2].height = 60
        title_Cell = sheet['A2']
        title_Cell.border = thick_border
        monthTitle = str(sheet.title)
        # print(monthTitle)
        title_Cell.value = monthTitle+" ELECTRICAL SYSTEMS INSTRUCTOR SCHEDULE"
        title_Cell.font = Font(size=60,name="Times New Roman",bold=True)

        

        #SET DATES
        for i in range(1,31):
            datecellref=sheet.cell(row=1, column=i)
            datecellref.fill = PatternFill("solid", fgColor="DDDDDD")
            
            if(i<10):
                datecellref.value="01/0"+str(i)+"/2021"
            else:
                datecellref.value="01/"+str(i)+"/2021"

        #SET DIVIDER      

        for i in range(1,31):
            datecellref2=sheet.cell(row=3, column=i)
            datecellref2.fill = PatternFill("solid", fgColor="000000")
            datecellref2.border = thick_border_blue_topBottom
            datecellref2.value="blank"  


        #GETTING INPUT VALUES FROM USER
        
        qualName = e2.get()
        if int(qualName) > 3:
            qualerrorLabel = Label(tab1,text = "Qual Num Doesn't Exist",font="Helvetica 10 bold", fg="red")
            qualerrorLabel.grid(row=11,column=0)
            e2.delete(0, END)
            return 
        blockName = e3.get()
        if int(blockName) > 6:
            blockerrorLabel = Label(tab1,text = "Block Number Doesn't Exist",font="Helvetica 10 bold", fg="red")
            blockerrorLabel.grid(row=11,column=0)
            e3.delete(0, END)
            return
        startDate = e4.get()
        startDate = startDate[3:5]
        print(str(startDate))
        if int(startDate) > 31:
            starterrorLabel = Label(tab1,text = "Start Date Number too big",font="Helvetica 10 bold", fg="red")
            starterrorLabel.grid(row=11,column=0)
            e4.delete(0, END)
            return
        # endDate = e5.get()
        # endDate = endDate[3:5]
        if int(endDate) > 31:
            enderrorLabel = Label(tab1,text = "End Date Number too big",font="Helvetica 10 bold", fg="red")
            enderrorLabel.grid(row=11,column=0)
            e5.delete(0, END)
            return

        classNum = e6.get()

        

        #Start Date
        col_index = int(startDate)
        row_index=4
        # print("First row index: "+str(row_index))

        #THE DASHED CELLS ITERATORS
        # blank_col_index = col_index
        # blank_row_index=row_index+1
        
        #TITLE CELL
        cellref = sheet.cell(row=row_index,column=col_index)
        cellref.border = thick_border
        blockcheckcell = sheet.cell(row=row_index, column = 1)
       

        #DASHED CELLS
        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
      

        #End Date
        #check if months match






        des_col=int(endDate)+1

        
        
        total_index = 0

      

   

      
    
        q1check = False



        #COLUMN ITERATOR LOOP
        while col_index < des_col:
            
            #IF THERES A VALUE IN THE CELL
            while cellref.value:
                

                #First: IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK IS LESS THAN OR EQUAL AND THE CLASS NUM ENTERED IS GREATER
                if(int(qualName) <= int(cellref.value[1]) and int(blockName) <= int(cellref.value[10]) and int(qualName) == 1 and int(blockName)==1):
                    print("Start Date: "+str(startDate))
                    print("Start Date Type: "+str(type(startDate)))
                    if startDate!="1" and (int(classNum) > int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999"):
                        sheet.insert_rows(4,5)
                        q1check = True
                        for rows in sheet.iter_rows(min_row=4,max_row=8,min_col=1,max_col=int(startDate)-1):
                            for cell in rows:
                                cell.value = "Q1  Block:1 Class:99999"
                                cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type = "solid")
                                cell.border = thick_border_blue
                        # row_index=4
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                    elif startDate!="1" and (int(classNum) < int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999"):
                        sheet.insert_rows(4,5)
                        q1check = True
                        for rows in sheet.iter_rows(min_row=4,max_row=8,min_col=1,max_col=int(startDate)-1):
                            for cell in rows:
                                cell.value = "Q1  Block:1 Class:99999"
                                cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type = "solid")
                                cell.border = thick_border_blue
                        row_index=4
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                    elif startDate=="1" and int(classNum) > int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999" :
                        sheet.insert_rows(4,5)
                        row_index=4
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                    elif startDate=="1" and int(classNum) < int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999" :
                        row_index+=5
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                    #Second:IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK NUM IS GREATER OR EQUAL 
                elif(int(qualName) <= int(cellref.value[1]) and int(blockName) > int(cellref.value[10])):
                    # print(cellref.value[18:23])
                    row_index+=5
                    print("Row index: " + str(row_index))
                    print(cellref.value)
                    cellref = sheet.cell(row=row_index,column=col_index)
                    blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                    blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                    blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                    blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                    # print("Skiped Title = " + cellref.coordinate)
                    # print("Skipped blank = " + blank_cellref.coordinate)

                    #Third:IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK IS LESS THAN OR EQUAL
                elif(int(qualName) <= int(cellref.value[1]) and int(blockName) <= int(cellref.value[10])):
                    if int(classNum) > int(cellref.value[18:23]):
                        print("Cell Coordinate: "+str(cellref))
                        print("Last Blank Cell Coordinate: "+str(blank_cellref4))
                        sheet.insert_rows(row_index,5)
                        workbook.save(filename=workbook_Title+".xlsx")
                        newrowref = cellref.row
                        print("Newrowref: "+str(newrowref))
                        row_index=newrowref-5
                        print("row index: "+str(row_index))
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                    else:

                        row_index+=5
                    # print("Row index: " + str(row_index))
                    # print(cellref.value)
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                    #FOURTH: IF THE ENTERED QUAL IS GREATER THAN THE CELL VALUE
                elif(int(qualName) > int(cellref.value[1])):
                    # print(cellref.value[18:23])
                    row_index+=5
                    # print("Row index: " + str(row_index))
                    # print(cellref.value)
                    cellref = sheet.cell(row=row_index,column=col_index)
                    blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                    blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                    blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                    blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                    # print("Skiped Title = " + cellref.coordinate)
                    # print("Skipped blank = " + blank_cellref.coordinate)   
                elif(int(qualName)<=int(cellref.value[1])):
                    sheet.move_range(""+str(cellref.coordinate)+":"+str(blank_cellref4.coordinate)+"", rows=5)
                    newrowref = cellref.row
                    row_index=newrowref-5
                    cellref = sheet.cell(row=row_index,column=col_index)
                    blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                    blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                    blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                    blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                      
            else:
                if(int(qualName)==1 and int(blockName)==1 and startDate!="1" and q1check ==False):
                    sheet.insert_rows(4,5)
                    print("Row Index: "+str(row_index))
                    # print("Start Date: "+str(int(startDate)))
                    # print("col index: "+str(col_index))
                    q1check =True
                    for rows in sheet.iter_rows(min_row=4,max_row=8,min_col=1,max_col=int(startDate)-1):
                            for cell in rows:
                            # print("cellref"+str(cellref.coordinate))
                            # print("\ncellref value: "+str(cellref.value))
                                cell.value = "Q1  Block:1 Class:99999"
                                cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type = "solid")
                                cell.border = thick_border_blue
                                print("col index: "+str(col_index))
                # row_index=4
                print("Row Index: "+str(row_index))
                cellref = sheet.cell(row=row_index,column=col_index)
                blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)    
                cellref.value="Q"+str(qualName)+" " +" Block:"+str(blockName)+" Class:"+str(classNum)
                print("col index: "+str(col_index))
                    #TITLE COLOR CODING
                    #QUAL 1
                cellref.border = thick_border
                if (qualName == "1" and blockName == "1"):
                    cellref.fill = PatternFill("solid", fgColor="00CCFF")
                elif (qualName == "1"and blockName == "2"):
                    cellref.fill = PatternFill("solid", fgColor="33CCCC")
                elif (qualName == "1" and blockName == "3"):
                    cellref.fill = PatternFill("solid", fgColor="92D050")
                elif (qualName == "1"and blockName == "4"):
                    cellref.fill = PatternFill("solid", fgColor="FFFF00")
                elif (qualName == "1" and blockName == "5"):
                    cellref.fill = PatternFill("solid", fgColor="FFC000")
                elif (qualName == "1"and blockName == "6"):
                    cellref.fill = PatternFill("solid", fgColor="FF0000")

                #QUAL2
                elif (qualName == "2" and blockName == "1"):
                    cellref.fill = PatternFill("solid", fgColor="00B0F0")
                elif (qualName == "2"and blockName == "2"):
                    cellref.fill = PatternFill("solid", fgColor="92D050")

                #QUAL3
                elif (qualName == "3" and blockName == "1"):
                    cellref.fill = PatternFill("solid", fgColor="FFFF00")
                elif (qualName == "3"and blockName == "2"):
                    cellref.fill = PatternFill("solid", fgColor="9BBB59")
                elif (qualName == "3"and blockName == "3"):
                    cellref.fill = PatternFill("solid", fgColor="FFC000")


                #SETTING BLANK VALUES BORDERS
                blank_cellref.border = thin_border_sides
                blank_cellref2.border = thin_border_sides
                blank_cellref3.border = thin_border_sides
                blank_cellref4.border = thin_border_sides_Bottom

                #SETTING BLANK VALUES
                blank_cellref.value="------"
                blank_cellref2.value="------"
                blank_cellref3.value="------"
                blank_cellref4.value="------"
                total_index+=1
                # print("Title = " + cellref.coordinate)
                # print("blank = " + blank_cellref.coordinate)
            col_index+=1
            # print("column index: "+str(col_index))   


            #This line controls if they shoot up the rows or nor
            # row_index=4
            # print("row index"+str(row_index))   
            cellref = sheet.cell(row=row_index,column=col_index)
            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
            print("Row Index: "+str(row_index))
            print("cellref: "+str(cellref))
            workbook.save(filename=workbook_Title+".xlsx")


            
            
            global submitTotal
            submittedLabel = Label(tab1,text = "Submitted ",font="Helvetica 10 bold", fg="grey35")
            datesLabel = Label(tab1,text = "Total Number of Dates Affected: "+str(submitTotal),font="Helvetica 8 bold", fg="grey35")
            submittedLabel.grid(row=11,column=0,ipadx=50)
            datesLabel.grid(row = 12, column=0)
            submitTotal+=1
            # e2.delete(0, END)
            # e3.delete(0, END)
            # e4.delete(0, END)
            # e5.delete(0, END)

        second_Month_Check() 




def second_Month_Check():
    if monthCheck == True:
        sheets = workbook.sheetnames
        # for i in sheets:
        #     print(i)
        global sheet
        print(str(sheet))
        sheet = workbook[sheets[sheetIndex+1]]
        print(str(sheet.title))
        # sheet = workbook.active
        global newstartDate
        newstartDate = "1"
        global newendDate
        newendDate = endDate = e5.get()
        newendDate = endDate[3:5]
        second_add_qual()
    else:
        None 

              
                
        



def second_add_qual():
        
    #NONE DECLARED ERROR LABELS USED
        global fileerrorLabel
        global qualerrorLabel
        global blockerrorLabel
        global starterrorLabel
        global enderrorLabel      
        
    #SET COLUMN WIDTH
        for col in range(1,31):
            column_letter = get_column_letter(col)
            # print(column_letter)
            sheet.column_dimensions[column_letter].width = 42

    #SET TITLE
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=31)
        sheet.row_dimensions[2].height = 60
        title_Cell = sheet['A2']
        title_Cell.border = thick_border
        monthTitle = str(sheet.title)
        # print(monthTitle)
        title_Cell.value = monthTitle+" ELECTRICAL SYSTEMS INSTRUCTOR SCHEDULE"
        title_Cell.font = Font(size=60,name="Times New Roman",bold=True)

        

        #SET DATES
        for i in range(1,31):
            datecellref=sheet.cell(row=1, column=i)
            datecellref.fill = PatternFill("solid", fgColor="DDDDDD")
            
            if(i<10):
                datecellref.value="01/0"+str(i)+"/2021"
            else:
                datecellref.value="01/"+str(i)+"/2021"

        #SET DIVIDER      

        for i in range(1,31):
            datecellref2=sheet.cell(row=3, column=i)
            datecellref2.fill = PatternFill("solid", fgColor="000000")
            datecellref2.value="blank"  


        #GETTING INPUT VALUES FROM USER
        
        qualName = e2.get()
        print("Qual Name: "+str(qualName))
        if int(qualName) > 3:
            qualerrorLabel = Label(tab1,text = "Qual Num Doesn't Exist",font="Helvetica 10 bold", fg="red")
            qualerrorLabel.grid(row=11,column=0)
            e2.delete(0, END)
            return 
        blockName = e3.get()
        print("Block Name: "+str(blockName))
        if int(blockName) > 6:
            blockerrorLabel = Label(tab1,text = "Block Number Doesn't Exist",font="Helvetica 10 bold", fg="red")
            blockerrorLabel.grid(row=11,column=0)
            e3.delete(0, END)
            return
        # startDate = e4.get()
        # startDate = startDate[3:5]
        # print(str(startDate))
        print("New Start Date: "+str(newstartDate))
        if int(newstartDate) > 31:
            starterrorLabel = Label(tab1,text = "Start Date Number too big",font="Helvetica 10 bold", fg="red")
            starterrorLabel.grid(row=11,column=0)
            e4.delete(0, END)
            return
        # endDate = e5.get()
        # endDate = endDate[3:5]
        print("New End Date: "+str(newendDate))
        if int(newendDate) > 31:
            enderrorLabel = Label(tab1,text = "End Date Number too big",font="Helvetica 10 bold", fg="red")
            enderrorLabel.grid(row=11,column=0)
            e5.delete(0, END)
            return

        classNum = e6.get()

        

        #Start Date
        col_index = int(newstartDate)
        row_index=4
        # print("First row index: "+str(row_index))

        #THE DASHED CELLS ITERATORS
        # blank_col_index = col_index
        # blank_row_index=row_index+1
        
        #TITLE CELL
        cellref = sheet.cell(row=row_index,column=col_index)
        cellref.border = thick_border
        blockcheckcell = sheet.cell(row=row_index, column = 1)
       

        #DASHED CELLS
        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
      

        #End Date
        #check if months match
        # startMonth = e4.get()
        # startMonth = startMonth[0:2]
        # endMonth = e5.get()
        # endMonth = endMonth[0:2]
        # monthCheck = False
        # if(startMonth == endMonth):
        #     None
        # elif(startMonth != endMonth):
        #     endDate = 30
        #     monthCheck =True





        des_col=int(newendDate)+1

        
        
        total_index = 0

      

   

      
    
        q1check = False



        #COLUMN ITERATOR LOOP
        while col_index < des_col:
            
            #IF THERES A VALUE IN THE CELL
            while cellref.value:
                

                #First: IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK IS LESS THAN OR EQUAL AND THE CLASS NUM ENTERED IS GREATER
                if(int(qualName) <= int(cellref.value[1]) and int(blockName) <= int(cellref.value[10]) and int(qualName) == 1 and int(blockName)==1):
                    print("Start Date: "+str(newstartDate))
                    print("Start Date Type: "+str(type(newstartDate)))
                    if newstartDate!="1" and (int(classNum) > int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999"):
                        sheet.insert_rows(4,5)
                        q1check = True
                        for rows in sheet.iter_rows(min_row=4,max_row=8,min_col=1,max_col=int(newstartDate)-1):
                            for cell in rows:
                                cell.value = "Q1  Block:1 Class:99999"
                                cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type = "solid")
                                cell.border = thick_border_blue
                        # row_index=4
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                    elif newstartDate!="1" and (int(classNum) < int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999"):
                        sheet.insert_rows(4,5)
                        q1check = True
                        for rows in sheet.iter_rows(min_row=4,max_row=8,min_col=1,max_col=int(newstartDate)-1):
                            for cell in rows:
                                cell.value = "Q1  Block:1 Class:99999"
                                cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type = "solid")
                                cell.border = thick_border_blue
                        row_index=4
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                    elif newstartDate=="1" and int(classNum) > int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999" :
                        sheet.insert_rows(4,5)
                        row_index=4
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                    elif newstartDate=="1" and int(classNum) < int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999" :
                        row_index+=5
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                    #Second:IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK NUM IS GREATER OR EQUAL 
                elif(int(qualName) <= int(cellref.value[1]) and int(blockName) > int(cellref.value[10])):
                    # print(cellref.value[18:23])
                    row_index+=5
                    print("Row index: " + str(row_index))
                    print(cellref.value)
                    cellref = sheet.cell(row=row_index,column=col_index)
                    blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                    blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                    blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                    blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                    # print("Skiped Title = " + cellref.coordinate)
                    # print("Skipped blank = " + blank_cellref.coordinate)

                    #Third:IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK IS LESS THAN OR EQUAL
                elif(int(qualName) <= int(cellref.value[1]) and int(blockName) <= int(cellref.value[10])):
                    if int(classNum) > int(cellref.value[18:23]):
                        print("Cell Coordinate: "+str(cellref))
                        print("Last Blank Cell Coordinate: "+str(blank_cellref4))
                        sheet.insert_rows(row_index,5)
                        workbook.save(filename=workbook_Title+".xlsx")
                        newrowref = cellref.row
                        print("Newrowref: "+str(newrowref))
                        row_index=newrowref-5
                        print("row index: "+str(row_index))
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                    else:

                        row_index+=5
                    # print("Row index: " + str(row_index))
                    # print(cellref.value)
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                    #FOURTH: IF THE ENTERED QUAL IS GREATER THAN THE CELL VALUE
                elif(int(qualName) > int(cellref.value[1])):
                    # print(cellref.value[18:23])
                    row_index+=5
                    # print("Row index: " + str(row_index))
                    # print(cellref.value)
                    cellref = sheet.cell(row=row_index,column=col_index)
                    blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                    blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                    blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                    blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                    # print("Skiped Title = " + cellref.coordinate)
                    # print("Skipped blank = " + blank_cellref.coordinate)   
                elif(int(qualName)<=int(cellref.value[1])):
                    sheet.move_range(""+str(cellref.coordinate)+":"+str(blank_cellref4.coordinate)+"", rows=5)
                    newrowref = cellref.row
                    row_index=newrowref-5
                    cellref = sheet.cell(row=row_index,column=col_index)
                    blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                    blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                    blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                    blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                      
            else:
                if(int(qualName)==1 and int(blockName)==1 and newstartDate!="1" and q1check ==False):
                    sheet.insert_rows(4,5)
                    print("Row Index: "+str(row_index))
                    # print("Start Date: "+str(int(startDate)))
                    # print("col index: "+str(col_index))
                    q1check =True
                    for rows in sheet.iter_rows(min_row=4,max_row=8,min_col=1,max_col=int(newstartDate)-1):
                            for cell in rows:
                            # print("cellref"+str(cellref.coordinate))
                            # print("\ncellref value: "+str(cellref.value))
                                cell.value = "Q1  Block:1 Class:99999"
                                cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type = "solid")
                                cell.border = thick_border_blue
                                print("col index: "+str(col_index))
                # row_index=4
                print("Row Index: "+str(row_index))
                cellref = sheet.cell(row=row_index,column=col_index)
                blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)    
                cellref.value="Q"+str(qualName)+" " +" Block:"+str(blockName)+" Class:"+str(classNum)
                print("col index: "+str(col_index))
                    #TITLE COLOR CODING
                    #QUAL 1
                cellref.border = thick_border
                if (qualName == "1" and blockName == "1"):
                    cellref.fill = PatternFill("solid", fgColor="00CCFF")
                elif (qualName == "1"and blockName == "2"):
                    cellref.fill = PatternFill("solid", fgColor="33CCCC")
                elif (qualName == "1" and blockName == "3"):
                    cellref.fill = PatternFill("solid", fgColor="92D050")
                elif (qualName == "1"and blockName == "4"):
                    cellref.fill = PatternFill("solid", fgColor="FFFF00")
                elif (qualName == "1" and blockName == "5"):
                    cellref.fill = PatternFill("solid", fgColor="FFC000")
                elif (qualName == "1"and blockName == "6"):
                    cellref.fill = PatternFill("solid", fgColor="FF0000")

                #QUAL2
                elif (qualName == "2" and blockName == "1"):
                    cellref.fill = PatternFill("solid", fgColor="00B0F0")
                elif (qualName == "2"and blockName == "2"):
                    cellref.fill = PatternFill("solid", fgColor="92D050")

                #QUAL3
                elif (qualName == "3" and blockName == "1"):
                    cellref.fill = PatternFill("solid", fgColor="FFFF00")
                elif (qualName == "3"and blockName == "2"):
                    cellref.fill = PatternFill("solid", fgColor="9BBB59")
                elif (qualName == "3"and blockName == "3"):
                    cellref.fill = PatternFill("solid", fgColor="FFC000")


                #SETTING BLANK VALUES BORDERS
                blank_cellref.border = thin_border_sides
                blank_cellref2.border = thin_border_sides
                blank_cellref3.border = thin_border_sides
                blank_cellref4.border = thin_border_sides_Bottom

                #SETTING BLANK VALUES
                blank_cellref.value="------"
                blank_cellref2.value="------"
                blank_cellref3.value="------"
                blank_cellref4.value="------"
                total_index+=1
                # print("Title = " + cellref.coordinate)
                # print("blank = " + blank_cellref.coordinate)
            col_index+=1
            # print("column index: "+str(col_index))   


            #This line controls if they shoot up the rows or nor
            # row_index=4
            # print("row index"+str(row_index))   
            cellref = sheet.cell(row=row_index,column=col_index)
            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
            print("Row Index: "+str(row_index))
            print("cellref: "+str(cellref))
            workbook.save(filename=workbook_Title+".xlsx")


            
            
            global submitTotal
            submittedLabel = Label(tab1,text = "Submitted ",font="Helvetica 10 bold", fg="grey35")
            datesLabel = Label(tab1,text = "Total Number of Dates Affected: "+str(submitTotal),font="Helvetica 8 bold", fg="grey35")
            submittedLabel.grid(row=11,column=0,ipadx=50)
            datesLabel.grid(row = 12, column=0)
            submitTotal+=1
            e2.delete(0, END)
            e3.delete(0, END)
            e4.delete(0, END)
            e5.delete(0, END)

        # if monthCheck == True:
        #     global sheet
        #     sheet = workbook.active[2]
        #     startDate = 1
        #     endDate = endDate = e5.get()
        #     endDate = endDate[3:5]
        #     second_add_qual()
        # else:
        #     None



#======================================================================================================#
#======================================================================================================#
#======================================================================================================#
#======================================================================================================#
#======================================================================================================#
#======================================================================================================#
#======================================================================================================#       




#MAIN PAGE
root = ThemedTk(theme="blue")

# ttk.Style().theme_use('black')

root.geometry("800x300")
root.title("Curtis Scheduling Tool "+"                                                                                          \u00A9" + " KandyKane Solutions  Ver.3.0.0*")
tab_parent = ttk.Notebook(root)
tab1 = ttk.Frame(tab_parent)
tab2 = ttk.Frame(tab_parent)
tab3 = ttk.Frame(tab_parent)

tab_parent.add(tab1,text="Use Existing")
tab_parent.add(tab2,text="New Workbook")
tab_parent.add(tab3,text="About")
tab_parent.pack(expand=1,fill='both')

# bg = PhotoImage(file="background5.png")
# my_label = Label(root,image=bg)
# my_label.place(x=0,y=0,relwidth=1,relheight=1)
p1 = PhotoImage(file = 'airforce.png')
root.iconphoto(False, p1)
# root.geometry("380x430")
# for i in range(3):
#     root.columnconfigure(i, weight=1)

# root.rowconfigure(1, weight=1)
# main_Label = Label(root,text="Curtis Schedule Tool",font='Helvetica 18 bold')
# myButton = Button(root,text="Add WorkBook",height = 2, width = 20,command=myClickNew, font='Helvetica 12 bold',bg="grey80")
# myButton2 = Button(root,text="Use Existing",height = 2, width = 20,command=myClick,font='Helvetica 12 bold',bg="grey80")
# myButton3 = Button(root,text="Exit",height = 2, width = 20,command=root.destroy,font='Helvetica 12 bold',bg="grey80")
# myButton4 = Button(root,text="About",height = 2, width = 20,command=myClickAbout,font='Helvetica 12 bold',bg="grey80")
# copyright_label = Label(tab1,text="\u00A9" + " KandyKane Solutions  Ver.3.0.0*",font='Helvetica 12 bold', fg="grey72")
# main_Label.pack()
# myButton.pack(fill=X, padx=80, pady=10)
# myButton2.pack(fill=X,padx=80, pady=10)
# myButton4.pack(fill=X,padx=80, pady=10)
# myButton3.pack(fill=X,padx=80, pady=10)
# changeOnHover(myButton, "aqua", "grey80")
# changeOnHover(myButton2, "aqua", "grey80")
# changeOnHover(myButton3, "aqua", "grey80")
# changeOnHover(myButton4, "aqua", "grey80")
# copyright_label.grid(row=19,column=0)



# top = Toplevel()
# top.title("Add Qual")
# top.geometry("350x300")


#TAB1==============================================================================================================TAB1
tab1Label0 = Label(tab1,text = "Existing Excel File",font="Helvetica 12 bold").grid(row=1,column=0)
tab1Label0 = Label(tab1,text = "File Name:").grid(row=2,column=0,pady=2)
tab1Label1 = Label(tab1,text = "Qual Num:").grid(row=3,column=0,pady=2)
tab1Label2 = Label(tab1,text = "Block Num:").grid(row=4,column=0,pady=2)
tab1Label3 = Label(tab1,text = "Start Date:").grid(row=5,column=0,pady=2)
tab1Label4 = Label(tab1,text = "End Date:").grid(row=6,column=0,pady=2)
tab1Label5 = Label(tab1,text = "Class Num:").grid(row=7,column=0,pady=2)



tab1Label2a = Label(tab1,text = "'Ex.test'",fg=("grey50")).grid(row=2,column=2,pady=2)
tab1Label3a = Label(tab1,text = "'Ex. 1-3'",fg=("grey50")).grid(row=3,column=2,pady=2)
tab1Label4a = Label(tab1,text = "'Ex. 1-6'",fg=("grey50")).grid(row=4,column=2,pady=2)
tab1Label5a = Label(tab1,text = "'01/01/21'",fg=("grey50")).grid(row=5,column=2,pady=2,padx=2)
tab1Label6a = Label(tab1,text = "'01/28/21'",fg=("grey50")).grid(row=6,column=2,pady=2,padx=2)
tab1Label7a = Label(tab1,text = "'20010'",fg=("grey50")).grid(row=7,column=2,pady=2)


global e1
global e2
global e3
global e4
global e5
global e6


e1 = Entry(tab1,width=10)
e1.grid(row=2,column=1)
e2 = Entry(tab1,width=5)
e2.grid(row=3,column=1)
e3 = Entry(tab1,width=5)
e3.grid(row=4,column=1)
e4 = Entry(tab1,width=10)
e4.grid(row=5,column=1)
e5 = Entry(tab1,width=10)
e5.grid(row=6,column=1)
e6 = Entry(tab1,width=8)
e6.grid(row=7,column=1)


myButton4 = Button(tab1,text="Submit",command=existing_WORKBOOK,bg="grey80")
myButton4.grid(row=10,column=0)
changeOnHover(myButton4, "aqua", "grey80")
global submitTotal
submitTotal = 1







#TAB2====================================================ADD NEW WORKBOOK======================================================TAB2
myLabel0 = Label(tab2,text = "New Excel File",font="Helvetica 12 bold").grid(row=1,column=0)
myLabel0 = Label(tab2,text = "File Name:").grid(row=2,column=0)
# myLabel1 = Label(top,text = "Qual Num:").grid(row=3,column=0)
# myLabel2 = Label(top,text = "Block Num:").grid(row=4,column=0)
# myLabel3 = Label(top,text = "Start Date:").grid(row=5,column=0)
# myLabel4 = Label(top,text = "End Date:").grid(row=6,column=0)
# myLabel5 = Label(top,text = "Class Num:").grid(row=7,column=0,pady=2)
# myLabel6 = Label(top,text = "Month:").grid(row=8,column=0,pady=2)

# myLabel2a = Label(top,text = "'Ex.test'",fg=("grey50")).grid(row=2,column=2,pady=2)
# myLabel3a = Label(top,text = "'Ex. 1-3'",fg=("grey50")).grid(row=3,column=2,pady=2)
# myLabel4a = Label(top,text = "'Ex. 1-6'",fg=("grey50")).grid(row=4,column=2,pady=2)
# myLabel5a = Label(top,text = "'1-31'",fg=("grey50")).grid(row=5,column=2,pady=2)
# myLabel6a = Label(top,text = "'1-31'",fg=("grey50")).grid(row=6,column=2,pady=2)
# myLabel7a = Label(top,text = "'20010'",fg=("grey50")).grid(row=7,column=2,pady=2)
# myLabel8a = Label(top,text = "'jan'",fg=("grey50")).grid(row=8,column=2,pady=2)
global newE1
global newe2
# global e3
# global e4
# global e5
# global e6
# global e7
newe1 = Entry(tab2,width=10)
newe1.grid(row=2,column=1)

# e3 = Entry(top,width=5)
# e3.grid(row=4,column=1)
# e4 = Entry(top,width=5)
# e4.grid(row=5,column=1)
# e5 = Entry(top,width=5)
# e5.grid(row=6,column=1)
# e6 = Entry(top,width=8)
# e6.grid(row=7,column=1)
# e7 = Entry(top,width=5)
# e7.grid(row=8,column=1)
createButton = Button(tab2,text="Create New",command=new_WORKBOOK,bg="grey80")
createButton.grid(row=4,column=0)
changeOnHover(myButton4, "aqua", "grey80")
# global submitTotal
# submitTotal = 1


#TAB3================================================ABOUT====================================================================TAB3
mylabel2 = Label(tab3,text="About",font='Helvetica 30 bold')
about = """This is a simple scheduler program for automatically creating\nand editing tasks on defined days on an excel sheet when \ngiven prescribed dates. Hope you enjoy! """
mylabel = Label(tab3,text=about,font='Helvetica 12 bold')
mylabel2 = Label(tab3,text="How to Use",font='Helvetica 12 bold')
mylabel7 = Label(tab3,text="-Put the excel files you wish to edit in the same folder as this programs .exe file\n-When entering numbers all entries must be single digit\n-Do not separate the .exe file from the images",font='Helvetica 12 bold')
mylabel3 = Label(tab3,text="Use Existing:",font='Helvetica 12 bold')
mylabel4 = Label(tab3,text="This is for adding to an existing file. Simply, enter\n the file name(no extension) and enter\n the rest of your information accordingly.",font='Helvetica 12 bold')
mylabel5 = Label(tab3,text="Add WorkBook:",font='Helvetica 12 bold')
mylabel6 = Label(tab3,text="This is for creating a new excel file. Simply, enter what\n you would like to call the file(no extension,no special characters) and enter\n the rest of your information accordingly.You must initilize it with a Qual entry",font='Helvetica 12 bold')
mylabel.pack(pady=(0,30))
mylabel2.pack()
mylabel7.pack()
mylabel3.pack()
mylabel4.pack()
mylabel5.pack()
mylabel6.pack()
    





root.mainloop()


