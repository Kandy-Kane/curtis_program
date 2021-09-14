from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell import cell
from openpyxl.cell.cell import ERROR_CODES
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.dimensions import SheetDimension
from openpyxl.utils import get_column_letter
from tkinter import * 
from tkinter import ttk
from ttkthemes import ThemedTk
import time
import os.path
from os import error, path
from PIL import Image
from tkinter import messagebox
import traceback
import sys
import datetime


thick_border = Border(left=Side(style='thick'), 
                     right=Side(style='thick'), 
                     top=Side(style='thick'), 
                     bottom=Side(style='thick'))

thick_border_blue = Border(left=Side(style='thick',color='0066CC'), 
                     right=Side(style='thick',color='0066CC')   
                  ) 

thick_border_blue_topBottom = Border(top=Side(style='thick',color='0066CC'), 
                     bottom=Side(style='thick',color='0066CC')   
                  )                    

thin_border_all = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

thin_border_all_grey = Border(left=Side(style='thin',color="DDDDDD"), 
                     right=Side(style='thin',color="DDDDDD"), 
                     top=Side(style='thin',color="DDDDDD"), 
                     bottom=Side(style='thin',color="DDDDDD"))

thin_border_sides = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style=None), 
                     bottom=Side(style=None))


thin_border_sides_Bottom = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style=None), 
                     bottom=Side(style='thin'))


#GLOBAL ERROR VARIABLES
fileerrorLabel = None
qualerrorLabel = None
blockerrorLabel = None
starterrorLabel = None
enderrorLabel = None
monthCheck = False

def changeOnHover(button, colorOnHover, colorOnLeave):
  
    # adjusting backgroung of the widget
    # background on entering widget
    button.bind("<Enter>", func=lambda e: button.config(
        background=colorOnHover))
  
    # background color on leving widget
    button.bind("<Leave>", func=lambda e: button.config(
        background=colorOnLeave))

#============================================================================================================================#
#============================================================================================================================#
#============================================================================================================================#
    
    
def Workbook_ReadIn():
    global workbook_Title
    workbook_Title = tab3e2.get()
    print(workbook_Title)
    global workbook
    if path.exists(str(tab3e2.get())+".xlsx") == True: 
        workbook = load_workbook(filename=workbook_Title+".xlsx")
        # workbook.add_named_style(date_style)
    else:
        fileerrorLabel = Label(tab3,text="File Does Not Exist",font="Helvetica 10 bold", fg="red")
        fileerrorLabel.grid(row=11,column=0)
        return
    # print(path.exists(str(e1.get())+".xlsx"))
    global sheet
    sheet = workbook.active
    print(workbook)
    global sheetIndex
    #Check Months
    startMonth = startdatecellref.value
    startMonth = startMonth[0:2]
    print(startMonth)
    endMonth = enddatecellref.value
    endMonth = endMonth[0:2]
    print(endMonth)
    global endDateReadIn
    endDateReadIn = enddatecellref.value
    endDateReadIn = endDateReadIn[3:5]
    print("FIRST END DATE READ IN: "+endDateReadIn)
    global monthCheck

    # month = str(e4.get())
    # print("Month: "+str(month[0:2]))
    print("Start Month: "+str(startMonth))
    if str(startMonth[0:2]) =="01":
        if'JAN' in workbook.sheetnames:
            sheet = workbook["JAN"]
            sheetIndex = 1
        elif 'Jan' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("JAN")
            sheet = ws1
            sheetIndex = 1

    if str(startMonth[0:2]) =="02":
        if'FEB' in workbook.sheetnames:
            sheet = workbook["FEB"]
            sheetIndex =2
        elif 'Feb' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("FEB")
            sheet = ws1
            sheetIndex = 2
    if str(startMonth[0:2]) =="03":
        if'MAR' in workbook.sheetnames:
            sheet = workbook["MAR"]
            sheetIndex = 3
        elif 'MAR' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("MAR")
            sheet = ws1
            sheetIndex = 3
    if str(startMonth[0:2]) =="04":
        if'APR' in workbook.sheetnames:
            sheet = workbook["APR"]
            sheetIndex = 4
        elif 'APR' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("APR")
            sheet = ws1
            sheetIndex = 4
    if str(startMonth[0:2]) =="05":
        if'MAY' in workbook.sheetnames:
            sheet = workbook["MAY"]
            sheetIndex = 5
        elif 'MAY' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("MAY")
            sheet = ws1
            sheetIndex = 5

    if str(startMonth[0:2]) =="06":
        if'JUNE' in workbook.sheetnames:
            sheet = workbook["JUNE"]
            sheetIndex = 6
        elif 'JUNE' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("JUNE")
            sheet = ws1
            sheetIndex = 6

    if str(startMonth[0:2]) =="07":
        if'JULY' in workbook.sheetnames:
            sheet = workbook["JULY"]
            sheetIndex = 7
        elif 'JULY' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("JULY")
            sheet = ws1
            sheetIndex = 7

    if str(startMonth[0:2]) =="08":
        if'AUG' in workbook.sheetnames:
            sheet = workbook["AUG"]
            sheetIndex = 8
        elif 'AUG' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("AUG")
            sheet = ws1
            sheetIndex = 8

    if str(startMonth[0:2]) =="09":
        if'SEPT' in workbook.sheetnames:
            sheet = workbook["SEPT"]
            sheetIndex = 9
        elif 'SEPT' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("SEPT")
            sheet = ws1
            sheetIndex = 9
    if str(startMonth[0:2]) =="10":
        if'OCT' in workbook.sheetnames:
            sheet = workbook["OCT"]
            sheetIndex = 10
        elif 'OCT' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("OCT")
            sheet = ws1
            sheetIndex = 10
    if str(startMonth[0:2]) =="11":
        if'NOV' in workbook.sheetnames:
            sheet = workbook["NOV"]
            sheetIndex = 11
        elif 'NOV' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("NOV")
            sheet = ws1
            sheetIndex = 11

    if str(startMonth[0:1]) =="12":
        if'DEC' in workbook.sheetnames:
            sheet = workbook["DEC"]
            sheetIndex = 12
        elif 'DEC' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("DEC")
            sheet = ws1
            sheetIndex = 12

    if(startMonth == endMonth):
        None
        tab3add_qual()
    elif(startMonth != endMonth):
        endDateReadIn = "30"
        monthCheck =True
        tab3add_qual() 


    



def readFromExcel():
    try:
        tab3Label3 = Label(tab3frame3,text = "FINISHED",fg="grey26",bg="grey26",font="Helvetica 10 bold").grid(row=4,column=0)
        global my_progress
        my_progress.grid(row=0,column=0,pady=(50,10))

        #GETTING READ FROM FILE
        global readExcelFile
        readExcelFile = tab3e1.get()
        print(readExcelFile)
        if path.exists(str(tab3e1.get())+".xlsx") == True: 
            readFromBook = load_workbook(filename =readExcelFile+ '.xlsx')
        else:
            fileerrorLabel = Label(tab3,text="File Does Not Exist",font="Helvetica 10 bold", fg="red")
            fileerrorLabel.grid(row=5,column=0)
            return

        
        sheet = readFromBook.active
        print(sheet.title)
        startingRow = 2
        startingCol = 1
        global qualcellref
        global blockcellref
        global startdatecellref
        global enddatecellref
        global classcellref
        qualcellref = sheet.cell(row=startingRow,column=startingCol)
        print(qualcellref.value)
        blockcellref = sheet.cell(row=startingRow,column=startingCol+1)
        print(blockcellref.value)
        startdatecellref = sheet.cell(row=startingRow,column=startingCol+2)
        print(startdatecellref.value)
        enddatecellref = sheet.cell(row=startingRow,column=startingCol+3)
        classcellref = sheet.cell(row=startingRow,column=startingCol+4)
        
        print(qualcellref.value)
        skittle = 1
        while qualcellref.value:
            
            my_progress['value']+=30
                
            tab3.update_idletasks() 
            print("QUALCELLREF VALUE: " +str(qualcellref.value))
            Workbook_ReadIn()
            print(skittle)
            startingRow+=1
            qualcellref = sheet.cell(row=startingRow,column=startingCol)
            print(qualcellref.value)
            blockcellref = sheet.cell(row=startingRow,column=startingCol+1)
            print(blockcellref.value)
            startdatecellref = sheet.cell(row=startingRow,column=startingCol+2)
            print(startdatecellref.value)
            enddatecellref = sheet.cell(row=startingRow,column=startingCol+3)
            classcellref = sheet.cell(row=startingRow,column=startingCol+4)
            skittle+=1
        tab3Label3 = Label(tab3frame3,text = "FINISHED",fg="white",bg="grey26",font="Helvetica 10 bold").grid(row=4,column=0)
        
        my_progress.stop()
    except():
        print(traceback.format_exc())
        messagebox.showwarning(title="Error Occured", message="something went wrong in READ FROM EXCEL. Check your entries and try again")
        if path.exists("errors.txt") == TRUE:
            ct = datetime.datetime.now() 
            with open("errors.txt", "a") as file:
                file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))
        else:
            ct = datetime.datetime.now() 
            with open("errors.txt", "x") as file:
                file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))
 
        




def existing_WORKBOOK():
    try:
        finishedLabel = Label(tab1frame3,text="FINISHED",font="Helvetica 10 bold", fg="grey26",bg="grey26").grid(row=4,column=0) 
        global my_progress2
        my_progress2.grid(row=0,column=0)
        global workbook_Title
        workbook_Title = e1.get()
        global workbook
        if path.exists(str(e1.get())+".xlsx") == True: 
            workbook = load_workbook(filename=workbook_Title+".xlsx")
            # workbook.add_named_style(date_style)
        else:
            fileerrorLabel = Label(tab1frame3,text="File Does Not Exist",font="Helvetica 10 bold", fg="red")
            fileerrorLabel.grid(row=2,column=0)
            return
        # print(path.exists(str(e1.get())+".xlsx"))
        global sheet
        sheet = workbook.active
        print(workbook)
        global sheetIndex
        #Check Months
        global startMonth
        startMonth = e4.get()
        startMonth = startMonth[0:2]
        endMonth = e5.get()
        endMonth = endMonth[0:2]
        global endDate
        endDate = e5.get()
        endDate = endDate[3:5]
        global monthCheck

        # month = str(e4.get())
        # print("Month: "+str(month[0:2]))
        print("Start Month: "+str(startMonth))
        tab1.update_idletasks() 
        if str(startMonth[0:2]) =="01":
            if'JAN' in workbook.sheetnames:
                sheet = workbook["JAN"]
                sheetIndex = 1
            elif 'Jan' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("JAN")
                sheet = ws1
                sheetIndex = 1

        if str(startMonth[0:2]) =="02":
            if'FEB' in workbook.sheetnames:
                sheet = workbook["FEB"]
                sheetIndex =2
            elif 'Feb' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("FEB")
                sheet = ws1
                sheetIndex = 2
        if str(startMonth[0:2]) =="03":
            if'MAR' in workbook.sheetnames:
                sheet = workbook["MAR"]
                sheetIndex = 3
            elif 'MAR' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("MAR")
                sheet = ws1
                sheetIndex = 3
        if str(startMonth[0:2]) =="04":
            if'APR' in workbook.sheetnames:
                sheet = workbook["APR"]
                sheetIndex = 4
            elif 'APR' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("APR")
                sheet = ws1
                sheetIndex = 4
        if str(startMonth[0:2]) =="05":
            if'MAY' in workbook.sheetnames:
                sheet = workbook["MAY"]
                sheetIndex = 5
            elif 'MAY' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("MAY")
                sheet = ws1
                sheetIndex = 5

        if str(startMonth[0:2]) =="06":
            if'JUNE' in workbook.sheetnames:
                sheet = workbook["JUNE"]
                sheetIndex = 6
            elif 'JUNE' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("JUNE")
                sheet = ws1
                sheetIndex = 6

        if str(startMonth[0:2]) =="07":
            if'JULY' in workbook.sheetnames:
                sheet = workbook["JULY"]
                sheetIndex = 7
            elif 'JULY' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("JULY")
                sheet = ws1
                sheetIndex = 7

        if str(startMonth[0:2]) =="08":
            if'AUG' in workbook.sheetnames:
                sheet = workbook["AUG"]
                sheetIndex = 8
            elif 'AUG' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("AUG")
                sheet = ws1
                sheetIndex = 8

        if str(startMonth[0:2]) =="09":
            if'SEPT' in workbook.sheetnames:
                sheet = workbook["SEPT"]
                sheetIndex = 9
            elif 'SEPT' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("SEPT")
                sheet = ws1
                sheetIndex = 9
        if str(startMonth[0:2]) =="10":
            if'OCT' in workbook.sheetnames:
                sheet = workbook["OCT"]
                sheetIndex = 10
            elif 'OCT' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("OCT")
                sheet = ws1
                sheetIndex = 10
        if str(startMonth[0:2]) =="11":
            if'NOV' in workbook.sheetnames:
                sheet = workbook["NOV"]
                sheetIndex = 11
            elif 'NOV' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("NOV")
                sheet = ws1
                sheetIndex = 11

        if str(startMonth[0:1]) =="12":
            if'DEC' in workbook.sheetnames:
                sheet = workbook["DEC"]
                sheetIndex = 12
            elif 'DEC' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("DEC")
                sheet = ws1
                sheetIndex = 12

        if(startMonth == endMonth):
            None
            add_qual() 
        elif(startMonth != endMonth):
            endDate = 30
            monthCheck =True
            add_qual()
        finishedLabel = Label(tab1frame3,text="FINISHED",font="Helvetica 10 bold", fg="white",bg="grey26").grid(row=4,column=0) 
        my_progress.stop() 
    except():
        print(traceback.format_exc())
        messagebox.showwarning(title="Error Occured", message="something went wrong in EXISTING BOOK. Check your entries and try again")
        if path.exists("errors.txt") == TRUE:
            ct = datetime.datetime.now() 
            with open("errors.txt", "a") as file:
                file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))
        else:
            ct = datetime.datetime.now() 
            with open("errors.txt", "x") as file:
                file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))



def new_WORKBOOK():
    try:
        global workbook_Title
        workbook_Title = newe1.get()
        if path.exists(str(newe1.get())+".xlsx") == TRUE:
            fileerrorLabel = Label(tab2mainframe,text="File Already Exists",font="Helvetica 10 bold", fg="red")
            fileerrorLabel.grid(row=5,column=0)
            return
        else:
            global workbook 
            workbook = Workbook()
            # sheet = workbook.active
            ws1 = workbook.create_sheet("JAN")
            ws2 = workbook.create_sheet("FEB")
            ws3 = workbook.create_sheet("MAR")
            ws4 = workbook.create_sheet("APR")
            ws5 = workbook.create_sheet("MAY")
            ws6 = workbook.create_sheet("JUNE")
            ws7 = workbook.create_sheet("JULY")
            ws8 = workbook.create_sheet("AUG")
            ws9 = workbook.create_sheet("SEPT")
            ws10 = workbook.create_sheet("OCT")
            ws11 = workbook.create_sheet("NOV")
            ws12 = workbook.create_sheet("DEC")

            #ADD DAYS TO THE SHEET
        
        
        def sheetDates():
            
            startmonth=""
            dayList = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
            global day_index
            day_index=4
            def addDaysAndDates():
                for col in range(1,32):
                    column_letter = get_column_letter(col)
                    # print(column_letter)
                    sheet.column_dimensions[column_letter].width = 42
                print(sheet.title)
                global day_index
                col_index=1
                print(endDay)
                print("Day INdex: "+str(day_index))
                while col_index<endDay:
                    if day_index ==6 and col_index<endDay:
                        daycellref=sheet.cell(row=2, column=col_index) 
                        daycellref.value = (dayList[day_index])
                        daycellref.border = thick_border_blue_topBottom
                        day_index = 0
                        col_index+=1
                        # workbook.save(filename=workbook_Title+".xlsx")
                    else:
                        daycellref=sheet.cell(row=2, column=col_index) 
                        daycellref.value = (dayList[day_index])
                        day_index+=1
                        col_index+=1
                        # workbook.save(filename=workbook_Title+".xlsx")
                    
                # workbook.save(filename=workbook_Title+".xlsx")
                
                for i in range(1,endDay):
                    print("Enday: "+str(endDay))
                    daycellref=sheet.cell(row=2, column=i)
                    daycellref.fill = PatternFill("solid", fgColor="DDDDDD")
                    if daycellref.value == "Saturday":
                        column_letter = get_column_letter(i)
                        sheet.column_dimensions[column_letter].width = 0.1
                        # workbook.save(filename=workbook_Title+".xlsx")
                    elif daycellref.value == "Sunday":
                        column_letter = get_column_letter(i)
                        sheet.column_dimensions[column_letter].width = 0.1
                        # workbook.save(filename=workbook_Title+".xlsx")

            for sheet in workbook.worksheets:
                endDay = 32
                print(sheet.title)
                if str(sheet.title)=="JAN" or str(sheet.title)=="MAR"or str(sheet.title)=="MAY"or str(sheet.title)=="JULY"or str(sheet.title)=="AUG"or str(sheet.title)=="OCT"or str(sheet.title)=="DEC":
                    endDay = 32
                    addDaysAndDates()
                elif str(sheet.title) == "APR"or str(sheet.title)=="JUNE"or str(sheet.title)=="SEPT"or str(sheet.title)=="NOV":
                    endDay = 31
                    addDaysAndDates()
                elif str(sheet.title) == "FEB":
                    endDay = 29
                    addDaysAndDates()
                else:
                    continue

                for i in range(1,endDay):
                    if str(sheet.title) == "JAN":
                        startmonth = "01"
                    elif str(sheet.title) == "FEB":
                        startmonth = "02"
                    elif str(sheet.title) == "MAR":
                        startmonth = "03"
                    elif str(sheet.title) == "APR":
                        startmonth = "04"
                    elif str(sheet.title) == "MAY":
                        startmonth = "05"
                    elif str(sheet.title) == "JUNE":
                        startmonth = "06"
                    elif str(sheet.title) == "JULY":
                        startmonth = "07"
                    elif str(sheet.title) == "AUG":
                        startmonth = "08"
                    elif str(sheet.title) == "SEPT":
                        startmonth = "09"
                    elif str(sheet.title) == "OCT":
                        startmonth = "10"
                    elif str(sheet.title) == "NOV":
                        startmonth = "11"
                    elif str(sheet.title) == "DEC":
                        startmonth = "12"
                    datecellref=sheet.cell(row=1, column=i)
                    datecellref.fill = PatternFill("solid", fgColor="DDDDDD")
                    if(i<10):
                        datecellref.value=str(startmonth)+"/0"+str(i)+"/21"
                    else:
                        datecellref.value=str(startmonth)+"/"+str(i)+"/21"

                #SETTING TITLE
                sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=32)
                sheet.row_dimensions[3].height = 60
                title_Cell = sheet['A3']
                title_Cell.border = thick_border
                monthTitle = str(sheet.title)
                # print(monthTitle)
                title_Cell.value = monthTitle+" ELECTRICAL SYSTEMS INSTRUCTOR SCHEDULE"
                title_Cell.font = Font(size=60,name="Times New Roman",bold=True)

                #SETTING DIVIDER
                for i in range(1,32):
                    datecellref2=sheet.cell(row=4, column=i)
                    datecellref2.fill = PatternFill("solid", fgColor="000000")
                    datecellref2.border = thick_border_blue_topBottom
                    datecellref2.value="blank"  
                    tab1.update_idletasks() 
                

        sheetDates()
        myLabel0 = Label(tab2mainframe,text = "Created",font='Helvetica 16 bold',bg="grey26").grid(row=5,column=0,ipadx=20)
        workbook.save(filename=workbook_Title+".xlsx")
    except:
            print(traceback.format_exc())
            messagebox.showwarning(title="Error Occured", message="something went wrong in NEW BOOK. Check your entries and try again")
            if path.exists("errors.txt") == TRUE:
                ct = datetime.datetime.now() 
                with open("errors.txt", "a") as file:
                    file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))
            else:
                ct = datetime.datetime.now() 
                with open("errors.txt", "x") as file:
                    file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))     



def tab3add_qual():
        try:
    #NONE DECLARED ERROR LABELS USED
            global fileerrorLabel
            global qualerrorLabel
            global blockerrorLabel
            global starterrorLabel
            global enderrorLabel      
            
        #SET COLUMN WIDTH
            # for col in range(1,31):
            #     column_letter = get_column_letter(col)
            #     # print(column_letter)
            #     sheet.column_dimensions[column_letter].width = 42

        #SET TITLE
            # sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=31)
            # sheet.row_dimensions[3].height = 60
            # title_Cell = sheet['A3']
            # title_Cell.border = thick_border
            # monthTitle = str(sheet.title)
            # # print(monthTitle)
            # title_Cell.value = monthTitle+" ELECTRICAL SYSTEMS INSTRUCTOR SCHEDULE"
            # title_Cell.font = Font(size=60,name="Times New Roman",bold=True)

            

            # #SET DATES
            # for i in range(1,31):
            #     datecellref=sheet.cell(row=1, column=i)
            #     datecellref.fill = PatternFill("solid", fgColor="DDDDDD")
                
            #     if(i<10):
            #         datecellref.value="01/0"+str(i)+"/2021"
            #     else:
            #         datecellref.value="01/"+str(i)+"/2021"

            #SET DIVIDER      

            for i in range(1,32):
                datecellref2=sheet.cell(row=4, column=i)
                datecellref2.fill = PatternFill("solid", fgColor="000000")
                datecellref2.border = thick_border_blue_topBottom
                datecellref2.value="blank"  


            #GETTING INPUT VALUES FROM USER
            print(qualcellref.value)
            qualName = str(qualcellref.value)
            if str(qualName) == "FAMILY DAY":
                None
            else:
                classNum = int(classcellref.value)
                if int(qualName) > 3:
                    qualerrorLabel = Label(tab3frame2,text = "Qual Num Doesn't Exist",font="Helvetica 10 bold", fg="red")
                    qualerrorLabel.grid(row=11,column=0)
                    e2.delete(0, END)
                    return 
                blockName = str(blockcellref.value)
                if int(blockName) > 6:
                    blockerrorLabel = Label(tab3frame2,text = "Block Number Doesn't Exist",font="Helvetica 10 bold", fg="red")
                    blockerrorLabel.grid(row=11,column=0)
                    e3.delete(0, END)
                    return
            startDate = str(startdatecellref.value)
            startDate = startDate[3:5]
            print(str(startDate))
            if int(startDate) > 31:
                starterrorLabel = Label(tab3frame2,text = "Start Date Number too big",font="Helvetica 10 bold", fg="red")
                starterrorLabel.grid(row=11,column=0)
                e4.delete(0, END)
                return
            # endDateReadIn = str(enddatecellref.value)
            # endDateReadIn = endDateReadIn[3:5]
            if int(endDateReadIn) > 31:
                enderrorLabel = Label(tab3frame2,text = "End Date Number too big",font="Helvetica 10 bold", fg="red")
                enderrorLabel.grid(row=11,column=0)
                e5.delete(0, END)
                return

            
            # print(classNum)
            

            #Start Date
            col_index = int(startDate)
            row_index=5
            # print("First row index: "+str(row_index))

            #THE DASHED CELLS ITERATORS
            # blank_col_index = col_index
            # blank_row_index=row_index+1
            
            #TITLE CELL
            cellref = sheet.cell(row=row_index,column=col_index)
            cellref.border = thick_border
            blockcheckcell = sheet.cell(row=row_index, column = 1)
        

            #DASHED CELLS
            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
        

            #End Date
            #check if months match






            des_col=int(endDateReadIn)+1

            
            
            total_index = 0

        

    

        
        
            q1check = False



            #COLUMN ITERATOR LOOP
            while col_index < des_col:
                
                #IF THERES A VALUE IN THE CELL
                while cellref.value:
                    if str(cellref.value) == "FAMILY DAY":
                        col_index+=1
                        cellref = sheet.cell(row=row_index,column=col_index)
                        None
                    
                    

                    #First: IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK IS LESS THAN OR EQUAL AND THE CLASS NUM ENTERED IS GREATER
                    elif(int(qualName) <= int(cellref.value[1]) and int(blockName) <= int(cellref.value[10]) and int(qualName) == 1 and int(blockName)==1):
                        print("Start Date: "+str(startDate))
                        print("Start Date Type: "+str(type(startDate)))
                        if startDate!="01" and (int(classNum) > int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999"):
                            sheet.insert_rows(5,5)
                            q1check = True
                            for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(startDate)-1):
                                for cell in rows:
                                    cell.value = "Q1  Block:1 Class:99999"
                                    cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type = "solid")
                                    cell.border = thick_border_blue
                            # row_index=4
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                        elif startDate!="01" and (int(classNum) < int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999"):
                            sheet.insert_rows(5,5)
                            q1check = True
                            for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(startDate)-1):
                                for cell in rows:
                                    cell.value = "Q1  Block:1 Class:99999"
                                    cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type = "solid")
                                    cell.border = thick_border_blue
                            row_index=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                        elif startDate=="1" and int(classNum) > int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999" :
                            sheet.insert_rows(5,5)
                            row_index=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        elif startDate=="1" and int(classNum) < int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999" :
                            row_index+=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                        #Second:IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK NUM IS GREATER OR EQUAL 
                    elif(int(qualName) <= int(cellref.value[1]) and int(blockName) > int(cellref.value[10])):
                        # print(cellref.value[18:23])
                        row_index+=5
                        print("Row index: " + str(row_index))
                        print(cellref.value)
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        # print("Skiped Title = " + cellref.coordinate)
                        # print("Skipped blank = " + blank_cellref.coordinate)

                        #Third:IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK IS LESS THAN OR EQUAL
                    elif(int(qualName) <= int(cellref.value[1]) and int(blockName) <= int(cellref.value[10])):
                        print(int(classNum))
                        print(int(cellref.value[18:23]))
                        if int(classNum) > int(cellref.value[18:23]):
                            print("Cell Coordinate: "+str(cellref))
                            print("Last Blank Cell Coordinate: "+str(blank_cellref4))
                            sheet.insert_rows(row_index,5)
                            workbook.save(filename=workbook_Title+".xlsx")
                            newrowref = cellref.row
                            print("Newrowref: "+str(newrowref))
                            row_index=newrowref-5
                            print("row index: "+str(row_index))
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        else:

                            row_index+=5
                        # print("Row index: " + str(row_index))
                        # print(cellref.value)
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        #FOURTH: IF THE ENTERED QUAL IS GREATER THAN THE CELL VALUE
                    elif(int(qualName) > int(cellref.value[1])):
                        # print(cellref.value[18:23])
                        row_index+=5
                        # print("Row index: " + str(row_index))
                        # print(cellref.value)
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        # print("Skiped Title = " + cellref.coordinate)
                        # print("Skipped blank = " + blank_cellref.coordinate)   
                    elif(int(qualName)<=int(cellref.value[1])):
                        sheet.move_range(""+str(cellref.coordinate)+":"+str(blank_cellref4.coordinate)+"", rows=5)
                        newrowref = cellref.row
                        row_index=newrowref-5
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        
                else:
                    if str(qualName)=="FAMILY DAY":
                        while row_index <=100:
                            cellref.value = "FAMILY DAY"
                            cellref.font = Font(color="FF0000")
                            cellref.fill = PatternFill("solid","FFFFFF")
                            cellref.border = thin_border_all_grey
                            row_index+=1
                            cellref = sheet.cell(row=row_index,column=col_index)
                    elif(int(qualName)==1 and int(blockName)==1 and startDate!="01" and q1check ==False):
                        sheet.insert_rows(5,5)
                        print("Row Index: "+str(row_index))
                        # print("Start Date: "+str(int(startDate)))
                        # print("col index: "+str(col_index))
                        q1check =True
                        for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(startDate)-1):
                                for cell in rows:
                                # print("cellref"+str(cellref.coordinate))
                                # print("\ncellref value: "+str(cellref.value))
                                    cell.value = "Q1  Block:1 Class:99999"
                                    cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type = "solid")
                                    cell.border = thick_border_blue
                                    print("col index: "+str(col_index))
                    # row_index=4
                    print("Row Index: "+str(row_index))
                    cellref = sheet.cell(row=row_index,column=col_index)
                    blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                    blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                    blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                    blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                    if str(qualName)=="FAMILY DAY":
                        None    
                    else:    
                        cellref.value="Q"+str(qualName)+" " +" Block:"+str(blockName)+" Class:"+str(classNum)
                        print("col index: "+str(col_index))
                            #TITLE COLOR CODING
                            #QUAL 1
                        cellref.border = thick_border
                        if (qualName == "1" and blockName == "1"):
                            cellref.fill = PatternFill("solid", fgColor="00CCFF")
                        elif (qualName == "1"and blockName == "2"):
                            cellref.fill = PatternFill("solid", fgColor="33CCCC")
                        elif (qualName == "1" and blockName == "3"):
                            cellref.fill = PatternFill("solid", fgColor="92D050")
                        elif (qualName == "1"and blockName == "4"):
                            cellref.fill = PatternFill("solid", fgColor="FFFF00")
                        elif (qualName == "1" and blockName == "5"):
                            cellref.fill = PatternFill("solid", fgColor="FFC000")
                        elif (qualName == "1"and blockName == "6"):
                            cellref.fill = PatternFill("solid", fgColor="FF0000")

                        #QUAL2
                        elif (qualName == "2" and blockName == "1"):
                            cellref.fill = PatternFill("solid", fgColor="00B0F0")
                        elif (qualName == "2"and blockName == "2"):
                            cellref.fill = PatternFill("solid", fgColor="92D050")

                        #QUAL3
                        elif (qualName == "3" and blockName == "1"):
                            cellref.fill = PatternFill("solid", fgColor="FFFF00")
                        elif (qualName == "3"and blockName == "2"):
                            cellref.fill = PatternFill("solid", fgColor="9BBB59")
                        elif (qualName == "3"and blockName == "3"):
                            cellref.fill = PatternFill("solid", fgColor="FFC000")


                        #SETTING BLANK VALUES BORDERS
                        blank_cellref.border = thin_border_sides
                        blank_cellref2.border = thin_border_sides
                        blank_cellref3.border = thin_border_sides
                        blank_cellref4.border = thin_border_sides_Bottom

                        #SETTING BLANK VALUES
                        blank_cellref.value="------"
                        blank_cellref2.value="------"
                        blank_cellref3.value="------"
                        blank_cellref4.value="------"
                    total_index+=1
                    # print("Title = " + cellref.coordinate)
                    # print("blank = " + blank_cellref.coordinate)
                col_index+=1
                # print("column index: "+str(col_index))   


                #This line controls if they shoot up the rows or nor
                # row_index=4
                # print("row index"+str(row_index))   
                cellref = sheet.cell(row=row_index,column=col_index)
                blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                print("Row Index: "+str(row_index))
                print("cellref: "+str(cellref))
                print(workbook_Title)
                workbook.save(filename=workbook_Title+".xlsx")


                
                
                global submitTotal
                submittedLabel = Label(tab3frame3,text = "Submitted ",font="Helvetica 10 bold", fg="black",bg="grey26")
                datesLabel = Label(tab3frame3,text = "Total Number of Dates Affected: "+str(submitTotal),font="Helvetica 10 bold", fg="black",bg="grey26")
                submittedLabel.grid(row=1,column=0)
                datesLabel.grid(row=2,column=0)
                submitTotal+=1
                e2.delete(0, END)
                e3.delete(0, END)
                e4.delete(0, END)
                e5.delete(0, END)

            second_Month_Check_ReadIn() 
        except:
            print(traceback.format_exc()) 
            messagebox.showwarning(title="Error Occured", message="something went wrong in Add Qual ReadIN. Check your entries and try again")
            if path.exists("errors.txt") == TRUE:
                ct = datetime.datetime.now() 
                with open("errors.txt", "a") as file:
                    file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))
            else:
                ct = datetime.datetime.now() 
                with open("errors.txt", "x") as file:
                    file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))
            





def add_qual():
        try:

        
    #NONE DECLARED ERROR LABELS USED
            global fileerrorLabel
            global qualerrorLabel
            global blockerrorLabel
            global starterrorLabel
            global enderrorLabel      
            sheet
        #SET COLUMN WIDTH
            # for col in range(1,32):
            #     column_letter = get_column_letter(col)
            #     # print(column_letter)
            #     sheet.column_dimensions[column_letter].width = 42

        #SET TITLE
            sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=32)
            sheet.row_dimensions[3].height = 60
            title_Cell = sheet['A3']
            title_Cell.border = thick_border
            monthTitle = str(sheet.title)
            # print(monthTitle)
            title_Cell.value = monthTitle+" ELECTRICAL SYSTEMS INSTRUCTOR SCHEDULE"
            title_Cell.font = Font(size=60,name="Times New Roman",bold=True)

            
            # global startMonth
            #SET DATES
            # for i in range(1,32):
            #     datecellref=sheet.cell(row=1, column=i)
            #     datecellref.fill = PatternFill("solid", fgColor="DDDDDD")
                
            #     if(i<10):
            #         datecellref.value=str(startMonth)+"/0"+str(i)+"/2021"
            #     else:
            #         datecellref.value=str(startMonth)+"/"+str(i)+"/2021"


            #SET DIVIDER      

            for i in range(1,32):
                datecellref2=sheet.cell(row=4, column=i)
                datecellref2.fill = PatternFill("solid", fgColor="000000")
                datecellref2.border = thick_border_blue_topBottom
                datecellref2.value="blank"  
                tab1.update_idletasks() 


            #GETTING INPUT VALUES FROM USER
            instructor = str(e7.get())
            qualName = e2.get()
            if str(qualName) == "FAMILY DAY":
                None
            else:
                qualName = e2.get()
                if int(qualName) > 3 and str(qualName) !="FAMILY DAY":
                    qualerrorLabel = Label(tab1frame2,text = "Qual Num Doesn't Exist",font="Helvetica 10 bold", fg="red")
                    qualerrorLabel.grid(row=11,column=0)
                    e2.delete(0, END)
                    return 
                blockName = e3.get()
                if int(blockName) > 6 and str(qualName) !="FAMILY DAY":
                    blockerrorLabel = Label(tab1frame2,text = "Block Number Doesn't Exist",font="Helvetica 10 bold", fg="red")
                    blockerrorLabel.grid(row=11,column=0)
                    e3.delete(0, END)
                    return
            startDate = e4.get()
            startDate = startDate[3:5]
            print(str(startDate))
            if int(startDate) > 32:
                starterrorLabel = Label(tab1frame2,text = "Start Date Number too big",font="Helvetica 10 bold", fg="red")
                starterrorLabel.grid(row=11,column=0)
                e4.delete(0, END)
                return
            if int(endDate) > 32:
                enderrorLabel = Label(tab1frame2,text = "End Date Number too big",font="Helvetica 10 bold", fg="red")
                enderrorLabel.grid(row=11,column=0)
                e5.delete(0, END)
                return

            classNum = e6.get()

            

            #Start Date
            col_index = int(startDate)
            row_index=5
            # print("First row index: "+str(row_index))

            #THE DASHED CELLS ITERATORS
            # blank_col_index = col_index
            # blank_row_index=row_index+1
            
            #TITLE CELL
            cellref = sheet.cell(row=row_index,column=col_index)
            cellref.border = thick_border
            blockcheckcell = sheet.cell(row=row_index, column = 1)
        

            #DASHED CELLS
            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
        

            #End Date
            #check if months match






            des_col=int(endDate)+1

            
            
            total_index = 0

        

    

        
        
            q1check = False



            #COLUMN ITERATOR LOOP
            while col_index < des_col: 
                
                
                #IF THERES A VALUE IN THE CELL
                while cellref.value:
                    print(str(cellref.value))
                    if str(cellref.value) == "FAMILY DAY":
                        col_index+=1
                        cellref = sheet.cell(row=row_index,column=col_index)
                        None
                    
                    elif str(cellref.value)!= "FAMILY DAY" and str(qualName)=="FAMILY DAY":
                        while row_index <=100:
                            cellref.value = "FAMILY DAY"
                            cellref.font = Font(color="FF0000")
                            cellref.fill = PatternFill("solid","FFFFFF")
                            cellref.border = thin_border_all_grey
                            row_index+=1
                            cellref = sheet.cell(row=row_index,column=col_index)

                    #First: IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK IS LESS THAN OR EQUAL AND THE CLASS NUM ENTERED IS GREATER
                    elif(int(qualName) <= int(cellref.value[1]) and int(blockName) <= int(cellref.value[10]) and int(qualName) == 1 and int(blockName)==1):
                        print("Start Date: "+str(startDate))
                        print("Start Date Type: "+str(type(startDate)))
                        if startDate!="01" and (int(classNum) >= int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999"):
                            sheet.insert_rows(5,5)
                            q1check = True
                            for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(startDate)-1):
                                for cell in rows:
                                    cell.value = "Q1  Block:1 Class:99999"
                                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid",fgColor="white")
                                    cell.border = thin_border_all_grey
                                    cell.font = Font(color="FFFFFF")
                            # row_index=4
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                        elif startDate!="01" and (int(classNum) < int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999"):
                            sheet.insert_rows(5,5)
                            q1check = True
                            for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(startDate)-1):
                                for cell in rows:
                                    cell.value = "Q1  Block:1 Class:99999"
                                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid",fgColor="white")
                                    cell.border = thin_border_all_grey
                                    cell.font = Font(color="FFFFFF")
                                    
                            row_index=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                        elif startDate=="1" and int(classNum) > int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999" :
                            sheet.insert_rows(5,5)
                            row_index=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        elif startDate=="1" and int(classNum) < int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999" :
                            row_index+=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                        #Second:IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK NUM IS GREATER OR EQUAL 
                    elif(int(qualName) <= int(cellref.value[1]) and int(blockName) > int(cellref.value[10])):
                        row_index+=5
                        print("Row index: " + str(row_index))
                        print(cellref.value)
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        

                        #Third:IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK IS LESS THAN OR EQUAL
                    elif(int(qualName) <= int(cellref.value[1]) and int(blockName) <= int(cellref.value[10])):
                        if int(classNum) > int(cellref.value[18:23]):
                            print("Cell Coordinate: "+str(cellref))
                            print("Last Blank Cell Coordinate: "+str(blank_cellref4))
                            sheet.insert_rows(row_index,5)
                            workbook.save(filename=workbook_Title+".xlsx")
                            newrowref = cellref.row
                            print("Newrowref: "+str(newrowref))
                            row_index=newrowref-5
                            print("row index: "+str(row_index))
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        else:

                            row_index+=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)


                        #FOURTH: IF THE ENTERED QUAL IS GREATER THAN THE CELL VALUE
                    elif(int(qualName) > int(cellref.value[1])):
                        row_index+=5
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)


                    elif(int(qualName)<=int(cellref.value[1])):
                        sheet.move_range(""+str(cellref.coordinate)+":"+str(blank_cellref4.coordinate)+"", rows=5)
                        newrowref = cellref.row
                        row_index=newrowref-5
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        
                else:
                    if str(qualName)=="FAMILY DAY":
                        while row_index <=100:
                            cellref.value = "FAMILY DAY"
                            cellref.font = Font(color="FF0000")
                            cellref.fill = PatternFill("solid","FFFFFF")
                            cellref.border = thin_border_all_grey
                            row_index+=1
                            cellref = sheet.cell(row=row_index,column=col_index)
                        
                    elif(int(qualName)==1 and int(blockName)==1 and startDate!="01" and q1check ==False):
                        sheet.insert_rows(5,5)
                        print("Row Index: "+str(row_index))
                        q1check =True
                        for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(startDate)-1):
                                for cell in rows:
                                    cell.value = "Q1  Block:1 Class:99999"
                                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid",fgColor="white")
                                    cell.border = thin_border_all_grey
                                    cell.font = Font(color="FFFFFF")
                                
                                    print("col index: "+str(col_index))
                    # row_index=4
                    print("Row Index: "+str(row_index))
                    cellref = sheet.cell(row=row_index,column=col_index)
                    blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                    blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                    blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                    blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                    if str(qualName)=="FAMILY DAY":
                        None    
                    else:
                        cellref.value="Q"+str(qualName)+" " +" Block:"+str(blockName)+" Class:"+str(classNum)
                    # print("col index: "+str(col_index))
                        #TITLE COLOR CODING
                        #QUAL 1
                        cellref.border = thick_border
                        if (qualName == "1" and blockName == "1"):
                            cellref.fill = PatternFill("solid", fgColor="00CCFF")
                        elif (qualName == "1"and blockName == "2"):
                            cellref.fill = PatternFill("solid", fgColor="33CCCC")
                        elif (qualName == "1" and blockName == "3"):
                            cellref.fill = PatternFill("solid", fgColor="92D050")
                        elif (qualName == "1"and blockName == "4"):
                            cellref.fill = PatternFill("solid", fgColor="FFFF00")
                        elif (qualName == "1" and blockName == "5"):
                            cellref.fill = PatternFill("solid", fgColor="FFC000")
                        elif (qualName == "1"and blockName == "6"):
                            cellref.fill = PatternFill("solid", fgColor="FF0000")

                        #QUAL2
                        elif (qualName == "2" and blockName == "1"):
                            cellref.fill = PatternFill("solid", fgColor="00B0F0")
                        elif (qualName == "2"and blockName == "2"):
                            cellref.fill = PatternFill("solid", fgColor="92D050")

                        #QUAL3
                        elif (qualName == "3" and blockName == "1"):
                            cellref.fill = PatternFill("solid", fgColor="FFFF00")
                        elif (qualName == "3"and blockName == "2"):
                            cellref.fill = PatternFill("solid", fgColor="9BBB59")
                        elif (qualName == "3"and blockName == "3"):
                            cellref.fill = PatternFill("solid", fgColor="FFC000")
                        


                        #SETTING BLANK VALUES BORDERS
                        blank_cellref.border = thin_border_sides
                        blank_cellref2.border = thin_border_sides
                        blank_cellref3.border = thin_border_sides
                        blank_cellref4.border = thin_border_sides_Bottom

                        #SETTING BLANK VALUES
                        blank_cellref.value="Instructor: "+instructor
                        blank_cellref.font = Font(bold=True)
                        blank_cellref2.value="------"
                        blank_cellref3.value="------"
                        blank_cellref4.value="MIRS"
                    total_index+=1
                    # print("Title = " + cellref.coordinate)
                    # print("blank = " + blank_cellref.coordinate)
                col_index+=1
                # print("column index: "+str(col_index))   


                #This line controls if they shoot up the rows or nor
                # row_index=4
                # print("row index"+str(row_index))   
                cellref = sheet.cell(row=row_index,column=col_index)
                blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                print("Row Index: "+str(row_index))
                print("cellref: "+str(cellref))
                workbook.save(filename=workbook_Title+".xlsx")


                
                
                global submitTotal
                submittedLabel = Label(tab1frame3,text = "Submitted ",font="Helvetica 10 bold", fg="black",bg="grey26")
                datesLabel = Label(tab1frame3,text = "Total Number of Dates Affected: "+str(submitTotal),font="Helvetica 8 bold", fg="black",bg="grey26")
                submittedLabel.grid(row=1,column=0,ipadx=50)
                datesLabel.grid(row = 2, column=0)
                submitTotal+=1
                # e2.delete(0, END)
                # e3.delete(0, END)
                # e4.delete(0, END)
                # e5.delete(0, END)
                my_progress2['value']+=30
                
                tab1.update_idletasks()

            second_Month_Check()
            my_progress2['value']=0
            
        except:
            print(traceback.format_exc())
            messagebox.showwarning(title="Error Occured", message="something went wrong in ADD QUAL.\nCheck your entries and try again.\nFor a more detailed explanation check errors txt file")
            if path.exists("errors.txt") == TRUE:
                ct = datetime.datetime.now() 
                with open("errors.txt", "a") as file:
                    file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))
            else:
                ct = datetime.datetime.now() 
                with open("errors.txt", "x") as file:
                    file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))




def second_Month_Check():
    if monthCheck == True:
        sheets = workbook.sheetnames
        # for i in sheets:
        #     print(i)
        global sheet
        print(str(sheet))
        sheet = workbook[sheets[sheetIndex+1]]
        print(str(sheet.title))
        # sheet = workbook.active
        global newstartDate
        newstartDate = "1"
        global newendDate
        newendDate = endDate = e5.get()
        newendDate = endDate[3:5]
        second_add_qual()
    else:
        None 

              

def second_Month_Check_ReadIn():
    if monthCheck == True:
        sheets = workbook.sheetnames
        # for i in sheets:
        #     print(i)
        global sheet
        # print(str(sheet))
        sheet = workbook[sheets[sheetIndex+1]]
        # print(str(sheet.title))
        # sheet = workbook.active
        global newstartDateReadIn
        newstartDateReadIn = "1"
        global newendDateReadIn
        newendDateReadIn = endDateReadIn=enddatecellref.value
        print("NEW END DATE Read In: "+str(endDateReadIn))
        newendDateReadIn = endDateReadIn[3:5]
        print(newendDateReadIn)
        second_add_qualReadIn()
    else:
        None                 
        

def second_add_qualReadIn():
        try:
        #NONE DECLARED ERROR LABELS USED
            global fileerrorLabel
            global qualerrorLabel
            global blockerrorLabel
            global starterrorLabel
            global enderrorLabel      
            
        #SET COLUMN WIDTH
            # for col in range(1,31):
            #     column_letter = get_column_letter(col)
            #     # print(column_letter)
            #     sheet.column_dimensions[column_letter].width = 42

        #SET TITLE
            sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=31)
            sheet.row_dimensions[3].height = 60
            title_Cell = sheet['A3']
            title_Cell.border = thick_border
            monthTitle = str(sheet.title)
            # print(monthTitle)
            title_Cell.value = monthTitle+" ELECTRICAL SYSTEMS INSTRUCTOR SCHEDULE"
            title_Cell.font = Font(size=60,name="Times New Roman",bold=True)

            

            #SET DATES
            for i in range(1,32):
                datecellref=sheet.cell(row=1, column=i)
                datecellref.fill = PatternFill("solid", fgColor="DDDDDD")
                
                if(i<10):
                    datecellref.value="01/0"+str(i)+"/2021"
                else:
                    datecellref.value="01/"+str(i)+"/2021"

            #SET DIVIDER      

            for i in range(1,32):
                datecellref2=sheet.cell(row=4, column=i)
                datecellref2.fill = PatternFill("solid", fgColor="000000")
                datecellref2.value="blank"  


            #GETTING INPUT VALUES FROM USER
            
            qualName = str(qualcellref.value)
            if str(qualName) == "FAMILY DAY":
                None
            else:
                print("Qual Name: "+str(qualName))
                if int(qualName) > 3:
                    qualerrorLabel = Label(tab1,text = "Qual Num Doesn't Exist",font="Helvetica 10 bold", fg="red")
                    qualerrorLabel.grid(row=11,column=0)
                    e2.delete(0, END)
                    return 
                blockName = str(blockcellref.value)
                print("Block Name: "+str(blockName))
                if int(blockName) > 6:
                    blockerrorLabel = Label(tab1,text = "Block Number Doesn't Exist",font="Helvetica 10 bold", fg="red")
                    blockerrorLabel.grid(row=11,column=0)
                    e3.delete(0, END)
                    return
                # startDate = e4.get()
                # startDate = startDate[3:5]
                # print(str(startDate))
                print("New Start Date: "+str(newstartDateReadIn))
            if int(newstartDateReadIn) > 32:
                starterrorLabel = Label(tab1,text = "Start Date Number too big",font="Helvetica 10 bold", fg="red")
                starterrorLabel.grid(row=11,column=0)
                e4.delete(0, END)
                return
            # endDate = e5.get()
            # endDate = endDate[3:5]
            print("New End Date: "+str(newendDateReadIn))
            if int(newendDateReadIn) > 32:
                enderrorLabel = Label(tab1,text = "End Date Number too big",font="Helvetica 10 bold", fg="red")
                enderrorLabel.grid(row=11,column=0)
                e5.delete(0, END)
                return

            classNum = int(classcellref.value)

            

            #Start Date
            col_index = int(newstartDateReadIn)
            row_index=5
            # print("First row index: "+str(row_index))

            #THE DASHED CELLS ITERATORS
            # blank_col_index = col_index
            # blank_row_index=row_index+1
            
            #TITLE CELL
            cellref = sheet.cell(row=row_index,column=col_index)
            cellref.border = thick_border
            blockcheckcell = sheet.cell(row=row_index, column = 1)
        

            #DASHED CELLS
            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
        

            #End Date
            #check if months match
            # startMonth = e4.get()
            # startMonth = startMonth[0:2]
            # endMonth = e5.get()
            # endMonth = endMonth[0:2]
            # monthCheck = False
            # if(startMonth == endMonth):
            #     None
            # elif(startMonth != endMonth):
            #     endDate = 30
            #     monthCheck =True





            des_col=int(newendDateReadIn)+1

            
            
            total_index = 0

        

    

        
        
            q1check = False



            #COLUMN ITERATOR LOOP
            while col_index < des_col:
                
                #IF THERES A VALUE IN THE CELL
                while cellref.value:
                    if str(cellref.value) == "FAMILY DAY":
                        col_index+=1
                        cellref = sheet.cell(row=row_index,column=col_index)
                        None
                    #First: IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK IS LESS THAN OR EQUAL AND THE CLASS NUM ENTERED IS GREATER
                    elif(int(qualName) <= int(cellref.value[1]) and int(blockName) <= int(cellref.value[10]) and int(qualName) == 1 and int(blockName)==1):
                        print("Start Date: "+str(newstartDate))
                        print("Start Date Type: "+str(type(newstartDate)))
                        if newstartDate!="1" and (int(classNum) > int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999"):
                            sheet.insert_rows(5,5)
                            q1check = True
                            for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(newstartDate)-1):
                                for cell in rows:
                                    cell.value = "Q1  Block:1 Class:99999"
                                    cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type = "solid")
                                    cell.border = thick_border_blue
                            # row_index=4
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                        elif newstartDate!="1" and (int(classNum) < int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999"):
                            sheet.insert_rows(5,5)
                            q1check = True
                            for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(newstartDate)-1):
                                for cell in rows:
                                    cell.value = "Q1  Block:1 Class:99999"
                                    cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type = "solid")
                                    cell.border = thick_border_blue
                            row_index=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                        elif newstartDate=="1" and int(classNum) > int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999" :
                            sheet.insert_rows(5,5)
                            row_index=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        elif newstartDate=="1" and int(classNum) < int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999" :
                            row_index+=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                        #Second:IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK NUM IS GREATER OR EQUAL 
                    elif(int(qualName) <= int(cellref.value[1]) and int(blockName) > int(cellref.value[10])):
                        # print(cellref.value[18:23])
                        row_index+=5
                        print("Row index: " + str(row_index))
                        print(cellref.value)
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        # print("Skiped Title = " + cellref.coordinate)
                        # print("Skipped blank = " + blank_cellref.coordinate)

                        #Third:IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK IS LESS THAN OR EQUAL
                    elif(int(qualName) <= int(cellref.value[1]) and int(blockName) <= int(cellref.value[10])):
                        if int(classNum) > int(cellref.value[18:23]):
                            print("Cell Coordinate: "+str(cellref))
                            print("Last Blank Cell Coordinate: "+str(blank_cellref4))
                            sheet.insert_rows(row_index,5)
                            workbook.save(filename=workbook_Title+".xlsx")
                            newrowref = cellref.row
                            print("Newrowref: "+str(newrowref))
                            row_index=newrowref-5
                            print("row index: "+str(row_index))
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        else:

                            row_index+=5
                        # print("Row index: " + str(row_index))
                        # print(cellref.value)
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        #FOURTH: IF THE ENTERED QUAL IS GREATER THAN THE CELL VALUE
                    elif(int(qualName) > int(cellref.value[1])):
                        # print(cellref.value[18:23])
                        row_index+=5
                        # print("Row index: " + str(row_index))
                        # print(cellref.value)
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        # print("Skiped Title = " + cellref.coordinate)
                        # print("Skipped blank = " + blank_cellref.coordinate)   
                    elif(int(qualName)<=int(cellref.value[1])):
                        sheet.move_range(""+str(cellref.coordinate)+":"+str(blank_cellref4.coordinate)+"", rows=5)
                        newrowref = cellref.row
                        row_index=newrowref-5
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                
                        
                else:
                    if str(qualName)=="FAMILY DAY":
                        while row_index <=100:
                            cellref.value = "FAMILY DAY"
                            cellref.font = Font(color="FF0000")
                            cellref.fill = PatternFill("solid","FFFFFF")
                            cellref.border = thin_border_all_grey
                            row_index+=1
                            cellref = sheet.cell(row=row_index,column=col_index)
                    elif(int(qualName)==1 and int(blockName)==1 and newstartDate!="1" and q1check ==False):
                        sheet.insert_rows(5,5)
                        print("Row Index: "+str(row_index))
                        # print("Start Date: "+str(int(startDate)))
                        # print("col index: "+str(col_index))
                        q1check =True
                        for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(newstartDate)-1):
                                for cell in rows:
                                # print("cellref"+str(cellref.coordinate))
                                # print("\ncellref value: "+str(cellref.value))
                                    cell.value = "Q1  Block:1 Class:99999"
                                    cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type = "solid")
                                    cell.border = thick_border_blue
                                    print("col index: "+str(col_index))
                    # row_index=4
                    print("Row Index: "+str(row_index))
                    cellref = sheet.cell(row=row_index,column=col_index)
                    blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                    blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                    blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                    blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                    if str(qualName)=="FAMILY DAY":
                        None    
                    else:    
                        cellref.value="Q"+str(qualName)+" " +" Block:"+str(blockName)+" Class:"+str(classNum)
                        print("col index: "+str(col_index))
                            #TITLE COLOR CODING
                            #QUAL 1
                        cellref.border = thick_border
                        if (qualName == "1" and blockName == "1"):
                            cellref.fill = PatternFill("solid", fgColor="00CCFF")
                        elif (qualName == "1"and blockName == "2"):
                            cellref.fill = PatternFill("solid", fgColor="33CCCC")
                        elif (qualName == "1" and blockName == "3"):
                            cellref.fill = PatternFill("solid", fgColor="92D050")
                        elif (qualName == "1"and blockName == "4"):
                            cellref.fill = PatternFill("solid", fgColor="FFFF00")
                        elif (qualName == "1" and blockName == "5"):
                            cellref.fill = PatternFill("solid", fgColor="FFC000")
                        elif (qualName == "1"and blockName == "6"):
                            cellref.fill = PatternFill("solid", fgColor="FF0000")

                        #QUAL2
                        elif (qualName == "2" and blockName == "1"):
                            cellref.fill = PatternFill("solid", fgColor="00B0F0")
                        elif (qualName == "2"and blockName == "2"):
                            cellref.fill = PatternFill("solid", fgColor="92D050")

                        #QUAL3
                        elif (qualName == "3" and blockName == "1"):
                            cellref.fill = PatternFill("solid", fgColor="FFFF00")
                        elif (qualName == "3"and blockName == "2"):
                            cellref.fill = PatternFill("solid", fgColor="9BBB59")
                        elif (qualName == "3"and blockName == "3"):
                            cellref.fill = PatternFill("solid", fgColor="FFC000")


                        #SETTING BLANK VALUES BORDERS
                        blank_cellref.border = thin_border_sides
                        blank_cellref2.border = thin_border_sides
                        blank_cellref3.border = thin_border_sides
                        blank_cellref4.border = thin_border_sides_Bottom

                        #SETTING BLANK VALUES
                        blank_cellref.value="------"
                        blank_cellref2.value="------"
                        blank_cellref3.value="------"
                        blank_cellref4.value="------"
                    total_index+=1
                    # print("Title = " + cellref.coordinate)
                    # print("blank = " + blank_cellref.coordinate)
                col_index+=1
                # print("column index: "+str(col_index))   


                #This line controls if they shoot up the rows or nor
                # row_index=4
                # print("row index"+str(row_index))   
                cellref = sheet.cell(row=row_index,column=col_index)
                blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                print("Row Index: "+str(row_index))
                print("cellref: "+str(cellref))
                workbook.save(filename=workbook_Title+".xlsx")


                
                
                global submitTotal
                submittedLabel = Label(tab3frame3,text = "Submitted ",font="Helvetica 10 bold", fg="black",bg="grey26")
                datesLabel = Label(tab3frame3,text = "Total Number of Dates Affected: "+str(submitTotal),font="Helvetica 10 bold", fg="black",bg="grey26")
                submittedLabel.grid(row=1,column=0,ipadx=50)
                datesLabel.grid(row = 2, column=0)
                submitTotal+=1
                e2.delete(0, END)
                e3.delete(0, END)
                e4.delete(0, END)
                e5.delete(0, END)

            # if monthCheck == True:
            #     global sheet
            #     sheet = workbook.active[2]
            #     startDate = 1
            #     endDate = endDate = e5.get()
            #     endDate = endDate[3:5]
            #     second_add_qual()
            # else:
            #     None
        except:
            print(traceback.format_exc())
            messagebox.showwarning(title="Error Occured", message="something went wrong in Second Qual Add ReadIN try again")
            if path.exists("errors.txt") == TRUE:
                ct = datetime.datetime.now() 
                with open("errors.txt", "a") as file:
                    file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))
            else:
                ct = datetime.datetime.now() 
                with open("errors.txt", "x") as file:
                    file.write("\n"+str(ct)+"\n"+str(traceback.format_exc())) 
            



def second_add_qual():
        try:
    #NONE DECLARED ERROR LABELS USED
            global fileerrorLabel
            global qualerrorLabel
            global blockerrorLabel
            global starterrorLabel
            global enderrorLabel      
            
        #SET COLUMN WIDTH
            # for col in range(1,31):
            #     column_letter = get_column_letter(col)
            #     # print(column_letter)
            #     sheet.column_dimensions[column_letter].width = 42

        #SET TITLE
            sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=32)
            sheet.row_dimensions[3].height = 60
            title_Cell = sheet['A3']
            title_Cell.border = thick_border
            monthTitle = str(sheet.title)
            # print(monthTitle)
            title_Cell.value = monthTitle+" ELECTRICAL SYSTEMS INSTRUCTOR SCHEDULE"
            title_Cell.font = Font(size=60,name="Times New Roman",bold=True)

            

            #SET DATES
            # for i in range(1,31):
            #     datecellref=sheet.cell(row=1, column=i)
            #     datecellref.fill = PatternFill("solid", fgColor="DDDDDD")
                
            #     if(i<10):
            #         datecellref.value="01/0"+str(i)+"/2021"
            #     else:
            #         datecellref.value="01/"+str(i)+"/2021"

            #SET DIVIDER      

            for i in range(1,32):
                datecellref2=sheet.cell(row=4, column=i)
                datecellref2.fill = PatternFill("solid", fgColor="000000")
                datecellref2.value="blank"  


            #GETTING INPUT VALUES FROM USER
            qualName = e2.get()
            if str(qualName) == "FAMILY DAY":
                None
            else:
                qualName = e2.get()
                print("Qual Name: "+str(qualName))
                if int(qualName) > 3:
                    qualerrorLabel = Label(tab1frame2,text = "Qual Num Doesn't Exist",font="Helvetica 10 bold", fg="red")
                    qualerrorLabel.grid(row=11,column=0)
                    e2.delete(0, END)
                    return 
                blockName = e3.get()
                print("Block Name: "+str(blockName))
                if int(blockName) > 6:
                    blockerrorLabel = Label(tab1frame2,text = "Block Number Doesn't Exist",font="Helvetica 10 bold", fg="red")
                    blockerrorLabel.grid(row=11,column=0)
                    e3.delete(0, END)
                    return
                # startDate = e4.get()
                # startDate = startDate[3:5]
                # print(str(startDate))
            print("New Start Date: "+str(newstartDate))
            if int(newstartDate) > 32:
                starterrorLabel = Label(tab1frame2,text = "Start Date Number too big",font="Helvetica 10 bold", fg="red")
                starterrorLabel.grid(row=11,column=0)
                e4.delete(0, END)
                return
            # endDate = e5.get()
            # endDate = endDate[3:5]
            print("New End Date: "+str(newendDate))
            if int(newendDate) > 32:
                enderrorLabel = Label(tab1frame2,text = "End Date Number too big",font="Helvetica 10 bold", fg="red")
                enderrorLabel.grid(row=11,column=0)
                e5.delete(0, END)
                return

            classNum = e6.get()

            

            #Start Date
            col_index = int(newstartDate)
            row_index=5
            # print("First row index: "+str(row_index))

            #THE DASHED CELLS ITERATORS
            # blank_col_index = col_index
            # blank_row_index=row_index+1
            
            #TITLE CELL
            cellref = sheet.cell(row=row_index,column=col_index)
            cellref.border = thick_border
            blockcheckcell = sheet.cell(row=row_index, column = 1)
        

            #DASHED CELLS
            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
        

            #End Date
            #check if months match
            # startMonth = e4.get()
            # startMonth = startMonth[0:2]
            # endMonth = e5.get()
            # endMonth = endMonth[0:2]
            # monthCheck = False
            # if(startMonth == endMonth):
            #     None
            # elif(startMonth != endMonth):
            #     endDate = 30
            #     monthCheck =True





            des_col=int(newendDate)+1

            
            
            total_index = 0

        

    

        
        
            q1check = False



            #COLUMN ITERATOR LOOP
            while col_index < des_col:
                
                #IF THERES A VALUE IN THE CELL
                while cellref.value:
                    print(str(cellref.value))
                    if str(cellref.value) == "FAMILY DAY":
                        col_index+=1
                        cellref = sheet.cell(row=row_index,column=col_index)
                        None
                    

                    #First: IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK IS LESS THAN OR EQUAL AND THE CLASS NUM ENTERED IS GREATER
                    elif(int(qualName) <= int(cellref.value[1]) and int(blockName) <= int(cellref.value[10]) and int(qualName) == 1 and int(blockName)==1):
                        print("Start Date: "+str(newstartDate))
                        print("Start Date Type: "+str(type(newstartDate)))
                        if newstartDate!="1" and (int(classNum) > int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999"):
                            sheet.insert_rows(5,5)
                            q1check = True
                            for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(newstartDate)-1):
                                for cell in rows:
                                    cell.value = "Q1  Block:1 Class:99999"
                                    cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type = "solid")
                                    cell.border = thick_border_blue
                            # row_index=4
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                        elif newstartDate!="1" and (int(classNum) < int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999"):
                            sheet.insert_rows(5,5)
                            q1check = True
                            for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(newstartDate)-1):
                                for cell in rows:
                                    cell.value = "Q1  Block:1 Class:99999"
                                    cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type = "solid")
                                    cell.border = thick_border_blue
                            row_index=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                        elif newstartDate=="1" and int(classNum) > int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999" :
                            sheet.insert_rows(5,5)
                            row_index=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        elif newstartDate=="1" and int(classNum) < int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999" :
                            row_index+=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                        #Second:IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK NUM IS GREATER OR EQUAL 
                    elif(int(qualName) <= int(cellref.value[1]) and int(blockName) > int(cellref.value[10])):
                        # print(cellref.value[18:23])
                        row_index+=5
                        print("Row index: " + str(row_index))
                        print(cellref.value)
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        # print("Skiped Title = " + cellref.coordinate)
                        # print("Skipped blank = " + blank_cellref.coordinate)

                        #Third:IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK IS LESS THAN OR EQUAL
                    elif(int(qualName) <= int(cellref.value[1]) and int(blockName) <= int(cellref.value[10])):
                        if int(classNum) > int(cellref.value[18:23]):
                            print("Cell Coordinate: "+str(cellref))
                            print("Last Blank Cell Coordinate: "+str(blank_cellref4))
                            sheet.insert_rows(row_index,5)
                            workbook.save(filename=workbook_Title+".xlsx")
                            newrowref = cellref.row
                            print("Newrowref: "+str(newrowref))
                            row_index=newrowref-5
                            print("row index: "+str(row_index))
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        else:

                            row_index+=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        #FOURTH: IF THE ENTERED QUAL IS GREATER THAN THE CELL VALUE
                    elif(int(qualName) > int(cellref.value[1])):
                        # print(cellref.value[18:23])
                        row_index+=5
                        # print("Row index: " + str(row_index))
                        # print(cellref.value)
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        # print("Skiped Title = " + cellref.coordinate)
                        # print("Skipped blank = " + blank_cellref.coordinate)   
                    elif(int(qualName)<=int(cellref.value[1])):
                        sheet.move_range(""+str(cellref.coordinate)+":"+str(blank_cellref4.coordinate)+"", rows=5)
                        newrowref = cellref.row
                        row_index=newrowref-5
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        
                else:
                    if str(qualName)=="FAMILY DAY":
                        while row_index <=100:
                            cellref.value = "FAMILY DAY"
                            cellref.font = Font(color="FF0000")
                            cellref.fill = PatternFill("solid","FFFFFF")
                            cellref.border = thin_border_all_grey
                            row_index+=1
                            cellref = sheet.cell(row=row_index,column=col_index)
                    elif(int(qualName)==1 and int(blockName)==1 and newstartDate!="1" and q1check ==False):
                        sheet.insert_rows(5,5)
                        print("Row Index: "+str(row_index))
                        # print("Start Date: "+str(int(startDate)))
                        # print("col index: "+str(col_index))
                        q1check =True
                        for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(newstartDate)-1):
                                for cell in rows:
                                # print("cellref"+str(cellref.coordinate))
                                # print("\ncellref value: "+str(cellref.value))
                                    cell.value = "Q1  Block:1 Class:99999"
                                    cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type = "solid")
                                    cell.border = thick_border_blue
                                    print("col index: "+str(col_index))
                    # row_index=4
                    print("Row Index: "+str(row_index))
                    cellref = sheet.cell(row=row_index,column=col_index)
                    blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                    blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                    blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                    blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                    if str(qualName)=="FAMILY DAY":
                        None    
                    else:    
                        cellref.value="Q"+str(qualName)+" " +" Block:"+str(blockName)+" Class:"+str(classNum)
                        print("col index: "+str(col_index))
                            #TITLE COLOR CODING
                            #QUAL 1
                        cellref.border = thick_border
                        if (qualName == "1" and blockName == "1"):
                            cellref.fill = PatternFill("solid", fgColor="00CCFF")
                        elif (qualName == "1"and blockName == "2"):
                            cellref.fill = PatternFill("solid", fgColor="33CCCC")
                        elif (qualName == "1" and blockName == "3"):
                            cellref.fill = PatternFill("solid", fgColor="92D050")
                        elif (qualName == "1"and blockName == "4"):
                            cellref.fill = PatternFill("solid", fgColor="FFFF00")
                        elif (qualName == "1" and blockName == "5"):
                            cellref.fill = PatternFill("solid", fgColor="FFC000")
                        elif (qualName == "1"and blockName == "6"):
                            cellref.fill = PatternFill("solid", fgColor="FF0000")

                        #QUAL2
                        elif (qualName == "2" and blockName == "1"):
                            cellref.fill = PatternFill("solid", fgColor="00B0F0")
                        elif (qualName == "2"and blockName == "2"):
                            cellref.fill = PatternFill("solid", fgColor="92D050")

                        #QUAL3
                        elif (qualName == "3" and blockName == "1"):
                            cellref.fill = PatternFill("solid", fgColor="FFFF00")
                        elif (qualName == "3"and blockName == "2"):
                            cellref.fill = PatternFill("solid", fgColor="9BBB59")
                        elif (qualName == "3"and blockName == "3"):
                            cellref.fill = PatternFill("solid", fgColor="FFC000")


                        #SETTING BLANK VALUES BORDERS
                        blank_cellref.border = thin_border_sides
                        blank_cellref2.border = thin_border_sides
                        blank_cellref3.border = thin_border_sides
                        blank_cellref4.border = thin_border_sides_Bottom

                        #SETTING BLANK VALUES
                        blank_cellref.value="------"
                        blank_cellref2.value="------"
                        blank_cellref3.value="------"
                        blank_cellref4.value="------"
                    total_index+=1
                    # print("Title = " + cellref.coordinate)
                    # print("blank = " + blank_cellref.coordinate)
                col_index+=1
                # print("column index: "+str(col_index))   


                #This line controls if they shoot up the rows or nor
                # row_index=4
                # print("row index"+str(row_index))   
                cellref = sheet.cell(row=row_index,column=col_index)
                blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                print("Row Index: "+str(row_index))
                print("cellref: "+str(cellref))
                workbook.save(filename=workbook_Title+".xlsx")


                
                
                global submitTotal
                submittedLabel = Label(tab1frame3,text = "Submitted ",font="Helvetica 10 bold", fg="black",bg="grey26")
                datesLabel = Label(tab1frame3,text = "Total Number of Dates Affected: "+str(submitTotal),font="Helvetica 8 bold", fg="black",bg="grey26")
                submittedLabel.grid(row=1,column=0,ipadx=50)
                datesLabel.grid(row = 2, column=0)
                submitTotal+=1
                e2.delete(0, END)
                e3.delete(0, END)
                e4.delete(0, END)
                e5.delete(0, END)

            # if monthCheck == True:
            #     global sheet
            #     sheet = workbook.active[2]
            #     startDate = 1
            #     endDate = endDate = e5.get()
            #     endDate = endDate[3:5]
            #     second_add_qual()
            # else:
            #     None
        except():
            print(traceback.format_exc())
            messagebox.showwarning(title="Error Occured", message="something went wrong in Second Add Month. Check your entries and try again")
            if path.exists("errors.txt") == TRUE:
                ct = datetime.datetime.now() 
                with open("errors.txt", "a") as file:
                    file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))
            else:
                ct = datetime.datetime.now() 
                with open("errors.txt", "x") as file:
                    file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))
            
             
            


#======================================================================================================#
#======================================================================================================#
#======================================================================================================#
#======================================================================================================#
#======================================================================================================#
#======================================================================================================#
#======================================================================================================#       




#MAIN PAGE
root = ThemedTk(theme="black")
root.geometry("800x500")
root.title("Curtis Scheduling Tool "+"                                                                                          \u00A9" + " KandyKane Solutions  Ver.4.0.0*")
tab_parent = ttk.Notebook(root)
tab1 = ttk.Frame(tab_parent)
tab2 = ttk.Frame(tab_parent)
tab3 = ttk.Frame(tab_parent)
tab4 = ttk.Frame(tab_parent)

tab_parent.add(tab1,text="Use Existing")
tab_parent.add(tab2,text="New Workbook")
tab_parent.add(tab3,text="Read From File")
tab_parent.add(tab4,text="About")
tab_parent.pack(expand=1,fill='both')


# bg = PhotoImage(file="background5.png")
# my_label = Label(root,image=bg)
# my_label.place(x=0,y=0,relwidth=1,relheight=1)
p1 = PhotoImage(file = 'airforce.png')
root.iconphoto(False, p1)




#TAB1==============================================================================================================TAB1
tab1titleframe=Frame(tab1,bg="grey26")
tab1titleframe.pack(side=TOP)
tab1mainframe = Frame(tab1,width=1000,bg="grey26", height=500,highlightbackground="black",highlightthickness=3)
tab1mainframe.pack(side=TOP,padx=(0,400))

tab1frame = Frame(tab1mainframe,bg="grey10")
tab1frame.pack()
tab1frame2 = Frame(tab1mainframe,bg="grey26")
tab1frame2.pack()
tab1frame3 = Frame(tab1mainframe,bg="grey26")
tab1frame3.pack()

tab1titlelabel = Label(tab1titleframe,text = "Existing Workbook File",fg="white",bg="grey26",font="Helvetica 36 bold").grid(row=1,column=0)
tab1Label0 = Label(tab1frame,text = "Entry",fg="white",bg="grey10",font="Helvetica 20 bold").grid(row=1,column=0,padx=140)
tab1Label0 = Label(tab1frame2,text = "File Name:",fg="white",bg="grey26",font="Helvetica 15 bold").grid(row=2,column=0,pady=2)
tab1Label1 = Label(tab1frame2,text = "Qual Num:",fg="white",bg="grey26",font="Helvetica 15 bold").grid(row=3,column=0,pady=2)
tab1Label2 = Label(tab1frame2,text = "Block Num:",fg="white",bg="grey26",font="Helvetica 15 bold").grid(row=4,column=0,pady=2)
tab1Label3 = Label(tab1frame2,text = "Start Date:",fg="white",bg="grey26",font="Helvetica 15 bold").grid(row=5,column=0,pady=2)
tab1Label4 = Label(tab1frame2,text = "End Date:",fg="white",bg="grey26",font="Helvetica 15 bold").grid(row=6,column=0,pady=2)
tab1Label5 = Label(tab1frame2,text = "Class Num:",fg="white",bg="grey26",font="Helvetica 15 bold").grid(row=7,column=0,pady=2)
tab1holidaylabel = Label(tab1frame2,text = "*For holiday and family\n days just put 'FAMILY DAY'\n in the qual entry",fg="white",bg="grey26",font="Helvetica 8 bold").grid(row=10,column=0,pady=2)
instructorLabel = Label(tab1frame2,text = "Instructor:",fg="white",bg="grey26",font="Helvetica 15 bold").grid(row=8,column=0,pady=2)


tab1Label2a = Label(tab1frame2,text = "'Ex.test'",fg=("grey50"),bg="grey26",font="Helvetica 15 bold").grid(row=2,column=2,pady=2)
tab1Label3a = Label(tab1frame2,text = "'Ex. 1-3'",fg=("grey50"),bg="grey26",font="Helvetica 15 bold").grid(row=3,column=2,pady=2)
tab1Label4a = Label(tab1frame2,text = "'Ex. 1-6'",fg=("grey50"),bg="grey26",font="Helvetica 15 bold").grid(row=4,column=2,pady=2)
tab1Label5a = Label(tab1frame2,text = "'01/01/21'",fg=("grey50"),bg="grey26",font="Helvetica 15 bold").grid(row=5,column=2,pady=2,padx=2)
tab1Label6a = Label(tab1frame2,text = "'01/28/21'",fg=("grey50"),bg="grey26",font="Helvetica 15 bold").grid(row=6,column=2,pady=2,padx=2)
tab1Label7a = Label(tab1frame2,text = "'200XX'",fg=("grey50"),bg="grey26",font="Helvetica 15 bold").grid(row=7,column=2,pady=2)
global my_progress2
my_progress2 = ttk.Progressbar(tab1frame3,orient=HORIZONTAL,length=200,mode="indeterminate")


#GIF
# file = 'loading.gif'
# info = Image.open(file)
# frames = info.n_frames
# print(frames)

# im = [PhotoImage(file=file,format=f"gif -index {i}") for i in range(frames)]

# count = 0
# anim = None
# def animation(count):
#     global anim
#     im2 = im[count]

#     gif_label.configure(image=im2)
#     count += 1
#     if count == frames:
#         count = 0
#     anim = tab1.after(50,lambda :animation(count))

# def stop_animation():
#     tab1.after_cancel(anim)

# gif_label = Label(tab1,image="",bg="grey26")
# gif_label.grid(row=4,column=5)

# start = Button(tab1,text="start",command=lambda :animation(count))
# start.grid(row=5,column=5)

# stop = Button(tab1,text="stop",command=stop_animation)
# stop.grid(row=6,column=5)
    

# gifframe = PhotoImage(file="loading-buffering.gif", format="gif -index 2")
# tab1Label7a = Label(tab1).grid(row=4,column=4,padx=50)
# tab1Label7a.configure(image=gifframe)


global e1
global e2
global e3
global e4
global e5
global e6
global e7



e1 = Entry(tab1frame2,width=10)
e1.grid(row=2,column=1)
e2 = Entry(tab1frame2,width=10)
e2.grid(row=3,column=1)
e3 = Entry(tab1frame2,width=10)
e3.grid(row=4,column=1)
e4 = Entry(tab1frame2,width=10)
e4.grid(row=5,column=1)
e5 = Entry(tab1frame2,width=10)
e5.grid(row=6,column=1)
e6 = Entry(tab1frame2,width=10)
e6.grid(row=7,column=1)
e7 = Entry(tab1frame2,width=10)
e7.grid(row=8,column=1)


myButton4 = Button(tab1frame2,text="Submit",command=existing_WORKBOOK,bg="grey80")
myButton4.grid(row=11,column=0)
changeOnHover(myButton4, "aqua", "grey80")
global submitTotal
submitTotal = 1




#TAB2====================================================ADD NEW WORKBOOK======================================================TAB2
tab2titleframe=Frame(tab2,bg="grey26")
tab2titleframe.pack()
tab2mainframe = Frame(tab2,bg="grey26")
tab2mainframe.pack()
myLabel0 = Label(tab2titleframe,text = "New Excel File",fg="white",bg="grey26",font="Helvetica 36 bold").grid(row=0,column=0)
myLabel0 = Label(tab2mainframe,text = "File Name:",fg="white",bg="grey26",font="Helvetica 15 bold").grid(row=0,column=0)
global newE1
global newe2
newe1 = Entry(tab2mainframe,width=20)
newe1.grid(row=2,column=0,pady=(0,10))
createButton = Button(tab2mainframe,text="Create New",command=new_WORKBOOK,bg="grey80")
createButton.grid(row=3,column=0)
changeOnHover(myButton4, "aqua", "grey80")



#TAB3============================READ FROM FILE========================================================TAB3#
tab3frame = Frame(tab3,bg="grey26")
tab3frame.grid(row=0,column=0)
tab3frame2 = Frame(tab3,bg="grey26")
tab3frame2.grid(row=0,column=1)
tab3frame3 = Frame(tab3,bg="grey26")
tab3frame3.grid(row=0,column=1)
tab3Label = Label(tab3frame,text = "Read from File",fg="white",bg="grey26",font="Helvetica 36 bold").pack()
tab3Labe2 = Label(tab3frame,text = "Read From File:",fg="white",bg="grey26",font="Helvetica 10 bold").pack()
global tab3e1
tab3e1 = Entry(tab3frame,width=20)
tab3e1.pack()
tab3Labe2 = Label(tab3frame,text = "Destination File:",fg="white",bg="grey26",font="Helvetica 10 bold").pack()
global tab3e2
tab3e2 = Entry(tab3frame,width=20)
tab3e2.pack()


global my_progress
my_progress = ttk.Progressbar(tab3frame3,orient=HORIZONTAL,length=300,mode="indeterminate")



tab3createButton = Button(tab3frame,text="Submit",command=readFromExcel,bg="grey80")
tab3createButton.pack(pady=(10))
changeOnHover(myButton4, "aqua", "grey80")



#TAB4================================================ABOUT====================================================================TAB3
mylabel2 = Label(tab4,text="About",font='Helvetica 30 bold')
about = """This is a simple scheduler program for automatically creating\nand editing tasks on defined days on an excel sheet when \ngiven prescribed dates. Hope you enjoy! """
mylabel = Label(tab4,text=about,font='Helvetica 12 bold')
mylabel2 = Label(tab4,text="How to Use",font='Helvetica 12 bold')
mylabel7 = Label(tab4,text="-Put the excel files you wish to edit in the same folder as this programs .exe file\n-When entering numbers all entries must be single digit\n-Do not separate the .exe file from the images",font='Helvetica 12 bold')
mylabel3 = Label(tab4,text="Use Existing:",font='Helvetica 12 bold')
mylabel4 = Label(tab4,text="This is for adding to an existing file. Simply, enter\n the file name(no extension) and enter\n the rest of your information accordingly.",font='Helvetica 12 bold')
mylabel5 = Label(tab4,text="Add WorkBook:",font='Helvetica 12 bold')
mylabel6 = Label(tab4,text="This is for creating a new excel file. Simply, enter what\n you would like to call the file(no extension,no special characters) and enter\n the rest of your information accordingly.You must initilize it with a Qual entry",font='Helvetica 12 bold')
mylabel.pack(pady=(0,30))
mylabel2.pack()
mylabel7.pack()
mylabel3.pack()
mylabel4.pack()
mylabel5.pack()
mylabel6.pack()
    





root.mainloop()


