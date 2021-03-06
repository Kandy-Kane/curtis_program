from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell.cell import ERROR_CODES
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.dimensions import SheetDimension
from openpyxl.utils import get_column_letter
from tkinter import *
import time
import os.path
from os import path


thick_border = Border(left=Side(style='thick'), 
                     right=Side(style='thick'), 
                     top=Side(style='thick'), 
                     bottom=Side(style='thick'))

thin_border_all = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

thin_border_sides = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style=None), 
                     bottom=Side(style=None))


thin_border_sides_Bottom = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style=None), 
                     bottom=Side(style='thin'))





def existing_WORKBOOK():
    global workbook_Title
    workbook_Title = e1.get()
    global workbook
    if path.exists(str(e1.get())+".xlsx") == True: 
        workbook = load_workbook(filename=workbook_Title+".xlsx")
    else:
        fileerrorLabel = Label(top,text="File Does Not Exist",font="Helvetica 10 bold", fg="red")
        fileerrorLabel.grid(row=8,column=0)
        return
    # print(path.exists(str(e1.get())+".xlsx"))
    global sheet
    sheet = workbook.active
    add_qual() 


#GLOBAL ERROR VARIABLES
fileerrorLabel = None
qualerrorLabel = None
blockerrorLabel = None
starterrorLabel = None
enderrorLabel = None

def add_qual():
        
    #NONE DECLARED ERROR LABELS USED
        global fileerrorLabel
        global qualerrorLabel
        global blockerrorLabel
        global starterrorLabel
        global enderrorLabel      
        sheet
    #SET COLUMN WIDTH
        for col in range(1,31):
            column_letter = get_column_letter(col)
            # print(column_letter)
            sheet.column_dimensions[column_letter].width = 42

    #SET TITLE
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=31)
        sheet.row_dimensions[2].height = 60
        title_Cell = sheet['A2']
        title_Cell.border = thick_border
        title_Cell.value = "SEPTEMBER ELECTRICAL SYSTEMS INSTRUCTOR SCHEDULE"
        title_Cell.font = Font(size=60,name="Times New Roman",bold=True)

        

        #SET DATES
        for i in range(1,31):
            datecellref=sheet.cell(row=1, column=i)
            datecellref.fill = PatternFill("solid", fgColor="DDDDDD")
            
            if(i<10):
                datecellref.value="01/0"+str(i)+"/2021"
            else:
                datecellref.value="01/"+str(i)+"/2021"

        #SET DIVIDER      

        for i in range(1,31):
            datecellref2=sheet.cell(row=3, column=i)
            datecellref2.fill = PatternFill("solid", fgColor="000000")
            datecellref2.value="blank"  


        #GETTING INPUT VALUES FROM USER
        
        qualName = e2.get()
        if int(qualName) > 3:
            qualerrorLabel = Label(top,text = "Qual Num Doesn't Exist",font="Helvetica 10 bold", fg="red")
            qualerrorLabel.grid(row=8,column=0)
            e2.delete(0, END)
            return 
        blockName = e3.get()
        if int(blockName) > 6:
            blockerrorLabel = Label(top,text = "Block Number Doesn't Exist",font="Helvetica 10 bold", fg="red")
            blockerrorLabel.grid(row=8,column=0)
            e3.delete(0, END)
            return
        startDate = e4.get()
        if int(startDate) > 31:
            starterrorLabel = Label(top,text = "Start Date Number too big",font="Helvetica 10 bold", fg="red")
            starterrorLabel.grid(row=8,column=0)
            e4.delete(0, END)
            return
        endDate = e5.get()
        if int(endDate) > 31:
            enderrorLabel = Label(top,text = "End Date Number too big",font="Helvetica 10 bold", fg="red")
            enderrorLabel.grid(row=8,column=0)
            e5.delete(0, END)
            return

        

        #Start Date
        col_index = int(startDate)
        row_index=4
        print("First row index: "+str(row_index))

        #THE DASHED CELLS ITERATORS
        # blank_col_index = col_index
        # blank_row_index=row_index+1
        
        #TITLE CELL
        cellref = sheet.cell(row=row_index,column=col_index)
        cellref.border = thick_border

       

        #DASHED CELLS
        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
        # print(cellref.coordinate)

        #End Date
        des_col=int(endDate)+1
        total_index = 0

        

        #COLUMN ITERATOR LOOP
        while col_index < des_col:
            
            #IF THERES A VALUE IN THE CELL
            while cellref.value:
                row_index+=1
                print(cellref.value[1])
                cellref = sheet.cell(row=row_index,column=col_index)
                blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                # print("Skiped Title = " + cellref.coordinate)
                # print("Skipped blank = " + blank_cellref.coordinate)            
            else:
                cellref.value="Q"+str(qualName)+" " +"Block:"+str(blockName)

                #TITLE COLOR CODING
                #QUAL 1
                cellref.border = thick_border
                if (qualName == "1" and blockName == "1"):
                    cellref.fill = PatternFill("solid", fgColor="00CCFF")
                elif (qualName == "1"and blockName == "2"):
                    cellref.fill = PatternFill("solid", fgColor="33CCCC")
                elif (qualName == "1" and blockName == "3"):
                    cellref.fill = PatternFill("solid", fgColor="92D050")
                elif (qualName == "1"and blockName == "4"):
                    cellref.fill = PatternFill("solid", fgColor="FFFF00")
                elif (qualName == "1" and blockName == "5"):
                    cellref.fill = PatternFill("solid", fgColor="FFC000")
                elif (qualName == "1"and blockName == "6"):
                    cellref.fill = PatternFill("solid", fgColor="FF0000")

                #QUAL2
                elif (qualName == "2" and blockName == "1"):
                    cellref.fill = PatternFill("solid", fgColor="00B0F0")
                elif (qualName == "2"and blockName == "2"):
                    cellref.fill = PatternFill("solid", fgColor="92D050")

                #QUAL3
                elif (qualName == "3" and blockName == "1"):
                    cellref.fill = PatternFill("solid", fgColor="FFFF00")
                elif (qualName == "3"and blockName == "2"):
                    cellref.fill = PatternFill("solid", fgColor="9BBB59")
                elif (qualName == "3"and blockName == "3"):
                    cellref.fill = PatternFill("solid", fgColor="FFC000")


                #SETTING BLANK VALUES BORDERS
                blank_cellref.border = thin_border_sides
                blank_cellref2.border = thin_border_sides
                blank_cellref3.border = thin_border_sides
                blank_cellref4.border = thin_border_sides_Bottom

                #SETTING BLANK VALUES
                blank_cellref.value="------"
                blank_cellref2.value="------"
                blank_cellref3.value="------"
                blank_cellref4.value="------"
                total_index+=1
                # print("Title = " + cellref.coordinate)
                # print("blank = " + blank_cellref.coordinate)
            col_index+=1
            print("column index: "+str(col_index))   

            row_index=4
            print("row index"+str(row_index))   
            cellref = sheet.cell(row=row_index,column=col_index)
            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
            workbook.save(filename=workbook_Title+".xlsx")


            
            
            global submitTotal
            submittedLabel = Label(top,text = "Submitted ",font="Helvetica 10 bold", fg="grey35")
            datesLabel = Label(top,text = "Total Number of Dates Affected: "+str(submitTotal),font="Helvetica 8 bold", fg="grey35")
            submittedLabel.grid(row=8,column=0,ipadx=30)
            datesLabel.grid(row = 9, column=0)
            submitTotal+=1
            e2.delete(0, END)
            e3.delete(0, END)
            e4.delete(0, END)
            e5.delete(0, END)
            
            
       



#======================================================================================================#
#======================================================================================================#


#======================================================================================================#

#======================================================================================================#

#======================================================================================================#
#======================================================================================================#
#======================================================================================================#
#======================================================================================================#
#======================================================================================================#
#======================================================================================================#
#======================================================================================================#       




def new_WORKBOOK():
    global workbook_Title
    workbook_Title = e1.get()
    if path.exists(str(e1.get())+".xlsx") == TRUE:
        fileerrorLabel = Label(top,text="File Already Exists",font="Helvetica 10 bold", fg="red")
        fileerrorLabel.grid(row=8,column=0)
        return
    else:
        global workbook 
        workbook = Workbook()
    global sheet
    sheet = workbook.active
    add_qual() 
 







def myClick():
    global top
    top = Toplevel()
    top.title("Add Qual")
    top.geometry("275x300")
    myLabel0 = Label(top,text = "Existing Excel File",font="Helvetica 12 bold").grid(row=1,column=0)
    myLabel0 = Label(top,text = "File Name:").grid(row=2,column=0,pady=2)
    myLabel1 = Label(top,text = "Qual Num:").grid(row=3,column=0,pady=2)
    myLabel2 = Label(top,text = "Block Num:").grid(row=4,column=0,pady=2)
    myLabel3 = Label(top,text = "Start Date:").grid(row=5,column=0,pady=2)
    myLabel4 = Label(top,text = "End Date:").grid(row=6,column=0,pady=2)
    global e1
    global e2
    global e3
    global e4
    global e5

    e1 = Entry(top,width=10)
    e1.grid(row=2,column=1)
    e2 = Entry(top,width=5)
    e2.grid(row=3,column=1)
    e3 = Entry(top,width=5)
    e3.grid(row=4,column=1)
    e4 = Entry(top,width=5)
    e4.grid(row=5,column=1)
    e5 = Entry(top,width=5)
    e5.grid(row=6,column=1)
    myButton4 = Button(top,text="Submit",command=existing_WORKBOOK,bg="grey80")
    myButton4.grid(row=7,column=0)
    changeOnHover(myButton4, "aqua", "grey80")
    global submitTotal
    submitTotal = 1




def myClickNew():
    global top
    top = Toplevel()
    top.title("Add Qual")
    top.geometry("275x300")
    myLabel0 = Label(top,text = "New Excel File",font="Helvetica 12 bold").grid(row=1,column=0)
    myLabel0 = Label(top,text = "File Name:").grid(row=2,column=0)
    myLabel1 = Label(top,text = "Qual Num:").grid(row=3,column=0)
    myLabel2 = Label(top,text = "Block Num:").grid(row=4,column=0)
    myLabel3 = Label(top,text = "Start Date:").grid(row=5,column=0)
    myLabel4 = Label(top,text = "End Date:").grid(row=6,column=0)
    global e1
    global e2
    global e3
    global e4
    global e5

    e1 = Entry(top,width=10)
    e1.grid(row=2,column=1)
    e2 = Entry(top,width=5)
    e2.grid(row=3,column=1)
    e3 = Entry(top,width=5)
    e3.grid(row=4,column=1)
    e4 = Entry(top,width=5)
    e4.grid(row=5,column=1)
    e5 = Entry(top,width=5)
    e5.grid(row=6,column=1)
    myButton4 = Button(top,text="Submit",command=new_WORKBOOK,bg="grey80")
    myButton4.grid(row=7,column=0)
    changeOnHover(myButton4, "aqua", "grey80")
    global submitTotal
    submitTotal = 1
    

def myClickAbout():
    top = Toplevel()
    top.title("About")
    top.geometry("650x380")
    mylabel2 = Label(top,text="About",font='Helvetica 30 bold')
    about = """This is a simple scheduler program for automatically creating\nand editing tasks on defined days on an excel sheet when \ngiven prescribed dates. Hope you enjoy! """
    mylabel = Label(top,text=about,font='Helvetica 12 bold')
    mylabel2 = Label(top,text="How to Use",font='Helvetica 12 bold')
    mylabel7 = Label(top,text="-Put the excel files you wish to edit in the same folder as this programs .exe file\n-When entering numbers all entries must be single digit\n-Do not separate the .exe file from the images",font='Helvetica 12 bold')
    mylabel3 = Label(top,text="Use Existing:",font='Helvetica 12 bold')
    mylabel4 = Label(top,text="This is for adding to an existing file. Simply, enter\n the file name(no extension) and enter\n the rest of your information accordingly.",font='Helvetica 12 bold')
    mylabel5 = Label(top,text="Add WorkBook:",font='Helvetica 12 bold')
    mylabel6 = Label(top,text="This is for creating a new excel file. Simply, enter what\n you would like to call the file(no extension,no special characters) and enter\n the rest of your information accordingly.You must initilize it with a Qual entry",font='Helvetica 12 bold')
    mylabel.pack(pady=(0,30))
    mylabel2.pack()
    mylabel7.pack()
    mylabel3.pack()
    mylabel4.pack()
    mylabel5.pack()
    mylabel6.pack()
    

def changeOnHover(button, colorOnHover, colorOnLeave):
  
    # adjusting backgroung of the widget
    # background on entering widget
    button.bind("<Enter>", func=lambda e: button.config(
        background=colorOnHover))
  
    # background color on leving widget
    button.bind("<Leave>", func=lambda e: button.config(
        background=colorOnLeave))

#MAIN PAGE
root = Tk()
bg = PhotoImage(file="background5.png")
my_label = Label(root,image=bg)
my_label.place(x=0,y=0,relwidth=1,relheight=1)
p1 = PhotoImage(file = 'bear.png')
root.iconphoto(False, p1)
root.geometry("380x430")
for i in range(3):
    root.columnconfigure(i, weight=1)

root.rowconfigure(1, weight=1)
main_Label = Label(root,text="Curtis Schedule Tool",font='Helvetica 18 bold')
myButton = Button(root,text="Add WorkBook",height = 2, width = 20,command=myClickNew, font='Helvetica 12 bold',bg="grey80")
myButton2 = Button(root,text="Use Existing",height = 2, width = 20,command=myClick,font='Helvetica 12 bold',bg="grey80")
myButton3 = Button(root,text="Exit",height = 2, width = 20,command=root.destroy,font='Helvetica 12 bold',bg="grey80")
myButton4 = Button(root,text="About",height = 2, width = 20,command=myClickAbout,font='Helvetica 12 bold',bg="grey80")
copyright_label = Label(root,text="\u00A9" + " KandyKane Solutions  Ver.1.0.0*",font='Helvetica 12 bold', fg="grey72")
main_Label.pack()
myButton.pack(fill=X, padx=80, pady=10)
myButton2.pack(fill=X,padx=80, pady=10)
myButton4.pack(fill=X,padx=80, pady=10)
myButton3.pack(fill=X,padx=80, pady=10)
changeOnHover(myButton, "aqua", "grey80")
changeOnHover(myButton2, "aqua", "grey80")
changeOnHover(myButton3, "aqua", "grey80")
changeOnHover(myButton4, "aqua", "grey80")
copyright_label.pack(side=BOTTOM, pady=(65,20))






root.mainloop()


