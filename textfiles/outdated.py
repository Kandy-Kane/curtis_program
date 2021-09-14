from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell import cell
from openpyxl.cell.cell import ERROR_CODES
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.dimensions import SheetDimension
from openpyxl.utils import get_column_letter
from tkinter import * 
import time
import os.path
from os import path
from openpyxl.styles import NamedStyle


# date_style = NamedStyle(name='custom_datetime', number_format='DD/MM/YYYY HH:MM:MM')
thick_border = Border(left=Side(style='thick'), 
                     right=Side(style='thick'), 
                     top=Side(style='thick'), 
                     bottom=Side(style='thick'))

thick_border_blue = Border(left=Side(style='thick',color='0066CC'), 
                     right=Side(style='thick',color='0066CC'), 
                     top=Side(style='thick',color='0066CC'), 
                     bottom=Side(style='thick',color='0066CC'))                     

thin_border_all = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

thin_border_sides = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style=None), 
                     bottom=Side(style=None))


thin_border_sides_Bottom = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style=None), 
                     bottom=Side(style='thin'))





def existing_WORKBOOK():
    global workbook_Title
    workbook_Title = e1.get()
    global workbook
    if path.exists(str(e1.get())+".xlsx") == True: 
        workbook = load_workbook(filename=workbook_Title+".xlsx")
        # workbook.add_named_style(date_style)
    else:
        fileerrorLabel = Label(top,text="File Does Not Exist",font="Helvetica 10 bold", fg="red")
        fileerrorLabel.grid(row=10,column=0)
        return
    # print(path.exists(str(e1.get())+".xlsx"))
    global sheet
    sheet = workbook.active
    if str(e7.get()) =="jan":
        sheet = workbook["Jan"]
    elif str(e7.get()) =="feb":
        sheet = workbook["Feb"]
    elif str(e7.get()) =="mar":
        sheet = workbook["Mar"]
    elif str(e7.get()) =="apr":
        sheet = workbook["Apr"]
    elif str(e7.get()) =="may":
        sheet = workbook["May"]
    elif str(e7.get()) =="jun":
        sheet = workbook["June"]
    elif str(e7.get()) =="jul":
        sheet = workbook["July"]
    elif str(e7.get()) =="aug":
        sheet = workbook["Aug"]
    elif str(e7.get()) =="sep":
        sheet = workbook["Sept"]
    elif str(e7.get()) =="oct":
        sheet = workbook["Oct"]
    elif str(e7.get()) =="nov":
        sheet = workbook["Nov"]
    elif str(e7.get()) =="dec":
        sheet = workbook["Dec"]
    add_qual() 






def new_WORKBOOK():
    global workbook_Title
    workbook_Title = e1.get()
    if path.exists(str(e1.get())+".xlsx") == TRUE:
        fileerrorLabel = Label(top,text="File Already Exists",font="Helvetica 10 bold", fg="red")
        fileerrorLabel.grid(row=10,column=0)
        return
    else:
        global workbook 
        workbook = Workbook()
        ws1 = workbook.create_sheet("Jan")
        ws2 = workbook.create_sheet("Feb")
        ws3 = workbook.create_sheet("Mar")
        ws4 = workbook.create_sheet("Apr")
        ws5 = workbook.create_sheet("May")
        ws6 = workbook.create_sheet("June")
        ws7 = workbook.create_sheet("July")
        ws8 = workbook.create_sheet("Aug")
        ws9 = workbook.create_sheet("Sept")
        ws10 = workbook.create_sheet("Oct")
        ws11 = workbook.create_sheet("Nov")
        ws12 = workbook.create_sheet("Dec")
    global sheet
    sheet = workbook.active
    if str(e7.get()) =="jan":
        sheet = ws1
    elif str(e7.get()) =="feb":
        sheet = ws2
    elif str(e7.get()) =="mar":
        sheet = ws3
    elif str(e7.get()) =="apr":
        sheet = ws4
    elif str(e7.get()) =="may":
        sheet = ws5
    elif str(e7.get()) =="june":
        sheet = ws6
    elif str(e7.get()) =="jul":
        sheet = ws7
    elif str(e7.get()) =="aug":
        sheet = ws8
    elif str(e7.get()) =="sep":
        sheet = ws9
    elif str(e7.get()) =="oct":
        sheet = ws10
    elif str(e7.get()) =="nov":
        sheet = ws11
    elif str(e7.get()) =="dec":
        sheet = ws12
    
    
    
    add_qual() 












#GLOBAL ERROR VARIABLES
fileerrorLabel = None
qualerrorLabel = None
blockerrorLabel = None
starterrorLabel = None
enderrorLabel = None

def add_qual():
        
    #NONE DECLARED ERROR LABELS USED
        global fileerrorLabel
        global qualerrorLabel
        global blockerrorLabel
        global starterrorLabel
        global enderrorLabel      
        sheet
    #SET COLUMN WIDTH
        for col in range(1,31):
            column_letter = get_column_letter(col)
            # print(column_letter)
            sheet.column_dimensions[column_letter].width = 42

    #SET TITLE
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=31)
        sheet.row_dimensions[2].height = 60
        title_Cell = sheet['A2']
        title_Cell.border = thick_border
        title_Cell.value = "SEPTEMBER ELECTRICAL SYSTEMS INSTRUCTOR SCHEDULE"
        title_Cell.font = Font(size=60,name="Times New Roman",bold=True)

        
        
        #SET DATES
        for i in range(1,31):
            datecellref=sheet.cell(row=1, column=i)
            datecellref.fill = PatternFill("solid", fgColor="DDDDDD")
            
            if(i<10):
                datecellref.value="01/0"+str(i)+"/2021"
            else:
                datecellref.value="01/"+str(i)+"/2021"
                

        #SET DIVIDER      

        for i in range(1,31):
            datecellref2=sheet.cell(row=3, column=i)
            datecellref2.fill = PatternFill("solid", fgColor="000000")
            datecellref2.value="blank"  


        #GETTING INPUT VALUES FROM USER
        
        qualName = e2.get()
        if int(qualName) > 3:
            qualerrorLabel = Label(top,text = "Qual Num Doesn't Exist",font="Helvetica 10 bold", fg="red")
            qualerrorLabel.grid(row=9,column=0)
            e2.delete(0, END)
            return 
        blockName = e3.get()
        if int(blockName) > 6:
            blockerrorLabel = Label(top,text = "Block Number Doesn't Exist",font="Helvetica 10 bold", fg="red")
            blockerrorLabel.grid(row=9,column=0)
            e3.delete(0, END)
            return
        startDate = e4.get()
        if int(startDate) > 31:
            starterrorLabel = Label(top,text = "Start Date Number too big",font="Helvetica 10 bold", fg="red")
            starterrorLabel.grid(row=9,column=0)
            e4.delete(0, END)
            return
        endDate = e5.get()
        if int(endDate) > 31:
            enderrorLabel = Label(top,text = "End Date Number too big",font="Helvetica 10 bold", fg="red")
            enderrorLabel.grid(row=9,column=0)
            e5.delete(0, END)
            return

        classNum = e6.get()

        

        #Start Date
        col_index = int(startDate)
        row_index=4
        # print("First row index: "+str(row_index))

        #THE DASHED CELLS ITERATORS
        # blank_col_index = col_index
        # blank_row_index=row_index+1
        
        #TITLE CELL
        cellref = sheet.cell(row=row_index,column=col_index)
        cellref.border = thick_border

        #DASHED CELLS
        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
        # print(cellref.coordinate)

        #End Date
        des_col=int(endDate)+1
        total_index = 0

      

        # cellref.value = sheet(row=row_index,column=col_index)
        # print("cellref coor: "+str(cellref.coordinate))
        # print ("cellref value: "+str(cellref.value))
        # print("blockref coor: "+str(blockcheckcell.coordinate))
        # print ("blockref value: "+str(blockcheckcell.value))

           

        







        #COLUMN ITERATOR LOOP
        while col_index < des_col:
            
            #IF THERES A VALUE IN THE CELL
            while cellref.value:
                

                #First: IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK IS LESS THAN OR EQUAL AND THE CLASS NUM ENTERED IS GREATER
                if(int(qualName) <= int(cellref.value[1]) and int(blockName) <= int(cellref.value[10]) and int(qualName) == 1 and int(blockName)==1):
                    print(str(cellref.value))
                    if int(classNum) > int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999":
                        sheet.insert_rows(4,5)
                        for rows in sheet.iter_rows(min_row=4,max_row=8,min_col=1,max_col=int(startDate)-1):
                            for cell in rows:
                            # print("cellref"+str(cellref.coordinate))
                            # print("\ncellref value: "+str(cellref.value))
                                cell.value = "Q1  Block:1 Class:99999"
                                cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type = "solid")
                                cell.border = thick_border_blue
                                # print("cell value: "+str(cell.value)+"\ncell coordinate: " + str(cell.coordinate))
                                # print("row index: "+str(row_index))
                        row_index=4
                        # for col in sheet.iter_cols(min_row=4, min_col = 1, max_col=des_col-1, max_row=9):
                        #     for cell in col:
                        #         cell.value = "Q1  Block:1 Class:99999"
                        #         cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type = "solid")
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index) 


                    #Second:IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK NUM IS GREATER OR EQUAL 
                elif(int(qualName) <= int(cellref.value[1]) and int(blockName) > int(cellref.value[10])):
                    # print(cellref.value[18:23])
                    row_index+=5
                    # print("Row index: " + str(row_index))
                    # print(cellref.value)
                    cellref = sheet.cell(row=row_index,column=col_index)
                    blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                    blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                    blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                    blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                    # print("Skiped Title = " + cellref.coordinate)
                    # print("Skipped blank = " + blank_cellref.coordinate)

                    #Third:IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK IS LESS THAN OR EQUAL
                elif(int(qualName) <= int(cellref.value[1]) and int(blockName) <= int(cellref.value[10])):
                    if int(classNum) > int(cellref.value[18:23]):
                        sheet.move_range(""+str(cellref.coordinate)+":"+str(blank_cellref4.coordinate)+"", rows=5)
                        newrowref = cellref.row
                        row_index=newrowref-5
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                    else:

                        row_index+=5
                    # print("Row index: " + str(row_index))
                    # print(cellref.value)
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                    #FOURTH: IF THE ENTERED QUAL IS GREATER THAN THE CELL VALUE
                elif(int(qualName) > int(cellref.value[1])):
                    # print(cellref.value[18:23])
                    row_index+=5
                    # print("Row index: " + str(row_index))
                    # print(cellref.value)
                    cellref = sheet.cell(row=row_index,column=col_index)
                    blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                    blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                    blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                    blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                    # print("Skiped Title = " + cellref.coordinate)
                    # print("Skipped blank = " + blank_cellref.coordinate)   
                elif(int(qualName)<=int(cellref.value[1])):
                    sheet.move_range(""+str(cellref.coordinate)+":"+str(blank_cellref4.coordinate)+"", rows=5)
                    newrowref = cellref.row
                    row_index=newrowref-5
                    cellref = sheet.cell(row=row_index,column=col_index)
                    blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                    blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                    blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                    blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                # elif(int(qualName) ==1 and int(blockName)==1):
                #     for col in sheet.iter_cols(min_row=4, min_col = 1, max_col=des_col-1, max_row=9):
                #         for cell in col:
                #             cell.value = "Q1  Block:1 Class:99999"
                #             cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type = "solid")



                # if(int(cellref.value[1]) <= int(qualName) and int(cellref.value[18:23]) > int(classNum)):
                #     print(cellref.value[18:23])
                #     row_index+=5
                #     print("Row index: " + str(row_index))
                #     print(cellref.value)
                #     cellref = sheet.cell(row=row_index,column=col_index)
                #     blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                #     blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                #     blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                #     blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                #     print("Skiped Title = " + cellref.coordinate)
                #     print("Skipped blank = " + blank_cellref.coordinate) 
                # elif(int(qualName)<int(cellref.value[1])):
                #     sheet.insert_rows(4,5)
                #     row_index=4
                #     cellref = sheet.cell(row=row_index,column=col_index)
                #     blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                #     blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                #     blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                #     blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)           
            else:
                if(int(qualName)==1 and int(blockName)==1):
                    print("col index: "+str(col_index))
                    for rows in sheet.iter_rows(min_row=4,max_row=8,min_col=1,max_col=int(startDate)-1):
                            for cell in rows:
                            # print("cellref"+str(cellref.coordinate))
                            # print("\ncellref value: "+str(cellref.value))
                                cell.value = "Q1  Block:1 Class:99999"
                                cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type = "solid")
                                cell.border = thick_border_blue
                                print("col index: "+str(col_index))
                cellref.value="Q"+str(qualName)+" " +" Block:"+str(blockName)+" Class:"+str(classNum)
                print("col index: "+str(col_index))
                    #TITLE COLOR CODING
                    #QUAL 1
                cellref.border = thick_border
                if (qualName == "1" and blockName == "1"):
                    cellref.fill = PatternFill("solid", fgColor="00CCFF")
                elif (qualName == "1"and blockName == "2"):
                    cellref.fill = PatternFill("solid", fgColor="33CCCC")
                elif (qualName == "1" and blockName == "3"):
                    cellref.fill = PatternFill("solid", fgColor="92D050")
                elif (qualName == "1"and blockName == "4"):
                    cellref.fill = PatternFill("solid", fgColor="FFFF00")
                elif (qualName == "1" and blockName == "5"):
                    cellref.fill = PatternFill("solid", fgColor="FFC000")
                elif (qualName == "1"and blockName == "6"):
                    cellref.fill = PatternFill("solid", fgColor="FF0000")

                #QUAL2
                elif (qualName == "2" and blockName == "1"):
                    cellref.fill = PatternFill("solid", fgColor="00B0F0")
                elif (qualName == "2"and blockName == "2"):
                    cellref.fill = PatternFill("solid", fgColor="92D050")

                #QUAL3
                elif (qualName == "3" and blockName == "1"):
                    cellref.fill = PatternFill("solid", fgColor="FFFF00")
                elif (qualName == "3"and blockName == "2"):
                    cellref.fill = PatternFill("solid", fgColor="9BBB59")
                elif (qualName == "3"and blockName == "3"):
                    cellref.fill = PatternFill("solid", fgColor="FFC000")


                #SETTING BLANK VALUES BORDERS
                blank_cellref.border = thin_border_sides
                blank_cellref2.border = thin_border_sides
                blank_cellref3.border = thin_border_sides
                blank_cellref4.border = thin_border_sides_Bottom

                #SETTING BLANK VALUES
                blank_cellref.value="------"
                blank_cellref2.value="------"
                blank_cellref3.value="------"
                blank_cellref4.value="------"
                total_index+=1
                # print("Title = " + cellref.coordinate)
                # print("blank = " + blank_cellref.coordinate)
            col_index+=1
            # print("column index: "+str(col_index))   


            #This line controls if they shoot up the rows or nor
            # row_index=4
            # print("row index"+str(row_index))   
            cellref = sheet.cell(row=row_index,column=col_index)
            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
            workbook.save(filename=workbook_Title+".xlsx")


            
            
            global submitTotal
            submittedLabel = Label(top,text = "Submitted ",font="Helvetica 10 bold", fg="grey35")
            datesLabel = Label(top,text = "Total Number of Dates Affected: "+str(submitTotal),font="Helvetica 8 bold", fg="grey35")
            submittedLabel.grid(row=10,column=0,ipadx=50)
            datesLabel.grid(row = 11, column=0)
            submitTotal+=1
            e2.delete(0, END)
            e3.delete(0, END)
            e4.delete(0, END)
            e5.delete(0, END)
                
                
        



#======================================================================================================#
#======================================================================================================#


#======================================================================================================#

#======================================================================================================#

#======================================================================================================#
#======================================================================================================#
#======================================================================================================#
#======================================================================================================#
#======================================================================================================#
#======================================================================================================#
#======================================================================================================#       





 







def myClick():
    global top
    top = Toplevel()
    top.title("Add Qual")
    top.geometry("275x300")
    myLabel0 = Label(top,text = "Existing Excel File",font="Helvetica 12 bold").grid(row=1,column=0)
    myLabel0 = Label(top,text = "File Name:").grid(row=2,column=0,pady=2)
    myLabel1 = Label(top,text = "Qual Num:").grid(row=3,column=0,pady=2)
    myLabel2 = Label(top,text = "Block Num:").grid(row=4,column=0,pady=2)
    myLabel3 = Label(top,text = "Start Date:").grid(row=5,column=0,pady=2)
    myLabel4 = Label(top,text = "End Date:").grid(row=6,column=0,pady=2)
    myLabel5 = Label(top,text = "Class Num:").grid(row=7,column=0,pady=2)
    myLabel6 = Label(top,text = "Month:").grid(row=8,column=0,pady=2)
   

    myLabel2a = Label(top,text = "'Ex.test'",fg=("grey50")).grid(row=2,column=2,pady=2)
    myLabel3a = Label(top,text = "'Ex. 1-3'",fg=("grey50")).grid(row=3,column=2,pady=2)
    myLabel4a = Label(top,text = "'Ex. 1-6'",fg=("grey50")).grid(row=4,column=2,pady=2)
    myLabel5a = Label(top,text = "'1-31'",fg=("grey50")).grid(row=5,column=2,pady=2)
    myLabel6a = Label(top,text = "'1-31'",fg=("grey50")).grid(row=6,column=2,pady=2)
    myLabel7a = Label(top,text = "'20010'",fg=("grey50")).grid(row=7,column=2,pady=2)
    myLabel8a = Label(top,text = "'jan'",fg=("grey50")).grid(row=8,column=2,pady=2)

    global e1
    global e2
    global e3
    global e4
    global e5
    global e6
    global e7

    e1 = Entry(top,width=10)
    e1.grid(row=2,column=1)
    e2 = Entry(top,width=5)
    e2.grid(row=3,column=1)
    e3 = Entry(top,width=5)
    e3.grid(row=4,column=1)
    e4 = Entry(top,width=5)
    e4.grid(row=5,column=1)
    e5 = Entry(top,width=5)
    e5.grid(row=6,column=1)
    e6 = Entry(top,width=8)
    e6.grid(row=7,column=1)
    e7 = Entry(top,width=5)
    e7.grid(row=8,column=1)

    myButton4 = Button(top,text="Submit",command=existing_WORKBOOK,bg="grey80")
    myButton4.grid(row=9,column=0)
    changeOnHover(myButton4, "aqua", "grey80")
    global submitTotal
    submitTotal = 1




def myClickNew():
    global top
    top = Toplevel()
    top.title("Add Qual")
    top.geometry("275x300")
    myLabel0 = Label(top,text = "New Excel File",font="Helvetica 12 bold").grid(row=1,column=0)
    myLabel0 = Label(top,text = "File Name:").grid(row=2,column=0)
    myLabel1 = Label(top,text = "Qual Num:").grid(row=3,column=0)
    myLabel2 = Label(top,text = "Block Num:").grid(row=4,column=0)
    myLabel3 = Label(top,text = "Start Date:").grid(row=5,column=0)
    myLabel4 = Label(top,text = "End Date:").grid(row=6,column=0)
    myLabel5 = Label(top,text = "Class Num:").grid(row=7,column=0,pady=2)
    myLabel6 = Label(top,text = "Month:").grid(row=8,column=0,pady=2)

    myLabel2a = Label(top,text = "'Ex.test'",fg=("grey50")).grid(row=2,column=2,pady=2)
    myLabel3a = Label(top,text = "'Ex. 1-3'",fg=("grey50")).grid(row=3,column=2,pady=2)
    myLabel4a = Label(top,text = "'Ex. 1-6'",fg=("grey50")).grid(row=4,column=2,pady=2)
    myLabel5a = Label(top,text = "'1-31'",fg=("grey50")).grid(row=5,column=2,pady=2)
    myLabel6a = Label(top,text = "'1-31'",fg=("grey50")).grid(row=6,column=2,pady=2)
    myLabel7a = Label(top,text = "'20010'",fg=("grey50")).grid(row=7,column=2,pady=2)
    myLabel8a = Label(top,text = "'jan'",fg=("grey50")).grid(row=8,column=2,pady=2)
    global e1
    global e2
    global e3
    global e4
    global e5
    global e6
    global e7
    e1 = Entry(top,width=10)
    e1.grid(row=2,column=1)
    e2 = Entry(top,width=5)
    e2.grid(row=3,column=1)
    e3 = Entry(top,width=5)
    e3.grid(row=4,column=1)
    e4 = Entry(top,width=5)
    e4.grid(row=5,column=1)
    e5 = Entry(top,width=5)
    e5.grid(row=6,column=1)
    e6 = Entry(top,width=8)
    e6.grid(row=7,column=1)
    e7 = Entry(top,width=5)
    e7.grid(row=8,column=1)
    myButton4 = Button(top,text="Submit",command=new_WORKBOOK,bg="grey80")
    myButton4.grid(row=9,column=0)
    changeOnHover(myButton4, "aqua", "grey80")
    global submitTotal
    submitTotal = 1
    

def myClickAbout():
    top = Toplevel()
    top.title("About")
    top.geometry("650x380")
    mylabel2 = Label(top,text="About",font='Helvetica 30 bold')
    about = """This is a simple scheduler program for automatically creating\nand editing tasks on defined days on an excel sheet when \ngiven prescribed dates. Hope you enjoy! """
    mylabel = Label(top,text=about,font='Helvetica 12 bold')
    mylabel2 = Label(top,text="How to Use",font='Helvetica 12 bold')
    mylabel7 = Label(top,text="-Put the excel files you wish to edit in the same folder as this programs .exe file\n-When entering numbers all entries must be single digit\n-Do not separate the .exe file from the images",font='Helvetica 12 bold')
    mylabel3 = Label(top,text="Use Existing:",font='Helvetica 12 bold')
    mylabel4 = Label(top,text="This is for adding to an existing file. Simply, enter\n the file name(no extension) and enter\n the rest of your information accordingly.",font='Helvetica 12 bold')
    mylabel5 = Label(top,text="Add WorkBook:",font='Helvetica 12 bold')
    mylabel6 = Label(top,text="This is for creating a new excel file. Simply, enter what\n you would like to call the file(no extension,no special characters) and enter\n the rest of your information accordingly.You must initilize it with a Qual entry",font='Helvetica 12 bold')
    mylabel.pack(pady=(0,30))
    mylabel2.pack()
    mylabel7.pack()
    mylabel3.pack()
    mylabel4.pack()
    mylabel5.pack()
    mylabel6.pack()
    

def changeOnHover(button, colorOnHover, colorOnLeave):
  
    # adjusting backgroung of the widget
    # background on entering widget
    button.bind("<Enter>", func=lambda e: button.config(
        background=colorOnHover))
  
    # background color on leving widget
    button.bind("<Leave>", func=lambda e: button.config(
        background=colorOnLeave))

#MAIN PAGE
root = Tk()
bg = PhotoImage(file="background5.png")
my_label = Label(root,image=bg)
my_label.place(x=0,y=0,relwidth=1,relheight=1)
p1 = PhotoImage(file = 'bear.png')
root.iconphoto(False, p1)
root.geometry("380x430")
for i in range(3):
    root.columnconfigure(i, weight=1)

root.rowconfigure(1, weight=1)
main_Label = Label(root,text="Curtis Schedule Tool",font='Helvetica 18 bold')
myButton = Button(root,text="Add WorkBook",height = 2, width = 20,command=myClickNew, font='Helvetica 12 bold',bg="grey80")
myButton2 = Button(root,text="Use Existing",height = 2, width = 20,command=myClick,font='Helvetica 12 bold',bg="grey80")
myButton3 = Button(root,text="Exit",height = 2, width = 20,command=root.destroy,font='Helvetica 12 bold',bg="grey80")
myButton4 = Button(root,text="About",height = 2, width = 20,command=myClickAbout,font='Helvetica 12 bold',bg="grey80")
copyright_label = Label(root,text="\u00A9" + " KandyKane Solutions  Ver.1.0.0*",font='Helvetica 12 bold', fg="grey72")
main_Label.pack()
myButton.pack(fill=X, padx=80, pady=10)
myButton2.pack(fill=X,padx=80, pady=10)
myButton4.pack(fill=X,padx=80, pady=10)
myButton3.pack(fill=X,padx=80, pady=10)
changeOnHover(myButton, "aqua", "grey80")
changeOnHover(myButton2, "aqua", "grey80")
changeOnHover(myButton3, "aqua", "grey80")
changeOnHover(myButton4, "aqua", "grey80")
copyright_label.pack(side=BOTTOM, pady=(65,20))






root.mainloop()


