from typing import List
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell import cell
from openpyxl.cell.cell import ERROR_CODES
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, alignment
from openpyxl.styles.fills import fills
from openpyxl.worksheet.dimensions import SheetDimension
from openpyxl.utils import get_column_letter
from tkinter import * 
from tkinter import ttk
from ttkthemes import ThemedTk
import time
import os.path
from os import error, path
from PIL import Image
from tkinter import messagebox
import traceback
import sys
import datetime

ft1 = Font(name='Arial',bold=True, size=14)
align = Alignment(horizontal='center')
thick_border = Border(left=Side(style='thick'), 
                     right=Side(style='thick'), 
                     top=Side(style='thick'), 
                     bottom=Side(style='thick'))

thick_border_blue = Border(left=Side(style='thick',color='0066CC'), 
                     right=Side(style='thick',color='0066CC')   
                  ) 

thick_border_blue_topBottom = Border(top=Side(style='thick',color='0066CC'), 
                     bottom=Side(style='thick',color='0066CC')   
                  )                    

thin_border_all = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

thin_border_all_grey = Border(left=Side(style='thin',color="DDDDDD"), 
                     right=Side(style='thin',color="DDDDDD"), 
                     top=Side(style='thin',color="DDDDDD"), 
                     bottom=Side(style='thin',color="DDDDDD"))

thin_border_sides = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style=None), 
                     bottom=Side(style=None))


thin_border_sides_Bottom = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style=None), 
                     bottom=Side(style='thin'))


#GLOBAL ERROR VARIABLES
fileerrorLabel = None
qualerrorLabel = None
blockerrorLabel = None
starterrorLabel = None
enderrorLabel = None
monthCheck = False

#GLOBAL MIRS
#QUAL1
global qual1mirs
qual1block1mirs=["3 MIRS / 1:6 / 3 HRS","0 MIRS","1 MIR / 1:12 / 3 HRS","0 MIRS"]
global qual2mirs
qual1block2mirs=["0 MIRS","0 MIRS","0 MIRS","1 MIR / 1:12 / 7 HRS","1 MIR / 1:12 / 4 HRS","1 MIR / 1:12 / 3 HRS","1 MIR / 1:12 / 2 HRS","1 MIR / 1:12 / 3.5 HRS","0 MIRS","1 MIR / 1:12 / 1.5HRS","1 MIR / 1:12 / 7.5HRS","1 MIR / 1:12 / 7.5HRS","1 MIR / 1:12 / 7.5HRS","0 MIRS","1 MIR / 1:12 / 7.5HRS","1 MIR / 1:12 / 7.5HRS","0 MIRS","0 MIRS"]
global qual1block3mirs
qual1block3mirs = [" 2 MIR / 1:8 / 4 HRS"," 2 MIR / 1:8 / 4.5 HRS"," 2 MIR / 1:8 / 5.5 HRS"," 2 MIR / 1:8 / 2.5 HRS"," 2 MIR / 1:8 / 4.5 HRS","0 MIRS"]
global qual1block4mirs
qual1block4mirs = ['2 MIRS / 1:8 / 6 HRS','2 MIRS/ 1:8 / 7 HRS','2 MIRS / 1:8 / 6 HRS','2 MIRS / 1:8 / 6.25 HRS','2 MIRS / 1:8 / 2 HRS','0 MIRS','2 MIRS / 1:8 / 5.5 HRS']
global qual1block5mirs
qual1block5mirs = ['0 MIRS','2 MIRS/ 1:8 / 3 HRS','2 MIRS/ 1:8 / 5.5 HRS','2 MIRS/ 1:8 / 5 HRS','2 MIRS/ 1:8 / 3.25 HRS','0 MIRS']
global qual1block6mirs
qual1block6mirs = ['1 MIRS / 1:12 / 3.5 HRS','1 MIRS / 1:12 / 5.75 HRS','1 MIRS / 1:12 / 4 HRS','1 MIRS / 1:12 / 6 HRS','1 MIRS / 1:12 / 5.25 HRS','1 MIRS / 1:12 / 4 HRS','1 MIRS / 1:12 / 6 HRS']
#QUAL2
global qual2block1mirs
qual2block1mirs =['1 MIR / 1:12 / 4 HRS','1 MIR / 1:12 / 5 HRS','3 MIRS / 1:6 / 8 HRS','3 MIRS / 1:6 / 8 HRS','3 MIRS / 1:6 / 8 HRS','3 MIRS / 1:6 / 8 HRS','3 MIRS / 1:6 / 8 HRS','2 MIRS / 1:8 / 8 HRS']
global qual2block2mirs
qual2block2mirs =['1 MIR / 1:12 / 3 HRS','1 MIR / 1:12 / 1 HR','3 MIRS / 1:6 / 3 HRS','3 MIRS / 1:6 / 8 HRS','3 MIRS / 1:6 / 8 HRS','3 MIR / 1:6 / 3 HRS','1 MIR / 1:12 / 2 HRS','2 MIRS / 1:8 / 6.5 HRS']
#QUAL3
global qual3block1mirs
qual3block1mirs = ['0 MIRS','1 MIRS / 1:12 / 4 HRS','1 MIRS / 1:12 / 8 HRS','1 MIRS / 1:12 / 3 HRS','1 MIRS / 1:12 / 5 HRS','2 MIRS / 1:8 / 8 HRS','2 MIRS / 1:8 / 8 HRS','1 MIRS / 1:12 / 8 HRS','1 MIRS / 1:12 / 8 HRS','1 MIRS / 1:12 / 6 HRS','1 MIRS / 1:12 / 6 HRS','1 MIRS / 1:12 / 1 HRS','0 MIRS']
global qual3block2mirs
qual3block2mirs =['1 MIRS / 1:12 / 3 HRS','1 MIRS / 1:12 / 8 HRS','1 MIRS / 1:12 / 8 HRS','1 MIRS / 1:12 / 8 HRS','1 MIRS / 1:12 / 8 HRS']
global qual3block3mirs
qual3block3mirs = ['1 MIRS / 1:12 / 1 HRS','1 MIRS / 1:12 / 8 HRS','1 MIRS / 1:12 / 5 HRS','1 MIRS / 1:12 / 8 HRS','1 MIRS / 1:12 / 4 HRS','2 MIRS / 1:8 / 6 HRS','2 MIRS / 1:8 / 8 HRS','2 MIRS / 1:8 / 4 HRS','0 MIRS']

def changeOnHover(button, colorOnHover, colorOnLeave):
  
    # adjusting backgroung of the widget
    # background on entering widget
    button.bind("<Enter>", func=lambda e: button.config(
        background=colorOnHover))
  
    # background color on leving widget
    button.bind("<Leave>", func=lambda e: button.config(
        background=colorOnLeave))

#============================================================================================================================#
#============================================================================================================================#
#============================================================================================================================#
    
    
def Workbook_ReadIn():
    global workbook_Title
    workbook_Title = tab3e2.get()
    print(workbook_Title)
    global workbook
    if path.exists(str(tab3e2.get())+".xlsx") == True: 
        workbook = load_workbook(filename=workbook_Title+".xlsx")
        # workbook.add_named_style(date_style)
    else:
        fileerrorLabel = Label(tab3,text="File Does Not Exist",font="Helvetica 10 bold", fg="red")
        fileerrorLabel.grid(row=11,column=0)
        return
    # print(path.exists(str(e1.get())+".xlsx"))
    global sheet
    sheet = workbook.active
    print(workbook)
    global sheetIndex
    #Check Months
    startMonth = startdatecellref.value
    startMonth = startMonth[0:2]
    print(startMonth)
    endMonth = enddatecellref.value
    endMonth = endMonth[0:2]
    print(endMonth)
    global endDateReadIn
    endDateReadIn = enddatecellref.value
    endDateReadIn = endDateReadIn[3:5]
    print("FIRST END DATE READ IN: "+endDateReadIn)
    global monthCheck

    # month = str(e4.get())
    # print("Month: "+str(month[0:2]))
    print("Start Month: "+str(startMonth))
    if str(startMonth[0:2]) =="01":
        if'JAN' in workbook.sheetnames:
            sheet = workbook["JAN"]
            sheetIndex = 1
        elif 'Jan' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("JAN")
            sheet = ws1
            sheetIndex = 1

    if str(startMonth[0:2]) =="02":
        if'FEB' in workbook.sheetnames:
            sheet = workbook["FEB"]
            sheetIndex =2
        elif 'Feb' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("FEB")
            sheet = ws1
            sheetIndex = 2
    if str(startMonth[0:2]) =="03":
        if'MAR' in workbook.sheetnames:
            sheet = workbook["MAR"]
            sheetIndex = 3
        elif 'MAR' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("MAR")
            sheet = ws1
            sheetIndex = 3
    if str(startMonth[0:2]) =="04":
        if'APR' in workbook.sheetnames:
            sheet = workbook["APR"]
            sheetIndex = 4
        elif 'APR' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("APR")
            sheet = ws1
            sheetIndex = 4
    if str(startMonth[0:2]) =="05":
        if'MAY' in workbook.sheetnames:
            sheet = workbook["MAY"]
            sheetIndex = 5
        elif 'MAY' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("MAY")
            sheet = ws1
            sheetIndex = 5

    if str(startMonth[0:2]) =="06":
        if'JUNE' in workbook.sheetnames:
            sheet = workbook["JUNE"]
            sheetIndex = 6
        elif 'JUNE' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("JUNE")
            sheet = ws1
            sheetIndex = 6

    if str(startMonth[0:2]) =="07":
        if'JULY' in workbook.sheetnames:
            sheet = workbook["JULY"]
            sheetIndex = 7
        elif 'JULY' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("JULY")
            sheet = ws1
            sheetIndex = 7

    if str(startMonth[0:2]) =="08":
        if'AUG' in workbook.sheetnames:
            sheet = workbook["AUG"]
            sheetIndex = 8
        elif 'AUG' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("AUG")
            sheet = ws1
            sheetIndex = 8

    if str(startMonth[0:2]) =="09":
        if'SEPT' in workbook.sheetnames:
            sheet = workbook["SEPT"]
            sheetIndex = 9
        elif 'SEPT' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("SEPT")
            sheet = ws1
            sheetIndex = 9
    if str(startMonth[0:2]) =="10":
        if'OCT' in workbook.sheetnames:
            sheet = workbook["OCT"]
            sheetIndex = 10
        elif 'OCT' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("OCT")
            sheet = ws1
            sheetIndex = 10
    if str(startMonth[0:2]) =="11":
        if'NOV' in workbook.sheetnames:
            sheet = workbook["NOV"]
            sheetIndex = 11
        elif 'NOV' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("NOV")
            sheet = ws1
            sheetIndex = 11

    if str(startMonth[0:1]) =="12":
        if'DEC' in workbook.sheetnames:
            sheet = workbook["DEC"]
            sheetIndex = 12
        elif 'DEC' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("DEC")
            sheet = ws1
            sheetIndex = 12

    if(startMonth == endMonth):
        None
        tab3add_qual()
    elif(startMonth != endMonth):
        endDateReadIn = "30"
        monthCheck =True
        tab3add_qual() 


    



def readFromExcel():
    try:
        tab3Label3 = Label(tab3frame3,text = "FINISHED",fg="grey26",bg="grey26",font="Helvetica 10 bold").grid(row=4,column=0)
        global my_progress
        my_progress.grid(row=0,column=0,pady=(50,10))

        #GETTING READ FROM FILE
        global readExcelFile
        readExcelFile = tab3e1.get()
        print(readExcelFile)
        if path.exists(str(tab3e1.get())+".xlsx") == True: 
            readFromBook = load_workbook(filename =readExcelFile+ '.xlsx')
        else:
            fileerrorLabel = Label(tab3,text="File Does Not Exist",font="Helvetica 10 bold", fg="red")
            fileerrorLabel.grid(row=5,column=0)
            return

        
        sheet = readFromBook.active
        print(sheet.title)
        startingRow = 2
        startingCol = 1
        global qualcellref
        global blockcellref
        global startdatecellref
        global enddatecellref
        global classcellref
        global instructorcellref
        qualcellref = sheet.cell(row=startingRow,column=startingCol)
        print(qualcellref.value)
        blockcellref = sheet.cell(row=startingRow,column=startingCol+1)
        print(blockcellref.value)
        startdatecellref = sheet.cell(row=startingRow,column=startingCol+2)
        print(startdatecellref.value)
        enddatecellref = sheet.cell(row=startingRow,column=startingCol+3)
        classcellref = sheet.cell(row=startingRow,column=startingCol+4)
        instructorcellref = sheet.cell(row=startingRow,column=startingCol+5)
        
        print(qualcellref.value)
        skittle = 1
        while qualcellref.value:
            
            my_progress['value']+=30
                
            tab3.update_idletasks() 
            print("QUALCELLREF VALUE: " +str(qualcellref.value))
            Workbook_ReadIn()
            print(skittle)
            startingRow+=1
            qualcellref = sheet.cell(row=startingRow,column=startingCol)
            print(qualcellref.value)
            blockcellref = sheet.cell(row=startingRow,column=startingCol+1)
            print(blockcellref.value)
            startdatecellref = sheet.cell(row=startingRow,column=startingCol+2)
            print(startdatecellref.value)
            enddatecellref = sheet.cell(row=startingRow,column=startingCol+3)
            classcellref = sheet.cell(row=startingRow,column=startingCol+4)
            instructorcellref = sheet.cell(row=startingRow,column=startingCol+5)
            skittle+=1
        tab3Label3 = Label(tab3frame3,text = "FINISHED",fg="white",bg="grey26",font="Helvetica 10 bold").grid(row=4,column=0)
        
        my_progress.stop()
    except():
        print(traceback.format_exc())
        messagebox.showwarning(title="Error Occured", message="something went wrong in READ FROM EXCEL. Check your entries and try again")
        if path.exists("errors.txt") == TRUE:
            ct = datetime.datetime.now() 
            with open("errors.txt", "a") as file:
                file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))
        else:
            ct = datetime.datetime.now() 
            with open("errors.txt", "x") as file:
                file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))
 
        




def existing_WORKBOOK():
    try:
        finishedLabel = Label(tab1frame3,text="FINISHED",font="Helvetica 10 bold", fg="grey26",bg="grey26").grid(row=4,column=0) 
        global my_progress2
        my_progress2.grid(row=0,column=0)
        global workbook_Title
        workbook_Title = e1.get()
        global workbook
        if path.exists(str(e1.get())+".xlsx") == True: 
            workbook = load_workbook(filename=workbook_Title+".xlsx")
            # workbook.add_named_style(date_style)
        else:
            fileerrorLabel = Label(tab1frame3,text="File Does Not Exist",font="Helvetica 10 bold", fg="red")
            fileerrorLabel.grid(row=2,column=0)
            return
        # print(path.exists(str(e1.get())+".xlsx"))
        global sheet
        sheet = workbook.active
        print(workbook)
        global sheetIndex
        #Check Months
        global startMonth
        startMonth = e4.get()
        startMonth = startMonth[0:2]
        endMonth = e5.get()
        endMonth = endMonth[0:2]
        global endDate
        endDate = e5.get()
        endDate = endDate[3:5]
        global monthCheck

        # month = str(e4.get())
        # print("Month: "+str(month[0:2]))
        print("Start Month: "+str(startMonth))
        tab1.update_idletasks() 
        if str(startMonth[0:2]) =="01":
            if'JAN' in workbook.sheetnames:
                sheet = workbook["JAN"]
                sheetIndex = 1
            elif 'Jan' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("JAN")
                sheet = ws1
                sheetIndex = 1

        if str(startMonth[0:2]) =="02":
            if'FEB' in workbook.sheetnames:
                sheet = workbook["FEB"]
                sheetIndex =2
            elif 'Feb' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("FEB")
                sheet = ws1
                sheetIndex = 2
        if str(startMonth[0:2]) =="03":
            if'MAR' in workbook.sheetnames:
                sheet = workbook["MAR"]
                sheetIndex = 3
            elif 'MAR' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("MAR")
                sheet = ws1
                sheetIndex = 3
        if str(startMonth[0:2]) =="04":
            if'APR' in workbook.sheetnames:
                sheet = workbook["APR"]
                sheetIndex = 4
            elif 'APR' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("APR")
                sheet = ws1
                sheetIndex = 4
        if str(startMonth[0:2]) =="05":
            if'MAY' in workbook.sheetnames:
                sheet = workbook["MAY"]
                sheetIndex = 5
            elif 'MAY' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("MAY")
                sheet = ws1
                sheetIndex = 5

        if str(startMonth[0:2]) =="06":
            if'JUNE' in workbook.sheetnames:
                sheet = workbook["JUNE"]
                sheetIndex = 6
            elif 'JUNE' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("JUNE")
                sheet = ws1
                sheetIndex = 6

        if str(startMonth[0:2]) =="07":
            if'JULY' in workbook.sheetnames:
                sheet = workbook["JULY"]
                sheetIndex = 7
            elif 'JULY' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("JULY")
                sheet = ws1
                sheetIndex = 7

        if str(startMonth[0:2]) =="08":
            if'AUG' in workbook.sheetnames:
                sheet = workbook["AUG"]
                sheetIndex = 8
            elif 'AUG' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("AUG")
                sheet = ws1
                sheetIndex = 8

        if str(startMonth[0:2]) =="09":
            if'SEPT' in workbook.sheetnames:
                sheet = workbook["SEPT"]
                sheetIndex = 9
            elif 'SEPT' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("SEPT")
                sheet = ws1
                sheetIndex = 9
        if str(startMonth[0:2]) =="10":
            if'OCT' in workbook.sheetnames:
                sheet = workbook["OCT"]
                sheetIndex = 10
            elif 'OCT' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("OCT")
                sheet = ws1
                sheetIndex = 10
        if str(startMonth[0:2]) =="11":
            if'NOV' in workbook.sheetnames:
                sheet = workbook["NOV"]
                sheetIndex = 11
            elif 'NOV' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("NOV")
                sheet = ws1
                sheetIndex = 11

        if str(startMonth[0:1]) =="12":
            if'DEC' in workbook.sheetnames:
                sheet = workbook["DEC"]
                sheetIndex = 12
            elif 'DEC' not in workbook.sheetnames:
                ws1 = workbook.create_sheet("DEC")
                sheet = ws1
                sheetIndex = 12

        if(startMonth == endMonth):
            None
            add_qual() 
        elif(startMonth != endMonth):
            endDate = 30
            monthCheck =True
            add_qual()
        finishedLabel = Label(tab1frame3,text="FINISHED",font="Helvetica 10 bold", fg="white",bg="grey26").grid(row=4,column=0) 
        my_progress.stop() 
    except():
        print(traceback.format_exc())
        messagebox.showwarning(title="Error Occured", message="something went wrong in EXISTING BOOK. Check your entries and try again")
        if path.exists("errors.txt") == TRUE:
            ct = datetime.datetime.now() 
            with open("errors.txt", "a") as file:
                file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))
        else:
            ct = datetime.datetime.now() 
            with open("errors.txt", "x") as file:
                file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))



def new_WORKBOOK():
    try:
        global workbook_Title
        workbook_Title = newe1.get()
        if path.exists(str(newe1.get())+".xlsx") == TRUE:
            fileerrorLabel = Label(tab2mainframe,text="File Already Exists",font="Helvetica 10 bold", fg="red")
            fileerrorLabel.grid(row=5,column=0)
            return
        else:
            global workbook 
            workbook = Workbook()
            # sheet = workbook.active
            ws1 = workbook.create_sheet("JAN")
            ws2 = workbook.create_sheet("FEB")
            ws3 = workbook.create_sheet("MAR")
            ws4 = workbook.create_sheet("APR")
            ws5 = workbook.create_sheet("MAY")
            ws6 = workbook.create_sheet("JUNE")
            ws7 = workbook.create_sheet("JULY")
            ws8 = workbook.create_sheet("AUG")
            ws9 = workbook.create_sheet("SEPT")
            ws10 = workbook.create_sheet("OCT")
            ws11 = workbook.create_sheet("NOV")
            ws12 = workbook.create_sheet("DEC")

            #ADD DAYS TO THE SHEET
        
        
        def sheetDates():
            
            startmonth=""
            dayList = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
            global day_index
            day_index=4
            def addDaysAndDates():
                for col in range(1,32):
                    column_letter = get_column_letter(col)
                    # print(column_letter)
                    sheet.column_dimensions[column_letter].width = 42
                print(sheet.title)
                global day_index
                col_index=1
                print(endDay)
                print("Day INdex: "+str(day_index))
                while col_index<endDay:
                    if day_index ==6 and col_index<endDay:
                        daycellref=sheet.cell(row=2, column=col_index) 
                        daycellref.value = (dayList[day_index])
                        daycellref.border = thick_border_blue_topBottom
                        day_index = 0
                        col_index+=1
                        # workbook.save(filename=workbook_Title+".xlsx")
                    else:
                        daycellref=sheet.cell(row=2, column=col_index) 
                        daycellref.value = (dayList[day_index])
                        day_index+=1
                        col_index+=1
                        # workbook.save(filename=workbook_Title+".xlsx")
                    
                # workbook.save(filename=workbook_Title+".xlsx")
                
                for i in range(1,endDay):
                    print("Enday: "+str(endDay))
                    daycellref=sheet.cell(row=2, column=i)
                    daycellref.fill = PatternFill("solid", fgColor="DDDDDD")
                    if daycellref.value == "Saturday":
                        column_letter = get_column_letter(i)
                        sheet.column_dimensions[column_letter].width = 0.1
                        for a in range(5,100):
                            daycellref2 = sheet.cell(row=a,column=i)
                            daycellref2.font = Font(size=1)
                        # workbook.save(filename=workbook_Title+".xlsx")
                    elif daycellref.value == "Sunday":
                        column_letter = get_column_letter(i)
                        sheet.column_dimensions[column_letter].width = 0.1
                        for a in range(5,100):
                            daycellref2 = sheet.cell(row=a,column=i)
                            daycellref2.font = Font(size=1)
                        # workbook.save(filename=workbook_Title+".xlsx")

            for sheet in workbook.worksheets:
                endDay = 32
                print(sheet.title)
                if str(sheet.title)=="JAN" or str(sheet.title)=="MAR"or str(sheet.title)=="MAY"or str(sheet.title)=="JULY"or str(sheet.title)=="AUG"or str(sheet.title)=="OCT"or str(sheet.title)=="DEC":
                    endDay = 32
                    addDaysAndDates()
                elif str(sheet.title) == "APR"or str(sheet.title)=="JUNE"or str(sheet.title)=="SEPT"or str(sheet.title)=="NOV":
                    endDay = 31
                    addDaysAndDates()
                elif str(sheet.title) == "FEB":
                    endDay = 29
                    addDaysAndDates()
                else:
                    continue

                for i in range(1,endDay):
                    if str(sheet.title) == "JAN":
                        startmonth = "01"
                    elif str(sheet.title) == "FEB":
                        startmonth = "02"
                    elif str(sheet.title) == "MAR":
                        startmonth = "03"
                    elif str(sheet.title) == "APR":
                        startmonth = "04"
                    elif str(sheet.title) == "MAY":
                        startmonth = "05"
                    elif str(sheet.title) == "JUNE":
                        startmonth = "06"
                    elif str(sheet.title) == "JULY":
                        startmonth = "07"
                    elif str(sheet.title) == "AUG":
                        startmonth = "08"
                    elif str(sheet.title) == "SEPT":
                        startmonth = "09"
                    elif str(sheet.title) == "OCT":
                        startmonth = "10"
                    elif str(sheet.title) == "NOV":
                        startmonth = "11"
                    elif str(sheet.title) == "DEC":
                        startmonth = "12"
                    datecellref=sheet.cell(row=1, column=i)
                    datecellref.fill = PatternFill("solid", fgColor="DDDDDD")
                    if(i<10):
                        datecellref.value=str(startmonth)+"/0"+str(i)+"/21"
                    else:
                        datecellref.value=str(startmonth)+"/"+str(i)+"/21"

                #SETTING TITLE
                sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=32)
                sheet.row_dimensions[3].height = 60
                title_Cell = sheet['A3']
                title_Cell.border = thick_border
                monthTitle = str(sheet.title)
                # print(monthTitle)
                title_Cell.value = monthTitle+" ELECTRICAL SYSTEMS INSTRUCTOR SCHEDULE"
                title_Cell.font = Font(size=60,name="Times New Roman",bold=True)

                #SETTING DIVIDER
                for i in range(1,32):
                    datecellref2=sheet.cell(row=4, column=i)
                    datecellref2.fill = PatternFill("solid", fgColor="000000")
                    datecellref2.border = thick_border_blue_topBottom
                    datecellref2.value="blank"  
                    tab1.update_idletasks() 


    #CEAT AND 7LVL
                navyUniqueRow = 20
                for i in range(1,32):
                                ceatRef = sheet.cell(row=navyUniqueRow,column=i) 
                                ceatRef.fill = PatternFill("solid", fgColor="B8CCE4")
                                ceatRef.value = "Navy Unique"

                ceatRow = navyUniqueRow+5
                for i in range(1,32):
                    ceatRef = sheet.cell(row=ceatRow,column=i) 
                    ceatRef.fill = PatternFill("solid", fgColor="B8CCE4")
                    ceatRef.value = "CEAT"
                ceatRow2 = ceatRow+5
                for i in range(1,32):
                    ceatRef2 = sheet.cell(row=ceatRow2,column=i) 
                    ceatRef2.fill = PatternFill("solid", fgColor="B8CCE4")
                    ceatRef2.value = "CEAT"
                distMxRow = ceatRow2+5
                for i in range(1,32):
                    ceatRef2 = sheet.cell(row=distMxRow,column=i) 
                    ceatRef2.fill = PatternFill("solid", fgColor="B8CCE4")
                    ceatRef2.value = "dISTR MX"
                sevenLRow = distMxRow+5
                for i in range(1,32):
                    ceatRef2 = sheet.cell(row=sevenLRow,column=i) 
                    ceatRef2.fill = PatternFill("solid", fgColor="B8CCE4")
                    ceatRef2.value = "7 LEVEL BLK 1/10"
                airfieldRow = sevenLRow+5
                for i in range(1,32):
                    ceatRef2 = sheet.cell(row=airfieldRow,column=i) 
                    ceatRef2.fill = PatternFill("solid", fgColor="B8CCE4")
                    ceatRef2.value = "AIRFIELD"
                #LEAVE
                tab5leaverow = airfieldRow+5
                for i in range(1,32):
                    leaveref = sheet.cell(row=tab5leaverow,column=i) 
                    leaveref.fill = PatternFill("solid", fgColor="E6B8B7")
                    leaveref.value = "LEAVE"
                #ADDITIONAL DUTIES
                additionaldutiesRow = tab5leaverow+7
                for i in range(1,32):
                    additionaldutiesref = sheet.cell(row=additionaldutiesRow,column=i) 
                    additionaldutiesref.fill = PatternFill("solid", fgColor="BFBFBF")
                    additionaldutiesref.value = "ADDITIONAL DUTIES"
                endRow = additionaldutiesRow+7
                for i in range(1,32):
                    additionaldutiesref = sheet.cell(row=endRow,column=i) 
                    additionaldutiesref.fill = PatternFill("solid", fgColor="000000")
                    additionaldutiesref.value = "END"


        sheetDates()
        myLabel0 = Label(tab2mainframe,text = "Created",font='Helvetica 16 bold',bg="grey26").grid(row=5,column=0,ipadx=20)
        workbook.save(filename=workbook_Title+".xlsx")
    except:
            print(traceback.format_exc())
            messagebox.showwarning(title="Error Occured", message="something went wrong in NEW BOOK. Check your entries and try again")
            if path.exists("errors.txt") == TRUE:
                ct = datetime.datetime.now() 
                with open("errors.txt", "a") as file:
                    file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))
            else:
                ct = datetime.datetime.now() 
                with open("errors.txt", "x") as file:
                    file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))     



def tab3add_qual():
        try:

        
    #NONE DECLARED ERROR LABELS USED
            global fileerrorLabel
            global qualerrorLabel
            global blockerrorLabel
            global starterrorLabel
            global enderrorLabel
            global qual1block1mirs
            global qual1block2mirs      
            sheet
        #SET COLUMN WIDTH
            # for col in range(1,32):
            #     column_letter = get_column_letter(col)
            #     # print(column_letter)
            #     sheet.column_dimensions[column_letter].width = 42

        #SET TITLE
            sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=32)
            sheet.row_dimensions[3].height = 60
            title_Cell = sheet['A3']
            title_Cell.border = thick_border
            monthTitle = str(sheet.title)
            # print(monthTitle)
            title_Cell.value = monthTitle+" ELECTRICAL SYSTEMS INSTRUCTOR SCHEDULE"
            title_Cell.font = Font(size=60,name="Times New Roman",bold=True)

            
            # global startMonth
            #SET DATES
            # for i in range(1,32):
            #     datecellref=sheet.cell(row=1, column=i)
            #     datecellref.fill = PatternFill("solid", fgColor="DDDDDD")
                
            #     if(i<10):
            #         datecellref.value=str(startMonth)+"/0"+str(i)+"/2021"
            #     else:
            #         datecellref.value=str(startMonth)+"/"+str(i)+"/2021"


            #SET DIVIDER      

            for i in range(1,32):
                datecellref2=sheet.cell(row=4, column=i)
                datecellref2.fill = PatternFill("solid", fgColor="000000")
                datecellref2.border = thick_border_blue_topBottom
                datecellref2.value="blank"  
                tab1.update_idletasks() 


            #GETTING INPUT VALUES FROM USER
            qualName = str(qualcellref.value)
            instructor = str(instructorcellref.value)
            if str(qualName) == "FAMILY DAY":
                None
            else:
                classNum = int(classcellref.value)
                if int(qualName) > 3 and str(qualName) !="FAMILY DAY":
                    qualerrorLabel = Label(tab1frame2,text = "Qual Num Doesn't Exist",font="Helvetica 10 bold", fg="red")
                    qualerrorLabel.grid(row=11,column=0)
                    e2.delete(0, END)
                    return 
                blockName = str(blockcellref.value)
                if int(blockName) > 6 and str(qualName) !="FAMILY DAY":
                    blockerrorLabel = Label(tab1frame2,text = "Block Number Doesn't Exist",font="Helvetica 10 bold", fg="red")
                    blockerrorLabel.grid(row=11,column=0)
                    e3.delete(0, END)
                    return
            startDate = str(startdatecellref.value)
            startDate = startDate[3:5]
            print(str(startDate))
            if int(startDate) > 32:
                starterrorLabel = Label(tab1frame2,text = "Start Date Number too big",font="Helvetica 10 bold", fg="red")
                starterrorLabel.grid(row=11,column=0)
                e4.delete(0, END)
                return
            if int(endDateReadIn) > 32:
                enderrorLabel = Label(tab1frame2,text = "End Date Number too big",font="Helvetica 10 bold", fg="red")
                enderrorLabel.grid(row=11,column=0)
                e5.delete(0, END)
                return


            

            #Start Date
            col_index = int(startDate)
            row_index=5
            # print("First row index: "+str(row_index))

            #THE DASHED CELLS ITERATORS
            # blank_col_index = col_index
            # blank_row_index=row_index+1
            
            #TITLE CELL
            cellref = sheet.cell(row=row_index,column=col_index)
            cellref.border = thick_border
            blockcheckcell = sheet.cell(row=row_index, column = 1)
        

            #DASHED CELLS
            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)


            #Setting Styles
            # cellref.font=ft1
            # cellref.alignment=align
            # blank_cellref.font=ft1
            # blank_cellref.alignment=align
            # blank_cellref2.font=ft1
            # blank_cellref2.alignment=align
            # blank_cellref3.font=ft1
            # blank_cellref3.alignment=align
            # blank_cellref4.font=ft1
            # blank_cellref4.alignment=align

            #End Date
            #check if months match


            #SETTING ACTIVE MIRS
            activeMirs=[]
            
            if qualName =="1" and blockName == "1":
                for items in qual1block1mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="2":
                for items in qual1block2mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="3":
                for items in qual1block3mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="4":
                for items in qual1block4mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="5":
                for items in qual1block5mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="6":
                for items in qual1block6mirs:
                    activeMirs.append(items)
            elif qualName =="2" and blockName =="1":
                for items in qual2block1mirs:
                    activeMirs.append(items)
            elif qualName =="2" and blockName =="2":
                for items in qual2block2mirs:
                    activeMirs.append(items)
            elif qualName =="3" and blockName =="1":
                for items in qual3block1mirs:
                    activeMirs.append(items)
            elif qualName =="3" and blockName =="2":
                for items in qual3block2mirs:
                    activeMirs.append(items)
            elif qualName =="3" and blockName =="3":
                for items in qual3block3mirs:
                    activeMirs.append(items)
        
                    



            des_col=int(endDateReadIn)+1

            
            
            total_index = 0

            mirsIndex = 0

            

            # print(tab5onleavelist[0])
            # print(tab5onleavelist[1])
            # print(tab5onleavelist[2])
            # for person in tab5onleavelist:
            #     print(person)
           
            
            
            
            
        
        
            q1check = False
            global colCheck
            colCheck = False


            #COLUMN ITERATOR LOOP
            while col_index < des_col:
                colCheck = False
                tab5leavecheck = False
                tab5onleavelist = []
                tab5leaverow = 5
                leavecellref = sheet.cell(row=tab5leaverow,column=col_index)
                while leavecellref.value !="LEAVE":
                    tab5leaverow+=1
                    leavecellref = sheet.cell(row=tab5leaverow,column=col_index)
                if leavecellref.value =="LEAVE":
                    tab5leaverow+=1
                    leavecellref = sheet.cell(row=tab5leaverow,column=col_index)
                while leavecellref.value:
                    person = str(leavecellref.value)
                    print(person)
                    tab5onleavelist.append(person)
                    tab5leaverow+=1
                    leavecellref = sheet.cell(row=tab5leaverow,column=col_index)

                for person in tab5onleavelist:
                    if instructor == person:
                        tab5leavecheck = True 
                
                #IF THERES A VALUE IN THE CELL
                while cellref.value:
                    colCheck = False
                    print(cellref.value)
                    if cellref.value == "Navy Unique":
                        sheet.insert_rows(row_index,5)
                        if col_index==1:
                            colCheck = True
                        else:
                            col_index-=1
                        # row_index-=5
                        cellref = sheet.cell(row=row_index,column=col_index)
                        # cellref.value="NEW"
                        # print(cellref.value)
                        workbook.save(filename=workbook_Title+".xlsx")
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        break
                    # print(str(cellref.value))
                    if str(cellref.value) == "FAMILY DAY":
                        col_index+=1
                        cellref = sheet.cell(row=row_index,column=col_index)
                        None
                    
                    elif str(cellref.value)!= "FAMILY DAY" and str(qualName)=="FAMILY DAY":
                        while row_index <=100:
                            cellref.value = "FAMILY DAY"
                            cellref.font = Font(color="FF0000")
                            cellref.fill = PatternFill("solid","FFFFFF")
                            cellref.border = thin_border_all_grey
                            row_index+=1
                            cellref = sheet.cell(row=row_index,column=col_index)

                    #First: IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK IS LESS THAN OR EQUAL AND THE CLASS NUM ENTERED IS GREATER
                    elif(int(qualName) <= int(cellref.value[1]) and int(blockName) <= int(cellref.value[10]) and int(qualName) == 1 and int(blockName)==1):
                        print("Start Date: "+str(startDate))
                        print("Start Date Type: "+str(type(startDate)))
                        if startDate!="01" and (int(classNum) >= int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999"):
                            sheet.insert_rows(5,5)
                            q1check = True
                            for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(startDate)-1):
                                for cell in rows:
                                    cell.value = "Q1  Block:1 Class:99999"
                                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid",fgColor="white")
                                    cell.border = thin_border_all_grey
                                    cell.font = Font(color="FFFFFF")
                            # row_index=4
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                        elif startDate!="01" and (int(classNum) < int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999"):
                            sheet.insert_rows(5,5)
                            q1check = True
                            for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(startDate)-1):
                                for cell in rows:
                                    cell.value = "Q1  Block:1 Class:99999"
                                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid",fgColor="white")
                                    cell.border = thin_border_all_grey
                                    cell.font = Font(color="FFFFFF")
                                    
                            row_index=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                        elif startDate=="1" and int(classNum) > int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999" :
                            sheet.insert_rows(5,5)
                            row_index=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        elif startDate=="1" and int(classNum) < int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999" :
                            row_index+=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                        #Second:IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK NUM IS GREATER OR EQUAL 
                    elif(int(qualName) <= int(cellref.value[1]) and int(blockName) > int(cellref.value[10])):
                        row_index+=5
                        print("Row index: " + str(row_index))
                        print(cellref.value)
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        

                        #Third:IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK IS LESS THAN OR EQUAL
                    elif(int(qualName) <= int(cellref.value[1]) and int(blockName) <= int(cellref.value[10])):
                        if int(classNum) > int(cellref.value[18:23]):
                            print("Cell Coordinate: "+str(cellref))
                            print("Last Blank Cell Coordinate: "+str(blank_cellref4))
                            sheet.insert_rows(row_index,5)
                            workbook.save(filename=workbook_Title+".xlsx")
                            newrowref = cellref.row
                            print("Newrowref: "+str(newrowref))
                            row_index=newrowref-5
                            print("row index: "+str(row_index))
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        else:

                            row_index+=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)


                        #FOURTH: IF THE ENTERED QUAL IS GREATER THAN THE CELL VALUE
                    elif(int(qualName) > int(cellref.value[1])):
                        row_index+=5
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)


                    elif(int(qualName)<=int(cellref.value[1])):
                        sheet.move_range(""+str(cellref.coordinate)+":"+str(blank_cellref4.coordinate)+"", rows=5)
                        newrowref = cellref.row
                        row_index=newrowref-5
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        
                else:
                    if str(qualName)=="FAMILY DAY":
                        while row_index <=100:
                            cellref.value = "FAMILY DAY"
                            cellref.font = Font(color="FF0000")
                            cellref.fill = PatternFill("solid","FFFFFF")
                            cellref.border = thin_border_all_grey
                            row_index+=1
                            cellref = sheet.cell(row=row_index,column=col_index)
                        
                    elif(int(qualName)==1 and int(blockName)==1 and startDate!="01" and q1check ==False):
                        sheet.insert_rows(5,5)
                        print("Row Index: "+str(row_index))
                        q1check =True
                        for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(startDate)-1):
                                for cell in rows:
                                    cell.value = "Q1  Block:1 Class:99999"
                                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid",fgColor="white")
                                    cell.border = thin_border_all_grey
                                    cell.font = Font(color="FFFFFF")
                                
                                    print("col index: "+str(col_index))
                    # row_index=4
                    print("Row Index: "+str(row_index))
                    cellref = sheet.cell(row=row_index,column=col_index)
                    blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                    blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                    blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                    blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                    if str(qualName)=="FAMILY DAY":
                        None    
                    else:
                        cellref.value="Q"+str(qualName)+" " +" Block:"+str(blockName)+" Class:"+str(classNum)
                        cellref.font = ft1
                        cellref.alignment=align
                    # print("col index: "+str(col_index))
                        #TITLE COLOR CODING
                        #QUAL 1
                        cellref.border = thick_border
                        if (qualName == "1" and blockName == "1"):
                            cellref.fill = PatternFill("solid", fgColor="00CCFF")
                        elif (qualName == "1"and blockName == "2"):
                            cellref.fill = PatternFill("solid", fgColor="33CCCC")
                        elif (qualName == "1" and blockName == "3"):
                            cellref.fill = PatternFill("solid", fgColor="92D050")
                        elif (qualName == "1"and blockName == "4"):
                            cellref.fill = PatternFill("solid", fgColor="FFFF00")
                        elif (qualName == "1" and blockName == "5"):
                            cellref.fill = PatternFill("solid", fgColor="FFC000")
                        elif (qualName == "1"and blockName == "6"):
                            cellref.fill = PatternFill("solid", fgColor="FF0000")

                        #QUAL2
                        elif (qualName == "2" and blockName == "1"):
                            cellref.fill = PatternFill("solid", fgColor="00B0F0")
                        elif (qualName == "2"and blockName == "2"):
                            cellref.fill = PatternFill("solid", fgColor="92D050")

                        #QUAL3
                        elif (qualName == "3" and blockName == "1"):
                            cellref.fill = PatternFill("solid", fgColor="FFFF00")
                        elif (qualName == "3"and blockName == "2"):
                            cellref.fill = PatternFill("solid", fgColor="9BBB59")
                        elif (qualName == "3"and blockName == "3"):
                            cellref.fill = PatternFill("solid", fgColor="FFC000")
                        


                        #SETTING BLANK VALUES BORDERS
                        blank_cellref.border = thin_border_sides
                        blank_cellref2.border = thin_border_sides
                        blank_cellref3.border = thin_border_sides
                        blank_cellref4.border = thin_border_sides_Bottom

                        #SETTING BLANK VALUES
                        if tab5leavecheck == True:
                            messagebox.showwarning(title="Instructor on Leave!", message="The entered instructor is on leave for one of the days")
                            blank_cellref.value="Instructor: Instructor is on leave"
                            tab5leavecheck ==False
                        else:
                            blank_cellref.value="Instructor: "+instructor
                        blank_cellref.font = ft1
                        blank_cellref.alignment=align
                        blank_cellref2.value="------"
                        blank_cellref2.font = ft1
                        blank_cellref2.alignment=align
                        blank_cellref3.value="------"
                        blank_cellref3.font = ft1
                        blank_cellref3.alignment=align
                        if mirsIndex > len(activeMirs)-1:
                            blank_cellref4.value = "NonPop"
                        else:
                            blank_cellref4.value=str(activeMirs[mirsIndex])
                        blank_cellref4.font = ft1
                        blank_cellref4.alignment=align
                    total_index+=1
                   
                    # print("Title = " + cellref.coordinate)
                    # print("blank = " + blank_cellref.coordinate)
                
                if colCheck ==True:
                    None
                else:
                    col_index+=1
                mirsIndex+=1
                # print("column index: "+str(col_index))   


                #This line controls if they shoot up the rows or nor
                # row_index=4
                # print("row index"+str(row_index))   
                cellref = sheet.cell(row=row_index,column=col_index)
                blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                print("Row Index: "+str(row_index))
                print("cellref: "+str(cellref))
                workbook.save(filename=workbook_Title+".xlsx")


                
                
                global submitTotal
                submittedLabel = Label(tab3frame3,text = "Submitted ",font="Helvetica 10 bold", fg="black",bg="grey26")
                datesLabel = Label(tab3frame3,text = "Total Number of Dates Affected: "+str(submitTotal),font="Helvetica 10 bold", fg="black",bg="grey26")
                submittedLabel.grid(row=1,column=0)
                datesLabel.grid(row=2,column=0)
                submitTotal+=1
                e2.delete(0, END)
                e3.delete(0, END)
                e4.delete(0, END)
                e5.delete(0, END)

            second_Month_Check_ReadIn() 
        except:
            print(traceback.format_exc()) 
            messagebox.showwarning(title="Error Occured", message="something went wrong in Add Qual ReadIN. Check your entries and try again")
            if path.exists("errors.txt") == TRUE:
                ct = datetime.datetime.now() 
                with open("errors.txt", "a") as file:
                    file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))
            else:
                ct = datetime.datetime.now() 
                with open("errors.txt", "x") as file:
                    file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))
            





def add_qual():
        try:

        
    #NONE DECLARED ERROR LABELS USED
            global fileerrorLabel
            global qualerrorLabel
            global blockerrorLabel
            global starterrorLabel
            global enderrorLabel
            global qual1block1mirs
            global qual1block2mirs      
            sheet
        #SET COLUMN WIDTH
            # for col in range(1,32):
            #     column_letter = get_column_letter(col)
            #     # print(column_letter)
            #     sheet.column_dimensions[column_letter].width = 42

        #SET TITLE
            sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=32)
            sheet.row_dimensions[3].height = 60
            title_Cell = sheet['A3']
            title_Cell.border = thick_border
            monthTitle = str(sheet.title)
            # print(monthTitle)
            title_Cell.value = monthTitle+" ELECTRICAL SYSTEMS INSTRUCTOR SCHEDULE"
            title_Cell.font = Font(size=60,name="Times New Roman",bold=True)

            
            # global startMonth
            #SET DATES
            # for i in range(1,32):
            #     datecellref=sheet.cell(row=1, column=i)
            #     datecellref.fill = PatternFill("solid", fgColor="DDDDDD")
                
            #     if(i<10):
            #         datecellref.value=str(startMonth)+"/0"+str(i)+"/2021"
            #     else:
            #         datecellref.value=str(startMonth)+"/"+str(i)+"/2021"


            #SET DIVIDER      

            for i in range(1,32):
                datecellref2=sheet.cell(row=4, column=i)
                datecellref2.fill = PatternFill("solid", fgColor="000000")
                datecellref2.border = thick_border_blue_topBottom
                datecellref2.value="blank"  
                tab1.update_idletasks() 


            #GETTING INPUT VALUES FROM USER
            instructor = str(e7.get())
            qualName = e2.get()
            if str(qualName) == "FAMILY DAY":
                None
            else:
                qualName = e2.get()
                if int(qualName) > 3 and str(qualName) !="FAMILY DAY":
                    qualerrorLabel = Label(tab1frame2,text = "Qual Num Doesn't Exist",font="Helvetica 10 bold", fg="red")
                    qualerrorLabel.grid(row=11,column=0)
                    e2.delete(0, END)
                    return 
                blockName = e3.get()
                if int(blockName) > 6 and str(qualName) !="FAMILY DAY":
                    blockerrorLabel = Label(tab1frame2,text = "Block Number Doesn't Exist",font="Helvetica 10 bold", fg="red")
                    blockerrorLabel.grid(row=11,column=0)
                    e3.delete(0, END)
                    return
            startDate = e4.get()
            startDate = startDate[3:5]
            print(str(startDate))
            if int(startDate) > 32:
                starterrorLabel = Label(tab1frame2,text = "Start Date Number too big",font="Helvetica 10 bold", fg="red")
                starterrorLabel.grid(row=11,column=0)
                e4.delete(0, END)
                return
            if int(endDate) > 32:
                enderrorLabel = Label(tab1frame2,text = "End Date Number too big",font="Helvetica 10 bold", fg="red")
                enderrorLabel.grid(row=11,column=0)
                e5.delete(0, END)
                return

            classNum = e6.get()

            

            #Start Date
            col_index = int(startDate)
            row_index=5
            # print("First row index: "+str(row_index))

            #THE DASHED CELLS ITERATORS
            # blank_col_index = col_index
            # blank_row_index=row_index+1
            
            #TITLE CELL
            cellref = sheet.cell(row=row_index,column=col_index)
            cellref.border = thick_border
            blockcheckcell = sheet.cell(row=row_index, column = 1)
        

            #DASHED CELLS
            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)


            #Setting Styles
            # cellref.font=ft1
            # cellref.alignment=align
            # blank_cellref.font=ft1
            # blank_cellref.alignment=align
            # blank_cellref2.font=ft1
            # blank_cellref2.alignment=align
            # blank_cellref3.font=ft1
            # blank_cellref3.alignment=align
            # blank_cellref4.font=ft1
            # blank_cellref4.alignment=align

            #End Date
            #check if months match


            #SETTING ACTIVE MIRS
            activeMirs=[]
            
            if qualName =="1" and blockName == "1":
                for items in qual1block1mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="2":
                for items in qual1block2mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="3":
                for items in qual1block3mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="4":
                for items in qual1block4mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="5":
                for items in qual1block5mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="6":
                for items in qual1block6mirs:
                    activeMirs.append(items)
            elif qualName =="2" and blockName =="1":
                for items in qual2block1mirs:
                    activeMirs.append(items)
            elif qualName =="2" and blockName =="2":
                for items in qual2block2mirs:
                    activeMirs.append(items)
            elif qualName =="3" and blockName =="1":
                for items in qual3block1mirs:
                    activeMirs.append(items)
            elif qualName =="3" and blockName =="2":
                for items in qual3block2mirs:
                    activeMirs.append(items)
            elif qualName =="3" and blockName =="3":
                for items in qual3block3mirs:
                    activeMirs.append(items)
        
                    



            des_col=int(endDate)+1

            
            
            total_index = 0

            mirsIndex = 0

            

            # print(tab5onleavelist[0])
            # print(tab5onleavelist[1])
            # print(tab5onleavelist[2])
            # for person in tab5onleavelist:
            #     print(person)
           
            
            
            
            
        
        
            q1check = False
            global colCheck
            colCheck = False


            #COLUMN ITERATOR LOOP
            while col_index < des_col:
                # print(col_index)
                # column_letter = get_column_letter(col_index)
                # print(column_letter)
                # if sheet.column_dimensions[column_letter].width < 1:
                #     col_index+=1
                #     continue
                colCheck = False
                tab5leavecheck = False
                tab5onleavelist = []
                tab5leaverow = 5
                leavecellref = sheet.cell(row=tab5leaverow,column=col_index)
                while leavecellref.value !="LEAVE":
                    tab5leaverow+=1
                    leavecellref = sheet.cell(row=tab5leaverow,column=col_index)
                if leavecellref.value =="LEAVE":
                    tab5leaverow+=1
                    leavecellref = sheet.cell(row=tab5leaverow,column=col_index)
                while leavecellref.value:
                    person = str(leavecellref.value)
                    print(person)
                    tab5onleavelist.append(person)
                    tab5leaverow+=1
                    leavecellref = sheet.cell(row=tab5leaverow,column=col_index)

                for person in tab5onleavelist:
                    if instructor == person:
                        tab5leavecheck = True 
                
                #IF THERES A VALUE IN THE CELL
                while cellref.value:
                    colCheck = False
                    print(cellref.value)
                    if cellref.value == "Navy Unique":
                        sheet.insert_rows(row_index,5)
                        if col_index==1:
                            colCheck = True
                        else:
                            col_index-=1
                        # row_index-=5
                        cellref = sheet.cell(row=row_index,column=col_index)
                        # cellref.value="NEW"
                        # print(cellref.value)
                        workbook.save(filename=workbook_Title+".xlsx")
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        break
                    # print(str(cellref.value))
                    if str(cellref.value) == "FAMILY DAY":
                        col_index+=1
                        cellref = sheet.cell(row=row_index,column=col_index)
                        None
                    
                    elif str(cellref.value)!= "FAMILY DAY" and str(qualName)=="FAMILY DAY":
                        while row_index <=100:
                            cellref.value = "FAMILY DAY"
                            cellref.font = Font(color="FF0000")
                            cellref.fill = PatternFill("solid","FFFFFF")
                            cellref.border = thin_border_all_grey
                            row_index+=1
                            cellref = sheet.cell(row=row_index,column=col_index)

                    #First: IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK IS LESS THAN OR EQUAL AND THE CLASS NUM ENTERED IS GREATER
                    elif(int(qualName) <= int(cellref.value[1]) and int(blockName) <= int(cellref.value[10]) and int(qualName) == 1 and int(blockName)==1):
                        print("Start Date: "+str(startDate))
                        print("Start Date Type: "+str(type(startDate)))
                        if startDate!="01" and (int(classNum) >= int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999"):
                            sheet.insert_rows(5,5)
                            q1check = True
                            for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(startDate)-1):
                                for cell in rows:
                                    cell.value = "Q1  Block:1 Class:99999"
                                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid",fgColor="white")
                                    cell.border = thin_border_all_grey
                                    cell.font = Font(color="FFFFFF")
                            # row_index=4
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                        elif startDate!="01" and (int(classNum) < int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999"):
                            sheet.insert_rows(5,5)
                            q1check = True
                            for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(startDate)-1):
                                for cell in rows:
                                    cell.value = "Q1  Block:1 Class:99999"
                                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid",fgColor="white")
                                    cell.border = thin_border_all_grey
                                    cell.font = Font(color="FFFFFF")
                                    
                            row_index=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                        elif startDate=="1" and int(classNum) > int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999" :
                            sheet.insert_rows(5,5)
                            row_index=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        elif startDate=="1" and int(classNum) < int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999" :
                            row_index+=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                        #Second:IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK NUM IS GREATER OR EQUAL 
                    elif(int(qualName) <= int(cellref.value[1]) and int(blockName) > int(cellref.value[10])):
                        row_index+=5
                        print("Row index: " + str(row_index))
                        print(cellref.value)
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        

                        #Third:IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK IS LESS THAN OR EQUAL
                    elif(int(qualName) <= int(cellref.value[1]) and int(blockName) <= int(cellref.value[10])):
                        if int(classNum) > int(cellref.value[18:23]):
                            print("Cell Coordinate: "+str(cellref))
                            print("Last Blank Cell Coordinate: "+str(blank_cellref4))
                            sheet.insert_rows(row_index,5)
                            workbook.save(filename=workbook_Title+".xlsx")
                            newrowref = cellref.row
                            print("Newrowref: "+str(newrowref))
                            row_index=newrowref-5
                            print("row index: "+str(row_index))
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        else:

                            row_index+=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)


                        #FOURTH: IF THE ENTERED QUAL IS GREATER THAN THE CELL VALUE
                    elif(int(qualName) > int(cellref.value[1])):
                        row_index+=5
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)


                    elif(int(qualName)<=int(cellref.value[1])):
                        sheet.move_range(""+str(cellref.coordinate)+":"+str(blank_cellref4.coordinate)+"", rows=5)
                        newrowref = cellref.row
                        row_index=newrowref-5
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        
                else:
                    if str(qualName)=="FAMILY DAY":
                        while row_index <=100:
                            cellref.value = "FAMILY DAY"
                            cellref.font = Font(color="FF0000")
                            cellref.fill = PatternFill("solid","FFFFFF")
                            cellref.border = thin_border_all_grey
                            row_index+=1
                            cellref = sheet.cell(row=row_index,column=col_index)
                        
                    elif(int(qualName)==1 and int(blockName)==1 and startDate!="01" and q1check ==False):
                        sheet.insert_rows(5,5)
                        print("Row Index: "+str(row_index))
                        q1check =True
                        for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(startDate)-1):
                                for cell in rows:
                                    cell.value = "Q1  Block:1 Class:99999"
                                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid",fgColor="white")
                                    cell.border = thin_border_all_grey
                                    cell.font = Font(color="FFFFFF")
                                
                                    print("col index: "+str(col_index))
                    # row_index=4
                    print("Row Index: "+str(row_index))
                    cellref = sheet.cell(row=row_index,column=col_index)
                    blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                    blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                    blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                    blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                    if str(qualName)=="FAMILY DAY":
                        None    
                    else:
                        cellref.value="Q"+str(qualName)+" " +" Block:"+str(blockName)+" Class:"+str(classNum)
                        cellref.font = ft1
                        cellref.alignment=align
                    # print("col index: "+str(col_index))
                        #TITLE COLOR CODING
                        #QUAL 1
                        cellref.border = thick_border
                        if (qualName == "1" and blockName == "1"):
                            cellref.fill = PatternFill("solid", fgColor="00CCFF")
                        elif (qualName == "1"and blockName == "2"):
                            cellref.fill = PatternFill("solid", fgColor="33CCCC")
                        elif (qualName == "1" and blockName == "3"):
                            cellref.fill = PatternFill("solid", fgColor="92D050")
                        elif (qualName == "1"and blockName == "4"):
                            cellref.fill = PatternFill("solid", fgColor="FFFF00")
                        elif (qualName == "1" and blockName == "5"):
                            cellref.fill = PatternFill("solid", fgColor="FFC000")
                        elif (qualName == "1"and blockName == "6"):
                            cellref.fill = PatternFill("solid", fgColor="FF0000")

                        #QUAL2
                        elif (qualName == "2" and blockName == "1"):
                            cellref.fill = PatternFill("solid", fgColor="00B0F0")
                        elif (qualName == "2"and blockName == "2"):
                            cellref.fill = PatternFill("solid", fgColor="92D050")

                        #QUAL3
                        elif (qualName == "3" and blockName == "1"):
                            cellref.fill = PatternFill("solid", fgColor="FFFF00")
                        elif (qualName == "3"and blockName == "2"):
                            cellref.fill = PatternFill("solid", fgColor="9BBB59")
                        elif (qualName == "3"and blockName == "3"):
                            cellref.fill = PatternFill("solid", fgColor="FFC000")
                        


                        #SETTING BLANK VALUES BORDERS
                        blank_cellref.border = thin_border_sides
                        blank_cellref2.border = thin_border_sides
                        blank_cellref3.border = thin_border_sides
                        blank_cellref4.border = thin_border_sides_Bottom

                        #SETTING BLANK VALUES
                        if tab5leavecheck == True:
                            messagebox.showwarning(title="Instructor on Leave!", message="The entered instructor is on leave for one of the days")
                            blank_cellref.value="Instructor: Instructor is on leave"
                            tab5leavecheck ==False
                        else:
                            blank_cellref.value="Instructor: "+instructor
                        blank_cellref.font = ft1
                        blank_cellref.alignment=align
                        blank_cellref2.value="------"
                        blank_cellref2.font = ft1
                        blank_cellref2.alignment=align
                        blank_cellref3.value="------"
                        blank_cellref3.font = ft1
                        blank_cellref3.alignment=align
                        if mirsIndex > len(activeMirs)-1:
                            blank_cellref4.value = "NonPop"
                        else:
                            blank_cellref4.value=str(activeMirs[mirsIndex])
                        blank_cellref4.font = ft1
                        blank_cellref4.alignment=align
                    total_index+=1
                   
                    # print("Title = " + cellref.coordinate)
                    # print("blank = " + blank_cellref.coordinate)
                
                if colCheck ==True:
                    None
                else:
                    col_index+=1
                mirsIndex+=1
                # print("column index: "+str(col_index))   


                #This line controls if they shoot up the rows or nor
                # row_index=4
                # print("row index"+str(row_index))   
                cellref = sheet.cell(row=row_index,column=col_index)
                blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                print("Row Index: "+str(row_index))
                print("cellref: "+str(cellref))
                workbook.save(filename=workbook_Title+".xlsx")


                
                
                global submitTotal
                submittedLabel = Label(tab1frame3,text = "Submitted ",font="Helvetica 10 bold", fg="black",bg="grey26")
                datesLabel = Label(tab1frame3,text = "Total Number of Dates Affected: "+str(submitTotal),font="Helvetica 8 bold", fg="black",bg="grey26")
                submittedLabel.grid(row=1,column=0,ipadx=50)
                datesLabel.grid(row = 2, column=0)
                submitTotal+=1
                # e2.delete(0, END)
                # e3.delete(0, END)
                # e4.delete(0, END)
                # e5.delete(0, END)
                my_progress2['value']+=30
                
                tab1.update_idletasks()

            second_Month_Check()
            my_progress2['value']=0
            
        except:
            print(traceback.format_exc())
            messagebox.showwarning(title="Error Occured", message="something went wrong in ADD QUAL.\nCheck your entries and try again.\nFor a more detailed explanation check errors txt file")
            if path.exists("errors.txt") == TRUE:
                ct = datetime.datetime.now() 
                with open("errors.txt", "a") as file:
                    file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))
            else:
                ct = datetime.datetime.now() 
                with open("errors.txt", "x") as file:
                    file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))




def second_Month_Check():
    if monthCheck == True:
        sheets = workbook.sheetnames
        # for i in sheets:
        #     print(i)
        global sheet
        print(str(sheet))
        sheet = workbook[sheets[sheetIndex+1]]
        print(str(sheet.title))
        # sheet = workbook.active
        global newstartDate
        newstartDate = "1"
        global newendDate
        newendDate = endDate = e5.get()
        newendDate = endDate[3:5]
        second_add_qual()
    else:
        None 

              

def second_Month_Check_ReadIn():
    if monthCheck == True:
        sheets = workbook.sheetnames
        # for i in sheets:
        #     print(i)
        global sheet
        # print(str(sheet))
        sheet = workbook[sheets[sheetIndex+1]]
        # print(str(sheet.title))
        # sheet = workbook.active
        global newstartDateReadIn
        newstartDateReadIn = "1"
        global newendDateReadIn
        newendDateReadIn = endDateReadIn=enddatecellref.value
        print("NEW END DATE Read In: "+str(endDateReadIn))
        newendDateReadIn = endDateReadIn[3:5]
        print(newendDateReadIn)
        second_add_qualReadIn()
    else:
        None                 
        

def second_add_qualReadIn():
        try:

        
    #NONE DECLARED ERROR LABELS USED
            global fileerrorLabel
            global qualerrorLabel
            global blockerrorLabel
            global starterrorLabel
            global enderrorLabel
            global qual1block1mirs
            global qual1block2mirs      
            sheet
        #SET COLUMN WIDTH
            # for col in range(1,32):
            #     column_letter = get_column_letter(col)
            #     # print(column_letter)
            #     sheet.column_dimensions[column_letter].width = 42

        #SET TITLE
            sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=32)
            sheet.row_dimensions[3].height = 60
            title_Cell = sheet['A3']
            title_Cell.border = thick_border
            monthTitle = str(sheet.title)
            # print(monthTitle)
            title_Cell.value = monthTitle+" ELECTRICAL SYSTEMS INSTRUCTOR SCHEDULE"
            title_Cell.font = Font(size=60,name="Times New Roman",bold=True)

            
            # global startMonth
            #SET DATES
            # for i in range(1,32):
            #     datecellref=sheet.cell(row=1, column=i)
            #     datecellref.fill = PatternFill("solid", fgColor="DDDDDD")
                
            #     if(i<10):
            #         datecellref.value=str(startMonth)+"/0"+str(i)+"/2021"
            #     else:
            #         datecellref.value=str(startMonth)+"/"+str(i)+"/2021"


            #SET DIVIDER      

            for i in range(1,32):
                datecellref2=sheet.cell(row=4, column=i)
                datecellref2.fill = PatternFill("solid", fgColor="000000")
                datecellref2.border = thick_border_blue_topBottom
                datecellref2.value="blank"  
                tab1.update_idletasks() 


            #GETTING INPUT VALUES FROM USER
            qualName = str(qualcellref.value)
            instructor = str(instructorcellref.value)
            if str(qualName) == "FAMILY DAY":
                None
            else:
                classNum = int(classcellref.value)
                if int(qualName) > 3 and str(qualName) !="FAMILY DAY":
                    qualerrorLabel = Label(tab1frame2,text = "Qual Num Doesn't Exist",font="Helvetica 10 bold", fg="red")
                    qualerrorLabel.grid(row=11,column=0)
                    e2.delete(0, END)
                    return 
                blockName = str(blockcellref.value)
                if int(blockName) > 6 and str(qualName) !="FAMILY DAY":
                    blockerrorLabel = Label(tab1frame2,text = "Block Number Doesn't Exist",font="Helvetica 10 bold", fg="red")
                    blockerrorLabel.grid(row=11,column=0)
                    e3.delete(0, END)
                    return
            # startDate = e4.get()
            # startDate = startDate[3:5]
            # print(str(startDate))
            if int(newstartDateReadIn) > 32:
                starterrorLabel = Label(tab1frame2,text = "Start Date Number too big",font="Helvetica 10 bold", fg="red")
                starterrorLabel.grid(row=11,column=0)
                e4.delete(0, END)
                return
            if int(newendDateReadIn) > 32:
                enderrorLabel = Label(tab1frame2,text = "End Date Number too big",font="Helvetica 10 bold", fg="red")
                enderrorLabel.grid(row=11,column=0)
                e5.delete(0, END)
                return

            classNum = e6.get()

            

            #Start Date
            col_index = int(newstartDateReadIn)
            row_index=5
            # print("First row index: "+str(row_index))

            #THE DASHED CELLS ITERATORS
            # blank_col_index = col_index
            # blank_row_index=row_index+1
            
            #TITLE CELL
            cellref = sheet.cell(row=row_index,column=col_index)
            cellref.border = thick_border
            blockcheckcell = sheet.cell(row=row_index, column = 1)
        

            #DASHED CELLS
            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)


            #Setting Styles
            # cellref.font=ft1
            # cellref.alignment=align
            # blank_cellref.font=ft1
            # blank_cellref.alignment=align
            # blank_cellref2.font=ft1
            # blank_cellref2.alignment=align
            # blank_cellref3.font=ft1
            # blank_cellref3.alignment=align
            # blank_cellref4.font=ft1
            # blank_cellref4.alignment=align

            #End Date
            #check if months match


            #SETTING ACTIVE MIRS
            activeMirs=[]
            
            if qualName =="1" and blockName == "1":
                for items in qual1block1mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="2":
                for items in qual1block2mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="3":
                for items in qual1block3mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="4":
                for items in qual1block4mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="5":
                for items in qual1block5mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="6":
                for items in qual1block6mirs:
                    activeMirs.append(items)
            elif qualName =="2" and blockName =="1":
                for items in qual2block1mirs:
                    activeMirs.append(items)
            elif qualName =="2" and blockName =="2":
                for items in qual2block2mirs:
                    activeMirs.append(items)
            elif qualName =="3" and blockName =="1":
                for items in qual3block1mirs:
                    activeMirs.append(items)
            elif qualName =="3" and blockName =="2":
                for items in qual3block2mirs:
                    activeMirs.append(items)
            elif qualName =="3" and blockName =="3":
                for items in qual3block3mirs:
                    activeMirs.append(items)
        
                    



            des_col=int(newendDateReadIn)+1

            
            
            total_index = 0

            mirsIndex = 0

            

            # print(tab5onleavelist[0])
            # print(tab5onleavelist[1])
            # print(tab5onleavelist[2])
            # for person in tab5onleavelist:
            #     print(person)
           
            
            
            
            
        
        
            q1check = False
            global colCheck
            colCheck = False


            #COLUMN ITERATOR LOOP
            while col_index < des_col:
                colCheck = False
                tab5leavecheck = False
                tab5onleavelist = []
                tab5leaverow = 5
                leavecellref = sheet.cell(row=tab5leaverow,column=col_index)
                while leavecellref.value !="LEAVE":
                    tab5leaverow+=1
                    leavecellref = sheet.cell(row=tab5leaverow,column=col_index)
                if leavecellref.value =="LEAVE":
                    tab5leaverow+=1
                    leavecellref = sheet.cell(row=tab5leaverow,column=col_index)
                while leavecellref.value:
                    person = str(leavecellref.value)
                    print(person)
                    tab5onleavelist.append(person)
                    tab5leaverow+=1
                    leavecellref = sheet.cell(row=tab5leaverow,column=col_index)

                for person in tab5onleavelist:
                    if instructor == person:
                        tab5leavecheck = True 
                
                #IF THERES A VALUE IN THE CELL
                while cellref.value:
                    colCheck = False
                    print(cellref.value)
                    if cellref.value == "Navy Unique":
                        sheet.insert_rows(row_index,5)
                        if col_index==1:
                            colCheck = True
                        else:
                            col_index-=1
                        # row_index-=5
                        cellref = sheet.cell(row=row_index,column=col_index)
                        # cellref.value="NEW"
                        # print(cellref.value)
                        workbook.save(filename=workbook_Title+".xlsx")
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        break
                    # print(str(cellref.value))
                    if str(cellref.value) == "FAMILY DAY":
                        col_index+=1
                        cellref = sheet.cell(row=row_index,column=col_index)
                        None
                    
                    elif str(cellref.value)!= "FAMILY DAY" and str(qualName)=="FAMILY DAY":
                        while row_index <=100:
                            cellref.value = "FAMILY DAY"
                            cellref.font = Font(color="FF0000")
                            cellref.fill = PatternFill("solid","FFFFFF")
                            cellref.border = thin_border_all_grey
                            row_index+=1
                            cellref = sheet.cell(row=row_index,column=col_index)

                    #First: IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK IS LESS THAN OR EQUAL AND THE CLASS NUM ENTERED IS GREATER
                    elif(int(qualName) <= int(cellref.value[1]) and int(blockName) <= int(cellref.value[10]) and int(qualName) == 1 and int(blockName)==1):
                        print("Start Date: "+str(newstartDate))
                        print("Start Date Type: "+str(type(newstartDate)))
                        if newstartDate!="01" and (int(classNum) >= int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999"):
                            sheet.insert_rows(5,5)
                            q1check = True
                            for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(newstartDate)-1):
                                for cell in rows:
                                    cell.value = "Q1  Block:1 Class:99999"
                                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid",fgColor="white")
                                    cell.border = thin_border_all_grey
                                    cell.font = Font(color="FFFFFF")
                            # row_index=4
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                        elif newstartDate!="01" and (int(classNum) < int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999"):
                            sheet.insert_rows(5,5)
                            q1check = True
                            for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(newstartDate)-1):
                                for cell in rows:
                                    cell.value = "Q1  Block:1 Class:99999"
                                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid",fgColor="white")
                                    cell.border = thin_border_all_grey
                                    cell.font = Font(color="FFFFFF")
                                    
                            row_index=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                        elif newstartDate=="1" and int(classNum) > int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999" :
                            sheet.insert_rows(5,5)
                            row_index=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        elif newstartDate=="1" and int(classNum) < int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999" :
                            row_index+=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                        #Second:IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK NUM IS GREATER OR EQUAL 
                    elif(int(qualName) <= int(cellref.value[1]) and int(blockName) > int(cellref.value[10])):
                        row_index+=5
                        print("Row index: " + str(row_index))
                        print(cellref.value)
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        

                        #Third:IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK IS LESS THAN OR EQUAL
                    elif(int(qualName) <= int(cellref.value[1]) and int(blockName) <= int(cellref.value[10])):
                        if int(classNum) > int(cellref.value[18:23]):
                            print("Cell Coordinate: "+str(cellref))
                            print("Last Blank Cell Coordinate: "+str(blank_cellref4))
                            sheet.insert_rows(row_index,5)
                            workbook.save(filename=workbook_Title+".xlsx")
                            newrowref = cellref.row
                            print("Newrowref: "+str(newrowref))
                            row_index=newrowref-5
                            print("row index: "+str(row_index))
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        else:

                            row_index+=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)


                        #FOURTH: IF THE ENTERED QUAL IS GREATER THAN THE CELL VALUE
                    elif(int(qualName) > int(cellref.value[1])):
                        row_index+=5
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)


                    elif(int(qualName)<=int(cellref.value[1])):
                        sheet.move_range(""+str(cellref.coordinate)+":"+str(blank_cellref4.coordinate)+"", rows=5)
                        newrowref = cellref.row
                        row_index=newrowref-5
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        
                else:
                    if str(qualName)=="FAMILY DAY":
                        while row_index <=100:
                            cellref.value = "FAMILY DAY"
                            cellref.font = Font(color="FF0000")
                            cellref.fill = PatternFill("solid","FFFFFF")
                            cellref.border = thin_border_all_grey
                            row_index+=1
                            cellref = sheet.cell(row=row_index,column=col_index)
                        
                    elif(int(qualName)==1 and int(blockName)==1 and newstartDate!="01" and q1check ==False):
                        sheet.insert_rows(5,5)
                        print("Row Index: "+str(row_index))
                        q1check =True
                        for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(newstartDate)-1):
                                for cell in rows:
                                    cell.value = "Q1  Block:1 Class:99999"
                                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid",fgColor="white")
                                    cell.border = thin_border_all_grey
                                    cell.font = Font(color="FFFFFF")
                                
                                    print("col index: "+str(col_index))
                    # row_index=4
                    print("Row Index: "+str(row_index))
                    cellref = sheet.cell(row=row_index,column=col_index)
                    blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                    blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                    blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                    blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                    if str(qualName)=="FAMILY DAY":
                        None    
                    else:
                        cellref.value="Q"+str(qualName)+" " +" Block:"+str(blockName)+" Class:"+str(classNum)
                        cellref.font = ft1
                        cellref.alignment=align
                    # print("col index: "+str(col_index))
                        #TITLE COLOR CODING
                        #QUAL 1
                        cellref.border = thick_border
                        if (qualName == "1" and blockName == "1"):
                            cellref.fill = PatternFill("solid", fgColor="00CCFF")
                        elif (qualName == "1"and blockName == "2"):
                            cellref.fill = PatternFill("solid", fgColor="33CCCC")
                        elif (qualName == "1" and blockName == "3"):
                            cellref.fill = PatternFill("solid", fgColor="92D050")
                        elif (qualName == "1"and blockName == "4"):
                            cellref.fill = PatternFill("solid", fgColor="FFFF00")
                        elif (qualName == "1" and blockName == "5"):
                            cellref.fill = PatternFill("solid", fgColor="FFC000")
                        elif (qualName == "1"and blockName == "6"):
                            cellref.fill = PatternFill("solid", fgColor="FF0000")

                        #QUAL2
                        elif (qualName == "2" and blockName == "1"):
                            cellref.fill = PatternFill("solid", fgColor="00B0F0")
                        elif (qualName == "2"and blockName == "2"):
                            cellref.fill = PatternFill("solid", fgColor="92D050")

                        #QUAL3
                        elif (qualName == "3" and blockName == "1"):
                            cellref.fill = PatternFill("solid", fgColor="FFFF00")
                        elif (qualName == "3"and blockName == "2"):
                            cellref.fill = PatternFill("solid", fgColor="9BBB59")
                        elif (qualName == "3"and blockName == "3"):
                            cellref.fill = PatternFill("solid", fgColor="FFC000")
                        


                        #SETTING BLANK VALUES BORDERS
                        blank_cellref.border = thin_border_sides
                        blank_cellref2.border = thin_border_sides
                        blank_cellref3.border = thin_border_sides
                        blank_cellref4.border = thin_border_sides_Bottom

                        #SETTING BLANK VALUES
                        if tab5leavecheck == True:
                            messagebox.showwarning(title="Instructor on Leave!", message="The entered instructor is on leave for one of the days")
                            blank_cellref.value="Instructor: Instructor is on leave"
                            tab5leavecheck ==False
                        else:
                            blank_cellref.value="Instructor: "+instructor
                        blank_cellref.font = ft1
                        blank_cellref.alignment=align
                        blank_cellref2.value="------"
                        blank_cellref2.font = ft1
                        blank_cellref2.alignment=align
                        blank_cellref3.value="------"
                        blank_cellref3.font = ft1
                        blank_cellref3.alignment=align
                        if mirsIndex > len(activeMirs)-1:
                            blank_cellref4.value = "NonPop"
                        else:
                            blank_cellref4.value=str(activeMirs[mirsIndex])
                        blank_cellref4.font = ft1
                        blank_cellref4.alignment=align
                    total_index+=1
                   
                    # print("Title = " + cellref.coordinate)
                    # print("blank = " + blank_cellref.coordinate)
                
                if colCheck ==True:
                    None
                else:
                    col_index+=1
                mirsIndex+=1
                # print("column index: "+str(col_index))   


                #This line controls if they shoot up the rows or nor
                # row_index=4
                # print("row index"+str(row_index))   
                cellref = sheet.cell(row=row_index,column=col_index)
                blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                print("Row Index: "+str(row_index))
                print("cellref: "+str(cellref))
                workbook.save(filename=workbook_Title+".xlsx")


                
                
                global submitTotal
                submittedLabel = Label(tab3frame3,text = "Submitted ",font="Helvetica 10 bold", fg="black",bg="grey26")
                datesLabel = Label(tab3frame3,text = "Total Number of Dates Affected: "+str(submitTotal),font="Helvetica 10 bold", fg="black",bg="grey26")
                submittedLabel.grid(row=1,column=0,ipadx=50)
                datesLabel.grid(row = 2, column=0)
                submitTotal+=1
                e2.delete(0, END)
                e3.delete(0, END)
                e4.delete(0, END)
                e5.delete(0, END)

            # if monthCheck == True:
            #     global sheet
            #     sheet = workbook.active[2]
            #     startDate = 1
            #     endDate = endDate = e5.get()
            #     endDate = endDate[3:5]
            #     second_add_qual()
            # else:
            #     None
        except:
            print(traceback.format_exc())
            messagebox.showwarning(title="Error Occured", message="something went wrong in Second Qual Add ReadIN try again")
            if path.exists("errors.txt") == TRUE:
                ct = datetime.datetime.now() 
                with open("errors.txt", "a") as file:
                    file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))
            else:
                ct = datetime.datetime.now() 
                with open("errors.txt", "x") as file:
                    file.write("\n"+str(ct)+"\n"+str(traceback.format_exc())) 
            



def second_add_qual():
        try:

        
    #NONE DECLARED ERROR LABELS USED
            global fileerrorLabel
            global qualerrorLabel
            global blockerrorLabel
            global starterrorLabel
            global enderrorLabel
            global qual1block1mirs
            global qual1block2mirs      
            sheet
        #SET COLUMN WIDTH
            # for col in range(1,32):
            #     column_letter = get_column_letter(col)
            #     # print(column_letter)
            #     sheet.column_dimensions[column_letter].width = 42

        #SET TITLE
            sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=32)
            sheet.row_dimensions[3].height = 60
            title_Cell = sheet['A3']
            title_Cell.border = thick_border
            monthTitle = str(sheet.title)
            # print(monthTitle)
            title_Cell.value = monthTitle+" ELECTRICAL SYSTEMS INSTRUCTOR SCHEDULE"
            title_Cell.font = Font(size=60,name="Times New Roman",bold=True)

            
            # global startMonth
            #SET DATES
            # for i in range(1,32):
            #     datecellref=sheet.cell(row=1, column=i)
            #     datecellref.fill = PatternFill("solid", fgColor="DDDDDD")
                
            #     if(i<10):
            #         datecellref.value=str(startMonth)+"/0"+str(i)+"/2021"
            #     else:
            #         datecellref.value=str(startMonth)+"/"+str(i)+"/2021"


            #SET DIVIDER      

            for i in range(1,32):
                datecellref2=sheet.cell(row=4, column=i)
                datecellref2.fill = PatternFill("solid", fgColor="000000")
                datecellref2.border = thick_border_blue_topBottom
                datecellref2.value="blank"  
                tab1.update_idletasks() 


            #GETTING INPUT VALUES FROM USER
            instructor = str(e7.get())
            qualName = e2.get()
            if str(qualName) == "FAMILY DAY":
                None
            else:
                qualName = e2.get()
                if int(qualName) > 3 and str(qualName) !="FAMILY DAY":
                    qualerrorLabel = Label(tab1frame2,text = "Qual Num Doesn't Exist",font="Helvetica 10 bold", fg="red")
                    qualerrorLabel.grid(row=11,column=0)
                    e2.delete(0, END)
                    return 
                blockName = e3.get()
                if int(blockName) > 6 and str(qualName) !="FAMILY DAY":
                    blockerrorLabel = Label(tab1frame2,text = "Block Number Doesn't Exist",font="Helvetica 10 bold", fg="red")
                    blockerrorLabel.grid(row=11,column=0)
                    e3.delete(0, END)
                    return
            # startDate = e4.get()
            # startDate = startDate[3:5]
            # print(str(startDate))
            if int(newstartDate) > 32:
                starterrorLabel = Label(tab1frame2,text = "Start Date Number too big",font="Helvetica 10 bold", fg="red")
                starterrorLabel.grid(row=11,column=0)
                e4.delete(0, END)
                return
            if int(newendDate) > 32:
                enderrorLabel = Label(tab1frame2,text = "End Date Number too big",font="Helvetica 10 bold", fg="red")
                enderrorLabel.grid(row=11,column=0)
                e5.delete(0, END)
                return

            classNum = e6.get()

            

            #Start Date
            col_index = int(newstartDate)
            row_index=5
            # print("First row index: "+str(row_index))

            #THE DASHED CELLS ITERATORS
            # blank_col_index = col_index
            # blank_row_index=row_index+1
            
            #TITLE CELL
            cellref = sheet.cell(row=row_index,column=col_index)
            cellref.border = thick_border
            blockcheckcell = sheet.cell(row=row_index, column = 1)
        

            #DASHED CELLS
            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)


            #Setting Styles
            # cellref.font=ft1
            # cellref.alignment=align
            # blank_cellref.font=ft1
            # blank_cellref.alignment=align
            # blank_cellref2.font=ft1
            # blank_cellref2.alignment=align
            # blank_cellref3.font=ft1
            # blank_cellref3.alignment=align
            # blank_cellref4.font=ft1
            # blank_cellref4.alignment=align

            #End Date
            #check if months match


            #SETTING ACTIVE MIRS
            activeMirs=[]
            
            if qualName =="1" and blockName == "1":
                for items in qual1block1mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="2":
                for items in qual1block2mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="3":
                for items in qual1block3mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="4":
                for items in qual1block4mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="5":
                for items in qual1block5mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="6":
                for items in qual1block6mirs:
                    activeMirs.append(items)
            elif qualName =="2" and blockName =="1":
                for items in qual2block1mirs:
                    activeMirs.append(items)
            elif qualName =="2" and blockName =="2":
                for items in qual2block2mirs:
                    activeMirs.append(items)
            elif qualName =="3" and blockName =="1":
                for items in qual3block1mirs:
                    activeMirs.append(items)
            elif qualName =="3" and blockName =="2":
                for items in qual3block2mirs:
                    activeMirs.append(items)
            elif qualName =="3" and blockName =="3":
                for items in qual3block3mirs:
                    activeMirs.append(items)
        
                    



            des_col=int(newendDate)+1

            
            
            total_index = 0

            mirsIndex = 0

            

            # print(tab5onleavelist[0])
            # print(tab5onleavelist[1])
            # print(tab5onleavelist[2])
            # for person in tab5onleavelist:
            #     print(person)
           
            
            
            
            
        
        
            q1check = False
            global colCheck
            colCheck = False


            #COLUMN ITERATOR LOOP
            while col_index < des_col:
                colCheck = False
                tab5leavecheck = False
                tab5onleavelist = []
                tab5leaverow = 5
                leavecellref = sheet.cell(row=tab5leaverow,column=col_index)
                while leavecellref.value !="LEAVE":
                    tab5leaverow+=1
                    leavecellref = sheet.cell(row=tab5leaverow,column=col_index)
                if leavecellref.value =="LEAVE":
                    tab5leaverow+=1
                    leavecellref = sheet.cell(row=tab5leaverow,column=col_index)
                while leavecellref.value:
                    person = str(leavecellref.value)
                    print(person)
                    tab5onleavelist.append(person)
                    tab5leaverow+=1
                    leavecellref = sheet.cell(row=tab5leaverow,column=col_index)

                for person in tab5onleavelist:
                    if instructor == person:
                        tab5leavecheck = True 
                
                #IF THERES A VALUE IN THE CELL
                while cellref.value:
                    colCheck = False
                    print(cellref.value)
                    if cellref.value == "Navy Unique":
                        sheet.insert_rows(row_index,5)
                        if col_index==1:
                            colCheck = True
                        else:
                            col_index-=1
                        # row_index-=5
                        cellref = sheet.cell(row=row_index,column=col_index)
                        # cellref.value="NEW"
                        # print(cellref.value)
                        workbook.save(filename=workbook_Title+".xlsx")
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        break
                    # print(str(cellref.value))
                    if str(cellref.value) == "FAMILY DAY":
                        col_index+=1
                        cellref = sheet.cell(row=row_index,column=col_index)
                        None
                    
                    elif str(cellref.value)!= "FAMILY DAY" and str(qualName)=="FAMILY DAY":
                        while row_index <=100:
                            cellref.value = "FAMILY DAY"
                            cellref.font = Font(color="FF0000")
                            cellref.fill = PatternFill("solid","FFFFFF")
                            cellref.border = thin_border_all_grey
                            row_index+=1
                            cellref = sheet.cell(row=row_index,column=col_index)

                    #First: IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK IS LESS THAN OR EQUAL AND THE CLASS NUM ENTERED IS GREATER
                    elif(int(qualName) <= int(cellref.value[1]) and int(blockName) <= int(cellref.value[10]) and int(qualName) == 1 and int(blockName)==1):
                        print("Start Date: "+str(newstartDate))
                        print("Start Date Type: "+str(type(newstartDate)))
                        if newstartDate!="01" and (int(classNum) >= int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999"):
                            sheet.insert_rows(5,5)
                            q1check = True
                            for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(newstartDate)-1):
                                for cell in rows:
                                    cell.value = "Q1  Block:1 Class:99999"
                                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid",fgColor="white")
                                    cell.border = thin_border_all_grey
                                    cell.font = Font(color="FFFFFF")
                            # row_index=4
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                        elif newstartDate!="01" and (int(classNum) < int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999"):
                            sheet.insert_rows(5,5)
                            q1check = True
                            for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(newstartDate)-1):
                                for cell in rows:
                                    cell.value = "Q1  Block:1 Class:99999"
                                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid",fgColor="white")
                                    cell.border = thin_border_all_grey
                                    cell.font = Font(color="FFFFFF")
                                    
                            row_index=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                        elif newstartDate=="1" and int(classNum) > int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999" :
                            sheet.insert_rows(5,5)
                            row_index=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        elif newstartDate=="1" and int(classNum) < int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999" :
                            row_index+=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                        #Second:IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK NUM IS GREATER OR EQUAL 
                    elif(int(qualName) <= int(cellref.value[1]) and int(blockName) > int(cellref.value[10])):
                        row_index+=5
                        print("Row index: " + str(row_index))
                        print(cellref.value)
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        

                        #Third:IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK IS LESS THAN OR EQUAL
                    elif(int(qualName) <= int(cellref.value[1]) and int(blockName) <= int(cellref.value[10])):
                        if int(classNum) > int(cellref.value[18:23]):
                            print("Cell Coordinate: "+str(cellref))
                            print("Last Blank Cell Coordinate: "+str(blank_cellref4))
                            sheet.insert_rows(row_index,5)
                            workbook.save(filename=workbook_Title+".xlsx")
                            newrowref = cellref.row
                            print("Newrowref: "+str(newrowref))
                            row_index=newrowref-5
                            print("row index: "+str(row_index))
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        else:

                            row_index+=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)


                        #FOURTH: IF THE ENTERED QUAL IS GREATER THAN THE CELL VALUE
                    elif(int(qualName) > int(cellref.value[1])):
                        row_index+=5
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)


                    elif(int(qualName)<=int(cellref.value[1])):
                        sheet.move_range(""+str(cellref.coordinate)+":"+str(blank_cellref4.coordinate)+"", rows=5)
                        newrowref = cellref.row
                        row_index=newrowref-5
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        
                else:
                    if str(qualName)=="FAMILY DAY":
                        while row_index <=100:
                            cellref.value = "FAMILY DAY"
                            cellref.font = Font(color="FF0000")
                            cellref.fill = PatternFill("solid","FFFFFF")
                            cellref.border = thin_border_all_grey
                            row_index+=1
                            cellref = sheet.cell(row=row_index,column=col_index)
                        
                    elif(int(qualName)==1 and int(blockName)==1 and newstartDate!="01" and q1check ==False):
                        sheet.insert_rows(5,5)
                        print("Row Index: "+str(row_index))
                        q1check =True
                        for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(newstartDate)-1):
                                for cell in rows:
                                    cell.value = "Q1  Block:1 Class:99999"
                                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid",fgColor="white")
                                    cell.border = thin_border_all_grey
                                    cell.font = Font(color="FFFFFF")
                                
                                    print("col index: "+str(col_index))
                    # row_index=4
                    print("Row Index: "+str(row_index))
                    cellref = sheet.cell(row=row_index,column=col_index)
                    blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                    blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                    blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                    blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                    if str(qualName)=="FAMILY DAY":
                        None    
                    else:
                        cellref.value="Q"+str(qualName)+" " +" Block:"+str(blockName)+" Class:"+str(classNum)
                        cellref.font = ft1
                        cellref.alignment=align
                    # print("col index: "+str(col_index))
                        #TITLE COLOR CODING
                        #QUAL 1
                        cellref.border = thick_border
                        if (qualName == "1" and blockName == "1"):
                            cellref.fill = PatternFill("solid", fgColor="00CCFF")
                        elif (qualName == "1"and blockName == "2"):
                            cellref.fill = PatternFill("solid", fgColor="33CCCC")
                        elif (qualName == "1" and blockName == "3"):
                            cellref.fill = PatternFill("solid", fgColor="92D050")
                        elif (qualName == "1"and blockName == "4"):
                            cellref.fill = PatternFill("solid", fgColor="FFFF00")
                        elif (qualName == "1" and blockName == "5"):
                            cellref.fill = PatternFill("solid", fgColor="FFC000")
                        elif (qualName == "1"and blockName == "6"):
                            cellref.fill = PatternFill("solid", fgColor="FF0000")

                        #QUAL2
                        elif (qualName == "2" and blockName == "1"):
                            cellref.fill = PatternFill("solid", fgColor="00B0F0")
                        elif (qualName == "2"and blockName == "2"):
                            cellref.fill = PatternFill("solid", fgColor="92D050")

                        #QUAL3
                        elif (qualName == "3" and blockName == "1"):
                            cellref.fill = PatternFill("solid", fgColor="FFFF00")
                        elif (qualName == "3"and blockName == "2"):
                            cellref.fill = PatternFill("solid", fgColor="9BBB59")
                        elif (qualName == "3"and blockName == "3"):
                            cellref.fill = PatternFill("solid", fgColor="FFC000")
                        


                        #SETTING BLANK VALUES BORDERS
                        blank_cellref.border = thin_border_sides
                        blank_cellref2.border = thin_border_sides
                        blank_cellref3.border = thin_border_sides
                        blank_cellref4.border = thin_border_sides_Bottom

                        #SETTING BLANK VALUES
                        if tab5leavecheck == True:
                            messagebox.showwarning(title="Instructor on Leave!", message="The entered instructor is on leave for one of the days")
                            blank_cellref.value="Instructor: Instructor is on leave"
                            tab5leavecheck ==False
                        else:
                            blank_cellref.value="Instructor: "+instructor
                        blank_cellref.font = ft1
                        blank_cellref.alignment=align
                        blank_cellref2.value="------"
                        blank_cellref2.font = ft1
                        blank_cellref2.alignment=align
                        blank_cellref3.value="------"
                        blank_cellref3.font = ft1
                        blank_cellref3.alignment=align
                        if mirsIndex > len(activeMirs)-1:
                            blank_cellref4.value = "NonPop"
                        else:
                            blank_cellref4.value=str(activeMirs[mirsIndex])
                        blank_cellref4.font = ft1
                        blank_cellref4.alignment=align
                    total_index+=1
                   
                    # print("Title = " + cellref.coordinate)
                    # print("blank = " + blank_cellref.coordinate)
                
                if colCheck ==True:
                    None
                else:
                    col_index+=1
                mirsIndex+=1
                # print("column index: "+str(col_index))   


                #This line controls if they shoot up the rows or nor
                # row_index=4
                # print("row index"+str(row_index))   
                cellref = sheet.cell(row=row_index,column=col_index)
                blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                print("Row Index: "+str(row_index))
                print("cellref: "+str(cellref))
                workbook.save(filename=workbook_Title+".xlsx")


                
                
                global submitTotal
                submittedLabel = Label(tab1frame3,text = "Submitted ",font="Helvetica 10 bold", fg="black",bg="grey26")
                datesLabel = Label(tab1frame3,text = "Total Number of Dates Affected: "+str(submitTotal),font="Helvetica 8 bold", fg="black",bg="grey26")
                submittedLabel.grid(row=1,column=0,ipadx=50)
                datesLabel.grid(row = 2, column=0)
                submitTotal+=1
                e2.delete(0, END)
                e3.delete(0, END)
                e4.delete(0, END)
                e5.delete(0, END)

            # if monthCheck == True:
            #     global sheet
            #     sheet = workbook.active[2]
            #     startDate = 1
            #     endDate = endDate = e5.get()
            #     endDate = endDate[3:5]
            #     second_add_qual()
            # else:
            #     None
        except():
            print(traceback.format_exc())
            messagebox.showwarning(title="Error Occured", message="something went wrong in Second Add Month. Check your entries and try again")
            if path.exists("errors.txt") == TRUE:
                ct = datetime.datetime.now() 
                with open("errors.txt", "a") as file:
                    file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))
            else:
                ct = datetime.datetime.now() 
                with open("errors.txt", "x") as file:
                    file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))
            
             
            


#======================================================================================================#
#======================================================================================================#
#======================================================================================================#
#======================================================================================================#
#======================================================================================================#
#======================================================================================================#
#======================================================================================================#       




#MAIN PAGE
root = ThemedTk(theme="black")
root.geometry("800x550")
root.title("Curtis Scheduling Tool "+"                                                                                          \u00A9" + " KandyKane Solutions  Ver.8.0.0*")
tab_parent = ttk.Notebook(root)
tab1 = ttk.Frame(tab_parent)
tab2 = ttk.Frame(tab_parent)
tab3 = ttk.Frame(tab_parent)
tab4 = ttk.Frame(tab_parent)
tab5 = ttk.Frame(tab_parent)
tab6 = ttk.Frame(tab_parent)

tab_parent.add(tab1,text="Use Existing")
tab_parent.add(tab2,text="New Workbook")
tab_parent.add(tab3,text="Read From File")
tab_parent.add(tab4,text="About")
tab_parent.add(tab5,text="Add MIRS")
tab_parent.add(tab6,text="Settings")
tab_parent.pack(expand=1,fill='both')


# bg = PhotoImage(file="background5.png")
# my_label = Label(root,image=bg)
# my_label.place(x=0,y=0,relwidth=1,relheight=1)
p1 = PhotoImage(file = 'airforce.png')
root.iconphoto(False, p1)




#TAB1==============================================================================================================TAB1
tab1titleframe=Frame(tab1,bg="grey26")
tab1titleframe.pack(side=TOP)
tab1mainframe = Frame(tab1,width=1000,bg="grey26", height=500,highlightbackground="black",highlightthickness=3)
tab1mainframe.pack(side=TOP,padx=(0,400))

tab1frame = Frame(tab1mainframe,bg="grey10")
tab1frame.pack()
tab1frame2 = Frame(tab1mainframe,bg="grey26")
tab1frame2.pack()
tab1frame3 = Frame(tab1mainframe,bg="grey26")
tab1frame3.pack()

tab1titlelabel = Label(tab1titleframe,text = "Existing Workbook File",fg="white",bg="grey26",font="Helvetica 36 bold").grid(row=1,column=0)
tab1Label0 = Label(tab1frame,text = "Entry",fg="white",bg="grey10",font="Helvetica 20 bold").grid(row=1,column=0,padx=140)
tab1Label0 = Label(tab1frame2,text = "File Name:",fg="white",bg="grey26",font="Helvetica 15 bold").grid(row=2,column=0,pady=2)
tab1Label1 = Label(tab1frame2,text = "Qual Num:",fg="white",bg="grey26",font="Helvetica 15 bold").grid(row=3,column=0,pady=2)
tab1Label2 = Label(tab1frame2,text = "Block Num:",fg="white",bg="grey26",font="Helvetica 15 bold").grid(row=4,column=0,pady=2)
tab1Label3 = Label(tab1frame2,text = "Start Date:",fg="white",bg="grey26",font="Helvetica 15 bold").grid(row=5,column=0,pady=2)
tab1Label4 = Label(tab1frame2,text = "End Date:",fg="white",bg="grey26",font="Helvetica 15 bold").grid(row=6,column=0,pady=2)
tab1Label5 = Label(tab1frame2,text = "Class Num:",fg="white",bg="grey26",font="Helvetica 15 bold").grid(row=7,column=0,pady=2)
tab1holidaylabel = Label(tab1frame2,text = "*For holiday and family\n days just put 'FAMILY DAY'\n in the qual entry",fg="white",bg="grey26",font="Helvetica 8 bold").grid(row=10,column=0,pady=2)
instructorLabel = Label(tab1frame2,text = "Instructor:",fg="white",bg="grey26",font="Helvetica 15 bold").grid(row=8,column=0,pady=2)


tab1Label2a = Label(tab1frame2,text = "'Ex.test'",fg=("grey50"),bg="grey26",font="Helvetica 15 bold").grid(row=2,column=2,pady=2)
tab1Label3a = Label(tab1frame2,text = "'Ex. 1-3'",fg=("grey50"),bg="grey26",font="Helvetica 15 bold").grid(row=3,column=2,pady=2)
tab1Label4a = Label(tab1frame2,text = "'Ex. 1-6'",fg=("grey50"),bg="grey26",font="Helvetica 15 bold").grid(row=4,column=2,pady=2)
tab1Label5a = Label(tab1frame2,text = "'01/01/21'",fg=("grey50"),bg="grey26",font="Helvetica 15 bold").grid(row=5,column=2,pady=2,padx=2)
tab1Label6a = Label(tab1frame2,text = "'01/28/21'",fg=("grey50"),bg="grey26",font="Helvetica 15 bold").grid(row=6,column=2,pady=2,padx=2)
tab1Label7a = Label(tab1frame2,text = "'200XX'",fg=("grey50"),bg="grey26",font="Helvetica 15 bold").grid(row=7,column=2,pady=2)
global my_progress2
my_progress2 = ttk.Progressbar(tab1frame3,orient=HORIZONTAL,length=200,mode="indeterminate")


global e1
global e2
global e3
global e4
global e5
global e6
global e7



e1 = Entry(tab1frame2,width=10)
e1.grid(row=2,column=1)
e2 = Entry(tab1frame2,width=10)
e2.grid(row=3,column=1)
e3 = Entry(tab1frame2,width=10)
e3.grid(row=4,column=1)
e4 = Entry(tab1frame2,width=10)
e4.grid(row=5,column=1)
e5 = Entry(tab1frame2,width=10)
e5.grid(row=6,column=1)
e6 = Entry(tab1frame2,width=10)
e6.grid(row=7,column=1)
e7 = Entry(tab1frame2,width=10)
e7.grid(row=8,column=1)


myButton4 = Button(tab1frame2,text="Submit",command=existing_WORKBOOK,bg="grey80")
myButton4.grid(row=11,column=0)
changeOnHover(myButton4, "aqua", "grey80")
global submitTotal
submitTotal = 1




#TAB2====================================================ADD NEW WORKBOOK======================================================TAB2
tab2titleframe=Frame(tab2,bg="grey26")
tab2titleframe.pack()
tab2mainframe = Frame(tab2,bg="grey26")
tab2mainframe.pack()
myLabel0 = Label(tab2titleframe,text = "New Excel File",fg="white",bg="grey26",font="Helvetica 36 bold").grid(row=0,column=0)
myLabel0 = Label(tab2mainframe,text = "File Name:",fg="white",bg="grey26",font="Helvetica 15 bold").grid(row=0,column=0)
global newE1
global newe2
newe1 = Entry(tab2mainframe,width=20)
newe1.grid(row=2,column=0,pady=(0,10))
createButton = Button(tab2mainframe,text="Create New",command=new_WORKBOOK,bg="grey80")
createButton.grid(row=3,column=0)
changeOnHover(myButton4, "aqua", "grey80")



#TAB3============================READ FROM FILE========================================================TAB3#
tab3frame = Frame(tab3,bg="grey26")
tab3frame.grid(row=0,column=0)
tab3frame2 = Frame(tab3,bg="grey26")
tab3frame2.grid(row=0,column=1)
tab3frame3 = Frame(tab3,bg="grey26")
tab3frame3.grid(row=0,column=1)
tab3Label = Label(tab3frame,text = "Read from File",fg="white",bg="grey26",font="Helvetica 36 bold").pack()
tab3Labe2 = Label(tab3frame,text = "Read From File:",fg="white",bg="grey26",font="Helvetica 15 bold").pack()
global tab3e1
tab3e1 = Entry(tab3frame,width=20)
tab3e1.pack()
tab3Labe2 = Label(tab3frame,text = "Destination File:",fg="white",bg="grey26",font="Helvetica 15 bold").pack()
global tab3e2
tab3e2 = Entry(tab3frame,width=20)
tab3e2.pack()


global my_progress
my_progress = ttk.Progressbar(tab3frame3,orient=HORIZONTAL,length=300,mode="indeterminate")



tab3createButton = Button(tab3frame,text="Submit",command=readFromExcel,bg="grey80")
tab3createButton.pack(pady=(10))
changeOnHover(myButton4, "aqua", "grey80")



#TAB4================================================ABOUT====================================================================TAB3
mylabel2 = Label(tab4,text="About",font='Helvetica 30 bold')
about = """This is a simple scheduler program for automatically creating\nand editing tasks on defined days on an excel sheet when \ngiven prescribed dates. Hope you enjoy! """
mylabel = Label(tab4,text=about,font='Helvetica 12 bold')
mylabel2 = Label(tab4,text="How to Use",font='Helvetica 12 bold')
mylabel7 = Label(tab4,text="-Put the excel files you wish to edit in the same folder as this programs .exe file\n-When entering numbers all entries must be single digit\n-Do not separate the .exe file from the images",font='Helvetica 12 bold')
mylabel3 = Label(tab4,text="Use Existing:",font='Helvetica 12 bold')
mylabel4 = Label(tab4,text="This is for adding to an existing file. Simply, enter\n the file name(no extension) and enter\n the rest of your information accordingly.",font='Helvetica 12 bold')
mylabel5 = Label(tab4,text="Add WorkBook:",font='Helvetica 12 bold')
mylabel6 = Label(tab4,text="This is for creating a new excel file. Simply, enter what\n you would like to call the file(no extension,no special characters) and enter\n the rest of your information accordingly.You must initilize it with a Qual entry",font='Helvetica 12 bold')
mylabel.pack(pady=(0,30))
mylabel2.pack()
mylabel7.pack()
mylabel3.pack()
mylabel4.pack()
mylabel5.pack()
mylabel6.pack()



#TAB5=================================================ADD MIRS==================================================#TAB5
maintitle = Frame(tab5,bg="grey26")
maintitle.pack()
titleframe = Frame(tab5,bg="grey26")
titleframe.pack()
listframe = Frame(tab5,bg="grey26")
listframe.pack(pady=(20))
buttonframe = Frame(tab5,bg="grey26")
buttonframe.pack(padx=(0,100))

title = Label(maintitle,text = "ADD MIRS",fg="white",bg="grey26",font="Helvetica 36 bold")
title.pack()

filename = Label(titleframe,text="File name:")
filename.grid(row=0,column=0,padx=5)
fileentry = Entry(titleframe,width=15)
fileentry.grid(row=0,column=1,padx=5)

datelabel = Label(titleframe,text='Date:')
datelabel.grid(row=0,column=2,padx=5)
dateentry = Entry(titleframe,width=15)
dateentry.grid(row=0,column=3,padx=5)

allpersons = Label(listframe,text="AVAILABLE EMPLOYESS")
onleave = Label(listframe,text="ON LEAVE")
daypreview = Label(listframe,text="DAY PREVIEW")
allpersons.grid(row=0,column=0)
daypreview.grid(row=0,column=1)
onleave.grid(row=0,column=2)

finishedLabel = Label(buttonframe,text='Finished',bg="grey26",fg="grey26")
finishedLabel.grid(row=1,column=0)

def rightarrow():
    tab5dayindex = dateentry.get()[3:5]
    tab5dayindex = int(tab5dayindex)
    newtab5dayindex = tab5dayindex+1
    dateentry.delete(3,5)
    if newtab5dayindex <=9:
        dateentry.insert(3,"0"+str(newtab5dayindex))
    else:
        dateentry.insert(3,str(newtab5dayindex))  

def leftarrow():
    tab5dayindex = dateentry.get()[3:5]
    tab5dayindex = int(tab5dayindex)
    newtab5dayindex = tab5dayindex-1
    dateentry.delete(3,5)
    if newtab5dayindex <=9:
        dateentry.insert(3,"0"+str(newtab5dayindex))
    else:
        dateentry.insert(3,str(newtab5dayindex))    

def populateList():
    finishedLabel = Label(buttonframe,text='Finished',bg="grey26",fg="grey26")
    finishedLabel.grid(row=1,column=0)
    
    global my_list2
    global listbox1
    listbox1.delete(0,END)
    listbox3.delete(0,END)
    


    listbox2.delete(0, END)
    global tab5monthindex
    global tab5dayindex
    tab5monthindex = dateentry.get()[0:2]
    tab5dayindex = dateentry.get()[3:5]
    tab5monthindex = int(tab5monthindex)
    tab5dayindex = int(tab5dayindex)
    #MONTH INDEX
    print(tab5monthindex)
    print(tab5dayindex)
    
    readbook = load_workbook(str(fileentry.get())+'.xlsx')
    sheets = readbook.sheetnames
    sheet = readbook[sheets[int(tab5monthindex)]]
    print(sheet.title)
    
    tab5row_index = 5 
    cellref = sheet.cell(row=tab5row_index,column=tab5dayindex)
    # print(cellref.value)
    while cellref.value != "END":
        listbox2.insert(END,cellref.value)
        if not cellref.value:
            listbox2.insert(END,"")
        elif "Q" in str(cellref.value):
            listbox2.itemconfig("end", bg = "red")
        tab5row_index+=1
        cellref =sheet.cell(row=tab5row_index,column=tab5dayindex)
        # print(cellref.value)
    
    tab5leavecheck = False
    tab5onleavelist = []
    tab5leaverow = 5
    leavecellref = sheet.cell(row=tab5leaverow,column=tab5dayindex)
    while leavecellref.value !="LEAVE":
        tab5leaverow+=1
        leavecellref = sheet.cell(row=tab5leaverow,column=tab5dayindex)
    if leavecellref.value =="LEAVE":
        tab5leaverow+=1
        leavecellref = sheet.cell(row=tab5leaverow,column=tab5dayindex)
    while leavecellref.value:
        person = str(leavecellref.value)
        person = person.upper()
        print(person)
        tab5onleavelist.append(person)
        tab5leaverow+=1
        leavecellref = sheet.cell(row=tab5leaverow,column=tab5dayindex)
    
    for person in tab5onleavelist:
        listbox3.insert(END,person)
    
    
    for item in mainemployeeslist:
        if item not in tab5onleavelist:
            listbox1.insert(END, item)

populate = Button(titleframe,text="Populate",command=populateList)
populate.grid(row=0,column=6,padx=5)

rightArrow = Button(titleframe,text=">",command=rightarrow)
rightArrow.grid(row=0,column=5)
leftArrow = Button(titleframe,text="<",command=leftarrow)
leftArrow.grid(row=0,column=4)


def finalize():
    listboxindex = 0
    readbook = load_workbook(str(fileentry.get())+'.xlsx')
    sheets = readbook.sheetnames
    sheet = readbook[sheets[int(tab5monthindex)]]
    finalrow_index = 5 
    cellref = sheet.cell(row=finalrow_index,column=tab5dayindex)
    # item = listbox2.get(0)
    # print(item)
    while listboxindex < listbox2.size():
        item = listbox2.get(listboxindex)
        cellref.value = str(item)
        finalrow_index+=1
        listboxindex+=1
        cellref = sheet.cell(row=finalrow_index,column=tab5dayindex)
    readbook.save(str(fileentry.get())+'.xlsx')
    finishedLabel = Label(buttonframe,text='Finished',bg="grey26",fg="white")
    finishedLabel.grid(row=1,column=0)


finalizeButton = Button(buttonframe,text="Finalize",command=finalize)
finalizeButton.grid(row=0,column=0)

scrollbar = Scrollbar(listframe,orient=VERTICAL)
listbox1 = Listbox(listframe, height=10,exportselection=0,font=("12"))
listbox1.grid(row=1, column=0,padx=(0,50))
listbox2 = Listbox(listframe, height=10,width=30, exportselection=0,yscrollcommand=scrollbar.set,font=("12"))
listbox2.grid(row=1,column=1)
listbox3 = Listbox(listframe, height=10,exportselection=0,font=("12"))
listbox3.grid(row=1,column=2,padx=(50,0))


scrollbar.config(command=listbox2.yview)
scrollbar.grid(row=1,column=2,rowspan=1,  sticky=N+S+W)
# my_list2 = ['Robert','Law','Ryo','Mario','Laura','Skittle','-----']



def select(event=None):
    index = str(listbox2.curselection())
    print(len(index))
    if len(index) ==5:
        index=str(index)[1:3]
        print(index)
    else:
        index=str(index)[1]
        print(index)
    listbox2.delete(index)
    listbox2.insert(index, listbox1.get(ANCHOR))
    listbox2.itemconfig(index, bg = "lightgreen")
    if listbox1.get(ANCHOR) =="-----":
        None
    else:
        listbox1.delete(ANCHOR)

    # listbox1.delete(0)
    # listbox2.insert(1, listbox1.get(ANCHOR))
   
    # listbox1.delete(ANCHOR)

moveButton = Button(listframe,text="Move",command=select)
moveButton.grid(row=2,column=1,pady=10)



 
# def move(e):
#     mylabel.config(text="coord"+str(e.x))



# list = ["1","2","3","4"]

# for item in list:
#     listbox.insert(END, item)

# def select(event=None):
#     listbox2.insert(END, listbox.get(ANCHOR))
#     listbox.delete(ANCHOR)

# def deselect(event=None):
#     listbox.insert(END, listbox2.get(ANCHOR))
#     listbox2.delete(ANCHOR)

# tab5.bind('<Right>', select)
# tab5.bind('<Left>', deselect)
# tab5.bind('<Right>', select)
# tab5.bind('<Left>', deselect)

#TAB6=================================================SETTINGS==================================================#TAB6
def addemployee():
    existlabel = Label(tab6frame,text="Employee is Already in list",bg="grey26",fg="grey26")
    existlabel.grid(row=5,column=0)
    
    f= open('employees.txt','a')
    employeeentry = str(tab6entry.get())
    employeeentry = employeeentry.upper()
    if employeeentry == "":
        existlabel = Label(tab6frame,text="BLANK ENTRY",bg="grey26",fg="white")
        existlabel.grid(row=5,column=0)
        return
    
    tab6listbox1index = 0
    employeeexistsflag = False
    while tab6listbox1index < tab6listbox1.size():
        listemployee = str(tab6listbox1.get(tab6listbox1index))
        if employeeentry == listemployee:
            print("EMPLOYEE ALREADY EXISTS")
            employeeexistsflag = True
        tab6listbox1index+=1

    if employeeexistsflag == False:
        tab6listbox1.insert(END,employeeentry)
        tab6listbox1.itemconfig(END, bg = "lightgreen")
        existlabel = Label(tab6frame,text="ADDED!",bg="grey26",fg="white")
        existlabel.grid(row=5,column=0)
        f.write(employeeentry+"\n")
        
    else:
        existlabel = Label(tab6frame,text="Employee is Already in list",bg="grey26",fg="white")
        existlabel.grid(row=5,column=0)
    f.close()

    
        
    
def deleteemployee():
    f= open('employees.txt','r+')
    lines = sorted(f.readlines())
    f.close()
    requestedemployee = str(tab6listbox1.get(ANCHOR))
    print(requestedemployee)
    new_f = open('employees.txt','w')
    for line in lines:
        print(line)
        if line.strip("\n") != requestedemployee:
            new_f.write(line)
    new_f.close()
    tab6listbox1.delete(ANCHOR)


    
# f= open('employees.txt','w')
# lines = sorted(f.readlines())
# f.close()
# # new_f = open('employees.txt','w')
# # for line in lines:
# #     new_f.write(line)
# # new_f.close()

global mainemployeeslist
with open('employees.txt') as f:
    mainemployeeslist = [line.rstrip() for line in sorted(f)]
for item in mainemployeeslist:
    listbox1.insert(END, item)

tab6frame = Frame(tab6,bg="grey26")
tab6frame.grid(row=0,column=0)
allworkerslabel = Label(tab6frame,text="ALL EMPLOYEES")
allworkerslabel.grid(row=0,column=0)
tab6scrollbar = Scrollbar(tab6frame,orient=VERTICAL)
tab6listbox1 = Listbox(tab6frame, height=10,exportselection=0,font=("12"),yscrollcommand=tab6scrollbar.set)
tab6listbox1.grid(row=1, column=0)


global existlabel
existlabel = Label(tab6frame,text="Employee is Already in list",bg="grey26",fg="grey26")
existlabel.grid(row=5,column=0)
tab6scrollbar.config(command=tab6listbox1.yview)
tab6scrollbar.grid(row=1,column=1,rowspan=1,  sticky=N+S+W)
tab6addbutton = Button(tab6frame,text="ADD",command=addemployee)
tab6addbutton.grid(row=2,column=0)
tab6deletebutton = Button(tab6frame,text="DELETE",command=deleteemployee)
tab6deletebutton.grid(row=2,column=1)
tab6entry = Entry(tab6frame)
tab6entry.grid(row=3,column=0)


for person in mainemployeeslist:
    tab6listbox1.insert(END,person)



root.mainloop()


