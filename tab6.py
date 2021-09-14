from typing import List
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell import cell
from openpyxl.cell.cell import ERROR_CODES
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, alignment
from openpyxl.styles.fills import fills
from openpyxl.worksheet.dimensions import SheetDimension
from openpyxl.utils import get_column_letter
from tkinter import * 
from tkinter import ttk
from ttkthemes import ThemedTk
import time
import os.path
from os import error, path
from PIL import Image
from tkinter import messagebox
import traceback
import sys
import datetime
# import ver8
# from ver8 import tab1frame2,tab7frame3,tab7my_progress,tab7e1,tab7e2,tab7
# from gui import tab7,tab7e2,tab7e1,tab7my_progress,tab7frame2,tab7frame3
import gui

ft1 = Font(name='Arial',bold=True, size=14)
align = Alignment(horizontal='center')
thick_border = Border(left=Side(style='thick'), 
                     right=Side(style='thick'), 
                     top=Side(style='thick'), 
                     bottom=Side(style='thick'))

thick_border_blue = Border(left=Side(style='thick',color='0066CC'), 
                     right=Side(style='thick',color='0066CC')   
                  ) 

thick_border_blue_topBottom = Border(top=Side(style='thick',color='0066CC'), 
                     bottom=Side(style='thick',color='0066CC')   
                  )                    

thin_border_all = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

thin_border_all_grey = Border(left=Side(style='thin',color="DDDDDD"), 
                     right=Side(style='thin',color="DDDDDD"), 
                     top=Side(style='thin',color="DDDDDD"), 
                     bottom=Side(style='thin',color="DDDDDD"))

thin_border_sides = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style=None), 
                     bottom=Side(style=None))


thin_border_sides_Bottom = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style=None), 
                     bottom=Side(style='thin'))


#GLOBAL ERROR VARIABLES
fileerrorLabel = None
qualerrorLabel = None
blockerrorLabel = None
starterrorLabel = None
enderrorLabel = None
monthCheck = False

#GLOBAL MIRS
#QUAL1
global qual1mirs
qual1block1mirs=["3 MIRS / 1:6 / 3 HRS","0 MIRS","1 MIR / 1:12 / 3 HRS","0 MIRS"]
global qual2mirs
qual1block2mirs=["0 MIRS","0 MIRS","0 MIRS","1 MIR / 1:12 / 7 HRS","1 MIR / 1:12 / 4 HRS","1 MIR / 1:12 / 3 HRS","1 MIR / 1:12 / 2 HRS","1 MIR / 1:12 / 3.5 HRS","0 MIRS","1 MIR / 1:12 / 1.5HRS","1 MIR / 1:12 / 7.5HRS","1 MIR / 1:12 / 7.5HRS","1 MIR / 1:12 / 7.5HRS","0 MIRS","1 MIR / 1:12 / 7.5HRS","1 MIR / 1:12 / 7.5HRS","0 MIRS","0 MIRS"]
global qual1block3mirs
qual1block3mirs = [" 2 MIR / 1:8 / 4 HRS"," 2 MIR / 1:8 / 4.5 HRS"," 2 MIR / 1:8 / 5.5 HRS"," 2 MIR / 1:8 / 2.5 HRS"," 2 MIR / 1:8 / 4.5 HRS","0 MIRS"]
global qual1block4mirs
qual1block4mirs = ['2 MIRS / 1:8 / 6 HRS','2 MIRS/ 1:8 / 7 HRS','2 MIRS / 1:8 / 6 HRS','2 MIRS / 1:8 / 6.25 HRS','2 MIRS / 1:8 / 2 HRS','0 MIRS','2 MIRS / 1:8 / 5.5 HRS']
global qual1block5mirs
qual1block5mirs = ['0 MIRS','2 MIRS/ 1:8 / 3 HRS','2 MIRS/ 1:8 / 5.5 HRS','2 MIRS/ 1:8 / 5 HRS','2 MIRS/ 1:8 / 3.25 HRS','0 MIRS']
global qual1block6mirs
qual1block6mirs = ['1 MIRS / 1:12 / 3.5 HRS','1 MIRS / 1:12 / 5.75 HRS','1 MIRS / 1:12 / 4 HRS','1 MIRS / 1:12 / 6 HRS','1 MIRS / 1:12 / 5.25 HRS','1 MIRS / 1:12 / 4 HRS','1 MIRS / 1:12 / 6 HRS']
#QUAL2
global qual2block1mirs
qual2block1mirs =['1 MIR / 1:12 / 4 HRS','1 MIR / 1:12 / 5 HRS','3 MIRS / 1:6 / 8 HRS','3 MIRS / 1:6 / 8 HRS','3 MIRS / 1:6 / 8 HRS','3 MIRS / 1:6 / 8 HRS','3 MIRS / 1:6 / 8 HRS','2 MIRS / 1:8 / 8 HRS']
global qual2block2mirs
qual2block2mirs =['1 MIR / 1:12 / 3 HRS','1 MIR / 1:12 / 1 HR','3 MIRS / 1:6 / 3 HRS','3 MIRS / 1:6 / 8 HRS','3 MIRS / 1:6 / 8 HRS','3 MIR / 1:6 / 3 HRS','1 MIR / 1:12 / 2 HRS','2 MIRS / 1:8 / 6.5 HRS']
#QUAL3
global qual3block1mirs
qual3block1mirs = ['0 MIRS','1 MIRS / 1:12 / 4 HRS','1 MIRS / 1:12 / 8 HRS','1 MIRS / 1:12 / 3 HRS','1 MIRS / 1:12 / 5 HRS','2 MIRS / 1:8 / 8 HRS','2 MIRS / 1:8 / 8 HRS','1 MIRS / 1:12 / 8 HRS','1 MIRS / 1:12 / 8 HRS','1 MIRS / 1:12 / 6 HRS','1 MIRS / 1:12 / 6 HRS','1 MIRS / 1:12 / 1 HRS','0 MIRS']
global qual3block2mirs
qual3block2mirs =['1 MIRS / 1:12 / 3 HRS','1 MIRS / 1:12 / 8 HRS','1 MIRS / 1:12 / 8 HRS','1 MIRS / 1:12 / 8 HRS','1 MIRS / 1:12 / 8 HRS']
global qual3block3mirs
qual3block3mirs = ['1 MIRS / 1:12 / 1 HRS','1 MIRS / 1:12 / 8 HRS','1 MIRS / 1:12 / 5 HRS','1 MIRS / 1:12 / 8 HRS','1 MIRS / 1:12 / 4 HRS','2 MIRS / 1:8 / 6 HRS','2 MIRS / 1:8 / 8 HRS','2 MIRS / 1:8 / 4 HRS','0 MIRS']


global qual1block1daytotal
qual1block1daytotal = 3

global qual1block2daytotal
qual1block2daytotal = 17

global qual1block3daytotal
qual1block3daytotal = 5

global qual1block4daytotal
qual1block4daytotal = 6

global qual1block5daytotal
qual1block5daytotal = 5

global qual1block6daytotal
qual1block6daytotal = 6


# global lastDay


def readFullQualIn():
    global lastDay
    global lastDate
    global workbook_Title
    workbook_Title = gui.tab7e2.get()
    print(workbook_Title)
    global workbook
    if path.exists(str(gui.tab7e2.get())+".xlsx") == True: 
        workbook = load_workbook(filename=workbook_Title+".xlsx")
        # workbook.add_named_style(date_style)
    else:
        fileerrorLabel = Label(gui.tab7,text="File Does Not Exist",font="Helvetica 10 bold", fg="red")
        fileerrorLabel.grid(row=11,column=0)
        return
    # print(path.exists(str(e1.get())+".xlsx"))
    global sheet
    sheet = workbook.active
    print(workbook)
    global sheetIndex
    #Check Months
    startMonth = startdatecellref.value
    print(startMonth)
    startMonth = startMonth[0:2]
    print(startMonth)
    global monthCheck

    # month = str(e4.get())
    # print("Month: "+str(month[0:2]))
    print("Start Month: "+str(startMonth))
    if str(startMonth[0:2]) =="01":
        if'JAN' in workbook.sheetnames:
            sheet = workbook["JAN"]
            sheetIndex = 1
            # lastDay =31
        elif 'Jan' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("JAN")
            sheet = ws1
            sheetIndex = 1
            # lastDay =31

    if str(startMonth[0:2]) =="02":
        if'FEB' in workbook.sheetnames:
            sheet = workbook["FEB"]
            sheetIndex =2
            # lastDay =28
        elif 'Feb' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("FEB")
            sheet = ws1
            sheetIndex = 2
            # lastDay =28
    if str(startMonth[0:2]) =="03":
        if'MAR' in workbook.sheetnames:
            sheet = workbook["MAR"]
            sheetIndex = 3
            lastDay = 31
        elif 'MAR' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("MAR")
            sheet = ws1
            sheetIndex = 3
            lastDay=31
    if str(startMonth[0:2]) =="04":
        if'APR' in workbook.sheetnames:
            sheet = workbook["APR"]
            sheetIndex = 4
            lastDay=30
        elif 'APR' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("APR")
            sheet = ws1
            sheetIndex = 4
            lastDay=30
    if str(startMonth[0:2]) =="05":
        if'MAY' in workbook.sheetnames:
            sheet = workbook["MAY"]
            sheetIndex = 5
            lastDay=31
        elif 'MAY' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("MAY")
            sheet = ws1
            sheetIndex = 5
            lastDay=31

    if str(startMonth[0:2]) =="06":
        if'JUNE' in workbook.sheetnames:
            sheet = workbook["JUNE"]
            sheetIndex = 6
            lastDay=30
        elif 'JUNE' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("JUNE")
            sheet = ws1
            sheetIndex = 6
            lastDay=30

    if str(startMonth[0:2]) =="07":
        if'JULY' in workbook.sheetnames:
            sheet = workbook["JULY"]
            sheetIndex = 7
            lastDay=31
        elif 'JULY' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("JULY")
            sheet = ws1
            sheetIndex = 7
            lastDay=31

    if str(startMonth[0:2]) =="08":
        if'AUG' in workbook.sheetnames:
            sheet = workbook["AUG"]
            sheetIndex = 8
            lastDay=31
        elif 'AUG' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("AUG")
            sheet = ws1
            sheetIndex = 8
            lastDay=31

    if str(startMonth[0:2]) =="09":
        if'SEPT' in workbook.sheetnames:
            sheet = workbook["SEPT"]
            sheetIndex = 9
            lastDay=30
        elif 'SEPT' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("SEPT")
            sheet = ws1
            sheetIndex = 9
            lastDay=30
    if str(startMonth[0:2]) =="10":
        if'OCT' in workbook.sheetnames:
            sheet = workbook["OCT"]
            sheetIndex = 10
            lastDay=31
        elif 'OCT' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("OCT")
            sheet = ws1
            sheetIndex = 10
            lastDay=31
    if str(startMonth[0:2]) =="11":
        if'NOV' in workbook.sheetnames:
            sheet = workbook["NOV"]
            sheetIndex = 11
            lastDay=30
        elif 'NOV' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("NOV")
            sheet = ws1
            sheetIndex = 11
            lastDay=30

    if str(startMonth[0:1]) =="12":
        if'DEC' in workbook.sheetnames:
            sheet = workbook["DEC"]
            sheetIndex = 12
            lastDay=31
        elif 'DEC' not in workbook.sheetnames:
            ws1 = workbook.create_sheet("DEC")
            sheet = ws1
            sheetIndex = 12
            lastDay=31

    # if(startMonth == endMonth):
    #     None
    #     tab7add_qual()
    # elif(startMonth != endMonth):
    #     endDateReadIn = "30"
    #     monthCheck =True
    # print(qualcellref.value)
    # print(type(qualcellref.value))
    qualValue = str(qualcellref.value)
    global secondMonthCheck
    secondMonthCheck = False
    
    if qualValue == "1":


        #CHECKING THE LAST DAY ISNT A WEEKEND=============================================
        column_index=1
        lastDay=sheet.cell(row=1,column=column_index)
        lastWeekDay = sheet.cell(row=2,column=column_index)
        while lastDay.value:
            lastDate = lastDay.value
            lastWeek = lastWeekDay.value
            # print(lastDate)
            lastDate=lastDate[3:5]
            column_index+=1
            lastDay=sheet.cell(row=1,column=column_index)
            lastWeekDay = sheet.cell(row=2,column=column_index)
            # print(lastDate)
            # print(lastWeek)
            lastDate=int(lastDate)
            
        if lastWeek =="Saturday":
            lastDate = lastDate-1
        elif lastWeek =="Sunday":
            lastDate=lastDate-2
        #============================================================================================================
        #block1
        print("LAST DATE: "+str(lastDate))
        startDate = str(startdatecellref.value)
        startDate = startDate[3:5]
        #=================CHECKING NUMBER OF WEEKENDS====================================
        weekendstart = int(startDate)
        print(weekendstart)
        weekendCounter = 0
        weekendcell = sheet.cell(row=2,column=weekendstart)
        endDateReadIn = int(startDate)+qual1block1daytotal
        print("END DATE: "+str(endDateReadIn))
        while weekendstart<=endDateReadIn:
            if weekendcell.value =="Saturday":
                weekendCounter+=1
            weekendstart+=1
            weekendcell = sheet.cell(row=2,column=weekendstart)

        print("Weekend COUnt: "+str(weekendCounter))
        #========================================================================#

        print(startDate)
        endDateReadIn = int(startDate)+qual1block1daytotal+(2*weekendCounter)
        print("END DATE: "+str(endDateReadIn))
        if endDateReadIn >lastDate:
            endDateReadIn = lastDate
            daytotal = endDateReadIn - int(startDate)
            print("DAY TOTAL: "+str(daytotal))
            qualName = "1"
            blockName = "1"
            tab7add_qual(qualName,blockName,startDate,endDateReadIn)

            sheets = workbook.sheetnames
            print(str(sheet))
            sheet = workbook[sheets[sheetIndex+1]]
            startDate = "01"
            if daytotal == 0:
                endDateReadIn = (qual1block1daytotal - 1)
                print(str(endDateReadIn))
            else:
                endDateReadIn = (qual1block1daytotal - daytotal)+1
                print(str(endDateReadIn))
            tab7add_qual(qualName,blockName,startDate,endDateReadIn)
        elif endDateReadIn ==lastDate:
            print(endDateReadIn)
            qualName = "1"
            blockName = "1"
            print(endDateReadIn)
            tab7add_qual(qualName,blockName,startDate,endDateReadIn)
            sheets = workbook.sheetnames
            print(str(sheet))
            sheet = workbook[sheets[sheetIndex+1]]
            print(sheet.title)
            startDate = "01"
            secondMonthCheck = True
        else:
            qualName = "1"
            blockName = "1"
            tab7add_qual(qualName,blockName,startDate,endDateReadIn)

        #============================================================================================# 
        #block2
        print("LAST DATE: "+str(lastDate))
        startDate = endDateReadIn+1
        #=================CHECKING NUMBER OF WEEKENDS====================================
        weekendstart = int(startDate)
        print(weekendstart)
        weekendCounter = 0
        weekendcell = sheet.cell(row=2,column=weekendstart)
        endDateReadIn = int(startDate)+qual1block2daytotal
        print("END DATE: "+str(endDateReadIn))
        while weekendstart<=endDateReadIn:
            if weekendcell.value =="Saturday":
                weekendCounter+=1
            weekendstart+=1
            weekendcell = sheet.cell(row=2,column=weekendstart)

        print("Weekend COUnt2: "+str(weekendCounter))
        #========================================================================#

        print("START DATEL "+str(startDate))
        endDateReadIn = int(startDate)+qual1block2daytotal+(2*weekendCounter)
        print("END DATE: "+str(endDateReadIn))
        if endDateReadIn >lastDate:
            newendDateReadIn = lastDate
            print("END DATE: "+str(newendDateReadIn))
            daytotal = newendDateReadIn - int(startDate)
            daysleft = endDateReadIn-newendDateReadIn
            qualName = "1"
            blockName = "2"
            tab7add_qual(qualName,blockName,startDate,newendDateReadIn)

            sheets = workbook.sheetnames
            print(str(sheet))
            sheet = workbook[sheets[sheetIndex+1]]
            startDate = "01"
            if daysleft == 0:
                endDateReadIn = (daysleft)
                print(str(endDateReadIn))
            else:
                endDateReadIn = (daysleft)
                print("NEW END DATE: "+str(endDateReadIn))
            tab7add_qual(qualName,blockName,startDate,endDateReadIn)
        elif endDateReadIn ==lastDate:
            print(endDateReadIn)
            qualName = "1"
            blockName = "2"
            print(endDateReadIn)
            tab7add_qual(qualName,blockName,startDate,endDateReadIn)
            sheets = workbook.sheetnames
            print(str(sheet))
            sheet = workbook[sheets[sheetIndex+1]]
            print(sheet.title)
            startDate = "01"
            secondMonthCheck = True
        else:
            qualName = "1"
            blockName = "2"
            tab7add_qual(qualName,blockName,startDate,endDateReadIn)

         #============================================================================================# 
        #block3
        print("LAST DATE: "+str(lastDate))
        startDate = endDateReadIn+1
        #=================CHECKING NUMBER OF WEEKENDS====================================
        weekendstart = int(startDate)
        print(weekendstart)
        weekendCounter = 0
        weekendcell = sheet.cell(row=2,column=weekendstart)
        endDateReadIn = int(startDate)+qual1block3daytotal
        print("END DATE: "+str(endDateReadIn))
        while weekendstart<=endDateReadIn:
            if weekendcell.value =="Saturday":
                weekendCounter+=1
            weekendstart+=1
            weekendcell = sheet.cell(row=2,column=weekendstart)

        print("Weekend COUnt2: "+str(weekendCounter))
        #========================================================================#

        print("START DATEL "+str(startDate))
        endDateReadIn = int(startDate)+qual1block3daytotal+(2*weekendCounter)
        print("END DATE: "+str(endDateReadIn))
        if endDateReadIn >lastDate:
            newendDateReadIn = lastDate
            print("END DATE: "+str(newendDateReadIn))
            daytotal = newendDateReadIn - int(startDate)
            daysleft = endDateReadIn-newendDateReadIn
            qualName = "1"
            blockName = "3"
            tab7add_qual(qualName,blockName,startDate,newendDateReadIn)

            sheets = workbook.sheetnames
            print(str(sheet))
            sheet = workbook[sheets[sheetIndex+1]]
            startDate = "01"
            if daysleft == 0:
                endDateReadIn = (daysleft)
                print(str(endDateReadIn))
            else:
                endDateReadIn = (daysleft)
                print("NEW END DATE: "+str(endDateReadIn))
            tab7add_qual(qualName,blockName,startDate,endDateReadIn)
        elif endDateReadIn ==lastDate:
            print(endDateReadIn)
            qualName = "1"
            blockName = "3"
            print(endDateReadIn)
            tab7add_qual(qualName,blockName,startDate,endDateReadIn)
            sheets = workbook.sheetnames
            print(str(sheet))
            sheet = workbook[sheets[sheetIndex+1]]
            print(sheet.title)
            startDate = "01"
            secondMonthCheck = True
        else:
            qualName = "1"
            blockName = "3"
            tab7add_qual(qualName,blockName,startDate,endDateReadIn)

         #============================================================================================# 
        #block4
        print("LAST DATE: "+str(lastDate))
        startDate = endDateReadIn+1
        #=================CHECKING NUMBER OF WEEKENDS====================================
        weekendstart = int(startDate)
        print(weekendstart)
        weekendCounter = 0
        weekendcell = sheet.cell(row=2,column=weekendstart)
        endDateReadIn = int(startDate)+qual1block4daytotal
        print("END DATE: "+str(endDateReadIn))
        while weekendstart<=endDateReadIn:
            if weekendcell.value =="Saturday":
                weekendCounter+=1
            weekendstart+=1
            weekendcell = sheet.cell(row=2,column=weekendstart)

        print("Weekend COUnt2: "+str(weekendCounter))
        #========================================================================#

        print("START DATEL "+str(startDate))
        endDateReadIn = int(startDate)+qual1block4daytotal+(2*weekendCounter)
        print("END DATE: "+str(endDateReadIn))
        if endDateReadIn >lastDate:
            newendDateReadIn = lastDate
            print("END DATE: "+str(newendDateReadIn))
            daytotal = newendDateReadIn - int(startDate)
            daysleft = endDateReadIn-newendDateReadIn
            qualName = "1"
            blockName = "4"
            tab7add_qual(qualName,blockName,startDate,newendDateReadIn)

            sheets = workbook.sheetnames
            print(str(sheet))
            sheet = workbook[sheets[sheetIndex+1]]
            startDate = "01"
            if daysleft == 0:
                endDateReadIn = (daysleft)
                print(str(endDateReadIn))
            else:
                endDateReadIn = (daysleft)
                print("NEW END DATE: "+str(endDateReadIn))
            tab7add_qual(qualName,blockName,startDate,endDateReadIn)
        elif endDateReadIn ==lastDate:
            print(endDateReadIn)
            qualName = "1"
            blockName = "4"
            print(endDateReadIn)
            tab7add_qual(qualName,blockName,startDate,endDateReadIn)
            sheets = workbook.sheetnames
            print(str(sheet))
            sheet = workbook[sheets[sheetIndex+1]]
            print(sheet.title)
            startDate = "01"
            secondMonthCheck = True
        else:
            qualName = "1"
            blockName = "4"
            tab7add_qual(qualName,blockName,startDate,endDateReadIn)
           


def addFullQual():
    try:
        tab7Label3 = Label(gui.tab7frame3,text = "FINISHED",fg="grey26",bg="grey26",font="Helvetica 10 bold").grid(row=4,column=0)
        global tab7my_progress
        gui.tab7my_progress.grid(row=0,column=0,pady=(50,10))

        #GETTING READ FROM FILE
        global tab7readExcelFile
        tab7readExcelFile = gui.tab7e1.get()
        # print(tab7readExcelFile)
        if path.exists(str(gui.tab7e1.get())+".xlsx") == True: 
            readFromBook = load_workbook(filename =tab7readExcelFile+ '.xlsx')
        else:
            fileerrorLabel = Label(gui.tab7,text="File Does Not Exist",font="Helvetica 10 bold", fg="red")
            fileerrorLabel.grid(row=5,column=0)
            return

        
        sheet = readFromBook.active
        # print(sheet.title)
        startingRow = 2
        startingCol = 1
        global qualcellref
        # global blockcellref
        global startdatecellref
        # global enddatecellref
        global classcellref
        # global instructorcellref
        qualcellref = sheet.cell(row=startingRow,column=startingCol)
        # print(qualcellref.value)
        # blockcellref = sheet.cell(row=startingRow,column=startingCol+1)
        # print(blockcellref.value)
        startdatecellref = sheet.cell(row=startingRow,column=startingCol+1)
        # print(startdatecellref.value)
        # enddatecellref = sheet.cell(row=startingRow,column=startingCol+3)
        classcellref = sheet.cell(row=startingRow,column=startingCol+3)
        # instructorcellref = sheet.cell(row=startingRow,column=startingCol+5)
        
        # print(qualcellref.value)
        while qualcellref.value:
            print(str(qualcellref.value))
            print(qualcellref.coordinate)
            print("CLASS: "+str(classcellref.value))
            print(str(classcellref.coordinate))
            gui.tab7my_progress['value']+=30
                
            gui.tab7.update_idletasks() 
            # print("QUALCELLREF VALUE: " +str(qualcellref.value))
            readFullQualIn()
            startingRow+=1
            qualcellref = sheet.cell(row=startingRow,column=startingCol)
            # print(qualcellref.value)
            # blockcellref = sheet.cell(row=startingRow,column=startingCol+1)
            # print(blockcellref.value)
            startdatecellref = sheet.cell(row=startingRow,column=startingCol+2)
            print(startdatecellref.value)
            # enddatecellref = sheet.cell(row=startingRow,column=startingCol+3)
            classcellref = sheet.cell(row=startingRow,column=startingCol+3)
            print("CLASS: "+str(classcellref.value))
            print(str(classcellref.coordinate))
            
            instructorcellref = sheet.cell(row=startingRow,column=startingCol+5)
        tab7Label3 = Label(gui.tab7frame3,text = "FINISHED",fg="white",bg="grey26",font="Helvetica 10 bold").grid(row=4,column=0)
        
        gui.tab7my_progress.stop()
    except():
        print(traceback.format_exc())
        messagebox.showwarning(title="Error Occured", message="something went wrong in READ FROM EXCEL. Check your entries and try again")
        if path.exists("errors.txt") == TRUE:
            ct = datetime.datetime.now() 
            with open("errors.txt", "a") as file:
                file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))
        else:
            ct = datetime.datetime.now() 
            with open("errors.txt", "x") as file:
                file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))  


def tab7add_qual(qualName,blockName,startDate,endDateReadIn):
        try:

        
    #NONE DECLARED ERROR LABELS USED
            global fileerrorLabel
            global qualerrorLabel
            global blockerrorLabel
            global starterrorLabel
            global enderrorLabel
            global qual1block1mirs
            global qual1block2mirs      
        #SET COLUMN WIDTH
            # for col in range(1,32):
            #     column_letter = get_column_letter(col)
            #     # print(column_letter)
            #     sheet.column_dimensions[column_letter].width = 42

        #SET TITLE
            sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=32)
            sheet.row_dimensions[3].height = 60
            title_Cell = sheet['A3']
            title_Cell.border = thick_border
            monthTitle = str(sheet.title)
            # print(monthTitle)
            title_Cell.value = monthTitle+" ELECTRICAL SYSTEMS INSTRUCTOR SCHEDULE"
            title_Cell.font = Font(size=60,name="Times New Roman",bold=True)

            
            # global startMonth
            #SET DATES
            # for i in range(1,32):
            #     datecellref=sheet.cell(row=1, column=i)
            #     datecellref.fill = PatternFill("solid", fgColor="DDDDDD")
                
            #     if(i<10):
            #         datecellref.value=str(startMonth)+"/0"+str(i)+"/2021"
            #     else:
            #         datecellref.value=str(startMonth)+"/"+str(i)+"/2021"


            #SET DIVIDER      

            for i in range(1,32):
                datecellref2=sheet.cell(row=4, column=i)
                datecellref2.fill = PatternFill("solid", fgColor="000000")
                datecellref2.border = thick_border_blue_topBottom
                datecellref2.value="blank"  
                gui.tab7.update_idletasks() 


            #GETTING INPUT VALUES FROM USER
            # qualName = str(qualcellref.value)
            # instructor = str(instructorcellref.value)
            if str(qualName) == "FAMILY DAY":
                None
            else:
                classNum = int(classcellref.value)
                if int(qualName) > 3 and str(qualName) !="FAMILY DAY":
                    qualerrorLabel = Label(gui.tab7frame2,text = "Qual Num Doesn't Exist",font="Helvetica 10 bold", fg="red")
                    qualerrorLabel.grid(row=11,column=0)
                    # e2.delete(0, END)
                    return 
                # blockName = str(blockcellref.value)
                if int(blockName) > 6 and str(qualName) !="FAMILY DAY":
                    blockerrorLabel = Label(gui.tab7frame2,text = "Block Number Doesn't Exist",font="Helvetica 10 bold", fg="red")
                    blockerrorLabel.grid(row=11,column=0)
                    # e3.delete(0, END)
                    return
            # startDate = str(startdatecellref.value)
            # startDate = startDate[3:5]
            # print(str(startDate))
            if int(startDate) > 32:
                starterrorLabel = Label(gui.tab7frame2,text = "Start Date Number too big",font="Helvetica 10 bold", fg="red")
                starterrorLabel.grid(row=11,column=0)
                # e4.delete(0, END)
                return
            if int(endDateReadIn) > 32:
                enderrorLabel = Label(gui.tab7frame2,text = "End Date Number too big",font="Helvetica 10 bold", fg="red")
                enderrorLabel.grid(row=11,column=0)
                # e5.delete(0, END)
                return


            

            #Start Date
            col_index = int(startDate)
            row_index=5
            # print("First row index: "+str(row_index))

            #THE DASHED CELLS ITERATORS
            # blank_col_index = col_index
            # blank_row_index=row_index+1
            
            #TITLE CELL
            cellref = sheet.cell(row=row_index,column=col_index)
            cellref.border = thick_border
            blockcheckcell = sheet.cell(row=row_index, column = 1)
        

            #DASHED CELLS
            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)


            #Setting Styles
            # cellref.font=ft1
            # cellref.alignment=align
            # blank_cellref.font=ft1
            # blank_cellref.alignment=align
            # blank_cellref2.font=ft1
            # blank_cellref2.alignment=align
            # blank_cellref3.font=ft1
            # blank_cellref3.alignment=align
            # blank_cellref4.font=ft1
            # blank_cellref4.alignment=align

            #End Date
            #check if months match


            #SETTING ACTIVE MIRS
            activeMirs=[]
            
            if qualName =="1" and blockName == "1":
                for items in qual1block1mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="2":
                for items in qual1block2mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="3":
                for items in qual1block3mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="4":
                for items in qual1block4mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="5":
                for items in qual1block5mirs:
                    activeMirs.append(items)
            elif qualName =="1" and blockName =="6":
                for items in qual1block6mirs:
                    activeMirs.append(items)
            elif qualName =="2" and blockName =="1":
                for items in qual2block1mirs:
                    activeMirs.append(items)
            elif qualName =="2" and blockName =="2":
                for items in qual2block2mirs:
                    activeMirs.append(items)
            elif qualName =="3" and blockName =="1":
                for items in qual3block1mirs:
                    activeMirs.append(items)
            elif qualName =="3" and blockName =="2":
                for items in qual3block2mirs:
                    activeMirs.append(items)
            elif qualName =="3" and blockName =="3":
                for items in qual3block3mirs:
                    activeMirs.append(items)
        
                    



            des_col=int(endDateReadIn)+1

            
            
            total_index = 0

            mirsIndex = 0

            

            # print(tab5onleavelist[0])
            # print(tab5onleavelist[1])
            # print(tab5onleavelist[2])
            # for person in tab5onleavelist:
            #     print(person)
           
            
            
            
            
        
        
            q1check = False
            global colCheck
            colCheck = False


            #COLUMN ITERATOR LOOP
            # print(type(des_col))
            # print(des_col)
            while col_index < des_col:
                # print("DES COL: "+str(des_col))
                # print("LAST DATE: "+str(lastDate))
                # dayName = sheet.cell(row=2,column=col_index)
                # print("DAY: "+str(dayName.value))
                # if dayName.value =="Saturday" and not des_col+2 > lastDate:
                #     col_index+=2
                #     des_col+=2
                #     print("DES COL: "+str(des_col))
                #     print("LAST DATE: "+str(lastDate))
                #     continue
                # # elif dayName.value == "Sunday"and not des_col+1 > lastDate:
                # #     col_index+=1
                # #     des_col+=1
                # #     continue
                # elif dayName.value =="Saturday" and des_col+2 > lastDate:
                #     col_index+=2
                #     print(des_col)
                #     print("LAST DATE: "+str(lastDate))
                #     continue
                # # elif dayName.value == "Sunday"and des_col+1 > lastDate:
                # #     col_index+=1
                # #     continue

                colCheck = False
                tab5leavecheck = False
                tab5onleavelist = []
                tab5leaverow = 5
                leavecellref = sheet.cell(row=tab5leaverow,column=col_index)
                while leavecellref.value !="LEAVE":
                    tab5leaverow+=1
                    leavecellref = sheet.cell(row=tab5leaverow,column=col_index)
                if leavecellref.value =="LEAVE":
                    tab5leaverow+=1
                    leavecellref = sheet.cell(row=tab5leaverow,column=col_index)
                while leavecellref.value:
                    person = str(leavecellref.value)
                    # print(person)
                    tab5onleavelist.append(person)
                    tab5leaverow+=1
                    leavecellref = sheet.cell(row=tab5leaverow,column=col_index)

                # for person in tab5onleavelist:
                #     if instructor == person:
                #         tab5leavecheck = True 
                
                #IF THERES A VALUE IN THE CELL
                while cellref.value:
                    colCheck = False
                    # print(cellref.value)
                    if cellref.value == "Navy Unique":
                        sheet.insert_rows(row_index,5)
                        if col_index==1:
                            colCheck = True
                        else:
                            col_index-=1
                        # row_index-=5
                        cellref = sheet.cell(row=row_index,column=col_index)
                        # cellref.value="NEW"
                        # print(cellref.value)
                        workbook.save(filename=workbook_Title+".xlsx")
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        break
                    # print(str(cellref.value))
                    if str(cellref.value) == "FAMILY DAY":
                        col_index+=1
                        cellref = sheet.cell(row=row_index,column=col_index)
                        None
                    
                    elif str(cellref.value)!= "FAMILY DAY" and str(qualName)=="FAMILY DAY":
                        while row_index <=100:
                            cellref.value = "FAMILY DAY"
                            cellref.font = Font(color="FF0000")
                            cellref.fill = PatternFill("solid","FFFFFF")
                            cellref.border = thin_border_all_grey
                            row_index+=1
                            cellref = sheet.cell(row=row_index,column=col_index)

                    #First: IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK IS LESS THAN OR EQUAL AND THE CLASS NUM ENTERED IS GREATER
                    elif(int(qualName) <= int(cellref.value[1]) and int(blockName) <= int(cellref.value[10]) and int(qualName) == 1 and int(blockName)==1):
                        # print("Start Date: "+str(startDate))
                        # print("Start Date Type: "+str(type(startDate)))
                        if startDate!="01" and (int(classNum) >= int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999"):
                            sheet.insert_rows(5,5)
                            q1check = True
                            for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(startDate)-1):
                                for cell in rows:
                                    cell.value = "Q1  Block:1 Class:99999"
                                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid",fgColor="white")
                                    cell.border = thin_border_all_grey
                                    cell.font = Font(color="FFFFFF")
                            # row_index=4
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                        elif startDate!="01" and (int(classNum) < int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999"):
                            sheet.insert_rows(5,5)
                            q1check = True
                            for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(startDate)-1):
                                for cell in rows:
                                    cell.value = "Q1  Block:1 Class:99999"
                                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid",fgColor="white")
                                    cell.border = thin_border_all_grey
                                    cell.font = Font(color="FFFFFF")
                                    
                            row_index=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                        elif startDate=="01" and int(classNum) > int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999" :
                            sheet.insert_rows(5,5)
                            row_index=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        elif startDate=="01" and (int(classNum) < int(cellref.value[18:23]) or cellref.value =="Q1  Block:1 Class:99999") :
                            row_index+=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)

                        #Second:IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK NUM IS GREATER OR EQUAL 
                    elif(int(qualName) <= int(cellref.value[1]) and int(blockName) > int(cellref.value[10])):
                        row_index+=5
                        # print("Row index: " + str(row_index))
                        # print(cellref.value)
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        

                        #Third:IF THE ENTERED QUAL IS LESS THAN OR EQUAL AND THE BLOCK IS LESS THAN OR EQUAL
                    elif(int(qualName) <= int(cellref.value[1]) and int(blockName) <= int(cellref.value[10])):
                        if int(classNum) > int(cellref.value[18:23]):
                            # print("Cell Coordinate: "+str(cellref))
                            # print("Last Blank Cell Coordinate: "+str(blank_cellref4))
                            sheet.insert_rows(row_index,5)
                            workbook.save(filename=workbook_Title+".xlsx")
                            newrowref = cellref.row
                            # print("Newrowref: "+str(newrowref))
                            row_index=newrowref-5
                            # print("row index: "+str(row_index))
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        else:

                            row_index+=5
                            cellref = sheet.cell(row=row_index,column=col_index)
                            blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                            blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                            blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                            blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)


                        #FOURTH: IF THE ENTERED QUAL IS GREATER THAN THE CELL VALUE
                    elif(int(qualName) > int(cellref.value[1])):
                        row_index+=5
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)


                    elif(int(qualName)<=int(cellref.value[1])):
                        sheet.move_range(""+str(cellref.coordinate)+":"+str(blank_cellref4.coordinate)+"", rows=5)
                        newrowref = cellref.row
                        row_index=newrowref-5
                        cellref = sheet.cell(row=row_index,column=col_index)
                        blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                        blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                        blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                        blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                        
                else:
                    if str(qualName)=="FAMILY DAY":
                        while row_index <=100:
                            cellref.value = "FAMILY DAY"
                            cellref.font = Font(color="FF0000")
                            cellref.fill = PatternFill("solid","FFFFFF")
                            cellref.border = thin_border_all_grey
                            row_index+=1
                            cellref = sheet.cell(row=row_index,column=col_index)
                        
                    elif(int(qualName)==1 and int(blockName)==1 and startDate!="01" and q1check ==False):
                        sheet.insert_rows(5,5)
                        # print("Row Index: "+str(row_index))
                        q1check =True
                        for rows in sheet.iter_rows(min_row=5,max_row=9,min_col=1,max_col=int(startDate)-1):
                                for cell in rows:
                                    cell.value = "Q1  Block:1 Class:99999"
                                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid",fgColor="white")
                                    cell.border = thin_border_all_grey
                                    cell.font = Font(color="FFFFFF")
                                
                                    # print("col index: "+str(col_index))
                    # row_index=4
                    # print("Row Index: "+str(row_index))
                    cellref = sheet.cell(row=row_index,column=col_index)
                    blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                    blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                    blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                    blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                    if str(qualName)=="FAMILY DAY":
                        None    
                    else:
                        cellref.value="Q"+str(qualName)+" " +" Block:"+str(blockName)+" Class:"+str(classNum)
                        cellref.font = ft1
                        cellref.alignment=align
                    # print("col index: "+str(col_index))
                        #TITLE COLOR CODING
                        #QUAL 1
                        cellref.border = thick_border
                        if (qualName == "1" and blockName == "1"):
                            cellref.fill = PatternFill("solid", fgColor="00CCFF")
                        elif (qualName == "1"and blockName == "2"):
                            cellref.fill = PatternFill("solid", fgColor="33CCCC")
                        elif (qualName == "1" and blockName == "3"):
                            cellref.fill = PatternFill("solid", fgColor="92D050")
                        elif (qualName == "1"and blockName == "4"):
                            cellref.fill = PatternFill("solid", fgColor="FFFF00")
                        elif (qualName == "1" and blockName == "5"):
                            cellref.fill = PatternFill("solid", fgColor="FFC000")
                        elif (qualName == "1"and blockName == "6"):
                            cellref.fill = PatternFill("solid", fgColor="FF0000")

                        #QUAL2
                        elif (qualName == "2" and blockName == "1"):
                            cellref.fill = PatternFill("solid", fgColor="00B0F0")
                        elif (qualName == "2"and blockName == "2"):
                            cellref.fill = PatternFill("solid", fgColor="92D050")

                        #QUAL3
                        elif (qualName == "3" and blockName == "1"):
                            cellref.fill = PatternFill("solid", fgColor="FFFF00")
                        elif (qualName == "3"and blockName == "2"):
                            cellref.fill = PatternFill("solid", fgColor="9BBB59")
                        elif (qualName == "3"and blockName == "3"):
                            cellref.fill = PatternFill("solid", fgColor="FFC000")
                        


                        #SETTING BLANK VALUES BORDERS
                        blank_cellref.border = thin_border_sides
                        blank_cellref2.border = thin_border_sides
                        blank_cellref3.border = thin_border_sides
                        blank_cellref4.border = thin_border_sides_Bottom

                        #SETTING BLANK VALUES
                        # if tab5leavecheck == True:
                        #     messagebox.showwarning(title="Instructor on Leave!", message="The entered instructor is on leave for one of the days")
                        #     blank_cellref.value="Instructor: Instructor is on leave"
                        #     tab5leavecheck ==False
                        # else:
                        #     blank_cellref.value="Instructor: "+instructor
                        blank_cellref.font = ft1
                        blank_cellref.alignment=align
                        blank_cellref2.value="------"
                        blank_cellref2.font = ft1
                        blank_cellref2.alignment=align
                        blank_cellref3.value="------"
                        blank_cellref3.font = ft1
                        blank_cellref3.alignment=align
                        if mirsIndex > len(activeMirs)-1:
                            blank_cellref4.value = "NonPop"
                        else:
                            blank_cellref4.value=str(activeMirs[mirsIndex])
                        blank_cellref4.font = ft1
                        blank_cellref4.alignment=align
                    total_index+=1
                   
                    # print("Title = " + cellref.coordinate)
                    # print("blank = " + blank_cellref.coordinate)
                
                if colCheck ==True:
                    None
                else:
                    col_index+=1
                mirsIndex+=1
                # print("column index: "+str(col_index))   


                #This line controls if they shoot up the rows or nor
                # row_index=4
                # print("row index"+str(row_index))   
                cellref = sheet.cell(row=row_index,column=col_index)
                blank_cellref = sheet.cell(row=row_index+1,column=col_index)
                blank_cellref2 = sheet.cell(row=row_index+2,column=col_index)
                blank_cellref3 = sheet.cell(row=row_index+3,column=col_index)
                blank_cellref4 = sheet.cell(row=row_index+4,column=col_index)
                # print("Row Index: "+str(row_index))
                # print("cellref: "+str(cellref))
                workbook.save(filename=workbook_Title+".xlsx")


                
                
                # gui.submitTotal
                submittedLabel = Label(gui.tab7frame3,text = "Submitted ",font="Helvetica 10 bold", fg="black",bg="grey26")
                datesLabel = Label(gui.tab7frame3,text = "Total Number of Dates Affected: "+str(gui.submitTotal),font="Helvetica 10 bold", fg="black",bg="grey26")
                submittedLabel.grid(row=1,column=0)
                datesLabel.grid(row=2,column=0)
                gui.submitTotal+=1
                # e2.delete(0, END)
                # e3.delete(0, END)
                # e4.delete(0, END)
                # e5.delete(0, END)

            # second_Month_Check_ReadIn() 
        except:
            print(traceback.format_exc()) 
            messagebox.showwarning(title="Error Occured", message="something went wrong in Add Qual ReadIN. Check your entries and try again")
            if path.exists("errors.txt") == TRUE:
                ct = datetime.datetime.now() 
                with open("errors.txt", "a") as file:
                    file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))
            else:
                ct = datetime.datetime.now() 
                with open("errors.txt", "x") as file:
                    file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))