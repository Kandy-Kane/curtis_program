from typing import List
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell import cell
from openpyxl.cell.cell import ERROR_CODES
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, alignment
from openpyxl.styles.fills import fills
from openpyxl.worksheet.dimensions import SheetDimension
from openpyxl.utils import get_column_letter
from tkinter import * 
from tkinter import filedialog
from pathlib import Path
from tkinter import ttk
from ttkthemes import ThemedTk
import time
import os.path
from os import error, path
# from PIL import Image
from tkinter import messagebox
import traceback
import sys
import datetime

ft1 = Font(name='Arial',bold=True, size=14)
align = Alignment(horizontal='center')
thick_border = Border(left=Side(style='thick'), 
                     right=Side(style='thick'), 
                     top=Side(style='thick'), 
                     bottom=Side(style='thick'))

thick_border_blue = Border(left=Side(style='thick',color='0066CC'), 
                     right=Side(style='thick',color='0066CC')   
                  ) 

thick_border_blue_topBottom = Border(top=Side(style='thick',color='0066CC'), 
                     bottom=Side(style='thick',color='0066CC')   
                  )                    

thin_border_all = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

thin_border_all_grey = Border(left=Side(style='thin',color="DDDDDD"), 
                     right=Side(style='thin',color="DDDDDD"), 
                     top=Side(style='thin',color="DDDDDD"), 
                     bottom=Side(style='thin',color="DDDDDD"))

thin_border_sides = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style=None), 
                     bottom=Side(style=None))


thin_border_sides_Bottom = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style=None), 
                     bottom=Side(style='thin'))


#GLOBAL ERROR VARIABLES
fileerrorLabel = None
qualerrorLabel = None
blockerrorLabel = None
starterrorLabel = None
enderrorLabel = None
monthCheck = False


#MAIN PAGE
root = ThemedTk(theme="black")
root.geometry("800x550")
root.title("Curtis Scheduling Tool "+"                                                                                          \u00A9" + " KandyKane Solutions  Ver.8.0.0*")
tab_parent = ttk.Notebook(root)
# tab1 = ttk.Frame(tab_parent)
# tab2 = ttk.Frame(tab_parent)
# tab3 = ttk.Frame(tab_parent)
# tab4 = ttk.Frame(tab_parent)
tab5 = ttk.Frame(tab_parent)
# tab6 = ttk.Frame(tab_parent)

# tab_parent.add(tab1,text="Use Existing")
# tab_parent.add(tab2,text="New Workbook")
# tab_parent.add(tab3,text="Read From File")
# tab_parent.add(tab4,text="About")
tab_parent.add(tab5,text="Add MIRS")
# tab_parent.add(tab6,text="Settings")
tab_parent.pack(expand=1,fill='both')


# bg = PhotoImage(file="background5.png")
# my_label = Label(root,image=bg)
# my_label.place(x=0,y=0,relwidth=1,relheight=1)
# p1 = PhotoImage(file = 'airforce.png')
# root.iconphoto(False, p1)








#TAB5=================================================ADD MIRS==================================================#TAB5
maintitle = Frame(tab5,bg="grey26")
maintitle.pack()
titleframe = Frame(tab5,bg="grey26")
titleframe.pack()
listframe = Frame(tab5,bg="grey26")
listframe.pack(pady=(20))
buttonframe = Frame(tab5,bg="grey26")
buttonframe.pack(padx=(0,100))

title = Label(maintitle,text = "ADD MIRS",fg="white",bg="grey26",font="Helvetica 36 bold")
title.pack()

filename = Label(titleframe,text="File name:")
filename.grid(row=0,column=0,padx=5)
fileentry = Entry(titleframe,width=15)
fileentry.grid(row=0,column=1,padx=5)




datelabel = Label(titleframe,text='Date:')
datelabel.grid(row=0,column=3,padx=5)
dateentry = Entry(titleframe,width=15)
dateentry.grid(row=0,column=4,padx=5)

allpersons = Label(listframe,text="AVAILABLE EMPLOYESS")
onleave = Label(listframe,text="ON LEAVE")
daypreview = Label(listframe,text="DAY PREVIEW")
allpersons.grid(row=0,column=0)
daypreview.grid(row=0,column=1)
onleave.grid(row=0,column=2)

finishedLabel = Label(buttonframe,text='Finished',bg="grey26",fg="grey26",font="Helvetica 36 bold")
finishedLabel.grid(row=1,column=0)

def dialog():
    filename =  filedialog.askopenfilename(initialdir = "./",title = "Select file",filetypes = (("excel files","*.xlsx"),("all files","*.*")))
    print(filename)
    # print(root.filename)
    filepath=str(filename)
    # print(filepath)
    path=Path(filepath)
    fileentry.insert(0,str(path.name))
    print(path.name)

directButton = Button(titleframe,text="^",command=dialog)
directButton.grid(row=0,column=2,padx=5)

def rightarrow():
    tab5dayindex = dateentry.get()[3:5]
    tab5dayindex = int(tab5dayindex)
    newtab5dayindex = tab5dayindex+1
    dateentry.delete(3,5)
    if newtab5dayindex <=9:
        dateentry.insert(3,"0"+str(newtab5dayindex))
    else:
        dateentry.insert(3,str(newtab5dayindex))
    populateList()  

def leftarrow():
    tab5dayindex = dateentry.get()[3:5]
    tab5dayindex = int(tab5dayindex)
    newtab5dayindex = tab5dayindex-1
    dateentry.delete(3,5)
    if newtab5dayindex <=9:
        dateentry.insert(3,"0"+str(newtab5dayindex))
    else:
        dateentry.insert(3,str(newtab5dayindex)) 
    populateList()     

def populateList():
    finishedLabel = Label(buttonframe,text='Finished',bg="grey26",fg="grey26")
    finishedLabel.grid(row=1,column=0)
    
    global my_list2
    global listbox1
    listbox1.delete(0,END)
    listbox3.delete(0,END)
    


    listbox2.delete(0, END)
    global tab5monthindex
    global tab5dayindex
    tab5monthindex = dateentry.get()[0:2]
    tab5dayindex = dateentry.get()[3:5]
    tab5monthindex = int(tab5monthindex)
    tab5dayindex = int(tab5dayindex)
    #MONTH INDEX
    print(tab5monthindex)
    print(tab5dayindex)
    
    readbook = load_workbook(str(fileentry.get()))
    sheets = readbook.sheetnames
    sheet = readbook[sheets[int(tab5monthindex)]]
    print(sheet.title)
    
    tab5row_index = 5 
    cellref = sheet.cell(row=tab5row_index,column=tab5dayindex)
    # print(cellref.value)
    while cellref.value != "END":
        listbox2.insert(END,cellref.value)
        if not cellref.value:
            listbox2.insert(END,"")
        elif "Class:99999" in str(cellref.value):
            listbox2.itemconfig("end", bg = "red")
        elif "Q" in str(cellref.value):
            listbox2.itemconfig("end", bg = "blue")

        tab5row_index+=1
        cellref =sheet.cell(row=tab5row_index,column=tab5dayindex)
        # print(cellref.value)
    
    tab5leavecheck = False
    tab5onleavelist = []
    tab5leaverow = 5
    leavecellref = sheet.cell(row=tab5leaverow,column=tab5dayindex)
    while leavecellref.value !="LEAVE":
        tab5leaverow+=1
        leavecellref = sheet.cell(row=tab5leaverow,column=tab5dayindex)
    if leavecellref.value =="LEAVE":
        tab5leaverow+=1
        leavecellref = sheet.cell(row=tab5leaverow,column=tab5dayindex)
    while leavecellref.value:
        person = str(leavecellref.value)
        person = person.upper()
        print(person)
        tab5onleavelist.append(person)
        tab5leaverow+=1
        leavecellref = sheet.cell(row=tab5leaverow,column=tab5dayindex)
    
    for person in tab5onleavelist:
        listbox3.insert(END,person)
    
    
    for item in mainemployeeslist:
        if item not in tab5onleavelist:
            listbox1.insert(END, item)

populate = Button(titleframe,text="Populate",command=populateList)
populate.grid(row=0,column=7,padx=5)

rightArrow = Button(titleframe,text=">",command=rightarrow)
rightArrow.grid(row=0,column=6)
leftArrow = Button(titleframe,text="<",command=leftarrow)
leftArrow.grid(row=0,column=5)


def finalize():
    
    listboxindex = 0
    readbook = load_workbook(str(fileentry.get()))
    sheets = readbook.sheetnames
    sheet = readbook[sheets[int(tab5monthindex)]]
    finalrow_index = 5 
    cellref = sheet.cell(row=finalrow_index,column=tab5dayindex)
    # item = listbox2.get(0)
    # print(item)
    while listboxindex < listbox2.size():
        item = listbox2.get(listboxindex)
        cellref.value = str(item)
        finalrow_index+=1
        listboxindex+=1
        cellref = sheet.cell(row=finalrow_index,column=tab5dayindex)
    readbook.save(str(fileentry.get()))
    finishedLabel = Label(buttonframe,text='Finished',bg="grey26",fg="white",font="Helvetica 36 bold")
    finishedLabel.grid(row=1,column=0,padx=(100,0))


finalizeButton = Button(buttonframe,text="Finalize",command=finalize,width=15,height=3)
finalizeButton.grid(row=0,column=0,padx=(100,0))

scrollbar = Scrollbar(listframe,orient=VERTICAL)
listbox1 = Listbox(listframe, height=10,exportselection=0,font=("12"))
listbox1.grid(row=1, column=0,padx=(0,50))
listbox2 = Listbox(listframe, height=10,width=30, exportselection=0,yscrollcommand=scrollbar.set,font=("12"))
listbox2.grid(row=1,column=1)
listbox3 = Listbox(listframe, height=10,exportselection=0,font=("12"))
listbox3.grid(row=1,column=2,padx=(50,0))


scrollbar.config(command=listbox2.yview)
scrollbar.grid(row=1,column=2,rowspan=1,  sticky=N+S+W)
# my_list2 = ['Robert','Law','Ryo','Mario','Laura','Skittle','-----']



def select(event=None):
    finishedLabel = Label(buttonframe,text='Finished',bg="grey26",fg="grey26",font="Helvetica 36 bold")
    finishedLabel.grid(row=1,column=0,padx=(100,0))
    index = str(listbox2.curselection())
    print(len(index))
    if len(index) ==5:
        index=str(index)[1:3]
        print(index)
    else:
        index=str(index)[1]
        print(index)
    listbox2.delete(index)
    listbox2.insert(index, listbox1.get(ANCHOR))
    listbox2.itemconfig(index, bg = "lightgreen")
    if listbox1.get(ANCHOR) =="-----":
        None
    else:
        listbox1.delete(ANCHOR)

    # listbox1.delete(0)
    # listbox2.insert(1, listbox1.get(ANCHOR))
   
    # listbox1.delete(ANCHOR)

moveButton = Button(listframe,text="Move",command=select)
moveButton.grid(row=2,column=1,pady=10)



 
# def move(e):
#     mylabel.config(text="coord"+str(e.x))



# list = ["1","2","3","4"]

# for item in list:
#     listbox.insert(END, item)

# def select(event=None):
#     listbox2.insert(END, listbox.get(ANCHOR))
#     listbox.delete(ANCHOR)

# def deselect(event=None):
#     listbox.insert(END, listbox2.get(ANCHOR))
#     listbox2.delete(ANCHOR)

# tab5.bind('<Right>', select)
# tab5.bind('<Left>', deselect)
# tab5.bind('<Right>', select)
# tab5.bind('<Left>', deselect)

#TAB6=================================================SETTINGS==================================================#TAB6
# def addemployee():
#     existlabel = Label(tab6frame,text="Employee is Already in list",bg="grey26",fg="grey26")
#     existlabel.grid(row=5,column=0)
    
#     f= open('employees.txt','a')
#     employeeentry = str(tab6entry.get())
#     employeeentry = employeeentry.upper()
#     if employeeentry == "":
#         existlabel = Label(tab6frame,text="BLANK ENTRY",bg="grey26",fg="white")
#         existlabel.grid(row=5,column=0)
#         return
    
#     tab6listbox1index = 0
#     employeeexistsflag = False
#     while tab6listbox1index < tab6listbox1.size():
#         listemployee = str(tab6listbox1.get(tab6listbox1index))
#         if employeeentry == listemployee:
#             print("EMPLOYEE ALREADY EXISTS")
#             employeeexistsflag = True
#         tab6listbox1index+=1

#     if employeeexistsflag == False:
#         tab6listbox1.insert(END,employeeentry)
#         tab6listbox1.itemconfig(END, bg = "lightgreen")
#         existlabel = Label(tab6frame,text="ADDED!",bg="grey26",fg="white")
#         existlabel.grid(row=5,column=0)
#         f.write(employeeentry+"\n")
        
#     else:
#         existlabel = Label(tab6frame,text="Employee is Already in list",bg="grey26",fg="white")
#         existlabel.grid(row=5,column=0)
#     f.close()

    
        
    
# def deleteemployee():
#     f= open('employees.txt','r+')
#     lines = sorted(f.readlines())
#     f.close()
#     requestedemployee = str(tab6listbox1.get(ANCHOR))
#     print(requestedemployee)
#     new_f = open('employees.txt','w')
#     for line in lines:
#         print(line)
#         if line.strip("\n") != requestedemployee:
#             new_f.write(line)
#     new_f.close()
#     tab6listbox1.delete(ANCHOR)


    
# # f= open('employees.txt','w')
# # lines = sorted(f.readlines())
# # f.close()
# # # new_f = open('employees.txt','w')
# # # for line in lines:
# # #     new_f.write(line)
# # # new_f.close()

global mainemployeeslist
with open('employees.txt') as f:
    mainemployeeslist = [line.rstrip() for line in sorted(f)]
for item in mainemployeeslist:
    listbox1.insert(END, item)

# tab6frame = Frame(tab6,bg="grey26")
# tab6frame.grid(row=0,column=0)
# allworkerslabel = Label(tab6frame,text="ALL EMPLOYEES")
# allworkerslabel.grid(row=0,column=0)
# tab6scrollbar = Scrollbar(tab6frame,orient=VERTICAL)
# tab6listbox1 = Listbox(tab6frame, height=10,exportselection=0,font=("12"),yscrollcommand=tab6scrollbar.set)
# tab6listbox1.grid(row=1, column=0)


# global existlabel
# existlabel = Label(tab6frame,text="Employee is Already in list",bg="grey26",fg="grey26")
# existlabel.grid(row=5,column=0)
# tab6scrollbar.config(command=tab6listbox1.yview)
# tab6scrollbar.grid(row=1,column=1,rowspan=1,  sticky=N+S+W)
# tab6addbutton = Button(tab6frame,text="ADD",command=addemployee)
# tab6addbutton.grid(row=2,column=0)
# tab6deletebutton = Button(tab6frame,text="DELETE",command=deleteemployee)
# tab6deletebutton.grid(row=2,column=1)
# tab6entry = Entry(tab6frame)
# tab6entry.grid(row=3,column=0)


# for person in mainemployeeslist:
#     tab6listbox1.insert(END,person)



root.mainloop()


