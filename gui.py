from typing import List
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell import cell
from openpyxl.cell.cell import ERROR_CODES
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, alignment
from openpyxl.styles.fills import fills
from openpyxl.worksheet.dimensions import SheetDimension
from openpyxl.utils import get_column_letter
from tkinter import * 
from tkinter import ttk
from ttkthemes import ThemedTk
import time
import os.path
from os import error, path
from PIL import Image
from tkinter import messagebox
import traceback
import sys
import datetime
from ver8 import*
import tab6
from tab6 import addFullQual
# from ver8 import *
from ver8 import existing_WORKBOOK
from ver8 import new_WORKBOOK
from ver8 import readFromExcel


ft1 = Font(name='Arial',bold=True, size=14)
align = Alignment(horizontal='center')
thick_border = Border(left=Side(style='thick'), 
                     right=Side(style='thick'), 
                     top=Side(style='thick'), 
                     bottom=Side(style='thick'))

thick_border_blue = Border(left=Side(style='thick',color='0066CC'), 
                     right=Side(style='thick',color='0066CC')   
                  ) 

thick_border_blue_topBottom = Border(top=Side(style='thick',color='0066CC'), 
                     bottom=Side(style='thick',color='0066CC')   
                  )                    

thin_border_all = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

thin_border_all_grey = Border(left=Side(style='thin',color="DDDDDD"), 
                     right=Side(style='thin',color="DDDDDD"), 
                     top=Side(style='thin',color="DDDDDD"), 
                     bottom=Side(style='thin',color="DDDDDD"))

thin_border_sides = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style=None), 
                     bottom=Side(style=None))


thin_border_sides_Bottom = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style=None), 
                     bottom=Side(style='thin'))


#GLOBAL ERROR VARIABLES
fileerrorLabel = None
qualerrorLabel = None
blockerrorLabel = None
starterrorLabel = None
enderrorLabel = None
monthCheck = False

def changeOnHover(button, colorOnHover, colorOnLeave):
  
    # adjusting backgroung of the widget
    # background on entering widget
    button.bind("<Enter>", func=lambda e: button.config(
        background=colorOnHover))
  
    # background color on leving widget
    button.bind("<Leave>", func=lambda e: button.config(
        background=colorOnLeave))



#MAIN PAGE
root = ThemedTk(theme="black")
root.geometry("800x550")
root.title("Curtis Scheduling Tool "+"                                                                                          \u00A9" + " KandyKane Solutions  Ver.8.0.0*")
tab_parent = ttk.Notebook(root)
tab1 = ttk.Frame(tab_parent)
tab2 = ttk.Frame(tab_parent)
tab3 = ttk.Frame(tab_parent)
tab4 = ttk.Frame(tab_parent)
tab5 = ttk.Frame(tab_parent)
tab6 = ttk.Frame(tab_parent)
tab7 = ttk.Frame(tab_parent)

tab_parent.add(tab1,text="USE EXISTING")
tab_parent.add(tab2,text="NEW WORKBOOK")
tab_parent.add(tab3,text="Read IN ENTRIES")
tab_parent.add(tab5,text="Add MIRS")
tab_parent.add(tab7,text="ADD FULL QUAL")
tab_parent.add(tab6,text="SETTINGS")
tab_parent.add(tab4,text="ABOUT")
tab_parent.pack(expand=1,fill='both')


# bg = PhotoImage(file="background5.png")
# my_label = Label(root,image=bg)
# my_label.place(x=0,y=0,relwidth=1,relheight=1)
p1 = PhotoImage(file = 'airforce.png')
root.iconphoto(False, p1)




#TAB1==============================================================================================================TAB1
tab1titleframe=Frame(tab1,bg="grey26")
tab1titleframe.pack(side=TOP)
tab1mainframe = Frame(tab1,width=1000,bg="grey26", height=500,highlightbackground="black",highlightthickness=3)
tab1mainframe.pack(side=TOP,padx=(0,400))

tab1frame = Frame(tab1mainframe,bg="grey10")
tab1frame.pack()
tab1frame2 = Frame(tab1mainframe,bg="grey26")
tab1frame2.pack()
tab1frame3 = Frame(tab1mainframe,bg="grey26")
tab1frame3.pack()

tab1titlelabel = Label(tab1titleframe,text = "SINGLE ENTRY",fg="white",bg="grey26",font="Helvetica 36 bold").grid(row=1,column=0)
tab1Label0 = Label(tab1frame,text = "Entry",fg="white",bg="grey10",font="Helvetica 20 bold").grid(row=1,column=0,padx=140)
tab1Label0 = Label(tab1frame2,text = "File Name:",fg="white",bg="grey26",font="Helvetica 15 bold").grid(row=2,column=0,pady=2)
tab1Label1 = Label(tab1frame2,text = "Qual Num:",fg="white",bg="grey26",font="Helvetica 15 bold").grid(row=3,column=0,pady=2)
tab1Label2 = Label(tab1frame2,text = "Block Num:",fg="white",bg="grey26",font="Helvetica 15 bold").grid(row=4,column=0,pady=2)
tab1Label3 = Label(tab1frame2,text = "Start Date:",fg="white",bg="grey26",font="Helvetica 15 bold").grid(row=5,column=0,pady=2)
tab1Label4 = Label(tab1frame2,text = "End Date:",fg="white",bg="grey26",font="Helvetica 15 bold").grid(row=6,column=0,pady=2)
tab1Label5 = Label(tab1frame2,text = "Class Num:",fg="white",bg="grey26",font="Helvetica 15 bold").grid(row=7,column=0,pady=2)
tab1holidaylabel = Label(tab1frame2,text = "*For holiday and family\n days just put 'FAMILY DAY'\n in the qual entry",fg="white",bg="grey26",font="Helvetica 8 bold").grid(row=10,column=0,pady=2)
instructorLabel = Label(tab1frame2,text = "Instructor:",fg="white",bg="grey26",font="Helvetica 15 bold").grid(row=8,column=0,pady=2)


tab1Label2a = Label(tab1frame2,text = "'Ex.test'",fg=("grey50"),bg="grey26",font="Helvetica 15 bold").grid(row=2,column=2,pady=2)
tab1Label3a = Label(tab1frame2,text = "'Ex. 1-3'",fg=("grey50"),bg="grey26",font="Helvetica 15 bold").grid(row=3,column=2,pady=2)
tab1Label4a = Label(tab1frame2,text = "'Ex. 1-6'",fg=("grey50"),bg="grey26",font="Helvetica 15 bold").grid(row=4,column=2,pady=2)
tab1Label5a = Label(tab1frame2,text = "'01/01/21'",fg=("grey50"),bg="grey26",font="Helvetica 15 bold").grid(row=5,column=2,pady=2,padx=2)
tab1Label6a = Label(tab1frame2,text = "'01/28/21'",fg=("grey50"),bg="grey26",font="Helvetica 15 bold").grid(row=6,column=2,pady=2,padx=2)
tab1Label7a = Label(tab1frame2,text = "'200XX'",fg=("grey50"),bg="grey26",font="Helvetica 15 bold").grid(row=7,column=2,pady=2)
global my_progress2
my_progress2 = ttk.Progressbar(tab1frame3,orient=HORIZONTAL,length=200,mode="indeterminate")


global e1
global e2
global e3
global e4
global e5
global e6
global e7



e1 = Entry(tab1frame2,width=10)
e1.grid(row=2,column=1)
e2 = Entry(tab1frame2,width=10)
e2.grid(row=3,column=1)
e3 = Entry(tab1frame2,width=10)
e3.grid(row=4,column=1)
e4 = Entry(tab1frame2,width=10)
e4.grid(row=5,column=1)
e5 = Entry(tab1frame2,width=10)
e5.grid(row=6,column=1)
e6 = Entry(tab1frame2,width=10)
e6.grid(row=7,column=1)
e7 = Entry(tab1frame2,width=10)
e7.grid(row=8,column=1)


myButton4 = Button(tab1frame2,text="Submit",command=existing_WORKBOOK,bg="grey80")
myButton4.grid(row=11,column=0)
changeOnHover(myButton4, "aqua", "grey80")
# global submitTotal
submitTotal = 1




#TAB2====================================================ADD NEW WORKBOOK======================================================TAB2
tab2titleframe=Frame(tab2,bg="grey26")
tab2titleframe.pack()
tab2mainframe = Frame(tab2,bg="grey26")
tab2mainframe.pack()
myLabel0 = Label(tab2titleframe,text = "CREATE NEW WOOKBOOK",fg="white",bg="grey26",font="Helvetica 36 bold").grid(row=0,column=0)
myLabel0 = Label(tab2mainframe,text = "File Name:",fg="white",bg="grey26",font="Helvetica 15 bold").grid(row=0,column=0)
global newE1
global newe2
newe1 = Entry(tab2mainframe,width=20)
newe1.grid(row=2,column=0,pady=(0,10))
createButton = Button(tab2mainframe,text="Create New",command=new_WORKBOOK,bg="grey80")
createButton.grid(row=3,column=0)
changeOnHover(myButton4, "aqua", "grey80")



#TAB3============================READ FROM FILE========================================================TAB3#
tab3frame = Frame(tab3,bg="grey26")
tab3frame.grid(row=0,column=0)
tab3frame2 = Frame(tab3,bg="grey26")
tab3frame2.grid(row=0,column=1)
tab3frame3 = Frame(tab3,bg="grey26")
tab3frame3.grid(row=0,column=1)
tab3Label = Label(tab3frame,text = "READ IN ENTRIES",fg="white",bg="grey26",font="Helvetica 36 bold").pack()
tab3Labe2 = Label(tab3frame,text = "Read From File:",fg="white",bg="grey26",font="Helvetica 15 bold").pack()
global tab3e1
tab3e1 = Entry(tab3frame,width=20)
tab3e1.pack()
tab3Labe2 = Label(tab3frame,text = "Destination File:",fg="white",bg="grey26",font="Helvetica 15 bold").pack()
global tab3e2
tab3e2 = Entry(tab3frame,width=20)
tab3e2.pack()


global my_progress
my_progress = ttk.Progressbar(tab3frame3,orient=HORIZONTAL,length=300,mode="indeterminate")



tab3createButton = Button(tab3frame,text="Submit",command=readFromExcel,bg="grey80")
tab3createButton.pack(pady=(10))
changeOnHover(myButton4, "aqua", "grey80")



#TAB4================================================ABOUT====================================================================TAB3
mylabel2 = Label(tab4,text="About",font='Helvetica 30 bold')
about = """This is a simple scheduler program for automatically creating\nand editing tasks on defined days on an excel sheet when \ngiven prescribed dates. Hope you enjoy! """
mylabel = Label(tab4,text=about,font='Helvetica 12 bold')
mylabel2 = Label(tab4,text="How to Use",font='Helvetica 12 bold')
mylabel7 = Label(tab4,text="-Put the excel files you wish to edit in the same folder as this programs .exe file\n-When entering numbers all entries must be single digit\n-Do not separate the .exe file from the images",font='Helvetica 12 bold')
mylabel3 = Label(tab4,text="Use Existing:",font='Helvetica 12 bold')
mylabel4 = Label(tab4,text="This is for adding to an existing file. Simply, enter\n the file name(no extension) and enter\n the rest of your information accordingly.",font='Helvetica 12 bold')
mylabel5 = Label(tab4,text="Add WorkBook:",font='Helvetica 12 bold')
mylabel6 = Label(tab4,text="This is for creating a new excel file. Simply, enter what\n you would like to call the file(no extension,no special characters) and enter\n the rest of your information accordingly.You must initilize it with a Qual entry",font='Helvetica 12 bold')
mylabel.pack(pady=(0,30))
mylabel2.pack()
mylabel7.pack()
mylabel3.pack()
mylabel4.pack()
mylabel5.pack()
mylabel6.pack()



#TAB5=================================================ADD MIRS==================================================#TAB5
maintitle = Frame(tab5,bg="grey26")
maintitle.pack()
titleframe = Frame(tab5,bg="grey26")
titleframe.pack()
listframe = Frame(tab5,bg="grey26")
listframe.pack(pady=(20))
buttonframe = Frame(tab5,bg="grey26")
buttonframe.pack(padx=(0,100))

title = Label(maintitle,text = "ADD MIRS",fg="white",bg="grey26",font="Helvetica 36 bold")
title.pack()

filename = Label(titleframe,text="File name:")
filename.grid(row=0,column=0,padx=5)
fileentry = Entry(titleframe,width=15)
fileentry.grid(row=0,column=1,padx=5)

datelabel = Label(titleframe,text='Date:')
datelabel.grid(row=0,column=2,padx=5)
dateentry = Entry(titleframe,width=15)
dateentry.grid(row=0,column=3,padx=5)

allpersons = Label(listframe,text="AVAILABLE EMPLOYESS")
onleave = Label(listframe,text="ON LEAVE")
daypreview = Label(listframe,text="DAY PREVIEW")
allpersons.grid(row=0,column=0)
daypreview.grid(row=0,column=1)
onleave.grid(row=0,column=2)

finishedLabel = Label(buttonframe,text='Finished',bg="grey26",fg="grey26")
finishedLabel.grid(row=1,column=0)

def rightarrow():
    tab5dayindex = dateentry.get()[3:5]
    tab5dayindex = int(tab5dayindex)
    newtab5dayindex = tab5dayindex+1
    dateentry.delete(3,5)
    if newtab5dayindex <=9:
        dateentry.insert(3,"0"+str(newtab5dayindex))
    else:
        dateentry.insert(3,str(newtab5dayindex))  

def leftarrow():
    tab5dayindex = dateentry.get()[3:5]
    tab5dayindex = int(tab5dayindex)
    newtab5dayindex = tab5dayindex-1
    dateentry.delete(3,5)
    if newtab5dayindex <=9:
        dateentry.insert(3,"0"+str(newtab5dayindex))
    else:
        dateentry.insert(3,str(newtab5dayindex))    

def populateList():
    finishedLabel = Label(buttonframe,text='Finished',bg="grey26",fg="grey26")
    finishedLabel.grid(row=1,column=0)
    
    global my_list2
    global listbox1
    listbox1.delete(0,END)
    listbox3.delete(0,END)
    


    listbox2.delete(0, END)
    global tab5monthindex
    global tab5dayindex
    tab5monthindex = dateentry.get()[0:2]
    tab5dayindex = dateentry.get()[3:5]
    tab5monthindex = int(tab5monthindex)
    tab5dayindex = int(tab5dayindex)
    #MONTH INDEX
    print(tab5monthindex)
    print(tab5dayindex)
    
    readbook = load_workbook(str(fileentry.get())+'.xlsx')
    sheets = readbook.sheetnames
    sheet = readbook[sheets[int(tab5monthindex)]]
    print(sheet.title)
    
    tab5row_index = 5 
    cellref = sheet.cell(row=tab5row_index,column=tab5dayindex)
    # print(cellref.value)
    while cellref.value != "END":
        listbox2.insert(END,cellref.value)
        if not cellref.value:
            listbox2.insert(END,"")
        elif "Q" in str(cellref.value):
            listbox2.itemconfig("end", bg = "red")
        tab5row_index+=1
        cellref =sheet.cell(row=tab5row_index,column=tab5dayindex)
        # print(cellref.value)
    
    tab5leavecheck = False
    tab5onleavelist = []
    tab5leaverow = 5
    leavecellref = sheet.cell(row=tab5leaverow,column=tab5dayindex)
    while leavecellref.value !="LEAVE":
        tab5leaverow+=1
        leavecellref = sheet.cell(row=tab5leaverow,column=tab5dayindex)
    if leavecellref.value =="LEAVE":
        tab5leaverow+=1
        leavecellref = sheet.cell(row=tab5leaverow,column=tab5dayindex)
    while leavecellref.value:
        person = str(leavecellref.value)
        person = person.upper()
        print(person)
        tab5onleavelist.append(person)
        tab5leaverow+=1
        leavecellref = sheet.cell(row=tab5leaverow,column=tab5dayindex)
    
    for person in tab5onleavelist:
        listbox3.insert(END,person)
    
    
    for item in mainemployeeslist:
        if item not in tab5onleavelist:
            listbox1.insert(END, item)

populate = Button(titleframe,text="Populate",command=populateList)
populate.grid(row=0,column=6,padx=5)

rightArrow = Button(titleframe,text=">",command=rightarrow)
rightArrow.grid(row=0,column=5)
leftArrow = Button(titleframe,text="<",command=leftarrow)
leftArrow.grid(row=0,column=4)


def finalize():
    listboxindex = 0
    readbook = load_workbook(str(fileentry.get())+'.xlsx')
    sheets = readbook.sheetnames
    sheet = readbook[sheets[int(tab5monthindex)]]
    finalrow_index = 5 
    cellref = sheet.cell(row=finalrow_index,column=tab5dayindex)
    # item = listbox2.get(0)
    # print(item)
    while listboxindex < listbox2.size():
        item = listbox2.get(listboxindex)
        cellref.value = str(item)
        finalrow_index+=1
        listboxindex+=1
        cellref = sheet.cell(row=finalrow_index,column=tab5dayindex)
    readbook.save(str(fileentry.get())+'.xlsx')
    finishedLabel = Label(buttonframe,text='Finished',bg="grey26",fg="white")
    finishedLabel.grid(row=1,column=0)


finalizeButton = Button(buttonframe,text="Finalize",command=finalize)
finalizeButton.grid(row=0,column=0)

scrollbar = Scrollbar(listframe,orient=VERTICAL)
listbox1 = Listbox(listframe, height=10,exportselection=0,font=("12"))
listbox1.grid(row=1, column=0,padx=(0,50))
listbox2 = Listbox(listframe, height=10,width=30, exportselection=0,yscrollcommand=scrollbar.set,font=("12"))
listbox2.grid(row=1,column=1)
listbox3 = Listbox(listframe, height=10,exportselection=0,font=("12"))
listbox3.grid(row=1,column=2,padx=(50,0))


scrollbar.config(command=listbox2.yview)
scrollbar.grid(row=1,column=2,rowspan=1,  sticky=N+S+W)
# my_list2 = ['Robert','Law','Ryo','Mario','Laura','Skittle','-----']



def select(event=None):
    index = str(listbox2.curselection())
    print(len(index))
    if len(index) ==5:
        index=str(index)[1:3]
        print(index)
    else:
        index=str(index)[1]
        print(index)
    listbox2.delete(index)
    listbox2.insert(index, listbox1.get(ANCHOR))
    listbox2.itemconfig(index, bg = "lightgreen")
    if listbox1.get(ANCHOR) =="-----":
        None
    else:
        listbox1.delete(ANCHOR)

    # listbox1.delete(0)
    # listbox2.insert(1, listbox1.get(ANCHOR))
   
    # listbox1.delete(ANCHOR)

moveButton = Button(listframe,text="Move",command=select)
moveButton.grid(row=2,column=1,pady=10)



 
# def move(e):
#     mylabel.config(text="coord"+str(e.x))



# list = ["1","2","3","4"]

# for item in list:
#     listbox.insert(END, item)

# def select(event=None):
#     listbox2.insert(END, listbox.get(ANCHOR))
#     listbox.delete(ANCHOR)

# def deselect(event=None):
#     listbox.insert(END, listbox2.get(ANCHOR))
#     listbox2.delete(ANCHOR)

# tab5.bind('<Right>', select)
# tab5.bind('<Left>', deselect)
# tab5.bind('<Right>', select)
# tab5.bind('<Left>', deselect)

#TAB6=================================================SETTINGS==================================================#TAB6
def addemployee():
    existlabel = Label(tab6frame,text="Employee is Already in list",bg="grey26",fg="grey26")
    existlabel.grid(row=5,column=0)
    
    f= open('employees.txt','a')
    employeeentry = str(tab6entry.get())
    employeeentry = employeeentry.upper()
    if employeeentry == "":
        existlabel = Label(tab6frame,text="BLANK ENTRY",bg="grey26",fg="white")
        existlabel.grid(row=5,column=0)
        return
    
    tab6listbox1index = 0
    employeeexistsflag = False
    while tab6listbox1index < tab6listbox1.size():
        listemployee = str(tab6listbox1.get(tab6listbox1index))
        if employeeentry == listemployee:
            print("EMPLOYEE ALREADY EXISTS")
            employeeexistsflag = True
        tab6listbox1index+=1

    if employeeexistsflag == False:
        tab6listbox1.insert(END,employeeentry)
        tab6listbox1.itemconfig(END, bg = "lightgreen")
        existlabel = Label(tab6frame,text="ADDED!",bg="grey26",fg="white")
        existlabel.grid(row=5,column=0)
        f.write(employeeentry+"\n")
        
    else:
        existlabel = Label(tab6frame,text="Employee is Already in list",bg="grey26",fg="white")
        existlabel.grid(row=5,column=0)
    f.close()

    
        
    
def deleteemployee():
    f= open('employees.txt','r+')
    lines = sorted(f.readlines())
    f.close()
    requestedemployee = str(tab6listbox1.get(ANCHOR))
    print(requestedemployee)
    new_f = open('employees.txt','w')
    for line in lines:
        print(line)
        if line.strip("\n") != requestedemployee:
            new_f.write(line)
    new_f.close()
    tab6listbox1.delete(ANCHOR)


    
# f= open('employees.txt','w')
# lines = sorted(f.readlines())
# f.close()
# # new_f = open('employees.txt','w')
# # for line in lines:
# #     new_f.write(line)
# # new_f.close()

global mainemployeeslist
with open('employees.txt') as f:
    mainemployeeslist = [line.rstrip() for line in sorted(f)]
for item in mainemployeeslist:
    listbox1.insert(END, item)

tab6frame = Frame(tab6,bg="grey26")
tab6frame.grid(row=0,column=0)
allworkerslabel = Label(tab6frame,text="ALL EMPLOYEES")
allworkerslabel.grid(row=0,column=0)
tab6scrollbar = Scrollbar(tab6frame,orient=VERTICAL)
tab6listbox1 = Listbox(tab6frame, height=10,exportselection=0,font=("12"),yscrollcommand=tab6scrollbar.set)
tab6listbox1.grid(row=1, column=0)


global existlabel
existlabel = Label(tab6frame,text="Employee is Already in list",bg="grey26",fg="grey26")
existlabel.grid(row=5,column=0)
tab6scrollbar.config(command=tab6listbox1.yview)
tab6scrollbar.grid(row=1,column=1,rowspan=1,  sticky=N+S+W)
tab6addbutton = Button(tab6frame,text="ADD",command=addemployee)
tab6addbutton.grid(row=2,column=0)
tab6deletebutton = Button(tab6frame,text="DELETE",command=deleteemployee)
tab6deletebutton.grid(row=2,column=1)
tab6entry = Entry(tab6frame)
tab6entry.grid(row=3,column=0)


for person in mainemployeeslist:
    tab6listbox1.insert(END,person)



#TAB7=======================================ADD FULL QUAL=================================================TAB7

tab7frame = Frame(tab7,bg="grey26")
tab7frame.grid(row=0,column=0)
tab7frame2 = Frame(tab7,bg="grey26")
tab7frame2.grid(row=0,column=1)
tab7frame3 = Frame(tab7,bg="grey26")
tab7frame3.grid(row=0,column=1)
tab7Label = Label(tab7frame,text = "Read from File",fg="white",bg="grey26",font="Helvetica 36 bold").pack()
tab7Labe2 = Label(tab7frame,text = "Read From File:",fg="white",bg="grey26",font="Helvetica 15 bold").pack()
global tab7e1
tab7e1 = Entry(tab7frame,width=20)
tab7e1.pack()
tab7Labe2 = Label(tab7frame,text = "Destination File:",fg="white",bg="grey26",font="Helvetica 15 bold").pack()
global tab7e2
tab7e2 = Entry(tab7frame,width=20)
tab7e2.pack()


global tab7my_progress
tab7my_progress = ttk.Progressbar(tab7frame3,orient=HORIZONTAL,length=300,mode="indeterminate")



tab7createButton = Button(tab7frame,text="Submit",command=addFullQual,bg="grey80")
tab7createButton.pack(pady=(10))
changeOnHover(myButton4, "aqua", "grey80")

root.mainloop()